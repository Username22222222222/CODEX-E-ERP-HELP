from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


@dataclass
class RowData:
    values: list
    styles: list
    height: float | None
    outline_level: int
    hidden: bool
    collapsed: bool


def capture_row(ws, row: int) -> RowData:
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    dim = ws.row_dimensions[row]
    return RowData(
        values=[ws.cell(row, col).value for col in range(1, ws.max_column + 1)],
        styles=styles,
        height=dim.height,
        outline_level=dim.outlineLevel,
        hidden=False,
        collapsed=False,
    )


def write_row(ws, row: int, data: RowData) -> None:
    for col, value in enumerate(data.values, start=1):
        cell = ws.cell(row, col)
        cell.value = value
        style = data.styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
    dim = ws.row_dimensions[row]
    dim.height = data.height
    dim.outlineLevel = data.outline_level
    dim.hidden = False
    dim.collapsed = False


def domain_start_rows(ws, thema_col: int, content_col: int, start: int, end: int) -> dict[str, int]:
    result = {}
    for row in range(start, end):
        if ws.cell(row, content_col).value == "HelpPage" and ws.row_dimensions[row].outlineLevel == 1:
            result[str(ws.cell(row, thema_col).value)] = row
    return result


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-swap-lager-laenderpakete-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    ansichten = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")
    faq = next(row for row in range(ansichten + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")
    starts = domain_start_rows(ws, thema_col, content_col, ansichten + 1, faq)
    laender = starts["Länderpakete"]
    lager = starts["Lager"]

    if not laender < lager:
        print("already ordered")
        wb.close()
        return

    next_after_lager = min(row for row in starts.values() if row > lager)
    laender_block = [capture_row(ws, row) for row in range(laender, lager)]
    lager_block = [capture_row(ws, row) for row in range(lager, next_after_lager)]
    combined = lager_block + laender_block
    for offset, row_data in enumerate(combined):
        write_row(ws, laender + offset, row_data)

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"swapped rows {laender}-{next_after_lager - 1}: Lager before Länderpakete")


if __name__ == "__main__":
    main()
