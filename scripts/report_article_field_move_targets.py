from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"

SECTION_TARGETS = {
    "artikel-uebersicht": ("ArticleEdit-Übersicht",),
    "artikel-details": ("ArticleEdit-Details",),
    "artikel-verkauf": ("ArticleEdit-Verkauf",),
    "artikel-beschaffung": ("ArticleEdit-Beschaffung",),
    "artikel-text": ("ArticleEdit-Text",),
    "artikel-bild": ("ArticleEdit-Bild",),
    "artikel-lagerung": ("ArticleEdit-Lagerung",),
    "artikel-lagerungshistorie": ("ArticleEdit-Lagerhistorie", "ArticleQuantityHistoryTableList"),
    "artikel-set": ("ArticleEdit-Set", "ArticleSetList"),
    "artikel-makro": ("ArticleEdit-Makro", "ArticleMacroList"),
    "artikel-produktion": ("ArticleEdit-Produktion", "ArticleProductList"),
    "artikel-zubehoer": ("ArticleEdit-Zubehör", "ArticleAccessoryList"),
    "artikel-kategorien": ("ArticleEdit-Kategorien", "ArticleCategoryList"),
    "artikel-katalognummern": ("ArticleEdit-Katalognummern", "PartnerCatalogNumberList"),
    "artikel-produktionsschritt-ressourcen": ("ArticleEdit-Produktionsschritt-Ressourcen", "ArticleProductionStepResourceList"),
    "artikel-produktionsschritt-artikel": ("ArticleEdit-Produktionsschritt-Artikel", "ArticleProductionStepComponentList"),
    "artikel-produktionszeit": ("ArticleEdit-Produktionszeit",),
    "artikel-verwendung-bei-sets": ("ArticleEdit-Verwendung in Sets", "ArticleUsageSetList"),
    "artikel-verwendung-bei-produktion": ("ArticleEdit-Verwendung in Produktion", "ArticleUsageProductionList"),
    "artikel-positionsliste": ("ArticleEdit-Positionsliste", "ArticleDocPositionList"),
    "artikel-umsatz": ("ArticleEdit-Umsatz", "ArticleRevenueList"),
    "artikel-vertraege": ("ArticleEdit-Verträge", "ArticleDocPositionList"),
    "artikel-lokal": ("ArticleEdit-Lokal",),
    "artikel-aenderungsprotokoll": ("ArticleEdit-Änderungsprotokoll", "ArticleHistoryList"),
    "artikel-anhaenge": ("ArticleEdit-Anhänge", "AttachmentList"),
}

ALIASES = {
    ("artikel-lagerungshistorie", "bestandszaehlung"): "inventur",
    ("artikel-lagerungshistorie", "dok"): "beleg",
    ("artikel-positionsliste", "dok-datum"): "belegdatum",
    ("artikel-positionsliste", "dok-nummer"): "belegnummer",
    ("artikel-positionsliste", "menge-offen"): "offene-menge",
    ("artikel-positionsliste", "bestandsstatus"): "lagerstatus",
    ("artikel-positionsliste", "lagerungsmenge"): "lagermenge",
    ("artikel-positionsliste", "gewinn-betrag"): "rohertrag-betrag",
    ("artikel-positionsliste", "gewinn-prozent"): "rohertrag",
    ("artikel-umsatz", "einkaufsumsatz"): "einkaufsbetrag",
    ("artikel-umsatz", "gewinn-betrag"): "rohertrag-betrag",
    ("artikel-umsatz", "gewinn-prozent"): "rohertrag",
    ("artikel-vertraege", "dok-datum"): "belegdatum",
    ("artikel-vertraege", "dok-nummer"): "belegnummer",
    ("artikel-vertraege", "menge-offen"): "offene-menge",
    ("artikel-verwendung-bei-sets", "set-matchcode"): "set-matchcode",
    ("artikel-verwendung-bei-sets", "name1"): "name-1",
    ("artikel-verwendung-bei-sets", "name2"): "name-2",
    ("artikel-verwendung-bei-produktion", "produktion-matchcode"): "produktions-matchcode",
    ("artikel-verwendung-bei-produktion", "name1"): "name-1",
    ("artikel-verwendung-bei-produktion", "name2"): "name-2",
    ("artikel-anhaenge", "erstellungsdatum"): "erstellt-am",
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def slug(value: str) -> str:
    value = value.replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue")
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    value = value.replace("%", "prozent")
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def field_slug_for(section_slug: str, field_slug: str) -> str:
    return ALIASES.get((section_slug, field_slug), field_slug)


def main() -> None:
    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    topic_col = headers["Thema"]
    url_col = headers["URL_PATH"]
    breadcrumb_col = headers["BREADCRUMB"]

    targets: dict[tuple[str, str], int] = {}
    for row in range(1, ws.max_row + 1):
        breadcrumb = cell_text(ws.cell(row, breadcrumb_col).value)
        if not breadcrumb.startswith("Ansichten > Artikel > ArticleEdit > "):
            continue
        parts = tuple(part.strip() for part in breadcrumb.split(">"))
        parent_tail = parts[3:-1]
        topic_slug = slug(cell_text(ws.cell(row, topic_col).value))
        for section_slug, target_parts in SECTION_TARGETS.items():
            if parent_tail == target_parts:
                targets[(section_slug, topic_slug)] = row

    matched: list[tuple[int, str, int]] = []
    unmatched: list[tuple[int, str, str]] = []
    for row in range(1, ws.max_row + 1):
        path = cell_text(ws.cell(row, url_col).value)
        if not path.startswith("Stammdaten/artikel/") or not path.endswith("/index.html"):
            continue
        level = int(ws.row_dimensions[row].outlineLevel or 0)
        if level < 3:
            continue
        parts = path.split("/")
        if len(parts) < 5:
            continue
        section_slug = parts[2]
        field_slug = parts[3]
        if section_slug not in SECTION_TARGETS:
            unmatched.append((row, path, "no-section-target"))
            continue
        target_key = (section_slug, field_slug_for(section_slug, field_slug))
        target_row = targets.get(target_key)
        if target_row:
            matched.append((row, path, target_row))
        else:
            unmatched.append((row, path, f"no-field-target:{target_key[1]}"))

    print(f"matched={len(matched)}")
    print(f"unmatched={len(unmatched)}")
    print("\nMATCHED")
    for source_row, path, target_row in matched:
        print(f"{source_row} -> {target_row} | {path}")
    print("\nUNMATCHED")
    for source_row, path, reason in unmatched:
        print(f"{source_row} | {reason} | {path}")
    wb.close()


if __name__ == "__main__":
    main()
