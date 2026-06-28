from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-outline-repair-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    breadcrumb_col = headers["BREADCRUMB"]
    topic_col = headers["Thema"]
    content_type_col = headers.get("CONTENT_TYPE")
    changed = 0

    for row in range(2, ws.max_row + 1):
        topic = cell_text(ws.cell(row, topic_col).value)
        breadcrumb = cell_text(ws.cell(row, breadcrumb_col).value)
        if not topic or not breadcrumb:
            continue
        level = max(0, len([part for part in breadcrumb.split(">") if part.strip()]) - 1)
        if int(ws.row_dimensions[row].outlineLevel or 0) != level:
            changed += 1
        ws.row_dimensions[row].outlineLevel = level
        content_type = cell_text(ws.cell(row, content_type_col).value) if content_type_col else ""
        ws.row_dimensions[row].hidden = level >= 2 and content_type != "HelpPageRoot"
        ws.row_dimensions[row].collapsed = False

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"outline_rows_repaired={changed}")


if __name__ == "__main__":
    main()
