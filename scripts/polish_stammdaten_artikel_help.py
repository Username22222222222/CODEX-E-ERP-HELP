from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
SOURCE_SCREENSHOT_DIR = PROJECT_ROOT / "reference/screenshots/stammdaten/artikel"
SHEET = "de-DE"

ARTICLE_PAGE = {
    "Thema": "Artikelstammdaten",
    "Beschreibung": """# Artikelstammdaten

Artikelstammdaten sind der zentrale Ort für alle Artikel, Leistungen, Handelswaren, Materialien, Sets und Produktionsartikel in X-ERP. Ein sauber gepflegter Artikelstamm sorgt dafür, dass Verkauf, Einkauf, Lager, Produktion, Webshop, Belege und Auswertungen mit denselben Informationen arbeiten.

![Artikelstammdaten in X-ERP: Übersicht der zentralen Artikeldaten](artikelstammdaten-uebersicht.webp)

## Wofür Artikelstammdaten genutzt werden

- Verkauf: Artikel werden in Angeboten, Aufträgen, Lieferscheinen, Rechnungen und Preislisten verwendet.
- Einkauf: Beschaffungspreise, Lieferanten, Katalognummern und Einkaufsregeln greifen auf denselben Artikel zu.
- Lager: Lagerführung, Mindestbestand, Maximalbestand, Standardlager und Lagerhistorie werden artikelbezogen gesteuert.
- Produktion: Produktionsartikel verbinden Stücklisten, Produktionsschritte, Ressourcen und Fertigungszeiten.
- Webshop und Kataloge: Bilder, Texte, Kategorien, GTIN und Shop-Kennzeichen bestimmen die externe Darstellung.
- Auswertungen: Umsatz, Bestand, Verwendung, Historie und Verträge werden über den Artikel zusammengeführt.

## Artikelnummer und Identifikation

Die Artikelnummer ist die wichtigste technische und organisatorische Kennung. Sie sollte eindeutig, dauerhaft und möglichst regelbasiert vergeben werden. Eine gute Nummernlogik erleichtert Suche, Import, Barcode-Nutzung, Katalogpflege und die Abstimmung zwischen Einkauf, Verkauf und Lager.

:::box Empfehlung zur Artikelnummernvergabe
Legen Sie früh fest, ob Artikelnummern rein fortlaufend, sprechend oder gruppenbezogen vergeben werden. Vermeiden Sie Bedeutungen, die sich später ändern können, zum Beispiel Lieferant, Lagerort oder Preisgruppe. Solche Informationen gehören in eigene Felder oder Kategorien.
:::

## Artikelarten und Besonderheiten

:::box Standardartikel
Standardartikel sind normale Verkaufs-, Einkaufs- oder Lagerartikel. Sie haben Bezeichnungen, Mengeneinheiten, Preise, Lagerinformationen und optional Bilder, Texte oder Kategorien.
:::

:::box Set-Artikel
Sets bündeln mehrere Artikel zu einem verkaufbaren Paket. Sie eignen sich für Zubehörpakete, Aktionsartikel, Bundles oder konfigurierte Zusammenstellungen. Entscheidend ist, ob das Set auf Belegen als ein Artikel erscheinen soll oder ob die Komponenten aufgelöst und einzeln gedruckt werden.
:::

:::box Produktionsartikel
Produktionsartikel beschreiben Artikel, die gefertigt werden. Sie verbinden die Stammdaten mit Komponenten, Produktionsschritten, Ressourcen und Zeiten. Dadurch wird aus dem Artikelstamm die Grundlage für Stückliste, Fertigungsauftrag und Nachkalkulation.
:::

:::box Makro-Artikel
Makro-Artikel automatisieren wiederkehrende Abläufe. Sie können Positionen, Mengen oder Beleglogik vorbereiten und eignen sich, wenn ein Artikel nicht nur ein Produkt, sondern auch eine Regel oder Aktion im Prozess auslösen soll.
:::

## Details gezielt pflegen

Die Detailseiten unterhalb dieser Seite folgen den Registerkarten und Feldern der Artikelmaske. Starten Sie mit den Grunddaten und ergänzen Sie anschließend nur die Bereiche, die für den jeweiligen Artikeltyp relevant sind: Verkauf, Beschaffung, Lagerung, Texte, Bilder, Sets, Produktion, Zubehör, Kategorien, Katalognummern und Historie.

## Kurzfassung für Schulung, Marketing und Video

Deutsch: Artikelstammdaten verbinden alle produktbezogenen Informationen in einem zentralen Datensatz. Ein Artikel kann Verkaufsartikel, Einkaufsartikel, Lagerartikel, Set, Produktionsartikel oder Makro sein und steuert dadurch viele Prozesse automatisch.

English: Article master data in X-ERP connects sales, purchasing, warehouse, production and e-commerce in one shared product record. One article can represent a sellable item, a purchased item, stock, a set, a manufactured product or an automation-driven macro item.

## So nutzen Sie diese Hilfe

Diese Startseite erklärt die wichtigsten Konzepte und Besonderheiten der Artikelstammdaten. Die darunterliegenden Detailseiten erklären die einzelnen Registerkarten und Felder, damit Anwender gezielt von der fachlichen Übersicht in die konkrete Pflege wechseln können.""",
    "TITLE": "Artikelstammdaten | X-ERP ERP Hilfe",
    "META_DESCRIPTION": "Artikelstammdaten in X-ERP: zentrale Pflege von Artikelnummern, Verkauf, Einkauf, Lager, Sets, Produktionsartikeln, Makros, Bildern, Texten und Kategorien.",
    "H1": "Artikelstammdaten",
    "PRIMARY_KEYWORD": "Artikelstammdaten ERP",
    "NAV_TITLE": "Artikelstammdaten",
    "BREADCRUMB": "Stammdaten > Artikelstammdaten",
    "IMAGE_ALT": "Artikelstammdaten in X-ERP mit Übersicht, Verkauf, Lagerung, Set, Produktion und Makro",
    "IMAGE_CAPTION": "Artikelstammdaten bündeln die produktbezogenen Informationen für Verkauf, Einkauf, Lager und Produktion.",
    "IMAGE_STATUS": "Screenshot eingebunden",
}

SCREENSHOTS = {
    "artikelstammdaten-uebersicht.webp": SOURCE_SCREENSHOT_DIR / "artikelstammdaten-uebersicht.webp",
}


def headers_for(ws) -> dict[str, int]:
    return {
        str(ws.cell(1, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value not in (None, "")
    }


def find_article_row(ws) -> int:
    headers = headers_for(ws)
    url_path_col = headers.get("URL_PATH")
    if not url_path_col:
        raise RuntimeError("Missing workbook header: URL_PATH")
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if value in {"Artikel", "Artikelstammdaten"}:
            path = ws.cell(row, url_path_col).value
            if path == "Stammdaten/artikel/index.html":
                return row
    raise RuntimeError("Artikelstammdaten row not found")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-stammdaten-artikel-polish-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False)
    ws = wb[SHEET]
    headers = headers_for(ws)
    row = find_article_row(ws)

    for name, value in ARTICLE_PAGE.items():
        col = headers.get(name)
        if not col:
            raise RuntimeError(f"Missing workbook header: {name}")
        ws.cell(row, col).value = value

    wb.save(WORKBOOK)
    wb.close()

    target_dir = HELP_ROOT / "Stammdaten/artikel"
    target_dir.mkdir(parents=True, exist_ok=True)
    copied = 0
    for target_name, source in SCREENSHOTS.items():
        if source.is_file():
            shutil.copy2(source, target_dir / target_name)
            copied += 1
        else:
            raise FileNotFoundError(source)

    print(f"backup={backup}")
    print(f"updated_row={row}")
    print(f"screenshots_copied={copied}")


if __name__ == "__main__":
    main()
