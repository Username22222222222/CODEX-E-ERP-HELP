from datetime import datetime
from pathlib import Path
import shutil

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-repair-child-outline-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    ansichten_row = next(
        row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten"
    )
    faq_row = next(
        row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ"
    )

    fixed = []
    for row in range(ansichten_row + 1, faq_row):
        thema = ws.cell(row, thema_col).value
        content_type = ws.cell(row, content_col).value
        level = ws.row_dimensions[row].outlineLevel
        if thema and content_type not in {"HelpPage", "View", "Wizard"} and level < 3:
            ws.row_dimensions[row].outlineLevel = 3
            ws.row_dimensions[row].hidden = False
            ws.row_dimensions[row].collapsed = False
            fixed.append((row, level, thema))

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"fixed={len(fixed)}")
    for row, old_level, thema in fixed:
        print(f"{row}: {thema} level {old_level} -> 3")


if __name__ == "__main__":
    main()
