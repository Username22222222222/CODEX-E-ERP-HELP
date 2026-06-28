from __future__ import annotations

import json
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
JSON_OUTPUT = HELP_ROOT / "xlsx-tree.json"
JS_OUTPUT = HELP_ROOT / "xlsx-tree.js"


def cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def first_markdown_heading(value: str) -> str:
    for line in (value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        line = line.strip()
        if line.startswith("#"):
            return line.lstrip("#").strip()
    return ""


def public_path(url_path: str) -> str:
    if not url_path:
        return ""
    url_path = url_path.strip().replace("\\", "/")
    if url_path.startswith("http://") or url_path.startswith("https://"):
        return url_path
    if url_path.startswith("/de/help/"):
        path = url_path
    else:
        path = "/de/help/" + url_path.lstrip("/")
    if path.endswith("index.html"):
        path = path[: -len("index.html")]
    elif path.endswith(".html"):
        # Non-index legacy pages can still be linked directly.
        pass
    elif not path.endswith("/"):
        path += "/"
    return quote(path, safe="/:-_.~%")


def make_node(row: int, level: int, values: dict[str, object]) -> dict[str, object]:
    description = cell_text(values.get("Beschreibung")) or cell_text(values.get("Original Text"))
    path_value = cell_text(values.get("URL_PATH"))
    title = (
        first_markdown_heading(description) if "/faq/" in public_path(path_value).casefold() and not path_value.endswith("index.html") else ""
    ) or (
        cell_text(values.get("NAV_TITLE"))
        or cell_text(values.get("Thema"))
        or f"Zeile {row}"
    )
    node: dict[str, object] = {
        "title": title,
        "pageTitle": description,
        "sourceRow": row,
        "sourceLevel": level,
        "children": [],
    }
    content_type = cell_text(values.get("CONTENT_TYPE"))
    if content_type:
        node["contentType"] = content_type
    path = public_path(path_value)
    if path:
        node["path"] = path
    topic = cell_text(values.get("Thema"))
    if topic and topic != title:
        node["sourceTitle"] = topic
    breadcrumb = cell_text(values.get("BREADCRUMB"))
    if breadcrumb:
        node["breadcrumb"] = breadcrumb
    return node


def prune_empty_children(node: dict[str, object]) -> None:
    children = node.get("children")
    if isinstance(children, list):
        for child in children:
            prune_empty_children(child)
        if not children:
            node.pop("children", None)


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }

    required = {"Thema"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"Missing required headers: {', '.join(missing)}")

    roots: list[dict[str, object]] = []
    stack: dict[int, dict[str, object]] = {}

    for row in range(2, ws.max_row + 1):
        topic = cell_text(ws.cell(row, headers["Thema"]).value)
        if not topic:
            continue

        level = int(ws.row_dimensions[row].outlineLevel or 0)
        values = {name: ws.cell(row, col).value for name, col in headers.items()}
        content_type = cell_text(values.get("CONTENT_TYPE"))
        if content_type == "InlineSection":
            for key in list(stack):
                if key >= level:
                    stack.pop(key, None)
            continue
        node = make_node(row, level, values)

        parent = stack.get(level - 1)
        if parent is None or level == 0:
            roots.append(node)
        else:
            parent.setdefault("children", []).append(node)

        for old_level in list(stack):
            if old_level >= level:
                stack.pop(old_level, None)
        stack[level] = node

    for root in roots:
        prune_empty_children(root)

    payload_pretty = json.dumps(roots, ensure_ascii=False, indent=2)
    JSON_OUTPUT.write_text(payload_pretty + "\n", encoding="utf-8-sig")
    JS_OUTPUT.write_text(
        "window.XERP_XLSX_TREE = "
        + json.dumps(roots, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    wb.close()

    print(f"roots={len(roots)}")
    print("top-level=" + ", ".join(str(root.get("title")) for root in roots))
    print(JSON_OUTPUT)
    print(JS_OUTPUT)


if __name__ == "__main__":
    main()
