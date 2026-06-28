from __future__ import annotations

import csv
import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
INVENTORY = ROOT / "outputs" / "program-audit" / "route-inventory.csv"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


MODULE_LABELS = {
    "Admin": "Administration",
    "AI": "KI-Assistent",
    "Article": "Artikel",
    "Attachment": "Anhänge",
    "Bulletin": "Bulletin",
    "Calendar": "Kalender",
    "CountryPackage": "Länderpakete",
    "Crm": "CRM",
    "Dashboard": "Dashboard",
    "Doc": "Belege",
    "Email": "E-Mail",
    "Employee": "Mitarbeiter",
    "ExtraModules": "Extra-Tabellen und Extra-Felder",
    "Finance": "Finanzwesen",
    "FormattedSearch": "Formatierte Suche",
    "Helpdesk": "Helpdesk",
    "Intercom": "Intercom",
    "Marketplace": "Marketplace",
    "OutputControl": "Ausgabesteuerung",
    "PackingList": "Packen",
    "Partner": "Partner",
    "Phone": "Telefon",
    "Procurement": "Beschaffungsassistent",
    "Production": "Produktion",
    "Project": "Projektmanagement",
    "Report": "Reportdesigner",
    "Resource": "Ressourcen",
    "Shipping": "Versand",
    "StickyNotes": "Elektronische Haftnotizen",
    "SupportSetting": "Support-Einstellungen",
    "SystemViews": "Systemansichten",
    "TextBlock": "Textblöcke",
    "TimeTracking": "Zeiterfassung",
    "Translation": "Übersetzungen",
    "UserLoginSecurity": "Benutzer und Sicherheit",
    "Warehouse": "Lager",
    "Webshop": "Webshop",
    "Wiki": "Wiki",
    "Workspace": "Desktopkonfigurator",
    "Sonstige": "Weitere Ansichten",
}

MODULE_ORDER = [
    "Admin",
    "Article",
    "Partner",
    "Employee",
    "Doc",
    "Procurement",
    "Shipping",
    "PackingList",
    "Warehouse",
    "Production",
    "Finance",
    "Crm",
    "Helpdesk",
    "Calendar",
    "Email",
    "Intercom",
    "Wiki",
    "Bulletin",
    "TimeTracking",
    "Project",
    "Resource",
    "TextBlock",
    "Attachment",
    "OutputControl",
    "ExtraModules",
    "FormattedSearch",
    "Dashboard",
    "Workspace",
    "UserLoginSecurity",
    "Report",
    "CountryPackage",
    "AI",
    "Webshop",
    "Marketplace",
    "Phone",
    "StickyNotes",
    "SupportSetting",
    "SystemViews",
    "Translation",
    "Sonstige",
]

PREFIX_MODULES = [
    ("Admin", "Admin"),
    ("Ai", "AI"),
    ("Article", "Article"),
    ("Attachment", "Attachment"),
    ("Bulletin", "Bulletin"),
    ("Calendar", "Calendar"),
    ("CountryPackage", "CountryPackage"),
    ("Crm", "Crm"),
    ("Dashboard", "Dashboard"),
    ("Doc", "Doc"),
    ("DeviceServiceDoc", "Doc"),
    ("Email", "Email"),
    ("Employee", "Employee"),
    ("Extra", "ExtraModules"),
    ("Finance", "Finance"),
    ("FinancialReports", "Finance"),
    ("FormattedSearch", "FormattedSearch"),
    ("Helpdesk", "Helpdesk"),
    ("Intercom", "Intercom"),
    ("Marketplace", "Marketplace"),
    ("OutputControl", "OutputControl"),
    ("PackingList", "PackingList"),
    ("Partner", "Partner"),
    ("Phone", "Phone"),
    ("Procurement", "Procurement"),
    ("Production", "Production"),
    ("Project", "Project"),
    ("Report", "Report"),
    ("Resource", "Resource"),
    ("Shipping", "Shipping"),
    ("StickyNote", "StickyNotes"),
    ("Support", "SupportSetting"),
    ("System", "SystemViews"),
    ("Authorization", "UserLoginSecurity"),
    ("Role", "UserLoginSecurity"),
    ("User", "UserLoginSecurity"),
    ("TextBlock", "TextBlock"),
    ("TimeTracking", "TimeTracking"),
    ("Translation", "Translation"),
    ("Warehouse", "Warehouse"),
    ("Webshop", "Webshop"),
    ("Wiki", "Wiki"),
    ("Workspace", "Workspace"),
]


@dataclass
class RowData:
    values: list
    styles: list
    height: float | None
    outline_level: int
    hidden: bool
    collapsed: bool


@dataclass
class Block:
    module: str
    source_row: int
    rows: list[RowData]
    name: str


def slugify(text: str) -> str:
    table = str.maketrans(
        {
            "ä": "ae",
            "ö": "oe",
            "ü": "ue",
            "Ä": "ae",
            "Ö": "oe",
            "Ü": "ue",
            "ß": "ss",
            "&": "und",
        }
    )
    text = text.translate(table).lower()
    chars = []
    previous_dash = False
    for char in text:
        if char.isalnum():
            chars.append(char)
            previous_dash = False
        elif not previous_dash:
            chars.append("-")
            previous_dash = True
    return "".join(chars).strip("-")


def read_inventory_modules() -> dict[str, str]:
    modules: dict[str, str] = {}
    with INVENTORY.open(encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        for row in csv.DictReader(handle, dialect=dialect):
            if row.get("Kind") in {"Edit", "Wizard"} and row.get("Name"):
                modules.setdefault(row["Name"], row.get("Module") or "Sonstige")
    return modules


def module_for_name(name: str, inventory_modules: dict[str, str]) -> str:
    if name in inventory_modules:
        return inventory_modules[name]
    base = name.split("_", 1)[0]
    if base in inventory_modules:
        return inventory_modules[base]
    for prefix, module in PREFIX_MODULES:
        if name.startswith(prefix):
            return module
    return "Sonstige"


def capture_row(ws, row: int) -> RowData:
    values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    dim = ws.row_dimensions[row]
    return RowData(values, styles, dim.height, dim.outlineLevel, dim.hidden, dim.collapsed)


def write_row(ws, row: int, data: RowData) -> None:
    for col, value in enumerate(data.values, start=1):
        cell = ws.cell(row, col)
        cell.value = value
        style = data.styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
    dim = ws.row_dimensions[row]
    dim.height = data.height
    dim.outlineLevel = data.outline_level
    dim.hidden = data.hidden
    dim.collapsed = data.collapsed


def make_module_row(template: RowData, headers: dict[str, int], module: str, order: int) -> RowData:
    row = RowData(
        values=list(template.values),
        styles=[copy(style) for style in template.styles],
        height=template.height,
        outline_level=1,
        hidden=False,
        collapsed=False,
    )
    label = MODULE_LABELS.get(module, module)
    slug = slugify(label)
    values = row.values
    for index in range(len(values)):
        values[index] = None
    def set_value(header: str, value):
        if header in headers:
            values[headers[header] - 1] = value
    set_value("Thema", label)
    set_value("Original Text", module)
    set_value("Beschreibung", f"Dieses Unterverzeichnis bündelt die EditViews und Wizards im Bereich {label}.")
    set_value("PAGE_ID", f"views-{slug}")
    set_value("SLUG", f"views/{slug}")
    set_value("TITLE", f"{label} Ansichten in X-ERP")
    set_value("META_DESCRIPTION", f"Übersicht der X-ERP EditViews und Wizards im Bereich {label}.")
    set_value("H1", f"{label} Ansichten")
    set_value("PRIMARY_KEYWORD", "ERP")
    set_value("CONTENT_TYPE", "HelpPage")
    set_value("STRUCTURED_DATA_TYPE", "TechArticle")
    set_value("CANONICAL_URL", f"/de/help/views/{slug}/")
    set_value("HREFLANG_GROUP", f"views-{slug}")
    set_value("SEO_STATUS", "draft")
    set_value("LAST_REVIEWED", "2026-06-27")
    set_value("DIRECTORY_PATH", f"views/{slug}")
    set_value("FILE_NAME", "index.html")
    set_value("URL_PATH", f"views/{slug}/index.html")
    set_value("STORAGE_PATH", f"views/{slug}")
    set_value("NAV_TITLE", label)
    set_value("BREADCRUMB", f"Ansichten > {label}")
    set_value("TOC_PARENT", "Ansichten")
    set_value("TOC_LEVEL", 1)
    set_value("TOC_ORDER", order)
    set_value("UNIQUE_PAGE_KEY", f"views:{slug}")
    return row


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-rebuild-views-outline-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    inventory_modules = read_inventory_modules()
    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = {cell.value: index + 1 for index, cell in enumerate(ws[1]) if cell.value}
    thema_col = headers["Thema"]
    content_col = headers["CONTENT_TYPE"]

    ansichten_row = next(
        row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten"
    )
    faq_row = next(row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")

    module_template = capture_row(ws, ansichten_row + 1)

    blocks: list[Block] = []
    row = ansichten_row + 1
    while row < faq_row:
        content_type = ws.cell(row, content_col).value
        name = ws.cell(row, thema_col).value
        if content_type in {"View", "Wizard"} and name:
            start = row
            row += 1
            while row < faq_row and ws.cell(row, content_col).value not in {"View", "Wizard"}:
                row += 1
            captured = []
            for source_row in range(start, row):
                value = ws.cell(source_row, thema_col).value
                if isinstance(value, str) and value.startswith("Nachtrag -"):
                    continue
                captured.append(capture_row(ws, source_row))
            blocks.append(
                Block(
                    module=module_for_name(str(name), inventory_modules),
                    source_row=start,
                    rows=captured,
                    name=str(name),
                )
            )
        else:
            row += 1

    module_index = {module: index for index, module in enumerate(MODULE_ORDER)}
    blocks.sort(key=lambda block: (module_index.get(block.module, 999), block.source_row))

    rebuilt: list[RowData] = []
    current_module = None
    module_order = 0
    for block in blocks:
        if block.module != current_module:
            module_order += 1
            rebuilt.append(make_module_row(module_template, headers, block.module, module_order))
            current_module = block.module

        for index, row_data in enumerate(block.rows):
            row_data.hidden = False
            row_data.collapsed = False
            if index == 0:
                row_data.outline_level = 2
                module_label = MODULE_LABELS.get(block.module, block.module)
                module_slug = slugify(module_label)
                file_name = row_data.values[headers["FILE_NAME"] - 1] or f"{slugify(block.name)}.html"
                row_data.values[headers["CONTENT_TYPE"] - 1] = "Wizard" if row_data.values[headers["CONTENT_TYPE"] - 1] == "Wizard" else "View"
                row_data.values[headers["DIRECTORY_PATH"] - 1] = f"views/{module_slug}"
                row_data.values[headers["URL_PATH"] - 1] = f"views/{module_slug}/{file_name}"
                row_data.values[headers["STORAGE_PATH"] - 1] = f"views/{module_slug}/{file_name}"
                row_data.values[headers["BREADCRUMB"] - 1] = f"Ansichten > {module_label} > {block.name}"
                row_data.values[headers["TOC_PARENT"] - 1] = module_label
                row_data.values[headers["TOC_LEVEL"] - 1] = 2
            else:
                if row_data.outline_level < 3:
                    row_data.outline_level = 3
            rebuilt.append(row_data)

    delete_count = faq_row - (ansichten_row + 1)
    ws.delete_rows(ansichten_row + 1, delete_count)
    ws.insert_rows(ansichten_row + 1, len(rebuilt))

    for offset, row_data in enumerate(rebuilt):
        write_row(ws, ansichten_row + 1 + offset, row_data)

    # Repair main trailing sections after the inserted view hierarchy.
    for row in range(1, ws.max_row + 1):
        dim = ws.row_dimensions[row]
        dim.hidden = False
        dim.collapsed = False

    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, thema_col).value
        if value in {"Ansichten", "FAQ", "Glossar"}:
            ws.row_dimensions[row].outlineLevel = 0
        elif value in {"Allgemeine Programmbedienung", "Datenübernahme / Migration", "Störung / Fehlermeldung", "Beleg", "Dokument (Dok)", "Geschäftspartner", "Partner"}:
            if row > ansichten_row:
                ws.row_dimensions[row].outlineLevel = 1

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    ws.freeze_panes = "A2"

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"blocks={len(blocks)} modules={len(set(block.module for block in blocks))} rebuilt_rows={len(rebuilt)}")


if __name__ == "__main__":
    main()
