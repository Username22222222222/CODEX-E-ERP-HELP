from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
ANSICHTEN_ROOT = HELP_ROOT / "Ansichten"


SOURCE_BY_PATH = {
    "Customizing/dynamische-doc-kette/index.html": "belege/doc-edit/DocEdit.webp",
    "Customizing/dynamische-doc-kette/belegart/index.html": "belege/doc-type-edit/DocTypeEdit.webp",
    "Customizing/dynamische-doc-kette/ueberfuehrungsregel/index.html": "belege/doc-type-edit/doc-type-edit-uebersicht/DocTypeEdit.webp",
    "Customizing/dynamische-doc-kette/felduebernahme/index.html": "belege/doc-xml-assignment-edit/DocXmlAssignmentEdit.webp",
    "Customizing/ausgabesteuerung/index.html": "ausgabesteuerung/output-control-format-edit/OutputControlFormatEdit.webp",
    "Customizing/ausgabesteuerung/ausgabekanal/index.html": "ausgabesteuerung/output-control-printer-configuration-edit/OutputControlPrinterConfigurationEdit.webp",
    "Customizing/ausgabesteuerung/ausgaberegel/index.html": "ausgabesteuerung/output-control-format-edit/OutputControlFormatEdit.webp",
    "Customizing/ausgabesteuerung/layoutzuordnung/index.html": "ausgabesteuerung/output-control-format-edit/output-control-format-edit-uebersicht/OutputControlFormatEdit.webp",
    "Customizing/desktopkonfigurator/index.html": "desktopkonfigurator/workspace-role-desktop-edit/WorkspaceRoleDesktopEdit.webp",
    "Customizing/desktopkonfigurator/menue-und-navigation/index.html": "desktopkonfigurator/workspace-role-menu-edit/WorkspaceRoleMenuEdit.webp",
    "Customizing/desktopkonfigurator/schaltflaechen/index.html": "desktopkonfigurator/workspace-role-menu-button-edit/WorkspaceRoleMenuButtonEdit.webp",
    "Customizing/desktopkonfigurator/rollenlayout/index.html": "desktopkonfigurator/workspace-role-desktop-edit/WorkspaceRoleDesktopEdit.webp",
    "Customizing/reportdesigner/index.html": "reportdesigner/reportdesigner/Reportdesigner.webp",
    "Customizing/reportdesigner/belegvorlage/index.html": "reportdesigner/report-template-edit/ReportTemplateEdit.webp",
    "Customizing/reportdesigner/corporate-design/index.html": "reportdesigner/reportdesigner/Reportdesigner.webp",
    "Customizing/reportdesigner/listendruck/index.html": "reportdesigner/report-template-edit/ReportTemplateEdit.webp",
    "Customizing/dashboarddesigner/index.html": "dashboard/dashboarddesigner/DashboardDesigner.webp",
    "Customizing/dashboarddesigner/kennzahl/index.html": "dashboard/dashboard-edit/DashboardEdit.webp",
    "Customizing/dashboarddesigner/diagramm/index.html": "dashboard/dashboard-edit/dashboard-edit-uebersicht/DashboardEdit.webp",
    "Customizing/dashboarddesigner/dashboardlayout/index.html": "dashboard/dashboarddesigner/DashboardDesigner.webp",
    "Customizing/extra-tabellen/index.html": "extra-tabellen/extra-table-edit/ExtraTableEdit.webp",
    "Customizing/extra-tabellen/tabellendefinition/index.html": "extra-tabellen/extra-table-edit/extra-table-edit-uebersicht/ExtraTableEdit.webp",
    "Customizing/extra-tabellen/datenpflege/index.html": "extra-tabellen/xxtable-edit/XXTableEdit.webp",
    "Customizing/extra-tabellen/verknuepfung/index.html": "extra-tabellen/extra-api-edit/ExtraApiEdit.webp",
    "Customizing/extra-felder/index.html": "extra-felder/extra-field-edit/ExtraFieldEdit.webp",
    "Customizing/extra-felder/feldtyp/index.html": "extra-felder/extra-field-edit/ExtraFieldEdit.webp",
    "Customizing/extra-felder/maskenzuordnung/index.html": "extra-felder/extra-field-edit/ExtraFieldEdit.webp",
    "Customizing/extra-felder/auswertbarkeit/index.html": "extra-felder/extra-field-edit/ExtraFieldEdit.webp",
    "Customizing/formatierte-suche/index.html": "formatierte-suche/formatted-search-edit/FormattedSearchEdit.webp",
    "Customizing/formatierte-suche/ausloeser/index.html": "formatierte-suche/formatted-search-edit/FormattedSearchEdit.webp",
    "Customizing/formatierte-suche/formel-und-abfrage/index.html": "formatierte-suche/formatted-search-edit/FormattedSearchEdit.webp",
    "Customizing/formatierte-suche/zielfeld/index.html": "formatierte-suche/formatted-search-edit/FormattedSearchEdit.webp",
}


NO_SAFE_SCREENSHOT = {
    "Customizing/index.html",
    "Customizing/vorschlagswerte/index.html",
    "Customizing/vorschlagswerte/feldvorbelegung/index.html",
    "Customizing/vorschlagswerte/gueltigkeitsbereich/index.html",
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_path(path: str) -> str:
    return path.replace("\\", "/").lstrip("/")


def file_safe_stem(value: str) -> str:
    stem = Path(value.replace("\\", "/")).stem
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_")
    return stem or "screenshot"


def clear_screenshot(ws, row: int, headers: dict[str, int], status: str) -> None:
    for column in ["Screenshot", "SCREENSHOT_WEB_PATH", "IMAGE_ALT", "IMAGE_CAPTION"]:
        if column in headers:
            ws.cell(row, headers[column]).value = None
    if "IMAGE_STATUS" in headers:
        ws.cell(row, headers["IMAGE_STATUS"]).value = status


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-customizing-screenshots-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False, keep_links=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }

    updated = 0
    copied = 0
    skipped: list[str] = []
    missing: list[str] = []
    unmapped: list[str] = []
    for row in range(2, ws.max_row + 1):
        url_path = normalize_path(cell_text(ws.cell(row, headers["URL_PATH"]).value))
        content_type = cell_text(ws.cell(row, headers["CONTENT_TYPE"]).value)
        if not url_path.startswith("Customizing/") or content_type not in {"HelpPage", "FAQ", "View"}:
            continue

        if url_path in NO_SAFE_SCREENSHOT:
            clear_screenshot(ws, row, headers, "no-safe-customizing-screenshot")
            skipped.append(url_path)
            continue

        source_rel = SOURCE_BY_PATH.get(url_path)
        if not source_rel:
            clear_screenshot(ws, row, headers, "unmapped-customizing-screenshot")
            unmapped.append(url_path)
            continue

        source_file = ANSICHTEN_ROOT / source_rel
        if not source_file.is_file():
            clear_screenshot(ws, row, headers, "missing-customizing-source")
            missing.append(f"{url_path} -> {source_rel}")
            continue

        target_dir = HELP_ROOT / url_path.removesuffix("index.html").rstrip("/")
        target_dir.mkdir(parents=True, exist_ok=True)
        target_name = file_safe_stem(source_file.name) + "_Customizing.webp"
        target_file = target_dir / target_name
        if not target_file.exists() or target_file.read_bytes() != source_file.read_bytes():
            shutil.copy2(source_file, target_file)
            copied += 1

        target_rel = target_file.relative_to(HELP_ROOT).as_posix()
        web_path = "/de/help/" + target_rel
        topic = cell_text(ws.cell(row, headers["Thema"]).value)
        ws.cell(row, headers["Screenshot"]).value = target_rel
        ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value = web_path
        ws.cell(row, headers["IMAGE_ALT"]).value = f"Screenshot {topic} in X-ERP"
        ws.cell(row, headers["IMAGE_CAPTION"]).value = f"{topic} im X-ERP Customizing"
        ws.cell(row, headers["IMAGE_STATUS"]).value = "local-copy:customizing-verified"
        updated += 1

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"updated_rows={updated}")
    print(f"copied_files={copied}")
    print(f"skipped_no_safe={len(skipped)}")
    for item in skipped:
        print(f"SKIPPED {item}")
    print(f"unmapped={len(unmapped)}")
    for item in unmapped:
        print(f"UNMAPPED {item}")
    print(f"missing_sources={len(missing)}")
    for item in missing:
        print(f"MISSING {item}")


if __name__ == "__main__":
    main()
