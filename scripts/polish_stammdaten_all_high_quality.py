from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"


PAGES = {
    "Stammdaten/index.html": {
        "title": "Stammdaten",
        "intro": "Stammdaten sind die dauerhafte Grundlage der täglichen Arbeit in X-ERP. Sie steuern, welche Artikel verkauft werden, welche Partner beliefert oder berechnet werden, wie Preise gefunden werden, wo Bestände liegen und welche Mitarbeiter, Texte, Projekte, Ressourcen und Anhänge in den Prozessen verwendet werden.",
        "flow": (
            "Stammdaten sauber aufbauen",
            [
                "Struktur klären: Legen Sie zuerst fest, welche Artikel, Partner, Lager, Preise und Verantwortlichkeiten wirklich benötigt werden.",
                "Pflichtdaten pflegen: Erfassen Sie eindeutige Nummern, Namen, Gruppen, Status und die wichtigsten kaufmännischen Einstellungen.",
                "Prozessdaten ergänzen: Aktivieren Sie Verkauf, Einkauf, Lager, Projekte oder Ressourcen nur dort, wo sie fachlich genutzt werden.",
                "Abhängigkeiten prüfen: Kontrollieren Sie Preislisten, Adressen, Lagerplätze, Finanzgruppen, Texte und Anhänge.",
                "Testen: Öffnen Sie einen Beispielbeleg oder eine typische Ansicht und prüfen Sie, ob Suche, Auswahl und Folgeprozesse stimmen.",
            ],
        ),
        "bullets": [
            "Stammdaten sollten eindeutig, aktuell und für Anwender verständlich sein.",
            "Nummern und Gruppen müssen stabil bleiben, auch wenn Lieferanten, Preise oder Verantwortliche wechseln.",
            "Spezialinformationen gehören in eigene Bereiche wie Preise, Banking, Lagerung, Anhänge oder Extra-Felder.",
            "Änderungen an Stammdaten können Belege, Auswertungen, Lager und Automatisierungen beeinflussen.",
        ],
        "box": ("Empfehlung", "Bearbeiten Sie Stammdaten nicht nur aus Sicht eines einzelnen Feldes. Entscheidend ist, ob der Datensatz später in Verkauf, Einkauf, Lager, Buchhaltung und Auswertung zuverlässig funktioniert."),
    },
    "Stammdaten/partner/index.html": {
        "title": "Partnerstammdaten",
        "intro": "Partnerstammdaten beschreiben Kunden, Lieferanten und weitere Geschäftspartner. Ein Partner kann mehrere Rollen haben: Er kann Käufer, Rechnungsempfänger, Lieferant, Ansprechpartner im CRM oder Portalbenutzer sein. Saubere Partnerdaten vermeiden falsche Lieferungen, falsche Rechnungen und unnötige Rückfragen.",
        "flow": (
            "Partner sicher anlegen",
            [
                "Identität pflegen: Name, Matchcode, Nummer und Status eindeutig erfassen.",
                "Rolle festlegen: Kundendetails, Lieferantendetails oder beide Rollen aktiv pflegen.",
                "Adressen prüfen: Hauptadresse, Lieferadressen und Rechnungsadressen sauber trennen.",
                "Kaufmännische Daten ergänzen: Zahlungsbedingungen, Lieferbedingungen, Finanzgruppen, Steuer- und Bankingdaten kontrollieren.",
                "Kommunikation anbinden: Ansprechpartner, E-Mail, CRM, Portal und Anhänge ergänzen.",
                "Testen: Partner in einem Beispielbeleg auswählen und prüfen, ob Adresse, Konditionen und Ansprechpartner stimmen.",
            ],
        ),
        "bullets": [
            "Der Matchcode sollte Anwendern helfen, den richtigen Partner schnell zu finden.",
            "Kunden- und Lieferanteninformationen gehören in getrennte Bereiche, auch wenn es derselbe Geschäftspartner ist.",
            "Rechnungs- und Lieferadressen sollten nur dann getrennt gepflegt werden, wenn sie fachlich abweichen.",
            "Offene Posten, Mahnungen, Umsatz und Kontenblatt helfen bei der laufenden Kontrolle.",
        ],
        "box": ("Kunde oder Lieferant?", "Ein Partner ist der gemeinsame Stammdatensatz. Die Rollen Kunde und Lieferant werden über die passenden Detailbereiche gesteuert. So bleiben Adresse und Kommunikation zentral, während Konditionen getrennt gepflegt werden."),
    },
    "Stammdaten/partner/partner-adresse/index.html": {
        "title": "Partner Adresse",
        "intro": "Die Adresse identifiziert den Partner und ist die Grundlage für Belege, Suche, Lieferungen, Rechnungen und Kommunikation. Fehler in diesem Bereich wirken sich sofort auf Ausdrucke, E-Mails und Folgeprozesse aus.",
        "flow": (
            "Adresse prüfen",
            [
                "Namen erfassen: Firmenname oder Personenname in den vorgesehenen Feldern pflegen.",
                "Suchbarkeit sichern: Matchcode, Nummer und Kurzbezeichnung eindeutig wählen.",
                "Anschrift erfassen: Straße, Postleitzahl, Ort und Land vollständig prüfen.",
                "Kommunikation ergänzen: Telefon, E-Mail und Webadresse eintragen, wenn sie zentral gelten.",
                "Status kontrollieren: Nur aktive Partner für neue Vorgänge verfügbar halten.",
            ],
        ),
        "bullets": [
            "Nutzen Sie Liefer- oder Rechnungsadressen nur für echte Abweichungen.",
            "Vermeiden Sie Zusatzinformationen im Namen, wenn dafür eigene Felder vorhanden sind.",
            "Bei internationalen Partnern sind Land und lokale Angaben besonders wichtig.",
        ],
        "box": ("Praxisregel", "Die Hauptadresse sollte den rechtlich oder organisatorisch richtigen Partner beschreiben. Zustell- und Abrechnungsvarianten gehören in Liefer- und Rechnungsadressen."),
    },
    "Stammdaten/partner/partner-optionen/index.html": {
        "title": "Partner Optionen",
        "intro": "Optionen steuern zusätzliche Eigenschaften des Partners. Sie ergänzen die Adresse um Kennzeichen, interne Zuordnungen und Verhaltensregeln, die in Belegen, Listen oder Automatisierungen berücksichtigt werden können.",
        "bullets": [
            "Pflegen Sie Optionen nur, wenn sie im Prozess tatsächlich ausgewertet werden.",
            "Status- und Sperrkennzeichen sollten eindeutig geregelt sein.",
            "Interne Gruppen und Kategorien erleichtern Filter, Auswertungen und Zuständigkeiten.",
            "Prüfen Sie nach Änderungen, ob bestehende Belege oder Workflows betroffen sind.",
        ],
        "box": ("Nicht als Notizfeld verwenden", "Optionen sollten strukturiert auswertbar bleiben. Freitextinformationen gehören in Info, CRM oder Anhänge."),
    },
    "Stammdaten/partner/partner-kundendetails/index.html": {
        "title": "Partner Kundendetails",
        "intro": "Kundendetails steuern, wie ein Partner im Verkauf behandelt wird. Hier werden Konditionen, Zuständigkeiten, Finanzgruppen und Regeln gepflegt, die Angebote, Aufträge, Lieferscheine, Rechnungen und Auswertungen beeinflussen.",
        "flow": (
            "Kunde verkaufsbereit machen",
            [
                "Kundennummer prüfen: Eindeutig und dauerhaft vergeben.",
                "Zahlungs- und Lieferbedingungen setzen: Standardwerte für neue Verkaufsbelege festlegen.",
                "Finanzgruppe wählen: Erlös-, Steuer- und Buchungslogik fachlich korrekt zuordnen.",
                "Zuständigkeit ergänzen: Verkäufer, Betreuer oder Gruppe pflegen, wenn damit gearbeitet wird.",
                "Kredit und Sperren prüfen: Risiko- und Freigabelogik bewusst setzen.",
            ],
        ),
        "bullets": [
            "Kundendetails sollten nur bei Partnern gepflegt werden, die wirklich als Kunde genutzt werden.",
            "Preis- und Rabattlogik muss zur späteren Verkaufspreisfindung passen.",
            "Elektronische Rechnungs- oder Portalregeln sollten früh geprüft werden.",
        ],
        "box": ("Auswirkung auf Belege", "Viele Werte aus den Kundendetails werden beim Erstellen eines neuen Verkaufsbelegs vorgeschlagen. Fehler werden deshalb oft erst im Auftrag oder in der Rechnung sichtbar."),
    },
    "Stammdaten/partner/partner-lieferantendetails/index.html": {
        "title": "Partner Lieferantendetails",
        "intro": "Lieferantendetails beschreiben, wie bei einem Partner eingekauft wird. Sie steuern Einkaufskonditionen, Lieferlogik, Zahlungsbedingungen, Fremdwährungen und Auswertungen.",
        "flow": (
            "Lieferant einkaufsbereit machen",
            [
                "Lieferantennummer prüfen: Den Partner eindeutig als Lieferant führen.",
                "Konditionen setzen: Zahlungsbedingung, Lieferbedingung und Lieferzeit realistisch pflegen.",
                "Finanzgruppe wählen: Aufwands-, Steuer- und Buchungslogik korrekt zuordnen.",
                "Preisbezug prüfen: Einkaufspreislisten und Katalognummern anbinden.",
                "Testbestellung prüfen: Einen Beispielartikel auswählen und Preis, Nummer und Lieferdaten kontrollieren.",
            ],
        ),
        "bullets": [
            "Lieferantendetails sind getrennt von Kundendetails zu betrachten.",
            "Fremdwährungen, Rabatte und Zusatzkosten müssen mit der Einkaufspreislogik zusammenpassen.",
            "Katalognummern helfen, Bestellungen für den Lieferanten eindeutig zu machen.",
        ],
        "box": ("Beschaffungssicherheit", "Je sauberer Lieferantendetails und Einkaufspreise gepflegt sind, desto zuverlässiger funktionieren Bestellungen, Beschaffungsplanung und Nachkalkulation."),
    },
    "Stammdaten/partner/partner-ansprechpartner/index.html": {
        "title": "Partner Ansprechpartner",
        "intro": "Ansprechpartner speichern Personen, Rollen und Kommunikationsdaten beim Partner. Sie helfen Verkauf, Einkauf, Service und Buchhaltung, die richtige Person direkt zu erreichen.",
        "flow": (
            "Ansprechpartner sinnvoll pflegen",
            [
                "Rolle klären: Einkauf, Verkauf, Technik, Buchhaltung oder Geschäftsführung erfassen.",
                "Kontaktdaten prüfen: E-Mail, Telefon und Mobilnummer aktuell halten.",
                "Standardkontakt markieren: Den wichtigsten Kontakt für typische Vorgänge erkennbar machen.",
                "Belegbezug prüfen: Kontrollieren, ob Ansprechpartner in Angeboten, Bestellungen oder E-Mails genutzt werden.",
            ],
        ),
        "bullets": [
            "Mehrere Ansprechpartner sind sinnvoll, wenn verschiedene Abteilungen beteiligt sind.",
            "Inaktive oder veraltete Kontakte sollten bereinigt werden.",
            "Rollen sind für gezielte Kommunikation wichtiger als lange Notizen.",
        ],
        "box": ("Qualität", "Ein guter Ansprechpartner-Datensatz beantwortet sofort: Wer ist zuständig, wofür ist die Person zuständig und wie erreiche ich sie?"),
    },
    "Stammdaten/partner/partner-lieferadressen/index.html": {
        "title": "Partner Lieferadressen",
        "intro": "Lieferadressen erfassen abweichende Zustellorte eines Partners. Sie werden genutzt, wenn Waren oder Leistungen nicht an die Hauptadresse geliefert werden sollen.",
        "flow": (
            "Lieferadresse anlegen",
            [
                "Zweck prüfen: Nur echte abweichende Lieferorte als Lieferadresse erfassen.",
                "Adresse vollständig pflegen: Name, Straße, PLZ, Ort, Land und Zusatzangaben kontrollieren.",
                "Standard setzen: Häufig verwendete Adresse als bevorzugten Vorschlag definieren.",
                "Beleg testen: Auftrag oder Lieferschein öffnen und prüfen, ob die richtige Adresse gezogen wird.",
            ],
        ),
        "bullets": [
            "Lieferadressen sind nicht für Rechnungsempfänger gedacht.",
            "Standorte, Baustellen oder Filialen sollten klar benannt werden.",
            "Veraltete Lieferorte sollten deaktiviert oder bereinigt werden.",
        ],
        "box": ("Abgrenzung", "Die Hauptadresse beschreibt den Partner. Die Lieferadresse beschreibt den Ort, an den geliefert wird."),
    },
    "Stammdaten/partner/partner-rechnungsadressen/index.html": {
        "title": "Partner Rechnungsadressen",
        "intro": "Rechnungsadressen erfassen abweichende Empfänger für Rechnungen. Sie sind wichtig, wenn Bestellung, Lieferung und Abrechnung über unterschiedliche Organisationseinheiten laufen.",
        "flow": (
            "Rechnungsadresse prüfen",
            [
                "Abweichung klären: Nur erfassen, wenn Rechnungsempfänger und Hauptadresse nicht identisch sind.",
                "Rechtliche Angaben prüfen: Firmierung, Anschrift und Land korrekt erfassen.",
                "Standard definieren: Häufige Rechnungsadresse als Vorschlag setzen.",
                "Rechnung testen: Beispielrechnung erstellen und Empfänger, Steuer- und Zahlungsdaten kontrollieren.",
            ],
        ),
        "bullets": [
            "Rechnungsadressen beeinflussen externe Dokumente unmittelbar.",
            "Sie sollten nicht als Lieferadressen zweckentfremdet werden.",
            "Bei Konzernen oder Filialstrukturen ist eine klare Benennung wichtig.",
        ],
        "box": ("Buchhaltung", "Falsche Rechnungsadressen führen schnell zu Rückfragen, Zahlungsverzug oder Korrekturrechnungen. Prüfen Sie diesen Bereich besonders sorgfältig."),
    },
    "Stammdaten/partner/partner-katalognummern/index.html": {
        "title": "Partner Katalognummern",
        "intro": "Katalognummern verbinden eigene Artikel mit Nummern des Partners. Dadurch können Kunden- oder Lieferantennummern genutzt werden, ohne die eigene Artikelnummer zu verändern.",
        "bullets": [
            "Lieferantenartikelnummern erleichtern Bestellungen und Wareneingang.",
            "Kundenartikelnummern helfen bei Angeboten, EDI oder Rahmenverträgen.",
            "Die eigene Artikelnummer bleibt stabil und intern eindeutig.",
            "Katalognummern sollten bei Dubletten oder Lieferantenwechsel geprüft werden.",
        ],
        "box": ("Keine Nummern vermischen", "Externe Nummern gehören in Katalognummern, nicht in die interne Artikelnummer oder Bezeichnung."),
    },
    "Stammdaten/partner/partner-einkaufspreislisten/index.html": {
        "title": "Partner Einkaufspreislisten",
        "intro": "Einkaufspreislisten speichern lieferantenspezifische Preise, Mengenstaffeln, Gültigkeiten und Konditionen. Sie sind die Grundlage für belastbare Einkaufspreise und Einstandskosten.",
        "flow": (
            "Einkaufspreise pflegen",
            [
                "Lieferant auswählen: Preis immer zum richtigen Partner erfassen.",
                "Artikel und Menge prüfen: Preis, Mengeneinheit und Staffel sauber eintragen.",
                "Gültigkeit setzen: Ab wann der Preis gilt und wann er ersetzt werden muss.",
                "Zusatzkosten bewerten: Einstandspreis und Nebenkosten prüfen, wenn sie kalkulationsrelevant sind.",
                "Bestellung testen: Artikel beim Lieferanten auswählen und Preisvorschlag kontrollieren.",
            ],
        ),
        "bullets": [
            "Preise ohne Gültigkeit sind schwer zu kontrollieren.",
            "Mengenstaffeln müssen zur Bestellmengeneinheit passen.",
            "Lieferantenwechsel sollte nicht über Artikelnummern, sondern über Preislisten und Katalognummern abgebildet werden.",
        ],
        "box": ("Kalkulation", "Für Entscheidungen ist oft nicht nur der reine Einkaufspreis wichtig, sondern der realistische Einstandspreis."),
    },
    "Stammdaten/partner/partner-umsatz/index.html": {
        "title": "Partner Umsatz",
        "intro": "Der Umsatzbereich zeigt, wie sich das Geschäft mit einem Partner entwickelt. Er unterstützt Vertrieb, Einkauf und Management bei Bewertung, Priorisierung und Nachverfolgung.",
        "bullets": [
            "Nutzen Sie Umsatzdaten zur Einordnung wichtiger Kunden oder Lieferanten.",
            "Prüfen Sie auffällige Veränderungen frühzeitig.",
            "Kombinieren Sie Umsatz mit offenen Posten, Margen oder Aktivitäten, wenn möglich.",
        ],
        "box": ("Analyse", "Umsatz allein erklärt nicht die Qualität einer Geschäftsbeziehung. Er wird aussagekräftiger, wenn Konditionen, Zahlungsweise, Reklamationen und Aufwand mit betrachtet werden."),
    },
    "Stammdaten/partner/partner-info/index.html": {
        "title": "Partner Info",
        "intro": "Info enthält ergänzende Hinweise zum Partner. Dieser Bereich eignet sich für kurze, wichtige Informationen, die Anwender beim Öffnen des Partners schnell verstehen sollen.",
        "bullets": [
            "Formulieren Sie Hinweise knapp und eindeutig.",
            "Zeitkritische oder personenbezogene Informationen sollten regelmäßig geprüft werden.",
            "Für Dateien, Nachweise oder längere Dokumente sind Anhänge besser geeignet.",
        ],
        "box": ("Kurz halten", "Info sollte nicht zum Sammelplatz für alte Gesprächsnotizen werden. Laufende Aktivitäten gehören eher in CRM oder Dokumentation."),
    },
    "Stammdaten/partner/partner-bild/index.html": {
        "title": "Partner Bild",
        "intro": "Das Partnerbild kann Logo, Foto oder visuelle Kennzeichnung des Partners enthalten. Es unterstützt die schnelle Orientierung und kann in passenden Oberflächen angezeigt werden.",
        "bullets": [
            "Verwenden Sie klare und aktuelle Bilder.",
            "Logos sollten nicht verzerrt oder abgeschnitten sein.",
            "Bei Personenbildern sind Datenschutz und Einwilligung zu beachten.",
        ],
        "box": ("Anwendungsfall", "Ein Bild ist hilfreich, wenn Anwender viele ähnliche Partner unterscheiden müssen oder Logos in der internen Orientierung genutzt werden."),
    },
    "Stammdaten/partner/partner-kategorie/index.html": {
        "title": "Partner Kategorie",
        "intro": "Kategorien ordnen Partner für Suche, Filter, Auswertungen und Zuständigkeiten. Sie helfen, große Partnerbestände strukturiert zu betrachten.",
        "bullets": [
            "Kategorien sollten aus Anwendersicht verständlich sein.",
            "Vermeiden Sie zu viele ähnliche Kategorien.",
            "Nutzen Sie Kategorien konsequent, damit Filter und Auswertungen stimmen.",
        ],
        "box": ("Struktur", "Eine Kategorie ist nur dann wertvoll, wenn sie dauerhaft gepflegt und tatsächlich ausgewertet wird."),
    },
    "Stammdaten/partner/partner-crm/index.html": {
        "title": "Partner CRM",
        "intro": "CRM verbindet den Partner mit Aktivitäten, Kontakten und Vertriebs- oder Servicevorgängen. Dadurch wird sichtbar, was mit dem Partner besprochen, geplant oder nachverfolgt wurde.",
        "flow": (
            "CRM sinnvoll nutzen",
            [
                "Aktivität erfassen: Gespräche, Aufgaben oder Termine zeitnah dokumentieren.",
                "Zuständigkeit setzen: Verantwortliche Person oder Gruppe klar zuordnen.",
                "Nächsten Schritt festlegen: Folgeaufgabe oder Wiedervorlage eintragen.",
                "Bezug herstellen: Wenn möglich zu Angebot, Auftrag oder Problem verknüpfen.",
            ],
        ),
        "bullets": [
            "CRM ersetzt keine Stammdatenpflege, sondern ergänzt sie um Verlauf und Aktivitäten.",
            "Kurze, konkrete Einträge sind im Alltag wertvoller als lange unstrukturierte Notizen.",
            "Wiedervorlagen helfen, offene Themen nicht zu verlieren.",
        ],
        "box": ("Vertrieb und Service", "CRM ist besonders nützlich, wenn mehrere Mitarbeiter mit demselben Partner arbeiten und den gleichen Informationsstand benötigen."),
    },
    "Stammdaten/partner/partner-e-mail/index.html": {
        "title": "Partner E-Mail",
        "intro": "Der E-Mail-Bereich zeigt oder steuert E-Mail-Kommunikation mit dem Partner. Er hilft, Nachrichten im Kontext des Partners wiederzufinden und Belegkommunikation nachvollziehbar zu halten.",
        "bullets": [
            "Zentrale E-Mail-Adressen sollten in Adresse oder Ansprechpartnern korrekt gepflegt sein.",
            "Für Belege ist wichtig, welche Adresse als Standard verwendet wird.",
            "Kommunikation sollte möglichst beim richtigen Partner und Ansprechpartner landen.",
        ],
        "box": ("Nachvollziehbarkeit", "Wenn E-Mails sauber zugeordnet werden, müssen Anwender Informationen nicht in persönlichen Postfächern suchen."),
    },
    "Stammdaten/partner/partner-dok-liste/index.html": {
        "title": "Partner Dokumentliste",
        "intro": "Die Dokumentliste zeigt Belege und Dokumente, die mit dem Partner verbunden sind. Sie ist ein schneller Einstieg in Angebote, Aufträge, Lieferungen, Rechnungen oder andere Vorgänge.",
        "bullets": [
            "Nutzen Sie die Liste, um Vorgänge beim Partner schnell nachzuvollziehen.",
            "Filter nach Datum, Dokumentart oder Status erleichtern die Kontrolle.",
            "Bei Rückfragen lässt sich der passende Beleg direkt aus dem Partnerkontext öffnen.",
        ],
        "box": ("Überblick", "Die Dokumentliste beantwortet: Was ist mit diesem Partner bereits gelaufen und welcher Vorgang ist aktuell relevant?"),
    },
    "Stammdaten/partner/partner-positionsliste/index.html": {
        "title": "Partner Positionsliste",
        "intro": "Die Positionsliste zeigt einzelne Belegpositionen zum Partner. Sie ist genauer als die Dokumentliste, weil sie Artikel, Mengen, Preise und Positionsdetails sichtbar macht.",
        "bullets": [
            "Prüfen Sie, welche Artikel ein Partner gekauft oder geliefert hat.",
            "Nutzen Sie Positionsdaten für Preisvergleiche und Reklamationen.",
            "Historische Mengen und Preise helfen bei neuen Angeboten oder Bestellungen.",
        ],
        "box": ("Detailanalyse", "Wenn die Frage nicht nur lautet welcher Beleg, sondern welcher Artikel zu welcher Menge, ist die Positionsliste der richtige Einstieg."),
    },
    "Stammdaten/partner/partner-mahnungsliste/index.html": {
        "title": "Partner Mahnungsliste",
        "intro": "Die Mahnungsliste zeigt mahnrelevante Vorgänge des Partners. Sie unterstützt die Buchhaltung bei der Kontrolle offener und überfälliger Forderungen.",
        "bullets": [
            "Prüfen Sie offene Beträge, Fälligkeit und Mahnstufe.",
            "Kontrollieren Sie vor einer Mahnung, ob Zahlungen oder Klärfälle vorliegen.",
            "Nutzen Sie Hinweise im Partner, wenn besondere Absprachen bestehen.",
        ],
        "box": ("Sorgfalt", "Mahnungen wirken direkt auf die Kundenbeziehung. Deshalb sollten Stammdaten, Rechnungsadresse und offene Posten vor dem Versand geprüft sein."),
    },
    "Stammdaten/partner/partner-offene-posten/index.html": {
        "title": "Partner Offene Posten",
        "intro": "Offene Posten zeigen unbezahlte oder noch nicht ausgeglichene Forderungen und Verbindlichkeiten zum Partner. Der Bereich ist wichtig für Zahlungsstatus, Liquidität und Klärfälle.",
        "bullets": [
            "Prüfen Sie Fälligkeit, Betrag und Belegbezug.",
            "Bei Kunden helfen offene Posten vor weiteren Lieferungen oder Freigaben.",
            "Bei Lieferanten unterstützen offene Posten die Zahlungsplanung.",
            "Abweichungen sollten mit Kontenblatt und Belegen abgeglichen werden.",
        ],
        "box": ("Kontrolle", "Offene Posten sind kein Stammdatenfeld, aber eine wichtige Sicht auf die Qualität der laufenden Geschäftsbeziehung."),
    },
    "Stammdaten/partner/partner-banking/index.html": {
        "title": "Partner Banking",
        "intro": "Banking enthält Bankverbindungen und zahlungsrelevante Informationen des Partners. Diese Daten werden für Zahlungsverkehr, Lastschriften, Überweisungen und Abgleiche benötigt.",
        "flow": (
            "Bankdaten prüfen",
            [
                "Kontoinhaber klären: Name und Partnerbezug prüfen.",
                "IBAN und BIC erfassen: Schreibweise und Plausibilität kontrollieren.",
                "Zahlungsart zuordnen: Überweisung, Lastschrift oder andere Verfahren bewusst setzen.",
                "Freigabe prüfen: Änderungen an Bankdaten intern kontrollieren.",
            ],
        ),
        "bullets": [
            "Bankdaten sind sensible Stammdaten und müssen besonders sorgfältig gepflegt werden.",
            "Veraltete Konten sollten deaktiviert oder klar gekennzeichnet werden.",
            "Bei Lastschriftverfahren sind Mandatsdaten und Freigaben relevant.",
        ],
        "box": ("Sicherheit", "Bankdaten sollten nie aus E-Mails ungeprüft übernommen werden. Änderungen brauchen eine klare interne Kontrollregel."),
    },
    "Stammdaten/partner/partner-portal/index.html": {
        "title": "Partner Portal",
        "intro": "Portal-Daten steuern, ob und wie ein Partner Zugriff auf Portal-Funktionen erhält. Das kann Kundenservice, Belegzugriff, Kommunikation oder Self-Service-Prozesse betreffen.",
        "bullets": [
            "Aktivieren Sie Portalzugriff nur für berechtigte Partner.",
            "Prüfen Sie, welche Kontakte oder Benutzer Zugriff erhalten sollen.",
            "Rechte und sichtbare Inhalte müssen zur Geschäftsbeziehung passen.",
            "Bei Änderungen an Rollen oder Ansprechpartnern sollte der Portalzugang mit geprüft werden.",
        ],
        "box": ("Zugriff", "Portalrechte sind Stammdaten mit Sicherheitswirkung. Sie sollten genauso bewusst gepflegt werden wie Banking- oder Buchhaltungsdaten."),
    },
    "Stammdaten/partner/partner-kontenblatt/index.html": {
        "title": "Partner Kontenblatt",
        "intro": "Das Kontenblatt zeigt buchhalterische Bewegungen zum Partner. Es hilft bei Abstimmung, Zahlungsprüfung, Reklamationen und Rückfragen zu Rechnungen oder Gutschriften.",
        "bullets": [
            "Nutzen Sie das Kontenblatt, wenn offene Posten oder Zahlungen unklar sind.",
            "Vergleichen Sie Belegdatum, Buchungsdatum, Betrag und Ausgleich.",
            "Bei Differenzen sollten Belegliste und offene Posten gemeinsam geprüft werden.",
        ],
        "box": ("Buchhalterische Sicht", "Während die Dokumentliste operative Vorgänge zeigt, zeigt das Kontenblatt die buchhalterische Bewegung."),
    },
    "Stammdaten/partner/partner-lokal/index.html": {
        "title": "Partner Lokal",
        "intro": "Lokale Angaben ergänzen den Partner um länder- oder regionalspezifische Informationen. Sie sind wichtig für Steuer, Außenhandel, Sprache, Formulare und lokale Geschäftsregeln.",
        "bullets": [
            "Prüfen Sie Land, Sprache, Steuer- und Registrierungsangaben.",
            "Bei internationalen Partnern können lokale Daten die Belegausgabe beeinflussen.",
            "Änderungen sollten mit Buchhaltung oder Exportverantwortlichen abgestimmt werden.",
        ],
        "box": ("International", "Lokale Stammdaten werden oft erst bei Belegausgabe, Steuerprüfung oder Außenhandel sichtbar. Pflegen Sie sie deshalb frühzeitig."),
    },
    "Stammdaten/partner/partner-aenderungsprotokoll/index.html": {
        "title": "Partner Änderungsprotokoll",
        "intro": "Das Änderungsprotokoll zeigt, welche Partnerdaten wann und von wem geändert wurden. Es unterstützt Nachvollziehbarkeit, Prüfung und Fehlersuche.",
        "bullets": [
            "Nutzen Sie das Protokoll bei unerwarteten Adress-, Konditions- oder Bankdatenänderungen.",
            "Prüfen Sie Zeitpunkt, Benutzer und geändertes Feld.",
            "Bei sensiblen Feldern wie Banking oder Sperren ist das Protokoll besonders wichtig.",
        ],
        "box": ("Audit", "Das Änderungsprotokoll erklärt nicht die fachliche Entscheidung, aber es zeigt, wann die Daten technisch geändert wurden."),
    },
    "Stammdaten/partner/partner-anhaenge/index.html": {
        "title": "Partner Anhänge",
        "intro": "Anhänge speichern Dateien direkt am Partner. Dazu gehören Verträge, Nachweise, Zertifikate, Freigaben, Gesprächsunterlagen oder kundenspezifische Dokumente.",
        "bullets": [
            "Benennen Sie Dateien so, dass Inhalt und Stand erkennbar sind.",
            "Veraltete Dokumente sollten entfernt oder eindeutig markiert werden.",
            "Sensible Dokumente benötigen passende Rechte und klare Ablage.",
        ],
        "box": ("Ablage", "Anhänge sind ideal für Unterlagen, die Anwender direkt im Partnerkontext benötigen und nicht erst in externen Ordnern suchen sollen."),
    },
    "Stammdaten/partner/partner-extra-felder/index.html": {
        "title": "Partner Extra-Felder",
        "intro": "Extra-Felder erweitern Partnerstammdaten um kundenspezifische Informationen. Welche Felder sichtbar sind, hängt von der Konfiguration ab.",
        "bullets": [
            "Nutzen Sie Extra-Felder für Informationen, die regelmäßig gesucht, gefiltert oder ausgewertet werden.",
            "Feldnamen sollten für Anwender eindeutig sein.",
            "Pflichtfelder sollten sparsam eingesetzt werden.",
            "Einheitliche Werte sind wichtig, wenn Auswertungen darauf basieren.",
        ],
        "box": ("Konfiguration", "Extra-Felder sind stark, wenn sie strukturiert genutzt werden. Für einmalige Hinweise sind Info, CRM oder Anhänge meist besser geeignet."),
    },
    "Stammdaten/preismanagement/index.html": {
        "title": "Preismanagement",
        "intro": "Preismanagement bündelt Einkaufs- und Verkaufspreise sowie preisrelevante Dokument- und Positionssichten. Es hilft, Preise nachvollziehbar zu pflegen, Gültigkeiten zu steuern und Preisentscheidungen aus realen Vorgängen abzuleiten.",
        "flow": (
            "Preise belastbar pflegen",
            [
                "Preisart klären: Einkaufspreis, Verkaufspreis, Staffel oder dokumentbezogener Preis unterscheiden.",
                "Bezug festlegen: Artikel, Partner, Preisgruppe, Menge, Währung und Einheit prüfen.",
                "Gültigkeit setzen: Ab wann ein Preis gilt und wann er abgelöst wird.",
                "Kalkulation prüfen: Einkaufspreis, Einstandspreis, Marge und Verkaufspreis zusammen betrachten.",
                "Belegtest durchführen: Preis in Angebot, Auftrag, Bestellung oder Rechnung kontrollieren.",
            ],
        ),
        "bullets": [
            "Preise ohne klare Gültigkeit sind schwer prüfbar.",
            "Mengenstaffeln müssen zur Mengeneinheit passen.",
            "Einkaufs- und Verkaufspreise sollten nicht isoliert betrachtet werden.",
            "Dokument- und Positionslisten helfen, Preisentscheidungen aus echten Vorgängen abzuleiten.",
        ],
        "box": ("Preisfindung", "Ein guter Preisstamm erklärt, welcher Preis wann, für wen, in welcher Einheit und aus welchem Grund gilt."),
    },
    "Stammdaten/preismanagement/preise-einkauf/index.html": {
        "title": "Preise Einkauf",
        "intro": "Einkaufspreise beschreiben, zu welchen Konditionen Artikel beschafft werden. Sie können lieferantenspezifisch, mengenabhängig, zeitlich gültig und in unterschiedlichen Währungen gepflegt werden.",
        "flow": (
            "Einkaufspreis anlegen",
            [
                "Artikel und Lieferant wählen: Preis dem richtigen Beschaffungskontext zuordnen.",
                "Menge und Einheit prüfen: Preisbasis, Preiseinheit und Staffel passend erfassen.",
                "Gültigkeit eintragen: Startdatum und bei Bedarf Ende oder Ablösung definieren.",
                "Einstand bewerten: Nebenkosten und Kalkulationsfaktoren berücksichtigen.",
                "Bestellung testen: Prüfen, ob der Preis im Einkauf richtig vorgeschlagen wird.",
            ],
        ),
        "bullets": [
            "Lieferantenpreise sollten mit Katalognummern abgestimmt sein.",
            "Staffeln müssen eindeutig und überschneidungsfrei sein.",
            "Bei Fremdwährungen sind Umrechnung und Gültigkeit besonders wichtig.",
        ],
        "box": ("Kostenblick", "Für Kalkulation und Marge ist der reine Einkaufspreis oft nicht ausreichend. Prüfen Sie, ob der Einstandspreis die bessere Entscheidungsgröße ist."),
    },
    "Stammdaten/preismanagement/preise-verkauf/index.html": {
        "title": "Preise Verkauf",
        "intro": "Verkaufspreise steuern, welche Preise in Angebot, Auftrag und Rechnung vorgeschlagen werden. Sie können nach Artikel, Preisgruppe, Partner, Menge, Währung und Gültigkeit differenziert werden.",
        "flow": (
            "Verkaufspreis anlegen",
            [
                "Artikel wählen: Preis dem richtigen Artikel und der richtigen Einheit zuordnen.",
                "Zielgruppe festlegen: Allgemeiner Preis, Preisgruppe oder kundenspezifischer Preis unterscheiden.",
                "Menge und Staffel prüfen: Mindestmenge, Preiseinheit und Staffel sauber pflegen.",
                "Gültigkeit setzen: Preisänderungen zeitlich kontrollierbar machen.",
                "Belegtest durchführen: Angebot oder Auftrag öffnen und Preisvorschlag prüfen.",
            ],
        ),
        "bullets": [
            "Verkaufspreise sollten zur Verkaufseinheit des Artikels passen.",
            "Sonderpreise brauchen klare Gültigkeit und eindeutigen Partner- oder Gruppenkontext.",
            "Preisänderungen sollten vor dem produktiven Einsatz in einem Testbeleg geprüft werden.",
        ],
        "box": ("Transparenz", "Je genauer der Preisbezug gepflegt ist, desto weniger manuelle Preisänderungen sind später im Beleg nötig."),
    },
    "Stammdaten/preismanagement/dok/index.html": {
        "title": "Dokumente im Preismanagement",
        "intro": "Dokumente zeigen preisrelevante Vorgänge auf Kopfebene. Sie helfen, Preisentscheidungen mit Angeboten, Aufträgen, Bestellungen, Rechnungen oder anderen Belegen zu verbinden.",
        "bullets": [
            "Nutzen Sie Dokumente, um Preisverlauf und Geschäftskontext zu verstehen.",
            "Prüfen Sie Status, Datum, Partner und Belegart.",
            "Für konkrete Artikelpreise ist die Positionssicht meist genauer.",
        ],
        "box": ("Kopf- und Positionssicht", "Der Dokumentkopf erklärt den Vorgang. Die Position erklärt den konkreten Preis."),
    },
    "Stammdaten/preismanagement/dok-positionen/index.html": {
        "title": "Dokumentpositionen im Preismanagement",
        "intro": "Dokumentpositionen zeigen, welche Artikel zu welchen Mengen und Preisen in realen Vorgängen verwendet wurden. Sie sind eine wichtige Grundlage für Preisprüfung, Nachkalkulation und Verhandlungen.",
        "bullets": [
            "Vergleichen Sie historische Preise mit aktuellen Preislisten.",
            "Prüfen Sie Mengen, Rabatte, Einheiten und Partnerbezug.",
            "Nutzen Sie Positionsdaten, wenn ein Preis im Beleg nicht nachvollziehbar ist.",
        ],
        "box": ("Praxisnutzen", "Positionsdaten zeigen nicht nur den gepflegten Preis, sondern den tatsächlich verwendeten Preis im Geschäftsprozess."),
    },
    "Stammdaten/preismanagement/dok-kette/index.html": {
        "title": "Dokumentkette",
        "intro": "Die Dokumentkette zeigt, wie Vorgänge miteinander verbunden sind. Sie macht nachvollziehbar, aus welchem Angebot, Auftrag, Lieferschein, Rechnung oder Einkaufsvorgang ein Preis- oder Mengenbezug entstanden ist.",
        "bullets": [
            "Nutzen Sie die Kette, um Ursprung und Folgebelege zu prüfen.",
            "Bei Preisabweichungen hilft die Kette, den ersten relevanten Beleg zu finden.",
            "Sie unterstützt Nachvollziehbarkeit über mehrere Prozessschritte hinweg.",
        ],
        "box": ("Rückverfolgung", "Wenn ein Preis oder eine Menge nicht plausibel ist, zeigt die Dokumentkette den Weg des Vorgangs."),
    },
    "Stammdaten/preismanagement/belegkette/index.html": {
        "title": "Belegkette",
        "intro": "Die Belegkette verbindet Belege über den Geschäftsprozess hinweg. Sie unterstützt Anwender dabei, den Weg von Anfrage, Angebot, Auftrag, Lieferung, Rechnung oder Beschaffung nachzuvollziehen.",
        "bullets": [
            "Prüfen Sie, welche Belege vor oder nach dem aktuellen Beleg entstanden sind.",
            "Nutzen Sie die Kette für Rückfragen, Reklamationen und Abstimmungen.",
            "Preis- und Mengenänderungen lassen sich entlang der Kette besser erklären.",
        ],
        "box": ("Prozesssicht", "Die Belegkette ist keine Preisstammdatenpflege, aber eine wichtige Kontrollsicht für preisrelevante Vorgänge."),
    },
    "Stammdaten/lagermanagement/index.html": {
        "title": "Lagermanagement",
        "intro": "Lagermanagement beschreibt die Stammdaten für Lager, Lagerplätze, Lagertypen und Lagermedien. Diese Struktur entscheidet, wie Bestände gefunden, bewegt, gezählt und ausgewertet werden können.",
        "flow": (
            "Lagerstruktur aufbauen",
            [
                "Lager definieren: Standorte oder organisatorische Lager klar trennen.",
                "Lagerplätze anlegen: Physische oder logische Plätze nachvollziehbar benennen.",
                "Typen festlegen: Lagerarten und Eigenschaften für Prozesse und Auswertungen setzen.",
                "Medien prüfen: Behälter, Paletten oder andere Lagermedien nur bei tatsächlicher Nutzung pflegen.",
                "Artikel testen: Lagergeführten Artikel buchen und Bestand auf Lager und Platz prüfen.",
            ],
        ),
        "bullets": [
            "Die Lagerstruktur sollte dem realen Prozess entsprechen, nicht nur der Gebäudeplanung.",
            "Zu detaillierte Lagerplätze erhöhen Pflegeaufwand.",
            "Zu grobe Lagerstrukturen erschweren Suche, Kommissionierung und Inventur.",
        ],
        "box": ("Bestandsqualität", "Gute Lagerstammdaten sind die Voraussetzung dafür, dass Bestände nicht nur rechnerisch, sondern auch praktisch auffindbar sind."),
    },
    "Stammdaten/lagermanagement/lager/index.html": {
        "title": "Lager",
        "intro": "Ein Lager beschreibt einen Bestandort oder eine organisatorische Bestandseinheit. Es ist die oberste Struktur für Lagerplätze, Bestände und Lagerbewegungen.",
        "bullets": [
            "Benennen Sie Lager so, dass Anwender den Standort eindeutig erkennen.",
            "Trennen Sie Lager nur, wenn Prozesse, Verantwortung oder Bestand wirklich getrennt sind.",
            "Standardlager sollten mit Artikel- und Belegprozessen zusammenpassen.",
        ],
        "box": ("Beispiel", "Ein Hauptlager, ein Sperrlager und ein Konsignationslager können unterschiedliche Prozesse haben und sollten deshalb getrennt geführt werden."),
    },
    "Stammdaten/lagermanagement/lagerplatz/index.html": {
        "title": "Lagerplatz",
        "intro": "Lagerplätze unterteilen ein Lager in konkrete Orte. Sie helfen beim Finden, Einlagern, Auslagern, Kommissionieren und Zählen von Beständen.",
        "flow": (
            "Lagerplatz sinnvoll anlegen",
            [
                "Ort bestimmen: Regal, Zone, Fach oder Fläche eindeutig benennen.",
                "Lager zuordnen: Platz immer dem richtigen Lager zuweisen.",
                "Eigenschaften prüfen: Sperre, Typ oder Kapazität erfassen, wenn relevant.",
                "Buchung testen: Artikel ein- oder umlagern und prüfen, ob der Platz korrekt erscheint.",
            ],
        ),
        "bullets": [
            "Lagerplatznamen sollten im Lageralltag lesbar und eindeutig sein.",
            "Zu kleinteilige Strukturen können die Arbeit verlangsamen.",
            "Für Inventur und Kommissionierung ist Konsistenz wichtiger als lange Beschreibungen.",
        ],
        "box": ("Praxisregel", "Ein Lagerplatz ist dann gut gepflegt, wenn ein Anwender den physischen Ort ohne Rückfragen findet."),
    },
    "Stammdaten/lagermanagement/lagertyp/index.html": {
        "title": "Lagertyp",
        "intro": "Lagertypen klassifizieren Lager oder Lagerplätze. Sie helfen, Eigenschaften wie Standardlager, Sperrlager, Versandlager oder Produktionslager fachlich zu unterscheiden.",
        "bullets": [
            "Nutzen Sie Lagertypen für Prozesslogik und Auswertungen.",
            "Die Typen sollten einfach und dauerhaft verständlich sein.",
            "Vermeiden Sie Typen, die nur einmalig oder uneinheitlich verwendet werden.",
        ],
        "box": ("Struktur", "Lagertypen sind Stammdaten für Ordnung und Steuerung. Sie ersetzen keine klare Lager- oder Lagerplatzbenennung."),
    },
    "Stammdaten/lagermanagement/lagermedium/index.html": {
        "title": "Lagermedium",
        "intro": "Lagermedien beschreiben Behälter, Paletten, Kisten oder andere Träger, auf denen Ware gelagert oder bewegt wird. Sie sind relevant, wenn Bestände nicht nur nach Ort, sondern auch nach Medium geführt werden.",
        "bullets": [
            "Pflegen Sie Lagermedien nur, wenn der Prozess sie wirklich unterscheidet.",
            "Bezeichnungen sollten im Lager eindeutig sein.",
            "Bei wiederverwendbaren Medien können Nummern oder Typen hilfreich sein.",
        ],
        "box": ("Einsatz", "Lagermedien sind besonders nützlich bei Verpackungseinheiten, Palettenlogik, Produktion oder detaillierter Rückverfolgung."),
    },
    "Stammdaten/mitarbeiter/index.html": {
        "title": "Mitarbeiter",
        "intro": "Mitarbeiterstammdaten beschreiben Personen, Gruppen, Provisionen und Urlaubsinformationen. Sie werden für Zuständigkeiten, Verkauf, Projekte, Zeiten, Freigaben und Auswertungen benötigt.",
        "flow": (
            "Mitarbeiter sauber anlegen",
            [
                "Person erfassen: Name, Kürzel, Kontaktdaten und Status pflegen.",
                "Gruppe zuordnen: Team, Abteilung oder Berechtigungsbezug sauber wählen.",
                "Rollen prüfen: Verkäufer, Projektmitarbeiter, Ressource oder Ansprechpartner unterscheiden.",
                "Provision und Urlaub ergänzen: Nur dort pflegen, wo diese Funktionen genutzt werden.",
                "Prozess testen: Mitarbeiter in Beleg, Projekt oder Aufgabe auswählen.",
            ],
        ),
        "bullets": [
            "Inaktive Mitarbeiter sollten nicht mehr als neue Zuständigkeit vorgeschlagen werden.",
            "Gruppen erleichtern Auswertungen und Filter.",
            "Mitarbeiterdaten können mit Projekten, CRM und Ressourcen zusammenhängen.",
        ],
        "box": ("Stammdaten und Rechte", "Mitarbeiterstammdaten sind nicht automatisch identisch mit Benutzerrechten. Prüfen Sie bei Bedarf beide Bereiche getrennt."),
    },
    "Stammdaten/mitarbeiter/mitarbeiter-stammdaten/index.html": {
        "title": "Mitarbeiter Stammdaten",
        "intro": "Die Mitarbeiter-Stammdaten enthalten die wichtigsten Angaben zur Person. Sie bilden die Grundlage für interne Zuordnung, Kommunikation, Zuständigkeit und Auswertungen.",
        "bullets": [
            "Name, Kürzel und Kontaktinformationen sollten eindeutig sein.",
            "Der Status entscheidet, ob ein Mitarbeiter aktiv verwendet werden kann.",
            "Gruppen und Rollen sollten zur tatsächlichen Organisation passen.",
            "Änderungen können Aufgaben, Projekte, Belege oder Auswertungen beeinflussen.",
        ],
        "box": ("Eindeutigkeit", "Ein Mitarbeiterdatensatz sollte so gepflegt sein, dass Anwender die richtige Person ohne Rückfrage auswählen können."),
    },
    "Stammdaten/mitarbeiter/mitarbeiter-gruppe/index.html": {
        "title": "Mitarbeiter Gruppe",
        "intro": "Mitarbeitergruppen fassen Personen fachlich zusammen. Sie können Teams, Abteilungen, Rollen oder Zuständigkeiten abbilden und helfen bei Filterung, Auswertung und Steuerung.",
        "bullets": [
            "Gruppen sollten nach realen Arbeitsabläufen benannt werden.",
            "Zu viele Gruppen machen die Pflege unübersichtlich.",
            "Gruppen sind besonders nützlich für Projekte, Aufgaben, Verkauf oder Service.",
        ],
        "box": ("Organisation", "Eine gute Gruppe beschreibt nicht nur, wo jemand sitzt, sondern wofür die Person im Prozess zuständig ist."),
    },
    "Stammdaten/mitarbeiter/mitarbeiter-provisionsgruppe/index.html": {
        "title": "Mitarbeiter Provisionsgruppe",
        "intro": "Provisionsgruppen ordnen Mitarbeiter für provisionsrelevante Auswertungen oder Berechnungen. Sie helfen, Vertriebsleistungen strukturiert zu bewerten.",
        "bullets": [
            "Provisionsgruppen müssen zur internen Vergütungslogik passen.",
            "Änderungen sollten zeitlich und fachlich abgestimmt werden.",
            "Prüfen Sie, ob Belege, Verkäuferzuordnung und Auswertungen die Gruppe korrekt verwenden.",
        ],
        "box": ("Sorgfalt", "Provisionsdaten können finanzielle Wirkung haben. Änderungen sollten nachvollziehbar und abgestimmt erfolgen."),
    },
    "Stammdaten/mitarbeiter/mitarbeiter-urlaub/index.html": {
        "title": "Mitarbeiter Urlaub",
        "intro": "Urlaubsinformationen zeigen Abwesenheiten oder urlaubsrelevante Daten eines Mitarbeiters. Sie unterstützen Planung, Projekte, Aufgaben und interne Abstimmung.",
        "bullets": [
            "Urlaub sollte zeitnah und vollständig gepflegt sein.",
            "Bei Projekt- oder Ressourcenplanung sind Abwesenheiten besonders wichtig.",
            "Vertretungen oder Hinweise sollten klar geregelt sein.",
        ],
        "box": ("Planung", "Urlaub ist mehr als eine Personalinformation, wenn Mitarbeiter in Projekten, Produktion oder Service eingeplant werden."),
    },
    "Stammdaten/textbloecke/index.html": {
        "title": "Textblöcke",
        "intro": "Textblöcke speichern wiederverwendbare Formulierungen für Belege, E-Mails, Artikeltexte oder interne Hinweise. Sie erhöhen Qualität und Geschwindigkeit, wenn häufig gleiche Texte benötigt werden.",
        "flow": (
            "Textblock vorbereiten",
            [
                "Zweck klären: Festlegen, wo der Text verwendet werden soll.",
                "Text sauber formulieren: Kurz, eindeutig und ohne veraltete Angaben schreiben.",
                "Gruppe zuordnen: Den Textblock in eine nachvollziehbare Struktur einsortieren.",
                "Variablen prüfen: Platzhalter nur verwenden, wenn sie im Zielkontext zuverlässig gefüllt werden.",
                "Ausgabe testen: Text in Beleg oder Maske einfügen und Darstellung prüfen.",
            ],
        ),
        "bullets": [
            "Textblöcke vermeiden uneinheitliche Formulierungen.",
            "Sie sind besonders nützlich für Lieferhinweise, Zahlungsinformationen, Standardtexte und Produktbeschreibungen.",
            "Regelmäßige Pflege verhindert veraltete Aussagen auf Belegen.",
        ],
        "box": ("Qualität", "Ein Textblock ist nur dann hilfreich, wenn Anwender ihn ohne Nacharbeit verwenden können."),
    },
    "Stammdaten/textbloecke/textbloecke-stammdaten/index.html": {
        "title": "Textblöcke Stammdaten",
        "intro": "Die Textblock-Stammdaten enthalten Bezeichnung, Inhalt, Status und grundlegende Steuerung des Textblocks. Hier wird festgelegt, was Anwender später einfügen können.",
        "bullets": [
            "Die Bezeichnung sollte den Inhalt sofort erkennen lassen.",
            "Texte sollten fachlich korrekt, aktuell und ohne unnötige Abkürzungen sein.",
            "Status und Einsatzbereich sollten verhindern, dass falsche Texte verwendet werden.",
        ],
        "box": ("Formulierung", "Schreiben Sie Textblöcke so, als würden sie direkt auf einem externen Dokument erscheinen. Das reduziert Nacharbeit und Fehler."),
    },
    "Stammdaten/textbloecke/textbloecke-gruppe/index.html": {
        "title": "Textblöcke Gruppe",
        "intro": "Textblockgruppen strukturieren wiederverwendbare Texte. Sie helfen Anwendern, den richtigen Text schnell zu finden und thematisch ähnliche Texte zusammenzuhalten.",
        "bullets": [
            "Gruppen sollten nach Einsatzbereich oder Thema benannt werden.",
            "Vermeiden Sie Gruppen mit nur einem selten genutzten Text.",
            "Eine klare Struktur erleichtert die Pflege veralteter Textblöcke.",
        ],
        "box": ("Auffindbarkeit", "Die beste Textvorlage hilft wenig, wenn Anwender sie nicht finden. Gruppen sind deshalb Teil der Bedienqualität."),
    },
    "Stammdaten/textbloecke/textbloecke-variablen/index.html": {
        "title": "Textblöcke Variablen",
        "intro": "Variablen sind Platzhalter, die beim Verwenden eines Textblocks durch echte Werte ersetzt werden können. Sie ermöglichen dynamische Texte, müssen aber sorgfältig getestet werden.",
        "flow": (
            "Variablen testen",
            [
                "Platzhalter wählen: Nur Variablen verwenden, die im Zielkontext verfügbar sind.",
                "Text lesen: Prüfen, ob der Satz mit eingesetztem Wert sprachlich funktioniert.",
                "Leerwerte beachten: Kontrollieren, was passiert, wenn ein Wert fehlt.",
                "Ausgabe prüfen: Text in einem echten Beleg oder Vorgang testen.",
            ],
        ),
        "bullets": [
            "Variablen sparen Pflegeaufwand, können aber falsche Texte erzeugen, wenn Werte fehlen.",
            "Technische Platzhalter sollten für Anwender nachvollziehbar dokumentiert sein.",
            "Nach Änderungen an Variablen ist ein Ausgabetest wichtig.",
        ],
        "box": ("Risiko", "Ein falsch getesteter Platzhalter kann auf vielen Belegen falsch erscheinen. Deshalb Variablen immer im Zielprozess prüfen."),
    },
    "Stammdaten/projektmanagement/index.html": {
        "title": "Projektmanagement",
        "intro": "Projektmanagement-Stammdaten verbinden Projekte, Aufgaben, Zeiten und Budgets. Sie helfen, Arbeit zu planen, Verantwortlichkeiten zu klären und Aufwand sowie Kosten nachvollziehbar zu machen.",
        "flow": (
            "Projektstruktur anlegen",
            [
                "Projekt definieren: Ziel, Partner, Verantwortliche und Zeitraum festlegen.",
                "Aufgaben gliedern: Arbeit in klare, prüfbare Schritte unterteilen.",
                "Zeiten erfassen: Aufwand regelmäßig und nachvollziehbar buchen.",
                "Budget prüfen: Planwerte, Istwerte und Abweichungen beobachten.",
                "Auswertung nutzen: Projektstatus anhand Aufgaben, Zeiten und Budget bewerten.",
            ],
        ),
        "bullets": [
            "Ein Projekt sollte nicht nur einen Namen, sondern einen klaren Zweck haben.",
            "Aufgaben machen Verantwortlichkeiten und Fortschritt sichtbar.",
            "Zeiten und Budgets sind nur belastbar, wenn sie regelmäßig gepflegt werden.",
        ],
        "box": ("Projektqualität", "Gute Projektdaten helfen nicht nur der Abrechnung, sondern auch der Steuerung laufender Arbeit."),
    },
    "Stammdaten/projektmanagement/projekt/index.html": {
        "title": "Projekt",
        "intro": "Das Projekt ist der zentrale Rahmen für Aufgaben, Zeiten, Budget und Auswertung. Es beschreibt, woran gearbeitet wird, für wen gearbeitet wird und wer verantwortlich ist.",
        "bullets": [
            "Projektname und Nummer sollten eindeutig sein.",
            "Partner, Verantwortliche und Zeitraum sollten früh gepflegt werden.",
            "Status und Priorität helfen bei Filterung und Steuerung.",
            "Projektstruktur sollte zur realen Arbeitsweise passen.",
        ],
        "box": ("Startpunkt", "Ein Projekt ist sinnvoll angelegt, wenn ein Anwender sofort erkennt, worum es geht, wer zuständig ist und wie der Fortschritt kontrolliert wird."),
    },
    "Stammdaten/projektmanagement/aufgabe/index.html": {
        "title": "Aufgabe",
        "intro": "Aufgaben unterteilen ein Projekt in konkrete Arbeitsschritte. Sie machen Verantwortlichkeiten, Termine, Status und offene Punkte sichtbar.",
        "flow": (
            "Aufgabe anlegen",
            [
                "Titel formulieren: Kurz und handlungsorientiert beschreiben.",
                "Verantwortung zuordnen: Mitarbeiter oder Gruppe eindeutig setzen.",
                "Termin und Priorität pflegen: Fälligkeit realistisch festlegen.",
                "Status aktualisieren: Fortschritt während der Bearbeitung nachführen.",
                "Abschluss prüfen: Ergebnis oder Folgeaufgabe dokumentieren.",
            ],
        ),
        "bullets": [
            "Eine gute Aufgabe beschreibt eine konkrete Handlung, nicht nur ein Thema.",
            "Offene Aufgaben sollten regelmäßig geprüft werden.",
            "Aufgaben können mit Zeiten, Projektbudget und Kommunikation zusammenhängen.",
        ],
        "box": ("Verbindlichkeit", "Aufgaben sind nur hilfreich, wenn Zuständigkeit und nächster Schritt klar sind."),
    },
    "Stammdaten/projektmanagement/projektzeiten/index.html": {
        "title": "Projektzeiten",
        "intro": "Projektzeiten dokumentieren den Aufwand, der für ein Projekt oder eine Aufgabe entstanden ist. Sie sind wichtig für Nachkalkulation, Abrechnung, Kapazitätsplanung und Projektsteuerung.",
        "bullets": [
            "Zeiten sollten zeitnah und mit klarer Tätigkeit erfasst werden.",
            "Die Zuordnung zu Projekt und Aufgabe muss stimmen.",
            "Regelmäßige Auswertung zeigt Abweichungen früh.",
            "Abrechenbare und interne Zeiten sollten sauber unterschieden werden.",
        ],
        "box": ("Nachkalkulation", "Projektzeiten sind nur dann belastbar, wenn sie vollständig und nachvollziehbar gepflegt werden."),
    },
    "Stammdaten/projektmanagement/projektbudget/index.html": {
        "title": "Projektbudget",
        "intro": "Das Projektbudget stellt geplante und tatsächliche Werte gegenüber. Es hilft, Kosten, Erlöse, Aufwand und Abweichungen im Blick zu behalten.",
        "bullets": [
            "Planwerte sollten vor Projektstart realistisch festgelegt werden.",
            "Istwerte hängen von Zeiten, Belegen und Kosten ab.",
            "Abweichungen sollten früh geprüft und begründet werden.",
            "Budgetdaten unterstützen Entscheidungen zu Umfang, Priorität und Abrechnung.",
        ],
        "box": ("Steuerung", "Budgetpflege ist kein reiner Abschlussbericht. Sie hilft während des Projekts, rechtzeitig zu reagieren."),
    },
    "Stammdaten/ressourcen/index.html": {
        "title": "Ressourcen",
        "intro": "Ressourcen beschreiben Kapazitäten, die in Projekten, Produktion oder Planung benötigt werden. Das können Mitarbeiter, Maschinen, Arbeitsplätze, Werkzeuge oder andere begrenzte Mittel sein.",
        "flow": (
            "Ressource vorbereiten",
            [
                "Art bestimmen: Mitarbeiter, Maschine, Arbeitsplatz oder sonstige Kapazität unterscheiden.",
                "Verfügbarkeit prüfen: Zeiten, Kalender oder Kapazitätsgrenzen realistisch pflegen.",
                "Kosten bewerten: Stundensätze oder Kostensätze ergänzen, wenn sie für Kalkulation genutzt werden.",
                "Zuordnung testen: Ressource in Projekt, Aufgabe oder Produktionsschritt verwenden.",
            ],
        ),
        "bullets": [
            "Ressourcen sind nur sinnvoll, wenn sie geplant oder ausgewertet werden.",
            "Zu viele Detailressourcen erhöhen den Pflegeaufwand.",
            "Für Produktion und Projektplanung sind realistische Verfügbarkeiten entscheidend.",
        ],
        "box": ("Kapazität", "Eine Ressource beschreibt nicht nur einen Namen, sondern eine nutzbare Kapazität im Prozess."),
    },
    "Stammdaten/anhaenge/index.html": {
        "title": "Anhänge",
        "intro": "Anhänge verbinden Dateien mit Stammdaten und Vorgängen. Sie sorgen dafür, dass wichtige Unterlagen direkt dort verfügbar sind, wo Anwender sie benötigen.",
        "flow": (
            "Anhänge sinnvoll pflegen",
            [
                "Zweck klären: Datei nur anhängen, wenn sie im Kontext wirklich benötigt wird.",
                "Datei benennen: Inhalt, Stand und Bezug erkennbar machen.",
                "Ablageort wählen: Datei am richtigen Artikel, Partner, Projekt oder Vorgang speichern.",
                "Aktualität prüfen: Veraltete Dateien entfernen oder klar kennzeichnen.",
                "Zugriff kontrollieren: Sensible Dokumente nur passenden Benutzern zugänglich machen.",
            ],
        ),
        "bullets": [
            "Anhänge vermeiden Suchaufwand in externen Ordnern.",
            "Dateinamen sollten ohne Öffnen der Datei verständlich sein.",
            "Versionen und Gültigkeiten sind besonders bei Zertifikaten, Verträgen und Zeichnungen wichtig.",
        ],
        "box": ("Kontext", "Ein Anhang ist am wertvollsten, wenn er direkt dort liegt, wo die fachliche Frage entsteht."),
    },
    "Stammdaten/anhaenge/anhaenge-stammdaten/index.html": {
        "title": "Anhänge Stammdaten",
        "intro": "Die Anhänge-Stammdaten beschreiben Grundinformationen zum Anhang. Dazu gehören Bezeichnung, Bezug, Status und gegebenenfalls weitere strukturierte Angaben.",
        "bullets": [
            "Die Bezeichnung sollte Inhalt und Zweck klar erkennen lassen.",
            "Der Bezug zum richtigen Stammdatensatz ist entscheidend.",
            "Status, Kategorie oder Gültigkeit helfen bei der späteren Pflege.",
        ],
        "box": ("Auffindbarkeit", "Gute Anhang-Stammdaten sorgen dafür, dass Dateien nicht nur gespeichert, sondern später auch gefunden und verstanden werden."),
    },
    "Stammdaten/anhaenge/anhaenge-dateien/index.html": {
        "title": "Anhänge Dateien",
        "intro": "Dateien sind der eigentliche Inhalt eines Anhangs. Sie können Zeichnungen, Verträge, Zertifikate, Bilder, Datenblätter, Nachweise oder andere Unterlagen enthalten.",
        "bullets": [
            "Prüfen Sie, ob die Datei vollständig und lesbar ist.",
            "Verwenden Sie klare Dateinamen und vermeiden Sie nichtssagende Scans.",
            "Bei aktualisierten Dokumenten sollte die alte Version entfernt oder kenntlich gemacht werden.",
            "Sensible Dateien benötigen passende Rechte und kontrollierte Ablage.",
        ],
        "box": ("Versionen", "Wenn mehrere Versionen einer Datei im Umlauf sind, muss der gültige Stand eindeutig erkennbar sein."),
    },
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def markdown_for(page: dict[str, object]) -> str:
    title = str(page["title"])
    parts = [f"# {title}", "", str(page["intro"]).strip()]
    flow = page.get("flow")
    if flow:
        flow_title, flow_items = flow
        parts.extend(["", f":::flow {flow_title}"])
        parts.extend(f"- {item}" for item in flow_items)
        parts.append(":::")
    bullets = page.get("bullets") or []
    if bullets:
        parts.extend(["", "## Worauf Sie achten sollten"])
        parts.extend(f"- {item}" for item in bullets)
    box = page.get("box")
    if box:
        box_title, box_body = box
        parts.extend(["", f":::box {box_title}", str(box_body).strip(), ":::"])
    return "\n".join(parts).strip() + "\n"


def meta(markdown: str, title: str) -> str:
    clean = re.sub(r"^#{1,6}\s*", "", markdown, flags=re.M)
    clean = re.sub(r"^:::(?:flow|box)\s+.*$", "", clean, flags=re.M)
    clean = clean.replace(":::", "")
    clean = re.sub(r"^[\s>*-]*[-*]\s*", "", clean, flags=re.M)
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        clean = f"{title} in der X-ERP Hilfe."
    return clean[:297].rstrip(" ,.;") + ("..." if len(clean) > 297 else "")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-stammdaten-quality-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False, keep_links=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    by_path = {
        cell_text(ws.cell(row, headers["URL_PATH"]).value).replace("\\", "/"): row
        for row in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row, headers["URL_PATH"]).value)
    }

    changed = 0
    missing: list[str] = []
    skipped_article = 0
    for path, page in PAGES.items():
        if path.startswith("Stammdaten/artikel/"):
            skipped_article += 1
            continue
        row = by_path.get(path)
        if not row:
            missing.append(path)
            continue
        title = str(page["title"])
        markdown = markdown_for(page)
        ws.cell(row, headers["Beschreibung"]).value = markdown
        for column, value in {
            "H1": title,
            "NAV_TITLE": title,
            "TITLE": f"{title} | X-ERP ERP Hilfe",
            "META_DESCRIPTION": meta(markdown, title),
            "PRIMARY_KEYWORD": f"{title} X-ERP",
        }.items():
            if column in headers:
                ws.cell(row, headers[column]).value = value
        changed += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"changed={changed}")
    print(f"skipped_article={skipped_article}")
    print(f"missing={len(missing)}")
    for path in missing:
        print(f"MISSING {path}")


if __name__ == "__main__":
    main()
