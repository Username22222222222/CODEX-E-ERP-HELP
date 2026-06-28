from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


SETUP_LEVELS = {
    "Setup": 0,
    "Server": 1,
    "Hardware": 2,  # first Hardware under Server; client hardware is handled below
    "Hardwarevorbereitung": 3,
    "Hardware-Empfehlung": 4,
    "Certificate": 4,
    "SQL Server 2025": 4,
    "Backup-System": 4,
    "Firewall": 4,
    "Antivirenprogramm": 4,
    "Servereinrichtung": 3,
    "Alle Updates installieren": 4,
    "IIS aktivieren": 4,
    ".NET Hosting Bundle installieren": 4,
    "MS SQL Server 2025 installieren": 4,
    "Setup starten": 4,
    "Instanz (X) installieren": 5,
    "Collation auswählen": 5,
    "Filestreaming aktivieren": 5,
    "XLD-Creator installieren": 4,
    "XLD-Datenbank installieren": 4,
    "X-ERP installieren": 4,
    "X-ERP-Configurator installieren": 4,
    "X-ERP konfigurieren": 4,
    "SQL Verbindung eintragen": 4,
    "Ordner auswählen": 5,
    "Lizenzdatei auswählen": 5,
    "Company-Creator installieren": 4,
    "Company-Datenbank installieren": 4,
    "Company aktivieren": 4,
    "X-ERP-Configurator öffnen": 4,
    "Certificate zuweisen": 5,
    "Binding hinzufügen": 5,
    "IIS (Internet Information Services)": 2,
    "SQL Server": 2,
    "Wartungsarbeiten": 2,
    "XLD (X-ERP Landscape Directory)": 2,
    "Backup-Service": 2,
    "Email-Service": 2,
    "Zertifikat": 2,
    "Updates": 2,
    "Client": 1,
    "Browser": 2,
}


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def set_level(ws, row: int, level: int, templates: dict[int, int]) -> None:
    if level in templates:
        copy_row_style(ws, templates[level], row)
    ws.row_dimensions[row].outlineLevel = level
    ws.row_dimensions[row].hidden = False
    ws.row_dimensions[row].collapsed = False


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-restore-setup-section-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    setup_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Setup")
    ansichten_row = next(row for row in range(setup_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")

    # Use current, verified visual templates from the workbook.
    faq_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")
    templates = {
        0: faq_row,
        1: faq_row + 1,
        2: 1116,  # view-style blue, indent 4
        3: 1117,  # register-style blue, indent 6
        4: 1118,  # field-style blue, indent 8
        5: 944 if setup_row <= 944 < ansichten_row else 1118,
    }

    seen_client = False
    seen_server = False
    fixed = []
    for row in range(setup_row, ansichten_row):
        title = ws.cell(row, thema_col).value
        if not title:
            continue

        if title == "Server":
            seen_server = True
            seen_client = False
        elif title == "Client":
            seen_client = True

        if title == "Hardware" and seen_client:
            level = 2
        elif title in SETUP_LEVELS:
            level = SETUP_LEVELS[title]
        else:
            # Keep unknown setup children under the current branch.
            level = max(ws.row_dimensions[row].outlineLevel, 2)

        set_level(ws, row, level, templates)
        if ws.cell(row, content_col).value != "HelpPage":
            ws.cell(row, content_col).value = "HelpPage"
        if "TOC_LEVEL" in col:
            ws.cell(row, col["TOC_LEVEL"]).value = level
        if "TOC_PARENT" in col:
            if level == 0:
                ws.cell(row, col["TOC_PARENT"]).value = None
            elif level == 1:
                ws.cell(row, col["TOC_PARENT"]).value = "Setup"
            else:
                ws.cell(row, col["TOC_PARENT"]).value = "setup"
        fixed.append((row, title, level))

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    if ws.sheet_view:
        ws.sheet_view.showOutlineSymbols = True

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"setup_row={setup_row}")
    print(f"ansichten_row={ansichten_row}")
    print(f"fixed={len(fixed)}")
    for item in fixed:
        print(f"{item[0]}: {item[1]} -> level {item[2]}")


if __name__ == "__main__":
    main()
