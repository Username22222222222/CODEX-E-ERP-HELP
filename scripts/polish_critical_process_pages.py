from __future__ import annotations

import datetime as dt
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"


MAIN_DESCRIPTIONS = {
    "HelpdeskEdit": "HelpdeskEdit ist die zentrale Ticketansicht fuer Service- und Supportfaelle. Anwender erfassen Partner, Ansprechpartner, Datum, Supporter, Status, Kategorie, Thema, Prioritaet und Bearbeitungszeiten. Die Register Dialog, Anhaenge und Extra-Felder dokumentieren Rueckfragen, Antworten, Dateien und Zusatzinformationen. Wichtig: Kategorie und Thema sind Pflichtangaben; die Hilfe muss den korrekten Auswahlweg klar erklaeren.",
    "HelpdeskDialogEdit": "HelpdeskDialogEdit erfasst einzelne Nachrichten, Rueckfragen und Antworten zu einem Helpdesk-Ticket. Die Ansicht verbindet den Dialog mit Ticket, Partner, Ansprechpartner und Supporter und dokumentiert Frage, Antwort, Antwortzeit und Bearbeitungsdauer. So entsteht eine nachvollziehbare Kommunikationshistorie zum Servicefall.",
    "DocEdit": "DocEdit ist die zentrale Belegansicht fuer Angebote, Auftraege, Lieferscheine, Rechnungen und weitere Dokumente. Die Maske verbindet Partner, Belegart, Adressen, Positionen, Zahlungs- und Lieferbedingungen, Archivierung und Folgebelege. Die Hilfe muss deutlich machen, dass offene Posten, Buchungen und Lagerbewegungen erst durch die passenden Folge-, Archivierungs- und Buchungsprozesse entstehen.",
    "DocEdit_ArchiveExternalFile": "Dieses Register verknuepft externe Belegdateien mit dem X-ERP-Beleg. Anwender archivieren hier Dokumente, die nicht direkt im ERP erzeugt wurden, etwa Eingangsrechnungen oder externe Nachweise. Pflichtangaben wie externe Belegnummern muessen vollstaendig sein, bevor ein Beleg revisionssicher weiterverarbeitet werden kann.",
    "FinancePostingDraftEdit": "FinancePostingDraftEdit zeigt Buchungsentwuerfe, bevor daraus gueltige Finanzbuchungen werden. Anwender pruefen Buchungsvorlage, Belegdatum, Periode, Belegnummer, Dok-Nummer, Betreff, Waehrung und Positionen. Ein Entwurf ist erst buchbar, wenn Soll und Haben fachlich ausgeglichen sind und alle Pflichtangaben stimmen.",
    "FinancePostingEdit": "FinancePostingEdit zeigt eine gueltige Finanzbuchung mit Buchungskopf und Positionen. Die Ansicht macht nachvollziehbar, welche Konten, Steuerwerte, Kostenstellen, Projekte und Gegenkonten gebucht wurden. Sie dient der Kontrolle, Nachverfolgung, Auswertung und gegebenenfalls Stornopruefung.",
    "FinanceOpenItemInternalReconciliationEdit": "Diese Ansicht dient dem internen Ausgleich offener Posten. Anwender ordnen offene Forderungen oder Verbindlichkeiten passenden Gegenpositionen zu, damit Salden korrekt ausgeglichen und Zahlungs- bzw. Mahnprozesse nachvollziehbar bleiben.",
    "FinancePaymentWizard": "Der Zahlungsassistent bereitet Zahlungen auf Basis offener Posten vor. Anwender waehlen Zahlungsausgaenge, Lieferanten, Skontoberuecksichtigung und Firmenbankkonto. Das Ergebnis ist ein kontrollierter Zahlungsvorschlag, der vor Ausfuehrung fachlich geprueft werden muss.",
    "WarehouseTransferEdit": "WarehouseTransferEdit erfasst Umlagerungen zwischen Lagerplaetzen. Quelle, Ziel, Mitarbeiter und Positionen legen fest, welche Artikelmenge von welchem Lagerplatz an welchen Zielplatz bewegt wird. Die Hilfe muss erklaeren, dass Historie und Bestand erst nach gespeicherten bzw. gebuchten Lagerbewegungen aussagekraeftig sind.",
    "WarehouseInventoryCountingEdit": "WarehouseInventoryCountingEdit dokumentiert eine Inventurzaehlung. Inventurleiter, Lager, Erstellzeitpunkt und Positionen bilden die Grundlage fuer den Abgleich zwischen gezahltem und systemischem Bestand. Die Positionen zeigen erst dann Werte, wenn eine Zaehlliste erzeugt oder bearbeitet wurde.",
    "WarehouseInventoryRevaluationEdit": "WarehouseInventoryRevaluationEdit dokumentiert eine Lagerbewertung oder Wertkorrektur. Anwender pruefen Bewertungsnummer, Verantwortlichen, Datum und Positionen, bevor Bestandswerte angepasst werden. Die Ansicht ist fuer nachvollziehbare Lagerbewertung und Finanzbezug wichtig.",
    "ProductionEdit": "ProductionEdit steuert einen Produktionsauftrag vom geplanten Artikel bis zur Rueckmeldung. Die Ansicht zeigt Betreff, Produktionsleiter, Planung, Laufzeiten, Start und Ende, Auftragsbezug, Liefertermin und Produktionslager. Positionen, BDE-Mengen, Zeiten und Kosten werden erst aussagekraeftig, wenn Stuecklisten, Rueckmeldungen und Kostenereignisse vorliegen.",
    "ProductionDataAcquisitionEdit": "ProductionDataAcquisitionEdit erfasst BDE-Rueckmeldungen zur Produktion. Anwender dokumentieren Aktion, Grundcode, Produktionsposition, Mitarbeiter, Arbeitsplatz, Ressource, Mengen, Ausschuss, Fortschritt, Start, Ende und Dauer. Diese Daten bilden die Grundlage fuer Produktionsfortschritt, Nachkalkulation und Auswertungen.",
    "WebshopSettingEdit": "WebshopSettingEdit steuert die zentralen Einstellungen des X-ERP.webshop. Dazu gehoeren Shopname, Kontakt, Impressum, Datenschutz, Sprache, Standort, Zahlungs- und Lieferbedingungen, Kundenanlage, Bewertungsfunktion, Gastmodus, Lageranzeige, E-Mail-Texte, Reportvorlage und Bild. Platzhaltertexte muessen vor finalen Screenshots durch fachlich plausible Demo-Inhalte ersetzt werden.",
    "WebshopActivityLogEdit": "WebshopActivityLogEdit zeigt protokollierte Ereignisse aus dem Webshop. Webshop, Aktion, Details und Erstellzeitpunkt helfen nachzuvollziehen, was im Portal passiert ist. Die Ansicht ist vor allem fuer Diagnose, Support und Prozesskontrolle relevant.",
    "CrmEdit": "CrmEdit buendelt die vertriebliche Nachverfolgung eines Kunden- oder Lieferantenkontakts. Partner, Ansprechpartner, Telefon, E-Mail, Mitarbeiter, Wiedervorlage, Status, Betreff, Opportunity und Abschlusswahrscheinlichkeit zeigen, welche Chance verfolgt wird und welcher naechste Schritt ansteht. Fuer finale Hilfe-Screenshots sollte der Demo-Text kein Lorem ipsum enthalten.",
}


FIELD_DESCRIPTIONS = {
    "Kategorie": "Ordnet den Vorgang einer fachlichen Kategorie zu. Im Helpdesk steuert sie, welche Themen zur Auswahl stehen und wie Tickets ausgewertet werden.",
    "Thema": "Praezisiert die Kategorie eines Tickets oder Vorgangs. Es hilft Supportern, den Inhalt schneller einzuordnen und passende Folgeaktionen zu waehlen.",
    "Supporter": "Legt fest, welcher Mitarbeiter den Vorgang betreut oder verantwortlich weiterbearbeitet.",
    "Prioritaet": "Legt die Dringlichkeit fest. Hohe Prioritaeten muessen im Tagesgeschaeft schneller sichtbar und bearbeitbar sein.",
    "Prio": "Legt die Dringlichkeit fest. Hohe Prioritaeten muessen im Tagesgeschaeft schneller sichtbar und bearbeitbar sein.",
    "Partner": "Verknuepft den Vorgang mit dem richtigen Geschaeftspartner. Dadurch passen Historie, Kommunikation, Belege und Auswertungen zusammen.",
    "Ansprechpartner": "Verknuepft den Vorgang mit der konkreten Person beim Partner, damit Rueckfragen und Kommunikation eindeutig adressiert sind.",
    "Helpdesk-Datum": "Dokumentiert den Zeitpunkt der Ticketanlage oder Bearbeitung und ist Grundlage fuer Fristen, Reaktionszeiten und Auswertungen.",
    "Geschätzte Bearbeitungszeit": "Schaetzt den erwarteten Aufwand eines Tickets oder Vorgangs und unterstuetzt Planung und Priorisierung.",
    "Bearbeitungszeit": "Dokumentiert die tatsaechliche Bearbeitungsdauer und kann fuer Auswertung, Nachkalkulation oder interne Steuerung genutzt werden.",
    "Rechnungsadresse": "Legt fest, an welche Adresse ein Beleg kaufmaennisch gerichtet ist. Diese Adresse ist fuer Rechnung, Archivierung und Buchhaltung relevant.",
    "Lieferadresse": "Legt fest, wohin Ware oder Leistung geliefert wird. Diese Adresse beeinflusst Versand, Lieferung und logistische Folgeprozesse.",
    "% Dok Rabatt": "Enthaelt den dokumentbezogenen Rabatt in Prozent. Der Rabatt wirkt auf die Belegsumme und muss fachlich vom Positionsrabatt unterschieden werden.",
    "Zahlungskonditionen": "Bestimmt Zahlungsziel, Skonto und weitere Zahlungsbedingungen des Belegs.",
    "Lieferkondition": "Bestimmt die Lieferbedingung fuer den Beleg und beeinflusst Versand, Logistik und Kundenkommunikation.",
    "Buchungsvorlage": "Waehlt eine Vorlage fuer wiederkehrende Buchungslogik und beschleunigt korrekte Kontierung.",
    "Buchungsperiode": "Ordnet die Buchung einer Periode zu. Diese Zuordnung ist fuer Abschluss, Auswertung und steuerliche Abgrenzung entscheidend.",
    "Soll": "Betrag oder Kontenseite im Soll. Eine Buchung ist nur stimmig, wenn Soll und Haben fachlich ausgeglichen sind.",
    "Haben": "Betrag oder Kontenseite im Haben. Eine Buchung ist nur stimmig, wenn Soll und Haben fachlich ausgeglichen sind.",
    "Steuerschlüssel": "Legt fest, welcher Steuersatz und welche Steuerlogik fuer die Buchungsposition verwendet wird.",
    "Kostenstelle": "Ordnet Kosten oder Erloese einer organisatorischen Einheit zu und ermoeglicht interne Auswertungen.",
    "Kostenträger": "Ordnet Kosten oder Erloese einem Auftrag, Projekt oder Kostentraeger zu.",
    "Produktionsleiter": "Legt fest, wer fuer Planung, Durchfuehrung und Kontrolle des Produktionsauftrags verantwortlich ist.",
    "Produktionsstart": "Zeigt oder setzt den geplanten bzw. tatsaechlichen Start der Produktion.",
    "Produktionsende": "Zeigt oder setzt das geplante bzw. tatsaechliche Ende der Produktion.",
    "Produktionslager": "Bestimmt das Lager, in dem Produktionsentnahmen oder Fertigmeldungen wirksam werden.",
    "Quantity": "Erfasste Menge der BDE-Rueckmeldung. Sie beeinflusst Produktionsfortschritt und Auswertung.",
    "Good Quantity": "Gutmenge der Rueckmeldung. Sie zeigt, wie viele Einheiten verwertbar produziert wurden.",
    "Scrap Quantity": "Ausschussmenge der Rueckmeldung. Sie dokumentiert nicht verwertbare Mengen fuer Analyse und Nachkalkulation.",
    "Actual Start": "Tatsaechlicher Startzeitpunkt der Rueckmeldung oder Arbeit.",
    "Actual End": "Tatsaechlicher Endzeitpunkt der Rueckmeldung oder Arbeit.",
    "Source Bin Location": "Quell-Lagerplatz der Umlagerung. Von hier wird Bestand abgebucht.",
    "Destination Bin Location": "Ziel-Lagerplatz der Umlagerung. Hier wird Bestand zugebucht.",
    "Inventory Manager": "Verantwortliche Person fuer Inventur oder Bewertung.",
    "Inventory Counting Number": "Eindeutige Nummer der Inventurzaehlung zur Nachverfolgung.",
    "Inventory Revaluation Number": "Eindeutige Nummer der Lagerbewertung oder Wertkorrektur.",
    "Action": "Beschreibt die im Webshop, in der BDE oder im Prozess ausgefuehrte Aktion.",
    "Details": "Enthaelt technische oder fachliche Zusatzinformationen zum protokollierten Ereignis.",
    "Opportunity": "Beschreibt die vertriebliche Chance oder den Anlass der CRM-Aktivitaet.",
    "Abschlusswahrscheinlichkeit %": "Bewertet die erwartete Erfolgswahrscheinlichkeit einer Vertriebschance in Prozent.",
    "Wiedervorlage": "Legt fest, wann der Vorgang erneut bearbeitet oder nachgefasst werden soll.",
}


META = {
    "HelpdeskEdit": "HelpdeskEdit in X-ERP: Tickets mit Partner, Kategorie, Thema, Prioritaet, Dialog, Anhaengen und Bearbeitungszeiten im ERP bearbeiten.",
    "DocEdit": "DocEdit in X-ERP: Belege mit Partner, Adressen, Positionen, Zahlungsbedingungen, Lieferbedingungen, Archivierung und Folgeprozessen erklaert.",
    "FinancePostingDraftEdit": "FinancePostingDraftEdit in X-ERP: Buchungsentwuerfe, Perioden, Konten, Steuern, Soll/Haben und Buchungsvoraussetzungen verstehen.",
    "WarehouseTransferEdit": "WarehouseTransferEdit in X-ERP: Umlagerungen zwischen Lagerplaetzen mit Quelle, Ziel, Mitarbeiter und Positionen nachvollziehbar erfassen.",
    "ProductionEdit": "ProductionEdit in X-ERP: Produktionsauftrag, Planung, Lager, BDE-Mengen, Zeiten, Positionen und Kosten im ERP verstehen.",
    "WebshopSettingEdit": "WebshopSettingEdit in X-ERP: Webshop-Stammdaten, Rechtstexte, Zahlungen, Lieferung, Kundenanlage, E-Mails und Lageranzeige konfigurieren.",
    "CrmEdit": "CrmEdit in X-ERP: CRM-Vorgaenge mit Partner, Wiedervorlage, Status, Opportunity, Abschlusswahrscheinlichkeit und Aktivitaeten steuern.",
}


def main() -> None:
    ARCHIVE.mkdir(exist_ok=True)
    shutil.copy2(WORKBOOK, ARCHIVE / f"X-ERP-HELP-before-critical-process-polish-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx")
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}
    changed = 0
    current_view = None
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, col["Thema"]).value
        if not name:
            continue
        content_type = ws.cell(row, col["CONTENT_TYPE"]).value if "CONTENT_TYPE" in col else None
        if content_type in ("View", "Wizard"):
            current_view = str(name)
        if name == "Belegrabatt (%)":
            ws.cell(row, col["Thema"]).value = "% Dok Rabatt"
            name = "% Dok Rabatt"
            changed += 1
        if name in MAIN_DESCRIPTIONS:
            ws.cell(row, col["Beschreibung"]).value = MAIN_DESCRIPTIONS[name]
            if name in META and "META_DESCRIPTION" in col:
                ws.cell(row, col["META_DESCRIPTION"]).value = META[name]
            changed += 1
        elif name in FIELD_DESCRIPTIONS and current_view in {
            "HelpdeskEdit", "HelpdeskDialogEdit", "DocEdit", "DocEdit_Details", "DocEdit_BillingAddress",
            "DocEdit_DeliveryAddress", "FinancePostingDraftEdit", "FinancePostingEdit",
            "FinancePostingDraftEdit_Position", "FinancePostingEdit_Position", "WarehouseTransferEdit",
            "WarehouseInventoryCountingEdit", "WarehouseInventoryRevaluationEdit", "ProductionEdit",
            "ProductionEdit_Overview", "ProductionDataAcquisitionEdit", "WebshopActivityLogEdit", "CrmEdit"
        }:
            ws.cell(row, col["Beschreibung"]).value = FIELD_DESCRIPTIONS[name]
            changed += 1
    wb.save(WORKBOOK)
    print(f"changed={changed}")


if __name__ == "__main__":
    main()
