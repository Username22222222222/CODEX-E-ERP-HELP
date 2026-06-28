from __future__ import annotations

import html
import json
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
TARGET_ROOT = HELP_ROOT / "Ansichten"
BASE_URL = "https://x-erp.de/de/help/Ansichten"


@dataclass
class Node:
    row: int
    level: int
    topic: str
    original: str = ""
    screenshot: str = ""
    field_name: str = ""
    description: str = ""
    page_id: str = ""
    slug: str = ""
    title: str = ""
    meta_description: str = ""
    h1: str = ""
    keyword: str = ""
    content_type: str = ""
    breadcrumb: str = ""
    toc_parent: str = ""
    toc_level: int = 0
    toc_order: str = ""
    unique_key: str = ""
    image_alt: str = ""
    image_caption: str = ""
    image_status: str = ""
    url_part: str = ""
    fs_dir: Path | None = None
    public_url: str = ""
    parent: "Node | None" = None
    children: list["Node"] = field(default_factory=list)


def get_cell(row: Any, key: str) -> str:
    value = row.get(key)
    if value is None:
        return ""
    return str(value).strip()


def slugify(value: str) -> str:
    replacements = {
        "Ä": "Ae",
        "Ö": "Oe",
        "Ü": "Ue",
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "seite"


def text(value: str) -> str:
    return html.escape(value or "", quote=True)


TOKEN_TRANSLATIONS = {
    "Account": "Konto",
    "Accessory": "Zubehör",
    "Address": "Adresse",
    "Admin": "Administration",
    "Appointment": "Termin",
    "Article": "Artikel",
    "Attachment": "Anhang",
    "Bank": "Bank",
    "Category": "Kategorie",
    "Company": "Firma",
    "Component": "Komponente",
    "Contact": "Kontakt",
    "Country": "Land",
    "Currency": "Währung",
    "Customer": "Kunde",
    "Date": "Datum",
    "Delivery": "Lieferung",
    "Document": "Dokument",
    "Doc": "Dok",
    "Edit": "bearbeiten",
    "Employee": "Mitarbeiter",
    "Field": "Feld",
    "Finance": "Finanzen",
    "Group": "Gruppe",
    "History": "Historie",
    "Invoice": "Rechnung",
    "Language": "Sprache",
    "List": "Liste",
    "Location": "Standort",
    "Macro": "Makro",
    "Name": "Name",
    "Order": "Auftrag",
    "Partner": "Partner",
    "Picture": "Bild",
    "Price": "Preis",
    "Procurement": "Beschaffung",
    "Product": "Produkt",
    "Production": "Produktion",
    "Project": "Projekt",
    "Purchase": "Einkauf",
    "Quantity": "Menge",
    "Resource": "Ressource",
    "Sales": "Verkauf",
    "Setting": "Einstellung",
    "Set": "Set",
    "Status": "Status",
    "Step": "Schritt",
    "Supplier": "Lieferant",
    "Text": "Text",
    "Time": "Zeit",
    "Translation": "Übersetzung",
    "User": "Benutzer",
    "Valid": "gültig",
    "Warehouse": "Lager",
    "Wizard": "Assistent",
    "Zip": "PLZ",
}

TAB_PURPOSES = {
    "Übersicht": "fasst die wichtigsten Stammdaten zusammen und dient als erster Kontrollpunkt für Suche, Identifikation und schnelle Bearbeitung.",
    "Details": "enthält ergänzende Fachinformationen, die für Auswertungen, Druck, Shop-Anbindung oder weitere Prozesse benötigt werden.",
    "Verkauf": "bündelt Preise, Verkaufskennzeichen, Finanzgruppen und die Verkaufspreislisten, die für Angebote, Aufträge und Rechnungen relevant sind.",
    "Beschaffung": "beschreibt Einkaufs- und Lieferantendaten, damit Beschaffungspreise, Finanzgruppen und Einkaufsprozesse nachvollziehbar bleiben.",
    "Text": "enthält beschreibende Texte für interne Hinweise, Bestellungen, Verkaufstexte oder weitere Dokumentausgaben.",
    "Bild": "verwaltet Bildinformationen, damit Anwender den Datensatz visuell erkennen und in passenden Ausgaben verwenden können.",
    "Lagerung": "zeigt lagerrelevante Werte wie Bestand, Mindestbestand, Lagerort und Bewegungen, die für Verfügbarkeit und Disposition wichtig sind.",
    "Set": "beschreibt Set-Komponenten und deren Mengen, damit zusammengehörige Artikel korrekt angeboten, aufgelöst oder verarbeitet werden.",
    "Makro": "fasst Makropositionen zusammen, mit denen wiederkehrende Artikel- oder Positionsstrukturen schneller angelegt werden.",
    "Produktion": "zeigt Produktionskomponenten, Arbeitsschritte und Abhängigkeiten, die für Fertigung und Materialbedarf relevant sind.",
    "Zubehör": "ordnet ergänzende Komponenten zu, die bei Verkauf, Einkauf oder Konfiguration zusätzlich vorgeschlagen werden können.",
}

FIELD_PURPOSES = [
    (("Matchcode",), "Der Matchcode ist der kurze Such- und Ordnungsbegriff. Er hilft Anwendern, den Datensatz schnell wiederzufinden."),
    (("Name", "Bezeichnung"), "Die Bezeichnung beschreibt den Datensatz für Listen, Auswertungen und die tägliche Suche."),
    (("Artikel", "Article"), "Das Feld verweist auf den zugehörigen Artikel und stellt den fachlichen Bezug innerhalb der Artikelansicht her."),
    (("Partner",), "Das Feld verweist auf den zugehörigen Partner, Kunden oder Lieferanten."),
    (("Preis", "Price"), "Der Preis wird für Kalkulation, Bewertung und nachgelagerte Verkaufs- oder Einkaufsprozesse verwendet."),
    (("Gültig ab", "Valid From", "ValidFrom"), "Gültig ab legt fest, ab welchem Datum der Eintrag in Preisen, Regeln oder Auswertungen verwendet wird."),
    (("Menge", "Quantity"), "Die Menge steuert, wie viele Einheiten in diesem Zusammenhang angesetzt, berechnet oder verarbeitet werden."),
    (("Finanzgruppe",), "Die Finanzgruppe ordnet den Eintrag der passenden Buchungs- und Auswertungslogik zu."),
    (("Aktiv", "Active"), "Aktiv legt fest, ob der Eintrag in Auswahlfeldern und Prozessen verwendet werden kann."),
    (("Favorit", "Favorite"), "Favorit kennzeichnet häufig genutzte Einträge und erleichtert die schnelle Auswahl."),
    (("Status",), "Der Status zeigt den aktuellen Bearbeitungs- oder Freigabezustand des Eintrags."),
    (("Datum", "Date"), "Das Datum ordnet den Eintrag zeitlich ein und ist wichtig für Historie, Filter und Auswertungen."),
    (("Info", "Beschreibung", "Text"), "Das Feld enthält ergänzende Informationen, die Anwender beim Prüfen oder Bearbeiten des Datensatzes unterstützen."),
]

PHRASE_PATTERNS = [
    (("Article", "Sales", "Price", "List"), "Verkaufspreislisten im Artikel"),
    (("Partner", "Purchase", "Price", "List"), "Einkaufspreislisten beim Partner"),
    (("Article", "Set"), "Set-Komponenten im Artikel"),
    (("Article", "Macro"), "Makropositionen im Artikel"),
    (("Article", "Product"), "Produktionskomponenten im Artikel"),
    (("Article", "Accessory"), "Zubehör im Artikel"),
    (("Article", "Category"), "Artikelkategorien"),
    (("Warehouse", "Quantity"), "Lagerbestände"),
    (("Contact", "Person"), "Ansprechpartner"),
    (("Customer",), "Kunden"),
    (("Supplier",), "Lieferanten"),
    (("Invoice",), "Rechnungen"),
    (("Project",), "Projekte"),
    (("Production", "Step"), "Produktionsschritte"),
]


def split_technical(value: str) -> list[str]:
    cleaned = re.sub(r"[_\-]+", " ", value or "")
    cleaned = re.sub(r"([a-z0-9])([A-ZÄÖÜ])", r"\1 \2", cleaned)
    cleaned = re.sub(r"([A-ZÄÖÜ]+)([A-ZÄÖÜ][a-zäöü])", r"\1 \2", cleaned)
    return [part for part in cleaned.split() if part]


def translated_tokens(value: str) -> str:
    parts = split_technical(value)
    part_set = set(parts)
    for needles, phrase in PHRASE_PATTERNS:
        if all(needle in part_set for needle in needles):
            if "Wizard" in part_set:
                return f"{phrase}-Assistent"
            if "Edit" in part_set:
                return f"{phrase} bearbeiten"
            if "List" in part_set:
                return phrase
            return phrase

    translated = [TOKEN_TRANSLATIONS.get(part, part) for part in parts]
    result = " ".join(translated)
    result = result.replace("Liste Liste", "Liste").replace("bearbeiten Liste", "Liste")
    result = result.replace("Preis Liste", "Preisliste").replace("Verkauf Preis", "Verkaufspreis")
    result = result.replace("Einkauf Preis", "Einkaufspreis").replace("Produktion Schritt", "Produktionsschritt")
    return result.strip()


def object_name_for_action(value: str) -> str:
    cleaned = value.replace(" anlegen/bearbeiten", "")
    cleaned = re.sub(r"Edit$", "", cleaned)
    phrase = translated_tokens(cleaned)
    return phrase.replace(" bearbeiten", "").strip()


def page_kind(node: Node) -> str:
    if node.level == 0:
        return "root"
    if node.level == 1:
        return "domain"
    if node.topic.endswith("Wizard"):
        return "wizard"
    if node.topic.endswith("List"):
        return "list"
    if " anlegen/bearbeiten" in node.topic:
        return "edit-action"
    if node.level == 3:
        return "tab"
    if node.topic.endswith("Edit") or "Edit-" in node.topic:
        return "edit"
    return "page"


def domain_name(node: Node) -> str:
    trail = ancestors(node) + [node]
    if len(trail) >= 2:
        return trail[1].topic
    return node.topic


def parent_page(node: Node) -> Node | None:
    current = node.parent
    while current:
        if is_page(current):
            return current
        current = current.parent
    return None


def friendly_name(value: str) -> str:
    if not value:
        return ""
    if "-" in value:
        return value.split("-")[-1].strip()
    return translated_tokens(value)


def display_heading(node: Node) -> str:
    kind = page_kind(node)
    if kind == "root":
        return "Ansichten in X-ERP"
    if kind == "domain":
        return f"{node.topic}: ERP-Ansichten im Überblick"
    if kind == "tab":
        tab = friendly_name(node.topic)
        return f"{node.topic}: {tab} in X-ERP"
    if kind == "list":
        readable = translated_tokens(node.topic)
        return f"{node.topic}: {readable}"
    if kind == "wizard":
        readable = translated_tokens(node.topic)
        return f"{node.topic}: {readable}"
    if kind == "edit-action":
        readable = object_name_for_action(node.topic)
        return f"{node.topic}: {readable} anlegen oder bearbeiten"
    if kind == "edit":
        readable = translated_tokens(node.topic)
        return f"{node.topic}: {readable}"
    return node.h1 or node.topic


def tab_purpose(node: Node) -> str | None:
    tab = friendly_name(node.topic)
    return TAB_PURPOSES.get(tab)


def clean_description(value: str, node: Node) -> str:
    result = re.sub(r"\s+", " ", value or "").strip()
    if not result:
        return ""
    result = result.replace("Die Erklärung ist für Anwender und die ERP-Dokumentation optimiert.", "")
    result = result.replace("Die Erklärung ist für Anwender und die ERP-Dokumentation optimiert", "")
    parent = parent_page(node)
    if parent:
        result = result.replace("Article-Bearbeitungsansicht", parent.topic)
    return re.sub(r"\s+", " ", result).strip(" ,.;")


def seo_description(node: Node) -> str:
    kind = page_kind(node)
    existing = clean_description(node.description, node)
    if kind == "tab":
        parent = parent_page(node)
        tab = friendly_name(node.topic)
        purpose = tab_purpose(node) or "ordnet die Felder, Listen und Aktionen dieses Bereichs fachlich ein."
        base = f"Das Register {tab} in {parent.topic if parent else domain_name(node)} {purpose}"
    elif existing and len(existing) >= 80 and "unterstützt Anwender" not in existing:
        base = existing
    else:
        domain = domain_name(node)
        if kind == "root":
            base = "Die X-ERP Hilfe erklärt alle ERP-Ansichten, Bearbeitungsmasken, Register, Felder und Aktionen in einer klaren, durchsuchbaren Struktur."
        elif kind == "domain":
            base = f"Der Bereich {node.topic} bündelt die zugehörigen ERP-Ansichten in X-ERP. Anwender finden hier Bearbeitungsmasken, Register, Felder und Aktionen in fachlicher Reihenfolge."
        elif kind == "tab":
            purpose = tab_purpose(node) or "ordnet die Felder und Funktionen dieses Registers fachlich ein."
            base = f"Das Register {node.topic} gehört zur Ansicht {parent_page(node).topic if parent_page(node) else domain} und {purpose}"
        elif kind == "list":
            parent = parent_page(node)
            base = f"Die eingebettete Liste {node.topic} zeigt zugehörige Datensätze innerhalb von {parent.topic if parent else domain}. Die Hilfe erklärt Zweck, Felder und Bearbeitung im ERP-Kontext."
        elif kind == "wizard":
            base = f"Der Assistent {node.topic} führt Anwender durch einen mehrstufigen ERP-Vorgang im Bereich {domain}. Die Hilfe erklärt Zweck, Eingaben und Ergebnis."
        elif kind == "edit-action":
            parent = parent_page(node)
            base = f"{node.topic} beschreibt, wie ein zugehöriger Datensatz aus {parent.topic if parent else domain} angelegt oder bearbeitet wird. Die Seite erklärt Felder, Pflichtinformationen und fachlichen Zusammenhang."
        elif kind == "edit":
            base = f"Die Bearbeitungsansicht {node.topic} dient zum Anlegen, Prüfen und Bearbeiten von Datensätzen im Bereich {domain}. Register, Felder und Aktionen werden in ERP-Reihenfolge erklärt."
        else:
            base = f"{node.topic} ist Teil der X-ERP Hilfe im Bereich {domain}. Die Seite erklärt Zweck, Struktur und Bedienung aus Anwendersicht."

    base = re.sub(r"\s+", " ", base).strip()
    base = base.strip()
    if len(base) > 300:
        return base[:297].rstrip(" ,.;") + "..."
    if not base.endswith("."):
        base += "."
    return base


def seo_title(node: Node) -> str:
    heading = display_heading(node)
    if len(heading) > 72:
        heading = heading[:69].rstrip(" -:") + "..."
    return f"{heading} | X-ERP ERP Hilfe"


def field_description(item: Node, current_page: Node) -> str:
    existing = item.description.strip()
    generic_markers = [
        "dient zur eindeutigen Zuordnung, Suche oder Identifikation",
        "zeigt einen Wert für Kalkulation",
        "Dieses Element gehört zur fachlichen Bearbeitung",
    ]
    if existing and not any(marker in existing for marker in generic_markers):
        return existing

    haystack = " ".join([item.topic, item.field_name, item.original])
    for needles, description in FIELD_PURPOSES:
        if any(needle.lower() in haystack.lower() for needle in needles):
            return description

    if item.field_name:
        return f"{item.topic} beschreibt den Wert des Feldes {item.field_name} in {current_page.topic} und unterstützt die fachlich korrekte Bearbeitung im ERP-Prozess."
    return f"{item.topic} gehört zu {current_page.topic} und unterstützt die fachliche Bearbeitung dieses Bereichs."


def read_nodes() -> list[Node]:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["de-DE"]
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}

    start = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Ansichten")
    root_level = ws.row_dimensions[start].outlineLevel
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if ws.row_dimensions[row].outlineLevel <= root_level:
            end = row - 1
            break

    nodes: list[Node] = []
    for row in range(start, end + 1):
        topic = ws.cell(row, 1).value
        if not isinstance(topic, str) or not topic.strip():
            continue
        data = {name: ws.cell(row, col).value for name, col in headers.items()}
        data_level = data.get("TOC_LEVEL")
        level = int(float(data_level)) if data_level not in (None, "") else ws.row_dimensions[row].outlineLevel
        node = Node(
            row=row,
            level=level,
            topic=topic.strip(),
            original=get_cell(data, "Original Text"),
            screenshot=get_cell(data, "Screenshot"),
            field_name=get_cell(data, "Feld"),
            description=get_cell(data, "Beschreibung"),
            page_id=get_cell(data, "PAGE_ID"),
            slug=get_cell(data, "SLUG"),
            title=get_cell(data, "TITLE") or topic.strip(),
            meta_description=get_cell(data, "META_DESCRIPTION"),
            h1=get_cell(data, "H1") or topic.strip(),
            keyword=get_cell(data, "PRIMARY_KEYWORD") or "ERP",
            content_type=get_cell(data, "CONTENT_TYPE") or "Hilfe",
            breadcrumb=get_cell(data, "BREADCRUMB"),
            toc_parent=get_cell(data, "TOC_PARENT"),
            toc_level=int(float(data.get("TOC_LEVEL") or 0)),
            toc_order=get_cell(data, "TOC_ORDER"),
            unique_key=get_cell(data, "UNIQUE_PAGE_KEY"),
            image_alt=get_cell(data, "IMAGE_ALT"),
            image_caption=get_cell(data, "IMAGE_CAPTION"),
            image_status=get_cell(data, "IMAGE_STATUS"),
        )
        nodes.append(node)

    stack: dict[int, Node] = {}
    for node in nodes:
        parent = stack.get(node.level - 1)
        node.parent = parent
        if parent:
            parent.children.append(node)
        for level in list(stack):
            if level >= node.level:
                stack.pop(level, None)
        stack[node.level] = node

    return nodes


def is_page(node: Node) -> bool:
    if node.level <= 3:
        return True
    if node.topic.endswith("List"):
        return True
    if " anlegen/bearbeiten" in node.topic:
        return True
    return False


def assign_paths(nodes: list[Node]) -> list[Node]:
    page_nodes = [node for node in nodes if is_page(node)]
    root = page_nodes[0]
    root.url_part = ""
    root.fs_dir = TARGET_ROOT
    root.public_url = f"{BASE_URL}/"

    for node in page_nodes[1:]:
        parts = []
        current: Node | None = node
        while current and current is not root:
            if is_page(current):
                parts.append(slugify(current.topic))
            current = current.parent
        parts.reverse()
        node.url_part = "/".join(parts)
        node.fs_dir = TARGET_ROOT.joinpath(*parts)
        node.public_url = f"{BASE_URL}/{node.url_part}/"

    seen: dict[str, Node] = {}
    for node in page_nodes:
        if node.fs_dir is None:
            continue
        key = str(node.fs_dir).lower()
        if key not in seen:
            seen[key] = node
            continue
        suffix = f"zeile-{node.row}"
        node.url_part = f"{node.url_part}-{suffix}" if node.url_part else suffix
        node.fs_dir = TARGET_ROOT.joinpath(*node.url_part.split("/"))
        node.public_url = f"{BASE_URL}/{node.url_part}/"
        seen[str(node.fs_dir).lower()] = node

    return page_nodes


def rel_link(from_node: Node, to_node: Node) -> str:
    if from_node.fs_dir is None or to_node.fs_dir is None:
        return "#"
    rel = Path(".") if from_node.fs_dir == to_node.fs_dir else Path(
        *Path(to_node.fs_dir).relative_to(TARGET_ROOT).parts
    )
    from_rel = Path(from_node.fs_dir).relative_to(TARGET_ROOT)
    depth = len(from_rel.parts)
    prefix = Path(*([".."] * depth)) if depth else Path(".")
    if to_node is page_root:
        return (prefix / "").as_posix() + "/"
    return (prefix / rel).as_posix() + "/"


def absolute_help_path(path: str) -> str:
    path = path.replace("\\", "/")
    if path.startswith("Screenshots/"):
        return "/de/help/" + path
    if path.startswith("/"):
        return path
    return "/de/help/" + path


def screenshot_file(path: str) -> Path:
    normalized = (path or "").replace("\\", "/").lstrip("/")
    if normalized.startswith("de/help/"):
        normalized = normalized[len("de/help/"):]
    return HELP_ROOT / normalized


def central_screenshot_file(path: str) -> Path:
    return HELP_ROOT / "Screenshots" / Path((path or "").replace("\\", "/")).name


def has_screenshot_file(node: Node) -> bool:
    return bool(node.screenshot) and screenshot_file(node.screenshot).is_file()


def descendants(node: Node) -> list[Node]:
    result: list[Node] = []
    stack = list(reversed(node.children))
    while stack:
        current = stack.pop()
        result.append(current)
        stack.extend(reversed(current.children))
    return result


def effective_screenshot_node(node: Node) -> Node | None:
    if has_screenshot_file(node):
        return node

    current = node.parent
    while current:
        if is_page(current) and has_screenshot_file(current):
            return current
        current = current.parent

    for child in descendants(node):
        if is_page(child) and has_screenshot_file(child):
            return child

    return None


def direct_page_children(node: Node) -> list[Node]:
    return [child for child in node.children if is_page(child)]


def direct_nonpage_children(node: Node) -> list[Node]:
    return [child for child in node.children if not is_page(child)]


def ancestors(node: Node) -> list[Node]:
    result: list[Node] = []
    current = node.parent
    while current:
        if is_page(current):
            result.append(current)
        current = current.parent
    return list(reversed(result))


def render_breadcrumb(node: Node) -> str:
    items = []
    for item in ancestors(node):
        items.append(f'<a href="{text(rel_link(node, item))}">{text(item.topic)}</a>')
    items.append(f"<span>{text(node.topic)}</span>")
    return '<nav class="view-breadcrumb" aria-label="Breadcrumb">' + '<span class="sep">/</span>'.join(items) + "</nav>"


def render_child_cards(node: Node) -> str:
    children = direct_page_children(node)
    if not children:
        return ""
    cards = []
    for child in children:
        desc = seo_description(child)
        cards.append(
            f"""
            <a class="view-card-link" href="{text(rel_link(node, child))}">
              <strong>{text(display_heading(child))}</strong>
              <span>{text(desc)}</span>
            </a>
            """
        )
    return f"""
    <section class="help-view-panel view-page-section">
      <div class="help-view-section-head">
        <h2>Unterseiten</h2>
      </div>
      <div class="view-card-grid">{''.join(cards)}</div>
    </section>
    """


def render_fields(node: Node) -> str:
    fields = direct_nonpage_children(node)
    if not fields:
        return ""
    rows = []
    for item in fields:
        field_code = f'<code>{text(item.field_name)}</code>' if item.field_name else ""
        original = f'<span class="view-muted">Original: {text(item.original)}</span>' if item.original and item.original != item.topic else ""
        rows.append(
            f"""
            <li>
              <h3>{text(item.topic)}</h3>
              <p>{text(field_description(item, node))}</p>
              <div class="view-field-meta">{field_code}{original}</div>
            </li>
            """
        )
    return f"""
    <section class="help-view-panel view-page-section">
      <div class="help-view-section-head">
        <h2>Felder und Elemente</h2>
        <p>Diese Einträge erscheinen direkt in diesem Bereich der Ansicht.</p>
      </div>
      <ol class="help-view-fields">{''.join(rows)}</ol>
    </section>
    """


def render_screenshot(node: Node) -> str:
    shot_node = effective_screenshot_node(node)
    if shot_node:
        src = absolute_help_path(shot_node.screenshot)
        alt = node.image_alt or f"Screenshot {display_heading(node)} in X-ERP"
        if shot_node is node:
            caption = node.image_caption or f"{node.topic} in X-ERP"
        elif shot_node in ancestors(node):
            caption = f"Orientierender Screenshot aus der übergeordneten Ansicht {shot_node.topic}."
        else:
            caption = f"Beispiel-Screenshot aus {shot_node.topic} im Bereich {domain_name(node)}."
        return f"""
        <figure class="help-view-shot view-screenshot-figure">
          <a class="lightbox" href="{text(src)}"><img src="{text(src)}" alt="{text(alt)}" loading="lazy"></a>
          <figcaption>{text(caption)}</figcaption>
        </figure>
        """
    return ""


def render_sibling_nav(node: Node) -> str:
    if not node.parent:
        return ""
    siblings = [child for child in node.parent.children if is_page(child)]
    if len(siblings) < 2:
        return ""
    index = siblings.index(node)
    previous = siblings[index - 1] if index > 0 else None
    next_item = siblings[index + 1] if index + 1 < len(siblings) else None
    prev_html = f'<a href="{text(rel_link(node, previous))}">Zurück: {text(previous.topic)}</a>' if previous else '<span></span>'
    next_html = f'<a href="{text(rel_link(node, next_item))}">Weiter: {text(next_item.topic)}</a>' if next_item else '<span></span>'
    return f'<nav class="view-sibling-nav" aria-label="Vorherige und nächste Seite">{prev_html}{next_html}</nav>'


def render_context_links(node: Node) -> str:
    links: list[tuple[str, Node, str]] = []
    parent = parent_page(node)
    if parent:
        links.append((
            f"Zurück zu {friendly_name(parent.topic) or parent.topic}",
            parent,
            f"{display_heading(parent)} ist der übergeordnete Kontext für {node.topic}.",
        ))

    for child in direct_page_children(node):
        if len(links) < 8:
            kind = page_kind(child)
            child_name = friendly_name(child.topic) or child.topic
            if kind == "tab":
                label = f"Register {child_name}"
                note = f"{child.topic} beschreibt den passenden Registerbereich innerhalb von {node.topic}."
            elif kind == "list":
                label = f"Liste {child_name}"
                note = f"{child.topic} zeigt die zugehörigen Datensätze, die direkt aus {node.topic} heraus gepflegt werden."
            elif kind == "edit-action":
                label = f"Dialog {child_name}"
                note = f"{child.topic} erklärt das Anlegen oder Bearbeiten eines Datensatzes aus diesem Abschnitt."
            else:
                label = f"Unterseite {child_name}"
                note = f"{child.topic} vertieft einen eigenen Teilbereich von {node.topic}."
            links.append((label, child, note))

    siblings = [child for child in parent.children if is_page(child)] if parent else []
    for sibling in siblings:
        if sibling is node or len(links) >= 8:
            continue
        sibling_name = friendly_name(sibling.topic) or sibling.topic
        links.append((
            f"Verwandte Ansicht {sibling_name}",
            sibling,
            f"{sibling.topic} liegt ebenfalls im Bereich {domain_name(node)} und kann für vergleichbare ERP-Aufgaben relevant sein.",
        ))

    if not links:
        return ""

    items = []
    used_labels: set[str] = set()
    used_notes: set[str] = set()
    for label, target, note in links:
        if label in used_labels:
            label = f"{label}: {target.topic}"
        if note in used_notes:
            note = f"{note} Bezug: {target.topic}."
        used_labels.add(label)
        used_notes.add(note)
        items.append(
            f"""
            <li>
              <span>{text(label)}</span>
              <a href="{text(rel_link(node, target))}">{text(display_heading(target))}</a>
              <p>{text(note)}</p>
            </li>
            """
        )

    return f"""
    <section class="help-view-panel view-page-section">
      <div class="help-view-section-head">
        <h2>Verwandte Hilfe-Seiten</h2>
        <p>Diese Links verbinden die aktuelle Seite mit ihrem fachlichen Umfeld in der ERP-Hilfe.</p>
      </div>
      <ul class="view-related-links">{''.join(items)}</ul>
    </section>
    """


def render_json_ld(node: Node) -> str:
    crumbs = []
    trail = ancestors(node) + [node]
    for index, item in enumerate(trail, start=1):
        crumbs.append({"@type": "ListItem", "position": index, "name": item.topic, "item": item.public_url})
    data = {
        "@context": "https://schema.org",
        "@type": "TechArticle",
        "headline": display_heading(node),
        "description": seo_description(node),
        "inLanguage": "de-DE",
        "about": node.keyword or "ERP",
        "keywords": ["ERP", "X-ERP", node.topic, domain_name(node)],
        "mainEntityOfPage": node.public_url,
        "isPartOf": {"@type": "CreativeWork", "name": "X-ERP Hilfe"},
        "breadcrumb": {"@type": "BreadcrumbList", "itemListElement": crumbs},
    }
    return json.dumps(data, ensure_ascii=False)


def render_page(node: Node) -> str:
    title = seo_title(node)
    desc = seo_description(node)
    h1 = display_heading(node)
    summary = desc
    hero_meta = [
        ("Bereich", node.content_type or "Ansicht"),
        ("Keyword", node.keyword or "ERP"),
        ("Ebene", str(node.toc_level)),
    ]
    if node.original and node.original != node.topic:
        hero_meta.append(("Original", node.original))
    if node.topic != h1:
        hero_meta.append(("Technischer Name", node.topic))
    meta_html = "".join(f'<span class="view-chip"><strong>{text(label)}:</strong> {text(value)}</span>' for label, value in hero_meta)
    canonical = node.public_url
    generated = datetime.now(timezone.utc).date().isoformat()
    shell_class = "help-page-shell has-view-screenshot" if effective_screenshot_node(node) else "help-page-shell"
    return f"""<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{text(title)}</title>
  <meta name="description" content="{text(desc)}">
  <meta name="robots" content="index, follow">
  <link rel="canonical" href="{text(canonical)}">
  <link rel="stylesheet" href="/de/help/styles.css">
  <link rel="stylesheet" href="/de/help/help.css?v=20260627-global-box-spacing">
  <link rel="stylesheet" href="/de/help/Ansichten/ansichten.css?v=20260627-screenshot-box-width">
  <style>.view-screenshot-figure{{max-width:1120px!important}}.view-screenshot-figure img{{max-width:100%!important}}</style>
  <script type="application/ld+json">{render_json_ld(node)}</script>
</head>
<body class="help-page">
  <div class="doc-layout">
    <aside class="doc-sidebar" id="sidebar">
      <div class="doc-sidebar-header">
        <h1><img src="/assets/logo/X-ERP.png" alt="X-ERP Logo"> X-ERP Hilfe</h1>
        <div class="doc-search"><input type="text" id="toc-search" placeholder="Suchen... (Strg+K)"></div>
      </div>
      <nav aria-label="Hilfe-Navigation"><ul class="toc-tree" id="toc"></ul></nav>
    </aside>
    <div class="overlay" id="overlay"></div>
    <main class="doc-content">
      <section class="{shell_class}">
        {render_breadcrumb(node)}
        <section class="help-view-hero">
          <div class="help-view-eyebrow">ERP Hilfe</div>
          <div class="help-view-hero-grid">
            <div>
              <h1>{text(h1)}</h1>
              <p class="help-view-summary">{text(summary)}</p>
              <div class="help-view-meta">{meta_html}</div>
            </div>
            <div class="help-view-hero-card">
              <h2>Einordnung</h2>
              <p>{text(node.breadcrumb or node.topic)}</p>
            </div>
          </div>
        </section>
        {render_screenshot(node)}
        {render_child_cards(node)}
        {render_fields(node)}
        {render_context_links(node)}
        {render_sibling_nav(node)}
        <p class="view-generated">Stand: {generated}. Quelle: X-ERP Help Excel, Zeile {node.row}.</p>
      </section>
    </main>
  </div>
  <button class="mobile-menu-btn" id="mobileMenuBtn" aria-label="Menü öffnen">&#9776;</button>
  <script src="/de/help/xlsx-tree.js?v=20260628-kommissionieren-packen-split"></script>
  <script src="/de/help/ansichten-tree.js?v=20260628-kommissionieren-packen-split"></script>
  <script src="/de/help/index-tree.js?v=20260628-kommissionieren-packen-split"></script>
  <script src="/de/help/app.js?v=20260628-kommissionieren-packen-split"></script>
</body>
</html>
"""


def write_css() -> None:
    css = """
.view-breadcrumb{display:flex;flex-wrap:wrap;gap:.45rem;align-items:center;color:var(--help-muted);font-size:.92rem;position:relative;z-index:1}
.view-breadcrumb a{color:var(--help-accent);text-decoration:none}
.view-breadcrumb a:hover{text-decoration:underline}
.view-breadcrumb .sep{opacity:.45}
.view-chip{display:inline-flex;gap:.35rem;align-items:center;border:1px solid rgba(141,148,158,.24);background:rgba(141,148,158,.09);border-radius:999px;padding:.42rem .7rem;color:var(--help-muted);font-size:.9rem}
.help-page-shell.has-view-screenshot{max-width:1480px}
.help-page-shell.has-view-screenshot>.view-breadcrumb,.help-page-shell.has-view-screenshot>.help-view-hero,.help-page-shell.has-view-screenshot>.help-view-panel,.help-page-shell.has-view-screenshot>.view-sibling-nav,.help-page-shell.has-view-screenshot>.view-generated{max-width:1120px;width:100%}
.view-screenshot-figure{margin:clamp(1.1rem,2vw,2rem) 0 clamp(1.4rem,2.6vw,2.4rem);position:relative;z-index:1;width:100%;max-width:1120px}
.view-screenshot-figure a{display:block;text-decoration:none}
body.help-page .doc-content .view-screenshot-figure a.lightbox img{display:block;width:100%!important;max-width:100%!important;border-radius:12px;border:1px solid rgba(255,255,255,.14);box-shadow:0 28px 80px -42px rgba(0,0,0,.95);background:#0f1115}
.view-screenshot-figure figcaption{margin-top:.65rem;color:var(--help-muted);font-size:.92rem;line-height:1.45}
.view-card-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:.85rem;position:relative;z-index:1}
.view-card-link{display:grid;gap:.45rem;padding:1rem;border-radius:14px;border:1px solid rgba(141,148,158,.18);background:rgba(255,255,255,.035);text-decoration:none;color:var(--help-text)}
.view-card-link:hover{border-color:rgba(141,148,158,.42);background:rgba(141,148,158,.10)}
.view-card-link strong{font-size:1.02rem}
.view-card-link span:last-child{color:var(--help-muted);line-height:1.55}
.view-field-meta{display:flex;flex-wrap:wrap;gap:.6rem;margin-top:.7rem;align-items:center}
.view-field-meta code{border:1px solid rgba(141,148,158,.22);background:rgba(0,0,0,.22);border-radius:8px;padding:.18rem .42rem;color:#dfe5ec}
.view-muted{color:var(--help-muted);font-size:.9rem}
.view-sibling-nav{display:grid;grid-template-columns:1fr 1fr;gap:1rem}
.view-sibling-nav a{display:block;border:1px solid rgba(141,148,158,.18);border-radius:14px;padding:1rem;background:rgba(255,255,255,.035);text-decoration:none;color:var(--help-text)}
.view-sibling-nav a:last-child{text-align:right}
.view-related-links{display:grid;gap:.7rem;list-style:none;margin:0;padding:0}
.view-related-links li{border:1px solid rgba(141,148,158,.16);background:rgba(255,255,255,.03);border-radius:12px;padding:.85rem}
.view-related-links span{display:block;color:var(--help-accent);font-size:.76rem;text-transform:uppercase;font-weight:700;letter-spacing:.1em;margin-bottom:.25rem}
.view-related-links a{color:var(--help-text);font-weight:700;text-decoration:none}
.view-related-links a:hover{text-decoration:underline}
.view-related-links p{color:var(--help-muted);margin:.35rem 0 0;line-height:1.5}
.view-generated{color:var(--help-muted);font-size:.86rem;margin:0}
@media(max-width:760px){.view-sibling-nav{grid-template-columns:1fr}.view-sibling-nav a:last-child{text-align:left}}
"""
    (TARGET_ROOT / "ansichten.css").write_text(css.strip() + "\n", encoding="utf-8")


def write_index_data(page_nodes: list[Node]) -> None:
    data = [
        {
            "title": node.topic,
            "h1": display_heading(node),
            "url": node.public_url,
            "path": node.url_part,
            "breadcrumb": node.breadcrumb,
            "description": seo_description(node),
            "screenshot": effective_screenshot_node(node).screenshot if effective_screenshot_node(node) else "",
            "row": node.row,
            "level": node.level,
        }
        for node in page_nodes
    ]
    (TARGET_ROOT / "index-data-ansichten.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_sitemap(page_nodes: list[Node]) -> None:
    entries = "\n".join(
        f"  <url><loc>{text(node.public_url)}</loc><changefreq>weekly</changefreq><priority>{'0.9' if node.level <= 2 else '0.7'}</priority></url>"
        for node in page_nodes
    )
    xml = f'<?xml version="1.0" encoding="UTF-8"?>\n<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n{entries}\n</urlset>\n'
    (TARGET_ROOT / "sitemap-ansichten.xml").write_text(xml, encoding="utf-8")


def clean_target_generated_pages() -> None:
    TARGET_ROOT.mkdir(parents=True, exist_ok=True)
    resolved = TARGET_ROOT.resolve()
    expected = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help\Ansichten").resolve()
    if resolved != expected:
        raise RuntimeError(f"Refusing to clean unexpected target: {resolved}")
    for item in TARGET_ROOT.iterdir():
        if item.name in {"ansichten.css", "index-data-ansichten.json", "sitemap-ansichten.xml"}:
            item.unlink(missing_ok=True)
        elif item.is_dir():
            shutil.rmtree(item)
        elif item.name.lower() == "index.html":
            item.unlink()


def collect_existing_local_screenshots() -> dict[str, bytes]:
    if not TARGET_ROOT.exists():
        return {}
    screenshots: dict[str, bytes] = {}
    for file in TARGET_ROOT.rglob("*.webp"):
        if file.is_file():
            screenshots[file.relative_to(HELP_ROOT).as_posix()] = file.read_bytes()
    return screenshots


def restore_local_screenshots(screenshots: dict[str, bytes]) -> None:
    for rel_path, data in screenshots.items():
        target = HELP_ROOT / rel_path
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(data)


def ensure_page_screenshots(page_nodes: list[Node]) -> tuple[int, int]:
    copied = 0
    missing = 0
    for node in page_nodes:
        if not node.screenshot:
            continue
        target = screenshot_file(node.screenshot)
        if target.is_file():
            continue
        source = central_screenshot_file(node.screenshot)
        if source.is_file():
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, target)
            copied += 1
        else:
            missing += 1
    return copied, missing


def main() -> None:
    nodes = read_nodes()
    global page_root
    page_nodes = assign_paths(nodes)
    page_root = page_nodes[0]

    local_screenshots = collect_existing_local_screenshots()
    clean_target_generated_pages()
    restore_local_screenshots(local_screenshots)
    copied_screenshots, missing_screenshots = ensure_page_screenshots(page_nodes)
    write_css()
    for node in page_nodes:
        assert node.fs_dir is not None
        node.fs_dir.mkdir(parents=True, exist_ok=True)
        (node.fs_dir / "index.html").write_text(render_page(node), encoding="utf-8")
    write_index_data(page_nodes)
    write_sitemap(page_nodes)
    print(f"pages={len(page_nodes)}")
    print(f"local_screenshots_restored={len(local_screenshots)}")
    print(f"local_screenshots_copied={copied_screenshots}")
    print(f"local_screenshots_missing={missing_screenshots}")
    print(TARGET_ROOT)


page_root: Node


if __name__ == "__main__":
    main()


