from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
ARCHIVE = Path(r"C:\Users\micha\Documents\X-ERP-HELP\ARCHIV")
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
SHEET = "de-DE"


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def rel_to_file(value: str) -> Path:
    value = value.replace("\\", "/").strip()
    if value.startswith("/de/help/"):
        value = value[len("/de/help/") :]
    if value.startswith("/"):
        value = value[1:]
    return HELP_ROOT / value


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-clear-missing-screenshots-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    ws = wb[SHEET]
    headers = {
        text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if text(ws.cell(1, col).value)
    }
    cols = ["Screenshot", "SCREENSHOT_REL_PATH", "SCREENSHOT_WEB_PATH", "IMAGE_STATUS"]
    missing = sorted(set(cols + ["Thema"]) - set(headers))
    if missing:
        raise RuntimeError(f"Missing required headers: {', '.join(missing)}")

    cleared = 0
    for row in range(2, ws.max_row + 1):
        raw = text(ws.cell(row, headers["SCREENSHOT_REL_PATH"]).value) or text(ws.cell(row, headers["Screenshot"]).value)
        if not raw:
            continue
        if not rel_to_file(raw).is_file():
            ws.cell(row, headers["Screenshot"]).value = None
            ws.cell(row, headers["SCREENSHOT_REL_PATH"]).value = None
            ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value = None
            ws.cell(row, headers["IMAGE_STATUS"]).value = "kein Screenshot"
            cleared += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"cleared_missing={cleared}")


if __name__ == "__main__":
    main()
