from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")


def u(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


UPDATES = {
    "WorkspaceRoleDesktopEdit": {
        "META_DESCRIPTION": u("WorkspaceRoleDesktopEdit in X-ERP: Desktop-Elemente und Startbereiche je Benutzerrolle f\\u00fcr den ERP-Arbeitsplatz konfigurieren."),
    },
    "WorkspaceRoleMenuEdit": {
        "META_DESCRIPTION": u("WorkspaceRoleMenuEdit in X-ERP: Men\\u00fcpunkte und Navigation je Benutzerrolle f\\u00fcr einen klaren ERP-Arbeitsplatz konfigurieren."),
    },
    "CrmEdit": {
        "Beschreibung": u("CrmEdit b\\u00fcndelt die vertriebliche Nachverfolgung eines Kunden- oder Lieferantenkontakts. Die Ansicht zeigt Partner, Ansprechpartner, Telefon, E-Mail, Wiedervorlage, Status, Opportunity, Abschlusswahrscheinlichkeit und Gespr\\u00e4chsnotizen. So behalten Vertrieb und Service offene Chancen, n\\u00e4chste Schritte und erledigte Aktivit\\u00e4ten im ERP nachvollziehbar im Blick."),
        "META_DESCRIPTION": u("CrmEdit in X-ERP: CRM-Vorg\\u00e4nge mit Partner, Status, Wiedervorlage, Opportunity, Abschlusswahrscheinlichkeit, Notizen, E-Mail, Dokumenten und ERP-Nachverfolgung."),
    },
    "FinancePostingDraftEdit": {
        "Beschreibung": u("FinancePostingDraftEdit zeigt und bearbeitet Buchungsentw\\u00fcrfe, bevor daraus g\\u00fcltige Finanzbuchungen werden. Anwender pr\\u00fcfen Buchungsvorlage, Belegdatum, Buchungsperiode, Belegnummer, Dok-Nummer, Betreff, Fremdw\\u00e4hrung und die Buchungspositionen. Der Status macht sichtbar, ob der Entwurf vollst\\u00e4ndig und buchbar ist oder noch ausgleichende Positionen fehlen."),
        "META_DESCRIPTION": u("FinancePostingDraftEdit in X-ERP: Buchungsentw\\u00fcrfe mit Periode, Belegnummer, Dok-Nummer, Sachkonto, Kundenkonto und Lieferantenkonto pr\\u00fcfen."),
    },
    "IntercomEdit": {
        "Beschreibung": u("IntercomEdit verwaltet interne Aufgaben und Nachrichten zwischen Anwendern. Empf\\u00e4nger, Absender, Status, Priorit\\u00e4t, Stichtag, Dauer, geplante Fertigstellung, Beschreibung und Antwort zeigen, wer was bis wann erledigen soll. Damit werden interne Abstimmungen im ERP nachvollziehbar statt in E-Mails oder Notizen verstreut."),
        "META_DESCRIPTION": u("IntercomEdit in X-ERP: interne Aufgaben und Nachrichten mit Empf\\u00e4nger, Status, Priorit\\u00e4t, Stichtag, Beschreibung und Antwort bearbeiten."),
    },
    "EmployeeEdit": {
        "Beschreibung": u("EmployeeEdit pflegt Mitarbeiterstammdaten f\\u00fcr Anmeldung, interne Zuordnung, Urlaub, Portalzugriff und Auswertungen. Die \\u00dcbersicht enth\\u00e4lt die wichtigsten Personen- und Organisationsdaten; weitere Register zeigen Details, Urlaubsanspruchshistorie, Portal, Anh\\u00e4nge und Extra-Felder. So bleiben Benutzer, Mitarbeiter und interne Prozesse sauber miteinander verbunden."),
        "META_DESCRIPTION": u("EmployeeEdit in X-ERP: Mitarbeiterstammdaten, Details, Urlaubsanspruch, Portalzugriff, Anh\\u00e4nge und Extra-Felder im ERP verwalten."),
    },
    "ProjectEdit": {
        "Beschreibung": u("ProjectEdit verwaltet ein Projekt vom Kundenbezug bis zu Terminen, Budget und Folgepositionen. Anwender sehen Projektname, Kunde, Projektdatum, Status, Sachbearbeiter, Stichtag, Abh\\u00e4ngigkeiten, Start- und Endzeit sowie Budgetwerte. Die Register Dokumente und Positionen zeigen, welche Belege und Leistungen zum Projekt geh\\u00f6ren."),
        "META_DESCRIPTION": u("ProjectEdit in X-ERP: Projekte mit Kunde, Status, Terminen, Abh\\u00e4ngigkeiten, Budget, Dokumenten und Positionen im ERP sicher verwalten."),
    },
    "ResourceEdit": {
        "Beschreibung": u("ResourceEdit beschreibt Maschinen, Arbeitspl\\u00e4tze oder andere Ressourcen f\\u00fcr Planung, Kalkulation und Produktion. Neben Ressourcennummer, Name und Gruppe werden technische Details, Preise, Zusatzinformationen und zugeordnete Produktionsschritte gepflegt. Dadurch kann X-ERP Ressourcen in Fertigungsabl\\u00e4ufen und Auswertungen korrekt verwenden."),
        "META_DESCRIPTION": u("ResourceEdit in X-ERP: Ressourcen, Maschinen, Preise, technische Daten und Produktionsschritte f\\u00fcr ERP-Produktion und Planung pflegen."),
    },
    "TimeTrackingEdit": {
        "Beschreibung": u("TimeTrackingEdit erfasst Arbeitszeiten mit Mitarbeiter-, Datums-, Projekt-, Beleg- oder Leistungsbezug. Die Ansicht unterst\\u00fctzt die Nachverfolgung geleisteter Arbeit und bildet die Grundlage f\\u00fcr Auswertung, Abrechnung oder interne Kontrolle. Anh\\u00e4nge und Extra-Felder erg\\u00e4nzen den Zeitdatensatz bei Bedarf."),
    },
}


def main() -> None:
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}
    changed = []
    for row in range(2, ws.max_row + 1):
        thema = ws.cell(row, col["Thema"]).value
        if thema not in UPDATES:
            continue
        for header, value in UPDATES[thema].items():
            ws.cell(row, col[header]).value = value
        changed.append((row, thema))
    wb.save(WORKBOOK)
    print(f"changed={changed}")


if __name__ == "__main__":
    main()
