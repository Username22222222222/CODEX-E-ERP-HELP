from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")

SOURCE_IMAGE = HELP_ROOT / "Ansichten" / "artikel" / "article-edit" / "article-edit-details" / "ArticleEdit_Details.webp"
TARGET_REL = Path("Stammdaten") / "artikel" / "artikel-details" / "ArticleEdit_Details_Stammdaten.webp"
TARGET_IMAGE = HELP_ROOT / TARGET_REL


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    if not SOURCE_IMAGE.is_file():
        raise FileNotFoundError(SOURCE_IMAGE)

    TARGET_IMAGE.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SOURCE_IMAGE, TARGET_IMAGE)

    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-artikel-details-screenshot-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    target_row = None
    for row in range(2, ws.max_row + 1):
        if cell_text(ws.cell(row, headers["BREADCRUMB"]).value) == "Stammdaten > Artikel > Artikel Details":
            target_row = row
            break
    if target_row is None:
        raise RuntimeError("Row not found: Stammdaten > Artikel > Artikel Details")

    web_path = "/de/help/" + TARGET_REL.as_posix()
    ws.cell(target_row, headers["Screenshot"]).value = TARGET_REL.as_posix()
    ws.cell(target_row, headers["SCREENSHOT_WEB_PATH"]).value = web_path
    ws.cell(target_row, headers["IMAGE_ALT"]).value = "Screenshot Artikel Details in X-ERP"
    ws.cell(target_row, headers["IMAGE_CAPTION"]).value = "Artikel Details in den X-ERP Stammdaten"
    ws.cell(target_row, headers["IMAGE_STATUS"]).value = "local-copy"

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"copied={TARGET_IMAGE}")
    print(f"row={target_row}")
    print(f"web_path={web_path}")


if __name__ == "__main__":
    main()
