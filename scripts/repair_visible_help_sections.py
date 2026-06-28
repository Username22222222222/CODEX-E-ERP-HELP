from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def set_row_visible(ws, row: int, level: int | None = None) -> None:
    dim = ws.row_dimensions[row]
    dim.hidden = False
    dim.collapsed = False
    if level is not None:
        dim.outlineLevel = level


def find_exact(ws, col: int, text: str) -> int | None:
    needle = text.casefold()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip().casefold() == needle:
            return row
    return None


def find_startswith(ws, col: int, text: str) -> int | None:
    needle = text.casefold()
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, col).value
        if isinstance(value, str) and value.strip().casefold().startswith(needle):
            return row
    return None


def main() -> None:
    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: idx + 1 for idx, header in enumerate(headers) if header}

    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]
    dir_col = col["DIRECTORY_PATH"]
    file_col = col["FILE_NAME"]
    url_col = col["URL_PATH"]
    storage_col = col["STORAGE_PATH"]
    nav_col = col["NAV_TITLE"]
    breadcrumb_col = col["BREADCRUMB"]
    toc_parent_col = col["TOC_PARENT"]
    toc_level_col = col["TOC_LEVEL"]

    ansichten_row = find_exact(ws, thema_col, "Ansichten")
    faq_row = find_startswith(ws, thema_col, "FAQ")
    glossar_row = find_exact(ws, thema_col, "Glossar")

    if not ansichten_row or not faq_row or not glossar_row:
        raise RuntimeError(
            f"Abschnitt nicht gefunden: Ansichten={ansichten_row}, FAQ={faq_row}, Glossar={glossar_row}"
        )

    # Make the workbook findable again in Excel. Keep outline levels, but do not
    # save the sheet with collapsed/hidden rows.
    for row in range(1, ws.max_row + 1):
        set_row_visible(ws, row)

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"

    # Main section styles from existing major help chapters.
    main_style_row = find_exact(ws, thema_col, "Stammdaten") or 66
    child_style_row = find_exact(ws, thema_col, "Voraussetzungen") or 5

    # Ansichten is a main chapter. Rows until FAQ belong to the views chapter.
    copy_row_style(ws, main_style_row, ansichten_row)
    set_row_visible(ws, ansichten_row, 0)
    ws.cell(ansichten_row, content_col).value = "HelpPage"
    ws.cell(ansichten_row, nav_col).value = "Ansichten"
    ws.cell(ansichten_row, toc_level_col).value = 0

    for row in range(ansichten_row + 1, faq_row):
        value = ws.cell(row, thema_col).value
        if not value:
            continue
        content_type = ws.cell(row, content_col).value
        if row == ansichten_row + 1 and content_type == "HelpPage":
            set_row_visible(ws, row, 1)
            copy_row_style(ws, child_style_row, row)
        elif content_type in {"View", "Wizard"}:
            set_row_visible(ws, row, 2)
        else:
            current_level = ws.row_dimensions[row].outlineLevel
            if current_level < 3:
                set_row_visible(ws, row, 3)
            else:
                set_row_visible(ws, row)

    # FAQ and Glossar are separate main chapters after the views chapter.
    section_updates = [
        (faq_row, "FAQ", "faq", "FAQ"),
        (glossar_row, "Glossar", "glossar", "Glossar"),
    ]
    for row, title, directory, breadcrumb in section_updates:
        copy_row_style(ws, main_style_row, row)
        set_row_visible(ws, row, 0)
        ws.cell(row, thema_col).value = title
        ws.cell(row, content_col).value = "HelpPage"
        ws.cell(row, dir_col).value = directory
        ws.cell(row, file_col).value = "index.html"
        ws.cell(row, url_col).value = f"{directory}/index.html"
        ws.cell(row, storage_col).value = directory
        ws.cell(row, nav_col).value = title
        ws.cell(row, breadcrumb_col).value = breadcrumb
        ws.cell(row, toc_parent_col).value = None
        ws.cell(row, toc_level_col).value = 0

    for row in range(faq_row + 1, glossar_row):
        if ws.cell(row, thema_col).value:
            set_row_visible(ws, row, 1)
            ws.cell(row, toc_parent_col).value = "FAQ"
            ws.cell(row, toc_level_col).value = 1
            if ws.cell(row, content_col).value == "HelpPage":
                ws.cell(row, dir_col).value = "faq"
                file_name = ws.cell(row, file_col).value or f"faq-{row}.html"
                ws.cell(row, url_col).value = f"faq/{file_name}"
                ws.cell(row, storage_col).value = f"faq/{file_name}"

    for row in range(glossar_row + 1, ws.max_row + 1):
        if ws.cell(row, thema_col).value:
            set_row_visible(ws, row, 1)
            ws.cell(row, toc_parent_col).value = "Glossar"
            ws.cell(row, toc_level_col).value = 1
            if ws.cell(row, content_col).value == "HelpPage":
                ws.cell(row, dir_col).value = "glossar"
                file_name = ws.cell(row, file_col).value or f"glossar-{row}.html"
                ws.cell(row, url_col).value = f"glossar/{file_name}"
                ws.cell(row, storage_col).value = f"glossar/{file_name}"

    wb.save(BOOK)
    wb.close()

    print(
        f"repaired_sections Ansichten={ansichten_row} FAQ={faq_row} Glossar={glossar_row} "
        f"rows={ws.max_row}"
    )


if __name__ == "__main__":
    main()
