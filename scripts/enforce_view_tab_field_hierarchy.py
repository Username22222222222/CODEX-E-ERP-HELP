from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


@dataclass
class RowData:
    values: list
    styles: list
    height: float | None
    outline_level: int
    hidden: bool = False
    collapsed: bool = False


def capture_row(ws, row: int) -> RowData:
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    dim = ws.row_dimensions[row]
    return RowData(
        values=[ws.cell(row, col).value for col in range(1, ws.max_column + 1)],
        styles=styles,
        height=dim.height,
        outline_level=dim.outlineLevel,
        hidden=False,
        collapsed=False,
    )


def write_row(ws, row: int, data: RowData) -> None:
    for col, value in enumerate(data.values, start=1):
        cell = ws.cell(row, col)
        cell.value = value
        style = data.styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
    dim = ws.row_dimensions[row]
    dim.height = data.height
    dim.outlineLevel = data.outline_level
    dim.hidden = False
    dim.collapsed = False


def make_overview_tab(template: RowData, headers: dict[str, int], view_name: str) -> RowData:
    tab = RowData(
        values=[None for _ in template.values],
        styles=[copy(style) for style in template.styles],
        height=template.height,
        outline_level=3,
    )
    tab.values[headers["Thema"] - 1] = "Übersicht"
    tab.values[headers["Beschreibung"] - 1] = (
        f"Das Register Übersicht enthält die zentralen Felder und Aktionen von {view_name}."
    )
    return tab


def is_register_row(row: RowData, headers: dict[str, int]) -> bool:
    thema = row.values[headers["Thema"] - 1]
    field = row.values[headers["Feld"] - 1]
    content_type = row.values[headers["CONTENT_TYPE"] - 1]
    return bool(thema) and field is None and content_type not in {"HelpPage", "View", "Wizard"}


def is_child_data_row(row: RowData, headers: dict[str, int]) -> bool:
    thema = row.values[headers["Thema"] - 1]
    field = row.values[headers["Feld"] - 1]
    content_type = row.values[headers["CONTENT_TYPE"] - 1]
    return content_type not in {"HelpPage", "View", "Wizard"} and (bool(thema) or field is not None)


def normalize_block(
    view_row: RowData,
    child_rows: list[RowData],
    headers: dict[str, int],
    register_template: RowData,
) -> tuple[list[RowData], bool]:
    view_name = str(view_row.values[headers["Thema"] - 1])
    view_row.outline_level = 2
    normalized = [view_row]

    real_children = [row for row in child_rows if is_child_data_row(row, headers)]
    if not real_children:
        return normalized, False

    segments: list[tuple[RowData, list[RowData]]] = []
    pre_fields: list[RowData] = []
    current_register: RowData | None = None
    current_fields: list[RowData] = []

    for row in real_children:
        if is_register_row(row, headers):
            if current_register is not None:
                segments.append((current_register, current_fields))
            current_register = row
            current_fields = []
        else:
            if current_register is None:
                pre_fields.append(row)
            else:
                current_fields.append(row)

    if current_register is not None:
        segments.append((current_register, current_fields))

    inserted_tab = False
    if pre_fields and segments:
        first_register, first_fields = segments[0]
        segments[0] = (first_register, pre_fields + first_fields)
    elif pre_fields:
        segments.insert(0, (make_overview_tab(register_template, headers, view_name), pre_fields))
        inserted_tab = True

    for register, fields in segments:
        register.outline_level = 3
        register.hidden = False
        register.collapsed = False
        normalized.append(register)
        for field in fields:
            field.outline_level = 4
            field.hidden = False
            field.collapsed = False
            normalized.append(field)

    return normalized, inserted_tab


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-enforce-view-tab-field-hierarchy-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    ansichten_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")
    faq_row = next(row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")

    register_template_row = next(
        (
            row
            for row in range(ansichten_row + 1, faq_row)
            if ws.row_dimensions[row].outlineLevel == 3
            and ws.cell(row, content_col).value is None
            and ws.cell(row, col["Feld"]).value is None
            and ws.cell(row, thema_col).value
        ),
        ansichten_row + 1,
    )
    register_template = capture_row(ws, register_template_row)

    rebuilt: list[RowData] = []
    inserted_tabs = 0
    normalized_blocks = 0
    row = ansichten_row + 1

    while row < faq_row:
        content_type = ws.cell(row, content_col).value
        if content_type == "HelpPage" and ws.row_dimensions[row].outlineLevel == 1:
            module_row = capture_row(ws, row)
            module_row.outline_level = 1
            rebuilt.append(module_row)
            row += 1
            continue

        if content_type in {"View", "Wizard"}:
            view_row = capture_row(ws, row)
            row += 1
            children = []
            while row < faq_row and ws.cell(row, content_col).value not in {"HelpPage", "View", "Wizard"}:
                children.append(capture_row(ws, row))
                row += 1

            normalized, inserted = normalize_block(view_row, children, col, register_template)
            rebuilt.extend(normalized)
            normalized_blocks += 1
            if inserted:
                inserted_tabs += 1
            continue

        # Keep unexpected rows, but never as a domain/view level.
        other = capture_row(ws, row)
        if other.outline_level < 4:
            other.outline_level = 4
        rebuilt.append(other)
        row += 1

    delete_count = faq_row - (ansichten_row + 1)
    ws.delete_rows(ansichten_row + 1, delete_count)
    ws.insert_rows(ansichten_row + 1, len(rebuilt))

    for offset, row_data in enumerate(rebuilt):
        write_row(ws, ansichten_row + 1 + offset, row_data)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].collapsed = False

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"normalized_blocks={normalized_blocks}")
    print(f"inserted_overview_tabs={inserted_tabs}")
    print(f"rows_before_ansichten_children={delete_count}")
    print(f"rows_after_ansichten_children={len(rebuilt)}")


if __name__ == "__main__":
    main()
