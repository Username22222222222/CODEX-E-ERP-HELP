from __future__ import annotations

import html
import json
import re
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
VERSION = "20260628-kommissionieren-packen-split"


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def esc(value: object) -> str:
    return html.escape(text(value), quote=True)


def public_path(url_path: str) -> str:
    url_path = text(url_path).replace("\\", "/")
    if not url_path:
        return "/de/help/glossar/"
    if url_path.startswith("/de/help/"):
        path = url_path
    else:
        path = "/de/help/" + url_path.lstrip("/")
    if path.endswith("index.html"):
        path = path[: -len("index.html")]
    elif not path.endswith("/") and not path.lower().endswith(".html"):
        path += "/"
    return path


def target_file_for_path(path: str) -> Path:
    rel = public_path(path).removeprefix("/de/help/").lstrip("/")
    if rel.endswith("/"):
        rel += "index.html"
    return HELP_ROOT / rel


def paragraphs(value: str) -> str:
    blocks = [block.strip() for block in value.replace("\r\n", "\n").split("\n\n") if block.strip()]
    if not blocks:
        return "<p>Dieser Abschnitt wird aus der Excel-Hilfe erzeugt.</p>"
    return "\n".join(f"<p>{html.escape(block).replace(chr(10), '<br>')}</p>" for block in blocks)


def normalized_plain(value: str) -> str:
    value = re.sub(r"<[^>]+>", " ", value or "")
    value = value.replace("...", "")
    value = re.sub(r"[.。…]+$", "", value)
    return re.sub(r"\s+", " ", value).strip().casefold()


def meta_description(value: str) -> str:
    clean = re.sub(r"\s+", " ", value or "").strip()
    if len(clean) <= 155:
        return clean
    return clean[:152].rsplit(" ", 1)[0].rstrip(".,;:") + "..."


def visible_summary(summary: str, body: str) -> str:
    summary_text = normalized_plain(summary)
    body_text = normalized_plain(body)
    if not summary_text:
        return ""
    if body_text.startswith(summary_text) or summary_text.startswith(body_text[: len(summary_text)]):
        return ""
    return f'<p class="help-view-summary">{esc(summary)}</p>'


def breadcrumb_items(title: str, breadcrumb: str, canonical_path: str) -> list[tuple[str, str]]:
    parts = [part.strip() for part in (breadcrumb or "").split(">") if part.strip()]
    items: list[tuple[str, str]] = [("X-ERP Hilfe", "/de/help/")]
    if any(part.casefold() == "glossar" for part in parts):
        items.append(("Glossar", "/de/help/glossar/"))
    elif title.casefold() != "glossar":
        items.append(("Glossar", "/de/help/glossar/"))
    if title.casefold() != "glossar":
        items.append((title, canonical_path))
    return items


def render_breadcrumb(items: list[tuple[str, str]]) -> str:
    rendered: list[str] = []
    for index, (item_title, path) in enumerate(items):
        if index < len(items) - 1:
            rendered.append(f'<a href="{html.escape(path)}">{html.escape(item_title)}</a><span class="sep">/</span>')
        else:
            rendered.append(f"<span>{html.escape(item_title)}</span>")
    return "".join(rendered)


def render_json_ld(title: str, description: str, canonical_path: str, items: list[tuple[str, str]]) -> str:
    meta_desc = meta_description(description)
    schema = [
        {
            "@context": "https://schema.org",
            "@type": "BreadcrumbList",
            "itemListElement": [
                {
                    "@type": "ListItem",
                    "position": index + 1,
                    "name": item_title,
                    "item": "https://x-erp.de" + path,
                }
                for index, (item_title, path) in enumerate(items)
            ],
        },
        {
            "@context": "https://schema.org",
            "@type": "TechArticle",
            "headline": f"{title} | X-ERP Hilfe",
            "description": meta_desc,
            "inLanguage": "de-DE",
            "about": "ERP",
            "mainEntityOfPage": "https://x-erp.de" + canonical_path,
            "isPartOf": {"@type": "CreativeWork", "name": "X-ERP Hilfe"},
        },
    ]
    return json.dumps(schema, ensure_ascii=False).replace("</", "<\\/")


def page_html(
    title: str,
    description: str,
    breadcrumb: str,
    canonical_path: str,
    children: list[dict[str, str]] | None = None,
) -> str:
    breadcrumb_list = breadcrumb_items(title, breadcrumb, canonical_path)
    breadcrumb_html = render_breadcrumb(breadcrumb_list)
    schema = render_json_ld(title, description, canonical_path, breadcrumb_list)
    summary_html = visible_summary(description, description)
    meta_desc = meta_description(description)
    child_html = ""
    if children:
        items = "\n".join(
            f'''            <a class="view-card-link" href="{esc(child["path"])}">
              <span class="view-card-kicker">Glossar</span>
              <strong>{esc(child["title"])}</strong>
              <span>{esc(child["description"])}</span>
            </a>'''
            for child in children
        )
        child_html = f"""
        <section class="help-view-panel">
          <div class="help-view-section-head">
            <h2>Begriffe</h2>
          </div>
          <div class="view-card-grid">
{items}
          </div>
        </section>"""

    return f"""<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(title)} | X-ERP Hilfe</title>
  <meta name="description" content="{esc(meta_desc)}">
  <meta name="robots" content="index, follow">
  <link rel="canonical" href="https://x-erp.de{esc(canonical_path)}">
  <link rel="stylesheet" href="/de/help/styles.css">
  <link rel="stylesheet" href="/de/help/help.css?v=20260627-global-box-spacing">
  <link rel="stylesheet" href="/de/help/Ansichten/ansichten.css">
  <script type="application/ld+json">{schema}</script>
</head>
<body class="help-page">
  <div class="doc-layout">
    <aside class="doc-sidebar" id="sidebar">
      <div class="doc-sidebar-header">
        <h1><img src="/assets/logo/X-ERP.png" alt="X-ERP Logo"> X-ERP Hilfe</h1>
        <div class="doc-search"><input type="text" id="toc-search" placeholder="Suchen... (Strg+K)"></div>
      </div>
      <nav aria-label="Hilfe-Navigation"><ul class="toc-tree" id="toc"></ul></nav>
    </aside>
    <div class="overlay" id="overlay"></div>
    <main class="doc-content">
      <section class="help-page-shell">
        <nav class="view-breadcrumb" aria-label="Breadcrumb">{breadcrumb_html}</nav>
        <section class="help-view-hero">
          <div class="help-view-eyebrow">Glossar</div>
          <div class="help-view-hero-grid">
            <div>
              <h1>{esc(title)}</h1>
              {summary_html}
              <div class="help-view-meta">
                <span class="view-chip"><strong>Bereich:</strong> Glossar</span>
                <span class="view-chip"><strong>Quelle:</strong> Excel-Hilfe</span>
              </div>
            </div>
            <div class="help-view-hero-card">
              <h2>Einordnung</h2>
              <p>{esc(" > ".join(item_title for item_title, _ in breadcrumb_list))}</p>
            </div>
          </div>
        </section>
        <section class="help-view-panel">
          <div class="help-view-section-head">
            <h2>Erklärung</h2>
          </div>
          {paragraphs(description)}
        </section>{child_html}
      </section>
    </main>
  </div>
  <button class="mobile-menu-btn" id="mobileMenuBtn" aria-label="Menü öffnen">&#9776;</button>
  <script src="/de/help/xlsx-tree.js?v={VERSION}"></script>
  <script src="/de/help/ansichten-tree.js?v={VERSION}"></script>
  <script src="/de/help/index-tree.js?v={VERSION}"></script>
  <script src="/de/help/app.js?v={VERSION}"></script>
</body>
</html>
"""


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {text(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1) if text(ws.cell(1, col).value)}

    thema_col = headers["Thema"]
    glossar_row = next(
        row for row in range(2, ws.max_row + 1)
        if text(ws.cell(row, thema_col).value) == "Glossar"
    )
    root_level = ws.row_dimensions[glossar_row].outlineLevel

    rows: list[dict[str, str]] = []
    for row in range(glossar_row, ws.max_row + 1):
        if row > glossar_row and ws.row_dimensions[row].outlineLevel <= root_level:
            break
        values = {name: text(ws.cell(row, col).value) for name, col in headers.items()}
        if values.get("CONTENT_TYPE") != "HelpPage":
            continue
        rows.append({
            "title": values.get("NAV_TITLE") or values.get("Thema") or f"Zeile {row}",
            "description": values.get("Beschreibung") or values.get("Original Text") or "",
            "breadcrumb": values.get("BREADCRUMB") or "Glossar",
            "file": values.get("FILE_NAME") or "index.html",
            "path": public_path(values.get("URL_PATH") or "glossar/index.html"),
            "level": str(ws.row_dimensions[row].outlineLevel),
        })

    root = rows[0]
    children = [row for row in rows[1:] if row["level"] == "1"]
    target = target_file_for_path(root["path"])
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        page_html(root["title"], root["description"], root["breadcrumb"], root["path"], children),
        encoding="utf-8",
    )
    for row in children:
        target = target_file_for_path(row["path"])
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_text(
            page_html(row["title"], row["description"], row["breadcrumb"], row["path"]),
            encoding="utf-8",
        )

    wb.close()
    print(f"glossar_pages={1 + len(children)}")
    print(HELP_ROOT / "Glossar")


if __name__ == "__main__":
    main()
