from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SCREENSHOT_DIR = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help\Screenshots")
REL_PREFIX = "Screenshots"


SPECIAL_BASENAMES = {
    "ArticleEdit-Lagerhistorie": "ArticleEdit_Lagerungshistorie",
    "ArticleEdit-Verwendung in Sets": "ArticleEdit_VerwendungbeiSets",
    "ArticleEdit-Verwendung in Produktion": "ArticleEdit_VerwendungbeiProduktion",
    "PartnerEdit-Adresse": "PartnerEdit_Adresse",
    "PartnerEdit-Optionen": "SupplierPartnerClicked_Optionen",
    "PartnerEdit-Kundendetails": "SupplierPartnerClicked_Kundendetails",
    "PartnerEdit-Lieferantendetails": "SupplierPartnerClicked_Lieferantendetails",
    "PartnerEdit-Ansprechpartner": "SupplierPartnerClicked_Ansprechpartner",
    "PartnerEdit-Lieferadressen": "SupplierPartnerClicked_Lieferadressen",
    "PartnerEdit-Rechnungsadressen": "SupplierPartnerClicked_Rechnungsadressen",
    "PartnerEdit-Katalognummern": "SupplierPartnerClicked_Katalognummern",
    "PartnerEdit-Einkaufspreislisten": "SupplierPartnerClicked_Einkaufspreislisten",
    "PartnerEdit-Umsatz": "SupplierPartnerClicked_Umsatz",
    "PartnerEdit-Info": "SupplierPartnerClicked_Info",
    "PartnerEdit-Bild": "SupplierPartnerClicked_Bild",
    "PartnerEdit-Kategorie": "SupplierPartnerClicked_Kategorie",
    "PartnerEdit-CRM": "SupplierPartnerClicked_CRM",
    "PartnerEdit-E-Mail": "SupplierPartnerClicked_E-Mail",
    "PartnerEdit-Positionsliste": "SupplierPartnerClicked_Positionsliste",
    "PartnerEdit-Mahnungsliste": "SupplierPartnerClicked_Mahnungsliste",
    "PartnerEdit-Offene Posten": "SupplierPartnerClicked_OffenePosten",
    "PartnerEdit-Banking": "SupplierPartnerClicked_Banking",
    "PartnerEdit-Portal": "SupplierPartnerClicked_Portal",
    "PartnerEdit-Kontenblatt": "SupplierPartnerClicked_Kontenblatt",
    "PartnerEdit-Änderungsprotokoll": "SupplierPartnerClicked_Aenderungsprotokoll",
    "PartnerEdit-Anhänge": "SupplierPartnerClicked_Anhaenge",
    "PartnerEdit-Extra-Felder": "SupplierPartnerClicked_Extra-Felder",
    "PartnerEdit-Belegliste": "SupplierPartnerClicked_Dok-Liste",
}


def de_ascii(value: str) -> str:
    replacements = {
        "Ä": "Ae",
        "Ö": "Oe",
        "Ü": "Ue",
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)
    return value


def compact(value: str) -> str:
    value = de_ascii(value)
    return re.sub(r"[^A-Za-z0-9_-]+", "", value)


def normalized_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", de_ascii(value).lower())


def candidate_basenames(topic: str) -> list[str]:
    bases: list[str] = []
    if topic in SPECIAL_BASENAMES:
        bases.append(SPECIAL_BASENAMES[topic])

    topic_ascii = de_ascii(topic)
    bases.append(compact(topic_ascii))

    if "-" in topic_ascii:
        left, right = topic_ascii.split("-", 1)
        right_no_spaces = compact(right)
        bases.append(f"{compact(left)}_{right_no_spaces}")
        bases.append(f"{compact(left)}-{right_no_spaces}")

    bases.append(compact(topic_ascii.replace(" ", "")))
    bases.append(compact(topic_ascii.replace("-", "_").replace(" ", "")))

    result: list[str] = []
    seen = set()
    for base in bases:
        if base and base not in seen:
            seen.add(base)
            result.append(base)
    return result


def row_candidate_basenames(topic: str, original: str | None) -> list[str]:
    bases = candidate_basenames(topic)

    if topic.endswith("-Übersicht"):
        bases.extend(candidate_basenames(topic.removesuffix("-Übersicht")))

    if topic.endswith(" anlegen/bearbeiten") and original:
        bases.extend(candidate_basenames(original))

    result: list[str] = []
    seen = set()
    for base in bases:
        if base and base not in seen:
            seen.add(base)
            result.append(base)
    return result


def build_screenshot_index() -> dict[str, Path]:
    files = [
        path
        for path in SCREENSHOT_DIR.rglob("*")
        if path.is_file()
        and path.suffix.lower() in {".webp", ".png", ".jpg", ".jpeg"}
        and "failed" not in path.name.lower()
    ]

    index: dict[str, Path] = {}
    for path in sorted(files, key=lambda p: (p.suffix.lower() != ".webp", p.name.lower())):
        index.setdefault(path.stem, path)
        index.setdefault(normalized_key(path.stem), path)
    return index


def relative_screenshot(path: Path) -> str:
    try:
        rel = path.relative_to(SCREENSHOT_DIR).as_posix()
    except ValueError:
        rel = path.name
    return f"{REL_PREFIX}/{rel}"


def find_ansichten(ws) -> tuple[int, int, int]:
    start = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Ansichten")
    level = ws.row_dimensions[start].outlineLevel
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if ws.row_dimensions[row].outlineLevel <= level:
            end = row - 1
            break
    return start, end, level


def main() -> None:
    if not SCREENSHOT_DIR.exists():
        raise FileNotFoundError(SCREENSHOT_DIR)

    archive = WORKBOOK.parent / "ARCHIV"
    archive.mkdir(exist_ok=True)
    backup = archive / f"X-ERP-HELP-before-screenshot-assignment-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    screenshot_index = build_screenshot_index()
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    start, end, ansichten_level = find_ansichten(ws)

    assigned = 0
    kept = 0
    missing: list[str] = []
    examples: list[tuple[int, str, str]] = []

    for row in range(start + 1, end + 1):
        level = ws.row_dimensions[row].outlineLevel
        topic = ws.cell(row, 1).value
        original = ws.cell(row, 2).value
        if not isinstance(topic, str) or not topic.strip():
            continue

        # Assign screenshots only to actual views/wizards/register tabs/list rows,
        # not to fields and buttons.
        if level < ansichten_level + 2 or level > ansichten_level + 5:
            continue
        if level >= ansichten_level + 4 and not (topic.endswith("List") or " anlegen/bearbeiten" in topic):
            continue

        current = ws.cell(row, 5).value
        found = None
        for base in row_candidate_basenames(topic, original if isinstance(original, str) else None):
            found = screenshot_index.get(base) or screenshot_index.get(normalized_key(base))
            if found:
                break

        if found:
            new_value = relative_screenshot(found)
            if current == new_value:
                kept += 1
            else:
                ws.cell(row, 5).value = new_value
                assigned += 1
                if len(examples) < 25:
                    examples.append((row, topic, new_value))
        elif level <= ansichten_level + 3:
            missing.append(topic)

    wb.save(WORKBOOK)

    print(f"assigned={assigned}")
    print(f"kept={kept}")
    print(f"missing_view_or_tab={len(missing)}")
    print(f"backup={backup}")
    for row, topic, screenshot in examples:
        print(f"example\t{row}\t{topic}\t{screenshot}")
    if missing:
        print("missing_sample=" + "; ".join(missing[:80]))


if __name__ == "__main__":
    main()
