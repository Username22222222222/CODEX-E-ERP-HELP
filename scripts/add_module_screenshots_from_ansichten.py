from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
ANSICHTEN_ROOT = HELP_ROOT / "Ansichten"


MODULE_DEFAULTS = {
    "verkauf": "belege/doc-edit/DocEdit.webp",
    "einkauf": "belege/doc-edit/DocEdit.webp",
    "kommissionieren": "lager/warehouse-quantity-edit/WarehouseQuantityEdit.webp",
    "packen": "packen/packing-list-box-edit/PackingListBoxEdit.webp",
    "versand": "versand/shipping-carrier-edit/ShippingCarrierEdit.webp",
    "produktion": "produktion/production-edit/ProductionEdit.webp",
    "beschaffungsassistent": "beschaffungsassistent/procurement-wizard/ProcurementWizard.webp",
    "aussendienst": "belege/field-service-resource-edit/FieldServiceResourceEdit.webp",
    "geraeteservice": "belege/device-service-doc-edit/DeviceServiceDocEdit.webp",
    "vertragsmanagement": "artikel/article-edit/article-edit-vertraege/ArticleEdit_Vertraege.webp",
    "verknuepfungsplan": "belege/doc-edit/DocEdit.webp",
    "genehmigungen": "benutzer/authorization-page-edit/AuthorizationPageEdit.webp",
    "crm": "crm/crm-edit/CrmEdit.webp",
    "intercom": "intercom/intercom-edit/IntercomEdit.webp",
    "wiki": "wiki/wiki-edit/WikiEdit.webp",
    "bulletin": "bulletin/bulletin-edit/BulletinEdit.webp",
    "zeiterfassung": "zeiterfassung/time-tracking-edit/TimeTrackingEdit.webp",
    "archivierung": "anhaenge/attachment-edit/AttachmentEdit.webp",
    "elektronische-haftnotizen": "elektronische-haftnotizen/sticky-note-edit/StickyNoteEdit.webp",
    "e-mail": "e-mail/email-edit/EmailEdit.webp",
    "kalender": "kalender/calendar-edit/CalendarEdit.webp",
    "ki-assistent": "ki-assistent/ai-text-info-edit/AiTextInfoEdit.webp",
    "finanzwesen": "finanzwesen/finance-open-item-edit/FinanceOpenItemEdit.webp",
}


SOURCE_BY_PATH = {
    "Module/index.html": "dashboard/dashboard-edit/DashboardEdit.webp",
    "Module/einkauf/bestellanforderung/index.html": "beschaffungsassistent/procurement-planning-ask-edit/ProcurementPlanningAskEdit.webp",
    "Module/einkauf/wareneingang/index.html": "lager/warehouse-quantity-edit/WarehouseQuantityEdit.webp",
    "Module/packen/packstueck/index.html": "packen/packing-list-box-edit/PackingListBoxEdit.webp",
    "Module/packen/packliste/index.html": "packen/packing-list-box-position-edit/PackingListBoxPositionEdit.webp",
    "Module/versand/frachtfuehrer/index.html": "versand/shipping-carrier-edit/ShippingCarrierEdit.webp",
    "Module/versand/versandauftrag/index.html": "versand/shipping-type-edit/ShippingTypeEdit.webp",
    "Module/versand/sendungsverfolgung/index.html": "versand/shipping-rate-edit/ShippingRateEdit.webp",
    "Module/versand/versandpapiere/index.html": "versand/shipping-zone-edit/ShippingZoneEdit.webp",
    "Module/produktion/stueckliste/index.html": "artikel/article-edit/article-edit-produktion/article-product-list/article-product-edit-anlegen-bearbeiten/ArticleProductEdit.webp",
    "Module/produktion/fertigungsauftrag/index.html": "produktion/production-edit/ProductionEdit.webp",
    "Module/produktion/materialentnahme/index.html": "artikel/article-edit/article-edit-produktionsschritt-artikel/article-production-step-component-list/article-production-step-component-edit-anlegen-bearbeiten/ArticleProductionStepComponentEdit.webp",
    "Module/produktion/rueckmeldung/index.html": "produktion/production-data-acquisition-edit/ProductionDataAcquisitionEdit.webp",
    "Module/beschaffungsassistent/bedarfsermittlung/index.html": "beschaffungsassistent/procurement-planning-ask-edit/ProcurementPlanningAskEdit.webp",
    "Module/beschaffungsassistent/bestellvorschlag/index.html": "beschaffungsassistent/procurement-planning-buy-edit/ProcurementPlanningBuyEdit.webp",
    "Module/beschaffungsassistent/beschaffungsart/index.html": "beschaffungsassistent/procurement-planning-make-edit/ProcurementPlanningMakeEdit.webp",
    "Module/geraeteservice/geraeteakte/index.html": "belege/device-service-doc-edit/DeviceServiceDocEdit.webp",
    "Module/geraeteservice/serviceauftrag/index.html": "belege/device-service-doc-edit/DeviceServiceDocEdit.webp",
    "Module/crm/kontakt/index.html": "crm/crm-edit/CrmEdit.webp",
    "Module/crm/aktivitaet/index.html": "crm/crm-activity-edit/CrmActivityEdit.webp",
    "Module/crm/verkaufschance/index.html": "crm/crm-edit/crm-edit-uebersicht/CrmEdit_Uebersicht.webp",
    "Module/crm/kampagne/index.html": "crm/crm-group-edit/CrmGroupEdit.webp",
    "Module/intercom/nachricht/index.html": "intercom/intercom-edit/IntercomEdit.webp",
    "Module/intercom/erwaehnung/index.html": "intercom/intercom-edit/intercom-edit-uebersicht/IntercomEdit_Uebersicht.webp",
    "Module/intercom/benachrichtigung/index.html": "intercom/intercom-recipient-status-edit/IntercomRecipientStatusEdit.webp",
    "Module/wiki/artikel/index.html": "wiki/wiki-edit/WikiEdit.webp",
    "Module/wiki/kategorie/index.html": "wiki/wiki-category-edit/WikiCategoryEdit.webp",
    "Module/wiki/suche/index.html": "wiki/wiki-status-edit/WikiStatusEdit.webp",
    "Module/bulletin/aushang/index.html": "bulletin/bulletin-edit/BulletinEdit.webp",
    "Module/bulletin/gueltigkeit/index.html": "bulletin/bulletin-edit/bulletin-edit-uebersicht/BulletinEdit_Uebersicht.webp",
    "Module/bulletin/kenntnisnahme/index.html": "bulletin/bulletin-category-edit/BulletinCategoryEdit.webp",
    "Module/zeiterfassung/zeitbuchung/index.html": "zeiterfassung/time-tracking-edit/TimeTrackingEdit.webp",
    "Module/zeiterfassung/projektzeit/index.html": "zeiterfassung/time-tracking-edit/time-tracking-edit-uebersicht/TimeTrackingEdit_Uebersicht.webp",
    "Module/zeiterfassung/zeitkonto/index.html": "zeiterfassung/time-tracking-setting-edit/TimeTrackingSettingEdit.webp",
    "Module/zeiterfassung/abwesenheit/index.html": "zeiterfassung/time-tracking-setting-edit/time-tracking-setting-edit-uebersicht/TimeTrackingSettingEdit.webp",
    "Module/archivierung/dokumentenarchiv/index.html": "anhaenge/attachment-edit/AttachmentEdit.webp",
    "Module/archivierung/revisionssicherheit/index.html": "anhaenge/attachment-status-edit/AttachmentStatusEdit.webp",
    "Module/archivierung/aufbewahrungsfrist/index.html": "belege/doc-edit/DocEdit.webp",
    "Module/e-mail/posteingang/index.html": "e-mail/email-edit/EmailEdit.webp",
    "Module/e-mail/belegversand/index.html": "e-mail/email-template-edit/EmailTemplateEdit.webp",
    "Module/e-mail/vorlage/index.html": "e-mail/email-template-edit/EmailTemplateEdit.webp",
    "Module/e-mail/zuordnung/index.html": "e-mail/email-rule-in-edit/EmailRuleInEdit.webp",
    "Module/kalender/termin/index.html": "kalender/calendar-appointment-edit/CalendarAppointmentEdit.webp",
    "Module/kalender/aufgabe/index.html": "kalender/calendar-edit/CalendarEdit.webp",
    "Module/kalender/wiedervorlage/index.html": "kalender/calendar-working-day-wizard/CalendarWorkingDayWizard.webp",
    "Module/kalender/freigabe-und-teilen/index.html": "kalender/calendar-name-edit/CalendarNameEdit.webp",
    "Module/ki-assistent/textunterstuetzung/index.html": "ki-assistent/ai-text-info-edit/AiTextInfoEdit.webp",
    "Module/ki-assistent/datenabfrage/index.html": "ki-assistent/ai-email-info-edit/AiEmailInfoEdit.webp",
    "Module/ki-assistent/vorschlaege/index.html": "ki-assistent/ai-text-info-edit/AiTextInfoEdit.webp",
    "Module/finanzwesen/op-verwaltung/index.html": "finanzwesen/finance-open-item-edit/FinanceOpenItemEdit.webp",
    "Module/finanzwesen/zahlungsverkehr/index.html": "finanzwesen/finance-payment-edit/FinancePaymentEdit.webp",
    "Module/finanzwesen/mahnwesen/index.html": "finanzwesen/finance-dunning-edit/FinanceDunningEdit.webp",
    "Module/finanzwesen/kostenrechnung/index.html": "finanzwesen/finance-cost-center-edit/FinanceCostCenterEdit.webp",
    "Module/finanzwesen/anlagenbuchhaltung/index.html": "finanzwesen/finance-fixed-asset-edit/FinanceFixedAssetEdit.webp",
    "Module/finanzwesen/kontenfindung/index.html": "finanzwesen/finance-account-determination-edit/FinanceAccountDeterminationEdit.webp",
    "Module/finanzwesen/kontinuierliche-bestandswertbuchungen/index.html": "finanzwesen/finance-group-article-edit/FinanceGroupArticleEdit.webp",
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_path(path: str) -> str:
    return path.replace("\\", "/").lstrip("/")


def file_safe_stem(value: str) -> str:
    stem = Path(value.replace("\\", "/")).stem
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_")
    return stem or "screenshot"


def module_key(url_path: str) -> str:
    parts = normalize_path(url_path).split("/")
    return parts[1] if len(parts) > 1 and parts[0] == "Module" else ""


def source_for(url_path: str) -> str:
    normalized = normalize_path(url_path)
    if normalized in SOURCE_BY_PATH:
        return SOURCE_BY_PATH[normalized]
    key = module_key(normalized)
    return MODULE_DEFAULTS.get(key, "")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-module-screenshots-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False, keep_links=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }

    updated = 0
    copied = 0
    missing: list[str] = []
    unmapped: list[str] = []
    for row in range(2, ws.max_row + 1):
        url_path = normalize_path(cell_text(ws.cell(row, headers["URL_PATH"]).value))
        content_type = cell_text(ws.cell(row, headers["CONTENT_TYPE"]).value)
        if not url_path.startswith("Module/") or content_type not in {"HelpPage", "FAQ"}:
            continue

        source_rel = source_for(url_path)
        if not source_rel:
            unmapped.append(url_path)
            continue
        source_file = ANSICHTEN_ROOT / source_rel
        if not source_file.is_file():
            missing.append(f"{url_path} -> {source_rel}")
            continue

        target_dir = HELP_ROOT / url_path.removesuffix("index.html").rstrip("/")
        target_dir.mkdir(parents=True, exist_ok=True)
        target_name = file_safe_stem(source_file.name) + "_Module.webp"
        target_file = target_dir / target_name
        if not target_file.exists() or target_file.read_bytes() != source_file.read_bytes():
            shutil.copy2(source_file, target_file)
            copied += 1

        target_rel = target_file.relative_to(HELP_ROOT).as_posix()
        web_path = "/de/help/" + target_rel
        topic = cell_text(ws.cell(row, headers["Thema"]).value)
        ws.cell(row, headers["Screenshot"]).value = target_rel
        ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value = web_path
        ws.cell(row, headers["IMAGE_ALT"]).value = f"Screenshot {topic} in X-ERP"
        ws.cell(row, headers["IMAGE_CAPTION"]).value = f"{topic} im X-ERP Modulbereich"
        ws.cell(row, headers["IMAGE_STATUS"]).value = "local-copy:module"
        updated += 1

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"updated_rows={updated}")
    print(f"copied_files={copied}")
    print(f"unmapped={len(unmapped)}")
    for item in unmapped:
        print(f"UNMAPPED {item}")
    print(f"missing_sources={len(missing)}")
    for item in missing:
        print(f"MISSING {item}")


if __name__ == "__main__":
    main()
