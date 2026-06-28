from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
ARCHIVE = Path(r"C:\Users\micha\Documents\X-ERP-HELP\ARCHIV")
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")
SHEET = "de-DE"


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_rel_path(value: str) -> str:
    value = value.strip().replace("\\", "/")
    if value.startswith("https://x-erp.de/de/help/"):
        value = value[len("https://x-erp.de/de/help/") :]
    if value.startswith("/de/help/"):
        value = value[len("/de/help/") :]
    if value.startswith("de/help/"):
        value = value[len("de/help/") :]
    if value.startswith("/"):
        value = value[1:]
    return value


def source_file(path_value: str) -> Path:
    rel = normalize_rel_path(path_value)
    return HELP_ROOT / rel


def page_dir(url_path: str) -> Path | None:
    rel = normalize_rel_path(url_path)
    if not rel:
        return None
    parts = [part for part in rel.split("/") if part]
    if not parts:
        return None
    if parts[0].casefold() == "ansichten":
        parts[0] = "Ansichten"
    if parts[-1].casefold() == "index.html" or "." in parts[-1]:
        parts = parts[:-1]
    if not parts:
        return None
    return HELP_ROOT.joinpath(*parts)


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-local-screenshot-paths-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    ws = wb[SHEET]
    headers = {
        text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if text(ws.cell(1, col).value)
    }
    required = {"Thema", "URL_PATH", "Screenshot", "SCREENSHOT_REL_PATH", "SCREENSHOT_WEB_PATH", "IMAGE_STATUS"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"Missing required headers: {', '.join(missing)}")

    copied = 0
    updated = 0
    missing_files: list[tuple[int, str, str]] = []
    skipped_no_url = 0
    operations: list[tuple[Path, Path]] = []

    for row in range(2, ws.max_row + 1):
        raw_shot = text(ws.cell(row, headers["SCREENSHOT_REL_PATH"]).value) or text(ws.cell(row, headers["Screenshot"]).value)
        if not raw_shot:
            continue
        target_dir = page_dir(text(ws.cell(row, headers["URL_PATH"]).value))
        if target_dir is None:
            skipped_no_url += 1
            continue
        src = source_file(raw_shot)
        if not src.is_file():
            missing_files.append((row, text(ws.cell(row, headers["Thema"]).value), raw_shot))
            continue
        dst = target_dir / src.name
        operations.append((src, dst))

        rel = dst.relative_to(HELP_ROOT).as_posix()
        ws.cell(row, headers["Screenshot"]).value = rel
        ws.cell(row, headers["SCREENSHOT_REL_PATH"]).value = rel
        ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value = "/de/help/" + rel
        if text(ws.cell(row, headers["IMAGE_STATUS"]).value).casefold() not in {"kein screenshot", "missing"}:
            ws.cell(row, headers["IMAGE_STATUS"]).value = "local"
        updated += 1

    for src, dst in operations:
        dst.parent.mkdir(parents=True, exist_ok=True)
        if not dst.exists() or src.read_bytes() != dst.read_bytes():
            shutil.copy2(src, dst)
            copied += 1

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"updated_rows={updated}")
    print(f"copied_files={copied}")
    print(f"missing_files={len(missing_files)}")
    print(f"skipped_no_url={skipped_no_url}")
    for row, topic, shot in missing_files[:50]:
        print(f"missing\t{row}\t{topic}\t{shot}")


if __name__ == "__main__":
    main()
