from __future__ import annotations

import copy
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"

SECTION_TARGETS = {
    "partner-adresse": ("Partner", "PartnerEdit", "PartnerEdit-Adresse"),
    "partner-optionen": ("Partner", "PartnerEdit", "PartnerEdit-Optionen"),
    "partner-kundendetails": ("Partner", "PartnerEdit", "PartnerEdit-Kundendetails"),
    "partner-lieferantendetails": ("Partner", "PartnerEdit", "PartnerEdit-Lieferantendetails"),
    "partner-ansprechpartner": ("Partner", "PartnerContactPersonEdit", "PartnerContactPersonEdit-Übersicht"),
    "partner-lieferadressen": ("Partner", "PartnerDeliveryAddressEdit", "PartnerDeliveryAddressEdit-Neuer Matchcode"),
    "partner-rechnungsadressen": ("Partner", "PartnerBillingAddressEdit", "PartnerBillingAddressEdit-Neuer Matchcode"),
    "partner-banking": ("Partner", "PartnerBankAccountEdit", "PartnerBankAccountEdit-Übersicht"),
    "partner-katalognummern": ("Partner", "PartnerCatalogNumberEdit", "PartnerCatalogNumberEdit-Übersicht"),
    "partner-einkaufspreislisten": ("Partner", "PartnerPurchasePriceListEdit", "PartnerPurchasePriceListEdit-Übersicht"),
    "partner-bild": ("Partner", "PartnerEdit", "PartnerEdit-Bild"),
    "partner-crm": ("Partner", "PartnerEdit", "PartnerEdit-CRM"),
    "partner-dok-liste": ("Partner", "PartnerEdit", "PartnerEdit-Belegliste"),
    "partner-e-mail": ("Partner", "PartnerEdit", "PartnerEdit-E-Mail"),
    "partner-extra-felder": ("Partner", "PartnerEdit", "PartnerEdit-Extra-Felder"),
    "partner-info": ("Partner", "PartnerEdit", "PartnerEdit-Info"),
    "partner-kategorie": ("Partner", "PartnerEdit", "PartnerEdit-Kategorie"),
    "partner-kontenblatt": ("Partner", "PartnerEdit", "PartnerEdit-Kontenblatt"),
    "partner-lokal": ("Partner", "PartnerEdit", "PartnerEdit-Lokal"),
    "partner-mahnungsliste": ("Partner", "PartnerEdit", "PartnerEdit-Mahnungsliste"),
    "partner-offene-posten": ("Partner", "PartnerEdit", "PartnerEdit-Offene Posten"),
    "partner-portal": ("Partner", "PartnerEdit", "PartnerEdit-Portal"),
    "partner-positionsliste": ("Partner", "PartnerEdit", "PartnerEdit-Positionsliste"),
    "partner-umsatz": ("Partner", "PartnerEdit", "PartnerEdit-Umsatz"),
    "partner-aenderungsprotokoll": ("Partner", "PartnerEdit", "PartnerEdit-Änderungsprotokoll"),
    "partner-anhaenge": ("Partner", "PartnerEdit", "PartnerEdit-Anhänge"),
    "preise-einkauf": ("Partner", "PartnerPurchasePriceListEdit", "PartnerPurchasePriceListEdit-Übersicht"),
    "preise-verkauf": ("Artikel", "ArticleSalesPriceListEdit", "ArticleSalesPriceListEdit-Fremdwährung"),
    "dok": ("Belege", "DocEdit_Overview", "DocEdit_Overview-Tel.1"),
    "dok-positionen": ("Belege", "DocEdit", "DocEdit-Positionen"),
    "dok-kette": ("Belege", "DocEdit", "DocEdit-Dok-Kette"),
    "belegkette": ("Belege", "DocEdit", "DocEdit-Belegkette"),
    "lager": ("Lager", "WarehouseEdit", "WarehouseEdit-Historie"),
    "lagerplatz": ("Lager", "WarehouseBinLocationEdit", "WarehouseBinLocationEdit-Uebersicht"),
    "lagertyp": ("Lager", "WarehouseTypeEdit", "WarehouseTypeEdit-Übersicht"),
    "lagermedium": ("Lager", "WarehouseMediumEdit", "WarehouseMediumEdit-Übersicht"),
    "mitarbeiter-stammdaten": ("Mitarbeiter", "EmployeeEdit", "EmployeeEdit-Übersicht"),
    "mitarbeiter-gruppe": ("Mitarbeiter", "EmployeeGroupEdit", "EmployeeGroupEdit-Übersicht"),
    "mitarbeiter-provisionsgruppe": (
        "Mitarbeiter",
        "EmployeeCommissionGroupEdit",
        "EmployeeCommissionGroupEdit-Übersicht",
    ),
    "mitarbeiter-urlaub": (
        "Mitarbeiter",
        "EmployeeVacationEntitlementEdit",
        "EmployeeVacationEntitlementEdit-EmployeeVacationEntitlementHistoryTableDetail",
    ),
    "textbloecke-stammdaten": ("Textblöcke", "TextBlockEdit", "TextBlockEdit-Übersicht"),
    "textbloecke-gruppe": ("Textblöcke", "TextBlockGroupEdit", "TextBlockGroupEdit-Übersicht"),
    "textbloecke-variablen": ("Textblöcke", "TextVariableEdit", "TextVariableEdit-Versand"),
    "anhaenge-stammdaten": ("Anhänge", "AttachmentEdit", "AttachmentEdit-Hochgeladene Dateien"),
    "anhaenge-dateien": ("Anhänge", "AttachmentEdit", "AttachmentEdit-Hochgeladene Dateien"),
}

FIELD_TARGET_OVERRIDES = {
    ("dok", "sachbearbeiter"): ("Belege", "DocEdit_Overview", "DocEdit_Overview-E-Mail"),
    ("dok", "verkaeufer"): ("Belege", "DocEdit_Overview", "DocEdit_Overview-E-Mail"),
    ("dok", "einkaeufer"): ("Belege", "DocEdit_Overview", "DocEdit_Overview-E-Mail"),
    ("dok", "abw-rechnungsempfaenger"): ("Belege", "DocEdit_Overview", "DocEdit_Overview-E-Mail"),
    ("dok", "interne-bemerkung"): ("Belege", "DocEdit_Overview", "DocEdit_Overview-E-Mail"),
    ("dok", "dok-typ"): ("Belege", "DocEdit_Overview", "DocEdit_Overview-E-Mail"),
}

ALIASES = {
    ("lager", "einlagersperre"): "eingangssperre",
    ("lager", "auslagersperre"): "ausgangssperre",
    ("lagerplatz", "eltern"): "parent",
    ("preise-einkauf", "lieferant"): "supplier",
    ("preise-einkauf", "artikel"): "article",
    ("preise-einkauf", "katalognummer"): "catalog-number",
    ("preise-einkauf", "liefertage"): "delivery-days",
    ("preise-einkauf", "mindestmenge"): "minimum-quantity",
    ("preise-einkauf", "losgroesse"): "lot-size",
    ("preise-einkauf", "kommentar"): "comment",
    ("preise-einkauf", "einkaufspreis"): "purchase-price",
    ("preise-einkauf", "rabatt"): "discount-prozent",
    ("preise-einkauf", "logistikkosten"): "logistic-costs",
    ("preise-einkauf", "zollkosten"): "customs-costs",
    ("preise-einkauf", "risikokosten"): "risk-costs",
    ("preise-einkauf", "einstandspreis"): "landed-cost",
    ("preise-einkauf", "preiseinheit"): "price-unit",
    ("preise-einkauf", "einheitspreis"): "unit-price",
    ("preise-einkauf", "mengenstaffel-a"): "quantity-level-a",
    ("preise-einkauf", "mengenstaffel-b"): "quantity-level-b",
    ("preise-einkauf", "mengenstaffel-c"): "quantity-level-c",
    ("preise-einkauf", "mengenrabatt-a"): "mengenrabatt-a-prozent",
    ("preise-einkauf", "mengenrabatt-b"): "mengenrabatt-b-prozent",
    ("preise-einkauf", "mengenrabatt-c"): "mengenrabatt-c-prozent",
    ("preise-einkauf", "mengenrabatt-d"): "mengenrabatt-d-prozent",
    ("preise-einkauf", "mengenrabatt-e"): "mengenrabatt-e-prozent",
    ("preise-verkauf", "gueltig-ab"): "gueltig-ab",
    ("partner-einkaufspreislisten", "rabatt"): "rabatt-prozent",
    ("partner-einkaufspreislisten", "mengenrabatt-a"): "mengenrabatt-a-prozent",
    ("partner-einkaufspreislisten", "mengenrabatt-b"): "mengenrabatt-b-prozent",
    ("partner-einkaufspreislisten", "mengenrabatt-c"): "mengenrabatt-c-prozent",
    ("partner-einkaufspreislisten", "mengenrabatt-d"): "mengenrabatt-d-prozent",
    ("partner-einkaufspreislisten", "mengenrabatt-e"): "mengenrabatt-e-prozent",
    ("partner-offene-posten", "steuer"): "steuer-prozent",
    ("partner-kontenblatt", "steuer"): "steuer-prozent",
    ("dok-positionen", "rabatt"): "rabatt-prozent",
    ("mitarbeiter-provisionsgruppe", "provision"): "provision-prozent",
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def slug(value: str) -> str:
    value = value.replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue")
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    value = value.replace("%", "prozent")
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "seite"


def headers_for(ws) -> dict[str, int]:
    return {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }


def source_fields(ws, headers: dict[str, int]) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    for row in range(2, ws.max_row + 1):
        path = cell_text(ws.cell(row, headers["URL_PATH"]).value)
        if not path.startswith("Stammdaten/") or not path.endswith("/index.html"):
            continue
        breadcrumb = cell_text(ws.cell(row, headers["BREADCRUMB"]).value)
        if not breadcrumb.startswith("Stammdaten >"):
            continue
        level = int(ws.row_dimensions[row].outlineLevel or 0)
        if level < 3:
            continue
        parts = path.split("/")
        if len(parts) < 5:
            continue
        section_slug = parts[2]
        field_slug = parts[3]
        if section_slug not in SECTION_TARGETS:
            continue
        result.append(
            {
                "row": row,
                "path": path,
                "section_slug": section_slug,
                "field_slug": field_slug,
                "target_slug": ALIASES.get((section_slug, field_slug), field_slug),
                "target_parts": FIELD_TARGET_OVERRIDES.get(
                    (section_slug, field_slug),
                    SECTION_TARGETS[section_slug],
                ),
            }
        )
    return result


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.font:
        dst.font = copy.copy(src.font)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.border:
        dst.border = copy.copy(src.border)
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.number_format:
        dst.number_format = src.number_format
    if src.protection:
        dst.protection = copy.copy(src.protection)


def last_descendant_row(ws, parent_row: int) -> int:
    parent_level = int(ws.row_dimensions[parent_row].outlineLevel or 0)
    row = parent_row + 1
    last = parent_row
    while row <= ws.max_row:
        level = int(ws.row_dimensions[row].outlineLevel or 0)
        topic = ws.cell(row, 1).value
        if topic not in (None, "") and level <= parent_level:
            break
        last = row
        row += 1
    return last


def breadcrumbs(ws, headers: dict[str, int]) -> dict[str, int]:
    return {
        cell_text(ws.cell(row, headers["BREADCRUMB"]).value): row
        for row in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row, headers["BREADCRUMB"]).value)
    }


def target_maps(ws, headers: dict[str, int]) -> tuple[dict[tuple[tuple[str, ...], str], int], dict[tuple[str, ...], int]]:
    topic_col = headers["Thema"]
    breadcrumb_col = headers["BREADCRUMB"]
    field_targets: dict[tuple[tuple[str, ...], str], int] = {}
    parent_targets: dict[tuple[str, ...], int] = {}
    for row in range(2, ws.max_row + 1):
        breadcrumb = cell_text(ws.cell(row, breadcrumb_col).value)
        if not breadcrumb.startswith("Ansichten > "):
            continue
        parts = tuple(part.strip() for part in breadcrumb.split(">"))
        if len(parts) < 2:
            continue
        tail = parts[1:]
        parent_targets[tail] = row
        parent_tail = parts[1:-1]
        if parent_tail:
            topic_slug = slug(cell_text(ws.cell(row, topic_col).value))
            field_targets[(parent_tail, topic_slug)] = row
    return field_targets, parent_targets


def normalize_relative_path(path: str) -> str:
    path = path.replace("\\", "/").strip()
    if path.startswith("/de/help/"):
        path = path[len("/de/help/") :]
    if path.startswith("de/help/"):
        path = path[len("de/help/") :]
    if path.startswith("ansichten/"):
        path = "Ansichten/" + path[len("ansichten/") :]
    if path.startswith("Ansichten/") and path.endswith("/"):
        path += "index.html"
    elif path.startswith("Ansichten/") and not path.endswith("index.html"):
        path = path.rstrip("/") + "/index.html"
    return path


def parent_url_path(ws, headers: dict[str, int], parent_row: int) -> str:
    path = normalize_relative_path(cell_text(ws.cell(parent_row, headers["URL_PATH"]).value))
    if not path:
        raise RuntimeError(f"Parent row {parent_row} has no URL_PATH")
    return path.removesuffix("index.html").rstrip("/")


def insert_copy_row(ws, headers: dict[str, int], source_row: int, parent_row: int, topic: str, content_type: str) -> int:
    insert_at = last_descendant_row(ws, parent_row) + 1
    ws.insert_rows(insert_at)
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(source_row, col), ws.cell(insert_at, col))
        ws.cell(insert_at, col).value = ws.cell(source_row, col).value
    ws.row_dimensions[insert_at].outlineLevel = int(ws.row_dimensions[parent_row].outlineLevel or 0) + 1
    ws.row_dimensions[insert_at].hidden = True
    ws.row_dimensions[insert_at].collapsed = False
    parent_bc = cell_text(ws.cell(parent_row, headers["BREADCRUMB"]).value)
    ws.cell(insert_at, headers["Thema"]).value = topic
    ws.cell(insert_at, headers["BREADCRUMB"]).value = f"{parent_bc} > {topic}"
    ws.cell(insert_at, headers["URL_PATH"]).value = None
    ws.cell(insert_at, headers["CONTENT_TYPE"]).value = content_type
    return insert_at


def set_if_present(ws, headers: dict[str, int], row: int, column: str, value: object) -> None:
    col = headers.get(column)
    if col:
        ws.cell(row, col).value = value


def ensure_parent(ws, headers: dict[str, int], source_row: int, target_parts: tuple[str, ...]) -> int:
    by_breadcrumb = breadcrumbs(ws, headers)
    current_bc = "Ansichten"
    parent_row = by_breadcrumb[current_bc]
    for part in target_parts:
        current_bc = f"{current_bc} > {part}"
        row = by_breadcrumb.get(current_bc)
        if row:
            parent_row = row
            continue
        row = insert_copy_row(ws, headers, source_row, parent_row, part, "HelpPage")
        configure_target(ws, headers, source_row, row, page_topic=part)
        by_breadcrumb = breadcrumbs(ws, headers)
        parent_row = row
    return parent_row


def nearest_url_parent(ws, headers: dict[str, int], row: int) -> int:
    breadcrumb = cell_text(ws.cell(row, headers["BREADCRUMB"]).value)
    parts = [part.strip() for part in breadcrumb.split(">") if part.strip()]
    by_breadcrumb = breadcrumbs(ws, headers)
    for end in range(len(parts) - 1, 0, -1):
        parent_bc = " > ".join(parts[:end])
        parent_row = by_breadcrumb.get(parent_bc)
        if parent_row and cell_text(ws.cell(parent_row, headers["URL_PATH"]).value):
            return parent_row
    raise RuntimeError(f"No URL parent found for row {row}")


def configure_target(ws, headers: dict[str, int], source_row: int, target_row: int, page_topic: str | None = None) -> str:
    topic = page_topic or cell_text(ws.cell(target_row, headers["Thema"]).value) or cell_text(ws.cell(source_row, headers["Thema"]).value)
    parent_row = nearest_url_parent(ws, headers, target_row)
    target_path = f"{parent_url_path(ws, headers, parent_row)}/{slug(topic)}/index.html"

    for column in ("Beschreibung", "PRIMARY_KEYWORD", "IMAGE_ALT", "IMAGE_CAPTION", "IMAGE_STATUS"):
        if column in headers:
            value = ws.cell(source_row, headers[column]).value
            if value not in (None, ""):
                ws.cell(target_row, headers[column]).value = value

    set_if_present(ws, headers, target_row, "URL_PATH", target_path)
    set_if_present(ws, headers, target_row, "CONTENT_TYPE", "HelpPage")
    set_if_present(ws, headers, target_row, "H1", topic)
    set_if_present(ws, headers, target_row, "NAV_TITLE", topic)
    set_if_present(ws, headers, target_row, "TITLE", f"{topic} | X-ERP ERP Hilfe")
    description = cell_text(ws.cell(target_row, headers["Beschreibung"]).value)
    meta = description.splitlines()[0].lstrip("# ").strip() if description else ""
    if not meta:
        meta = f"{topic} in der passenden X-ERP Ansicht."
    set_if_present(ws, headers, target_row, "META_DESCRIPTION", meta[:300])
    return target_path


def retire_source(ws, headers: dict[str, int], source_row: int, target_path: str) -> None:
    topic = cell_text(ws.cell(source_row, headers["Thema"]).value)
    set_if_present(ws, headers, source_row, "URL_PATH", None)
    set_if_present(ws, headers, source_row, "CONTENT_TYPE", "InlineSection")
    set_if_present(ws, headers, source_row, "Beschreibung", f"Verschoben nach {target_path}.")
    for column in ("H1", "TITLE", "META_DESCRIPTION"):
        set_if_present(ws, headers, source_row, column, None)
    set_if_present(ws, headers, source_row, "NAV_TITLE", topic)


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-move-remaining-stammdaten-fields-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = headers_for(ws)

    sources = source_fields(ws, headers)
    created_parents = 0
    created_fields = 0
    for source in sources:
        parent_row = ensure_parent(ws, headers, int(source["row"]), tuple(source["target_parts"]))
        _, parent_targets = target_maps(ws, headers)
        if tuple(source["target_parts"]) not in parent_targets:
            raise RuntimeError(f"Missing parent after create: {source['target_parts']}")
        field_targets, _ = target_maps(ws, headers)
        key = (tuple(source["target_parts"]), str(source["target_slug"]))
        if key not in field_targets:
            insert_copy_row(
                ws,
                headers,
                int(source["row"]),
                parent_row,
                cell_text(ws.cell(int(source["row"]), headers["Thema"]).value),
                "HelpPage",
            )
            created_fields += 1

    moved = 0
    skipped: list[str] = []
    for source in source_fields(ws, headers):
        field_targets, _ = target_maps(ws, headers)
        key = (tuple(source["target_parts"]), str(source["target_slug"]))
        target_row = field_targets.get(key)
        if not target_row:
            skipped.append(str(source["path"]))
            continue
        target_path = configure_target(ws, headers, int(source["row"]), target_row)
        retire_source(ws, headers, int(source["row"]), target_path)
        moved += 1

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"created_parent_rows={created_parents}")
    print(f"created_field_rows={created_fields}")
    print(f"moved_fields={moved}")
    print(f"skipped={len(skipped)}")
    for path in skipped:
        print(f"SKIPPED {path}")


if __name__ == "__main__":
    main()
