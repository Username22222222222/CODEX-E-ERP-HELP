from __future__ import annotations

import datetime as dt
import re
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"


def copy_cell_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)


def apply_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed


def is_junk_register(name: str) -> bool:
    return bool(
        re.search(r"-(right|md-\d+|tabs-fill|tabpane-fill)$", name)
        or name in {"right", "md-7", "md-4", "tabs-fill", "tabpane-fill"}
    )


def main() -> None:
    ARCHIVE.mkdir(exist_ok=True)
    shutil.copy2(WORKBOOK, ARCHIVE / f"X-ERP-HELP-before-repair-extended-outline-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx")
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}

    # Remove CSS/layout fragments that were extracted as pseudo registers.
    remove_rows = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, col["Thema"]).value
        if isinstance(name, str) and is_junk_register(name):
            remove_rows.append(row)
    for row in sorted(remove_rows, reverse=True):
        ws.delete_rows(row, 1)

    # Stable style source rows from the validated part of the sheet.
    style = {
        "group": 1141,
        "view": 1142,
        "tab": 1224,
        "field4": 1225,
        "field3": 1206,
    }

    start = None
    end = None
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, col["Thema"]).value
        if name == "Nachtrag - Dashboard" and start is None:
            start = row
        if name == "FAQ (Häufig gestellte Fragen)" and start is not None:
            end = row
            break
    if start is None or end is None:
        raise RuntimeError("Could not locate extended view block")

    # Determine which views have explicit register rows.
    has_tabs: dict[str, bool] = {}
    current_view = None
    for row in range(start, end):
        name = ws.cell(row, col["Thema"]).value
        content_type = ws.cell(row, col["CONTENT_TYPE"]).value
        if content_type in ("View", "Wizard"):
            current_view = str(name)
            has_tabs[current_view] = False
        elif current_view and isinstance(name, str) and name.startswith(f"{current_view}-"):
            has_tabs[current_view] = True

    current_view = None
    current_has_tabs = False
    for row in range(start, end):
        name = ws.cell(row, col["Thema"]).value
        content_type = ws.cell(row, col["CONTENT_TYPE"]).value
        if not name:
            continue
        if isinstance(name, str) and name.startswith("Nachtrag - "):
            apply_row_style(ws, style["group"], row)
            current_view = None
            current_has_tabs = False
        elif content_type in ("View", "Wizard"):
            apply_row_style(ws, style["view"], row)
            current_view = str(name)
            current_has_tabs = has_tabs.get(current_view, False)
        elif current_view and isinstance(name, str) and name.startswith(f"{current_view}-"):
            apply_row_style(ws, style["tab"], row)
        else:
            apply_row_style(ws, style["field4" if current_has_tabs else "field3"], row)

    wb.save(WORKBOOK)
    print(f"removed_junk_rows={len(remove_rows)}")


if __name__ == "__main__":
    main()
