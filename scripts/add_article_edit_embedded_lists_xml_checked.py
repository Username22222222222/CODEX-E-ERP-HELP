from __future__ import annotations

import html
import re
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from add_article_edit_embedded_lists_checked import SPECS


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS = {"m": MAIN_NS}


def colnum(ref: str) -> int:
    col = "".join(ch for ch in ref if ch.isalpha())
    num = 0
    for ch in col:
        num = num * 26 + ord(ch.upper()) - 64
    return num


def load_shared_values(root: ET.Element) -> list[str]:
    values: list[str] = []
    for si in root.findall("m:si", NS):
        values.append("".join((t.text or "") for t in si.iter(f"{{{MAIN_NS}}}t")))
    return values


def cell_value(cell: ET.Element, shared_values: list[str]) -> str:
    v = cell.find("m:v", NS)
    if v is None:
        return ""
    raw = v.text or ""
    if cell.attrib.get("t") == "s":
        return shared_values[int(raw)]
    return raw


def read_rows(sheet_xml: str, shared_xml: str) -> tuple[dict[int, dict], int, int, int]:
    shared_root = ET.fromstring(shared_xml.encode("utf-8"))
    shared_values = load_shared_values(shared_root)
    sheet_root = ET.fromstring(sheet_xml.encode("utf-8"))
    rows: dict[int, dict] = {}
    for row in sheet_root.findall("m:sheetData/m:row", NS):
        row_num = int(row.attrib["r"])
        values: dict[int, str] = {}
        for cell in row.findall("m:c", NS):
            values[colnum(cell.attrib["r"])] = cell_value(cell, shared_values)
        rows[row_num] = {
            "level": int(row.attrib.get("outlineLevel", "0")),
            "values": values,
        }

    article_row = next(r for r, data in rows.items() if data["values"].get(1) == "ArticleEdit")
    article_level = rows[article_row]["level"]
    article_end = max(rows)
    for r in range(article_row + 1, max(rows) + 1):
        if r in rows and rows[r]["level"] <= article_level:
            article_end = r - 1
            break
    return rows, article_row, article_end, article_level


def child_end(rows: dict[int, dict], article_end: int, parent_row: int) -> int:
    parent_level = rows[parent_row]["level"]
    end = parent_row
    for row in range(parent_row + 1, article_end + 1):
        if row not in rows:
            continue
        if rows[row]["level"] <= parent_level:
            break
        end = row
    return end


def describe_column(label: str, field: str, list_name: str) -> str:
    key = field.lower()
    if any(part in key for part in ("price", "cost", "amount", "revenue", "profit", "landed")):
        return f"{label} zeigt einen Wert für Kalkulation, Preisbewertung oder Auswertung in {list_name}."
    if any(part in key for part in ("quantity", "stock", "available", "minimum", "maximum", "threshold", "packed", "ordered", "announced")):
        return f"{label} zeigt eine Menge oder Bestandsinformation in {list_name}."
    if any(part in key for part in ("date", "time", "year", "month")):
        return f"{label} zeigt den zeitlichen Bezug des Listeneintrags."
    if any(part in key for part in ("id", "number", "matchcode", "code")):
        return f"{label} dient zur eindeutigen Zuordnung, Suche oder Identifikation des Listeneintrags."
    return f"{label} zeigt eine fachliche Information des Listeneintrags in {list_name}."


def build_rows(spec: dict) -> list[tuple[int, dict[str, str]]]:
    rows: list[tuple[int, dict[str, str]]] = [
        (
            1,
            {
                "A": spec["list"],
                "B": spec["list"],
                "H": f"Die eingebettete Liste {spec['list']} {spec['purpose']} Sie macht die zugehörigen Datensätze direkt im Register sichtbar.",
            },
        )
    ]
    for label, original, field in spec["columns"]:
        rows.append(
            (
                2,
                {
                    "A": label,
                    "B": original,
                    "F": field,
                    "H": describe_column(label, field, spec["list"]),
                },
            )
        )
    if spec["edit"]:
        rows.append(
            (
                2,
                {
                    "A": f"{spec['edit']} anlegen/bearbeiten",
                    "B": spec["edit"],
                    "F": spec["edit"],
                    "H": f"Über Neu oder Bearbeiten öffnet X-ERP die Maske {spec['edit']}. Dort werden die einzelnen Datensätze dieser Liste gepflegt.",
                },
            )
        )
    return rows


def append_shared_strings(shared_xml: str, strings: list[str]) -> tuple[str, dict[str, int]]:
    count_match = re.search(r'\bcount="(\d+)"', shared_xml)
    unique_match = re.search(r'\buniqueCount="(\d+)"', shared_xml)
    if not count_match or not unique_match:
        raise RuntimeError("sharedStrings counters not found")
    start_index = int(unique_match.group(1))
    mapping: dict[str, int] = {}
    additions = []
    current = start_index
    for text in strings:
        if text in mapping:
            continue
        mapping[text] = current
        current += 1
        additions.append(f"<si><t>{html.escape(text)}</t></si>")
    shared_xml = shared_xml.replace("</sst>", "".join(additions) + "</sst>")
    shared_xml = re.sub(r'\bcount="\d+"', f'count="{int(count_match.group(1)) + len(strings)}"', shared_xml, count=1)
    shared_xml = re.sub(r'\buniqueCount="\d+"', f'uniqueCount="{int(unique_match.group(1)) + len(additions)}"', shared_xml, count=1)
    return shared_xml, mapping


def shift_row_refs(xml: str, insert_at: int, delta: int) -> str:
    def row_repl(match: re.Match[str]) -> str:
        number = int(match.group(2))
        if number >= insert_at:
            number += delta
        return f"{match.group(1)}{number}{match.group(3)}"

    def cell_repl(match: re.Match[str]) -> str:
        number = int(match.group(2))
        if number >= insert_at:
            number += delta
        return f"{match.group(1)}{number}"

    xml = re.sub(r'(<row\b[^>]*\br=")(\d+)(")', row_repl, xml)
    xml = re.sub(r"\b([A-Z]{1,3})(\d+)\b", cell_repl, xml)
    xml = re.sub(
        r'(<dimension\s+ref="A1:[A-Z]+)(\d+)(")',
        lambda m: f"{m.group(1)}{int(m.group(2)) + delta}{m.group(3)}",
        xml,
        count=1,
    )
    return xml


def make_row(row_number: int, outline_level: int, values: dict[str, str], shared_map: dict[str, int]) -> str:
    row_style = 'spans="1:101" ht="16.899999999999999" customHeight="1"'
    cells = []
    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        style = "11" if outline_level >= 5 and col == "A" else ("10" if col in {"A", "B"} else "9")
        text = values.get(col, "")
        if text:
            cells.append(f'<c r="{col}{row_number}" s="{style}" t="s"><v>{shared_map[text]}</v></c>')
        else:
            cells.append(f'<c r="{col}{row_number}" s="{style}"/>')
    return f'<row r="{row_number}" {row_style} outlineLevel="{outline_level}">' + "".join(cells) + "</row>"


def validate_workbook() -> list[str]:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["de-DE"]
    start = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "ArticleEdit")
    article_level = ws.row_dimensions[start].outlineLevel
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if ws.row_dimensions[row].outlineLevel <= article_level:
            end = row - 1
            break

    problems: list[str] = []
    for row in range(start + 1, end + 1):
        value = ws.cell(row, 1).value
        level = ws.row_dimensions[row].outlineLevel
        if isinstance(value, str) and value.startswith("ArticleEdit-") and level != article_level + 1:
            problems.append(f"{row}: {value} Level {level}, erwartet {article_level + 1}")

    for spec in SPECS:
        parent = None
        for row in range(start + 1, end + 1):
            if ws.cell(row, 1).value == spec["parent"] and ws.row_dimensions[row].outlineLevel == article_level + 1:
                parent = row
                break
        if parent is None:
            problems.append(f"{spec['parent']} fehlt")
            continue
        parent_level = ws.row_dimensions[parent].outlineLevel
        segment_end = parent
        for row in range(parent + 1, end + 1):
            if ws.row_dimensions[row].outlineLevel <= parent_level:
                break
            segment_end = row
        list_rows = [
            row
            for row in range(parent + 1, segment_end + 1)
            if ws.cell(row, 1).value == spec["list"]
        ]
        if not list_rows:
            problems.append(f"{spec['list']} unter {spec['parent']} fehlt")
        elif ws.row_dimensions[list_rows[0]].outlineLevel != parent_level + 1:
            problems.append(f"{list_rows[0]}: {spec['list']} falsches Level")
    return problems


def main() -> None:
    archive_dir = WORKBOOK.parent / "ARCHIV"
    archive_dir.mkdir(exist_ok=True)
    backup = archive_dir / f"X-ERP-HELP-before-article-edit-lists-xml-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    with zipfile.ZipFile(WORKBOOK, "r") as zin:
        sheet_xml = zin.read("xl/worksheets/sheet1.xml").decode("utf-8")
        shared_xml = zin.read("xl/sharedStrings.xml").decode("utf-8")

    rows, _, article_end, article_level = read_rows(sheet_xml, shared_xml)
    plans: list[tuple[int, int, dict, list[tuple[int, dict[str, str]]]]] = []
    for spec in SPECS:
        parent_row = None
        for row_num, data in rows.items():
            if data["values"].get(1) == spec["parent"] and data["level"] == article_level + 1:
                parent_row = row_num
                break
        if parent_row is None:
            continue
        end_row = child_end(rows, article_end, parent_row)
        if any(rows[row]["values"].get(1) == spec["list"] for row in range(parent_row + 1, end_row + 1) if row in rows):
            continue
        new_rows = build_rows(spec)
        plans.append((end_row + 1, rows[parent_row]["level"], spec, new_rows))

    if not plans:
        print("added=0")
        return

    all_strings = [
        text
        for _, _, _, new_rows in plans
        for _, values in new_rows
        for text in values.values()
    ]
    shared_xml, shared_map = append_shared_strings(shared_xml, all_strings)

    total_added = 0
    for insert_at, parent_level, spec, new_rows in sorted(plans, key=lambda item: item[0], reverse=True):
        sheet_xml = shift_row_refs(sheet_xml, insert_at, len(new_rows))
        rendered = []
        for offset, (relative_level, values) in enumerate(new_rows):
            rendered.append(make_row(insert_at + offset, parent_level + relative_level, values, shared_map))
        marker = re.search(rf'<row\b[^>]*\br="{insert_at - 1}"[\s\S]*?</row>', sheet_xml)
        if not marker:
            raise RuntimeError(f"marker row {insert_at - 1} not found for {spec['list']}")
        sheet_xml = sheet_xml[: marker.end()] + "".join(rendered) + sheet_xml[marker.end() :]
        total_added += len(new_rows)

    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td) / WORKBOOK.name
        with zipfile.ZipFile(WORKBOOK, "r") as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    data = sheet_xml.encode("utf-8")
                elif item.filename == "xl/sharedStrings.xml":
                    data = shared_xml.encode("utf-8")
                zout.writestr(item, data)
        shutil.move(str(tmp), WORKBOOK)

    # Normalize OpenXML package after direct insertion.
    wb = load_workbook(WORKBOOK)
    wb.save(WORKBOOK)

    problems = validate_workbook()
    if problems:
        shutil.copy2(backup, WORKBOOK)
        raise RuntimeError("Validierung fehlgeschlagen, Backup wiederhergestellt:\n" + "\n".join(problems[:40]))

    print(f"added={total_added}")
    print(f"backup={backup}")


if __name__ == "__main__":
    main()
