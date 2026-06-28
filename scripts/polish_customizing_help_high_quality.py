from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"


PAGES = {
    "Customizing/index.html": {
        "title": "Customizing",
        "intro": "Customizing passt X-ERP an Arbeitsweise, Rollen, Beleglogik, Ausgabe, Zusatzfelder, Dashboards und Auswertungen an. Dieser Bereich richtet sich an erfahrene Anwender und Administratoren, weil Änderungen nicht nur eine Maske betreffen, sondern oft ganze Folgeprozesse verändern.",
        "flow": ("Customizing kontrolliert ändern", [
            "Ziel klären: Beschreiben Sie zuerst, welches fachliche Problem gelöst werden soll.",
            "Auswirkung prüfen: Belege, Rollen, Ausgabe, Suche, Reports und Auswertungen als Folgeprozesse mitdenken.",
            "Testumgebung nutzen: Kritische Einstellungen vor produktiver Nutzung mit Beispieldaten prüfen.",
            "Änderung dokumentieren: Zweck, Regel, Verantwortliche und Gültigkeit nachvollziehbar festhalten.",
            "Anwender informieren: Betroffene Rollen und Abläufe nach der Änderung gezielt prüfen lassen.",
        ]),
        "bullets": [
            "Customizing sollte nie nur aus Sicht eines einzelnen Feldes geändert werden.",
            "Viele Einstellungen wirken erst beim Erstellen, Drucken, Überführen oder Auswerten eines Vorgangs.",
            "Für komplexe Regeln sind Testfälle wichtiger als lange Beschreibungen.",
            "Bei Unsicherheit ist eine kleine, überprüfbare Änderung besser als eine breite Umstellung.",
        ],
        "box": ("Sorgfalt", "Customizing ist mächtig, aber auch riskant. Eine falsche Regel kann viele Belege, Benutzer oder Auswertungen betreffen."),
    },
    "Customizing/dynamische-doc-kette/index.html": {
        "title": "Dynamische Doc-Kette",
        "intro": "Die dynamische Doc-Kette steuert, wie Belege ineinander überführt werden und welche Werte dabei übernommen werden. Sie ist relevant, wenn Standard-Belegketten nicht ausreichen oder besondere Geschäftsabläufe abgebildet werden müssen.",
        "flow": ("Belegkette vorbereiten", [
            "Belegarten bestimmen: Quell- und Zielbeleg fachlich eindeutig festlegen.",
            "Überführungsregel definieren: Entscheiden, wann und wie ein Beleg fortgeführt werden darf.",
            "Feldübernahme prüfen: Werte wie Partner, Adresse, Preise, Status und Zusatzfelder kontrolliert übertragen.",
            "Testfall ausführen: Ein Beispiel vom Quellbeleg bis zum Zielbeleg komplett prüfen.",
            "Abweichungen dokumentieren: Sonderfälle, Sperren und manuelle Eingriffe klar beschreiben.",
        ]),
        "bullets": [
            "Eine Doc-Kette muss fachlich zur realen Beleglogik passen.",
            "Feldübernahmen können Datenqualität verbessern, aber auch falsche Werte weiterreichen.",
            "Prüfen Sie immer auch Storno, Teilmengen und Sonderfälle.",
        ],
        "box": ("Wichtig", "Änderungen an Doc-Ketten wirken auf Folgebelege. Testen Sie deshalb nicht nur die Maske, sondern den kompletten Belegfluss."),
    },
    "Customizing/dynamische-doc-kette/belegart/index.html": {
        "title": "Belegart",
        "intro": "Die Belegart legt fest, welche Dokumenttypen in einer dynamischen Doc-Kette beteiligt sind. Sie entscheidet, aus welchem Vorgang ein Folgebeleg entstehen kann und welche fachliche Bedeutung dieser Schritt hat.",
        "flow": ("Belegart prüfen", [
            "Quellbeleg wählen: Den Ausgangspunkt der Überführung eindeutig bestimmen.",
            "Zielbeleg wählen: Festlegen, welcher Folgebeleg entstehen soll.",
            "Fachlichen Zweck klären: Nur Belegarten verbinden, die im Prozess wirklich zusammengehören.",
            "Berechtigung und Status prüfen: Sicherstellen, dass nur zulässige Vorgänge fortgeführt werden.",
        ]),
        "bullets": [
            "Belegarten sollten nicht technisch, sondern fachlich verbunden werden.",
            "Nicht jede theoretische Überführung ist im Geschäftsprozess sinnvoll.",
            "Die Belegart ist die Grundlage für Überführungsregeln und Feldübernahmen.",
        ],
        "box": ("Praxisregel", "Wenn Anwender später nicht erklären können, warum ein Folgebeleg entsteht, ist die Belegart-Zuordnung wahrscheinlich zu unklar."),
    },
    "Customizing/dynamische-doc-kette/ueberfuehrungsregel/index.html": {
        "title": "Überführungsregel",
        "intro": "Überführungsregeln bestimmen, unter welchen Bedingungen ein Beleg in einen anderen Beleg überführt werden darf. Sie schützen vor falschen Folgebelegen und automatisieren wiederkehrende Prozesslogik.",
        "flow": ("Regel sicher definieren", [
            "Bedingung formulieren: Status, Belegart, Rolle, Menge oder andere Kriterien klar festlegen.",
            "Ausnahmefälle prüfen: Teilmengen, Sperren, Storno und manuelle Korrekturen berücksichtigen.",
            "Reihenfolge beachten: Regeln dürfen sich nicht widersprechen oder gegenseitig überdecken.",
            "Testbelege verwenden: Erfolgsfall und Ablehnungsfall getrennt prüfen.",
        ]),
        "bullets": [
            "Eine gute Regel ist eindeutig, testbar und für Fachanwender erklärbar.",
            "Zu allgemeine Regeln führen schnell zu unerwarteten Folgebelegen.",
            "Dokumentieren Sie, warum eine Regel existiert und wer sie verantwortet.",
        ],
        "box": ("Testfall", "Prüfen Sie nicht nur, ob die Regel auslöst, sondern auch, ob sie in unzulässigen Fällen nicht auslöst."),
    },
    "Customizing/dynamische-doc-kette/felduebernahme/index.html": {
        "title": "Feldübernahme",
        "intro": "Feldübernahmen steuern, welche Werte aus einem Quellbeleg in den Zielbeleg übernommen werden. Dazu können Partnerdaten, Adressen, Positionen, Preise, Texte, Statuswerte oder Zusatzfelder gehören.",
        "flow": ("Felder kontrolliert übernehmen", [
            "Quelle festlegen: Das Feld im Ausgangsbeleg eindeutig bestimmen.",
            "Ziel festlegen: Das passende Feld im Folgebeleg zuordnen.",
            "Wertlogik prüfen: Entscheiden, ob der Wert kopiert, berechnet, überschrieben oder leer bleiben soll.",
            "Sonderfälle testen: Leere Werte, manuelle Änderungen und abweichende Partnerdaten prüfen.",
        ]),
        "bullets": [
            "Feldübernahmen sparen Arbeit, können aber falsche Daten vervielfältigen.",
            "Nicht jedes Feld aus dem Quellbeleg ist im Zielbeleg fachlich richtig.",
            "Bei Preisen, Steuern und Adressen ist besondere Sorgfalt nötig.",
        ],
        "box": ("Datenqualität", "Eine Feldübernahme ist nur gut, wenn der Zielbeleg danach ohne stille Korrekturen weiterverarbeitet werden kann."),
    },
    "Customizing/ausgabesteuerung/index.html": {
        "title": "Ausgabesteuerung",
        "intro": "Die Ausgabesteuerung legt fest, wie Dokumente, Belege und Informationen ausgegeben werden. Sie verbindet Ausgabekanal, Regeln, Formate, Drucker, E-Mail und Layoutzuordnung.",
        "flow": ("Ausgabe steuern", [
            "Ausgabefall bestimmen: Belegart, Empfänger, Kanal und Zweck klären.",
            "Kanal wählen: Druck, E-Mail, Datei oder andere Ausgabewege fachlich zuordnen.",
            "Regel prüfen: Bedingungen für automatische oder manuelle Ausgabe definieren.",
            "Layout zuordnen: Passende Vorlage, Sprache und Corporate Design berücksichtigen.",
            "Ausgabe testen: Dokument erzeugen und Empfänger, Darstellung sowie Anhänge prüfen.",
        ]),
        "bullets": [
            "Ausgabe ist sichtbar für Kunden, Lieferanten und interne Anwender.",
            "Falsche Regeln können Dokumente an falsche Empfänger oder im falschen Layout erzeugen.",
            "Testen Sie Ausgaben immer mit realistischen Beispieldaten.",
        ],
        "box": ("Außenwirkung", "Ausgabesteuerung ist Customizing mit direkter Außenwirkung. Prüfen Sie deshalb Inhalt, Empfänger und Layout gemeinsam."),
    },
    "Customizing/ausgabesteuerung/ausgabekanal/index.html": {
        "title": "Ausgabekanal",
        "intro": "Der Ausgabekanal beschreibt, über welchen Weg ein Dokument ausgegeben wird. Beispiele sind Druck, E-Mail, Dateiablage oder ein anderer definierter Ausgabepfad.",
        "flow": ("Kanal auswählen", [
            "Zweck klären: Entscheiden, ob der Empfänger ein Dokument sehen, erhalten oder intern archiviert werden soll.",
            "Technischen Weg prüfen: Drucker, E-Mail-Konto, Dateiablage oder Schnittstelle kontrollieren.",
            "Empfängerlogik prüfen: Sicherstellen, dass der Kanal zur Partner- und Ansprechpartnerlogik passt.",
            "Probeausgabe durchführen: Ergebnis und Status kontrollieren.",
        ]),
        "bullets": [
            "Ein Kanal ist nicht nur Technik, sondern Teil des Geschäftsprozesses.",
            "E-Mail-Ausgaben brauchen saubere Empfänger und Vorlagen.",
            "Druck- und Datei-Ausgaben sollten klar benannt und auffindbar sein.",
        ],
        "box": ("Abgrenzung", "Der Kanal legt den Weg fest. Layout und Regel entscheiden, wie und wann dieser Weg genutzt wird."),
    },
    "Customizing/ausgabesteuerung/ausgaberegel/index.html": {
        "title": "Ausgaberegel",
        "intro": "Ausgaberegeln entscheiden, wann welche Ausgabe erzeugt wird. Sie können nach Belegart, Partner, Status, Sprache, Benutzerrolle oder anderen Bedingungen greifen.",
        "flow": ("Regel prüfen", [
            "Auslöser bestimmen: Festlegen, bei welchem Vorgang die Regel greifen soll.",
            "Bedingungen formulieren: Belegart, Partner, Sprache, Kanal oder Status eindeutig setzen.",
            "Priorität prüfen: Überschneidungen mit anderen Regeln vermeiden.",
            "Testausgabe erzeugen: Mindestens einen Trefferfall und einen Nicht-Trefferfall prüfen.",
        ]),
        "bullets": [
            "Regeln sollten möglichst spezifisch und nachvollziehbar sein.",
            "Mehrere ähnliche Regeln können unerwartete Ergebnisse erzeugen.",
            "Dokumentieren Sie, für welchen Geschäftsfall die Regel erstellt wurde.",
        ],
        "box": ("Sicherheit", "Eine Ausgaberegel sollte immer mit einem realen Dokument getestet werden, nicht nur anhand der Einstellung selbst."),
    },
    "Customizing/ausgabesteuerung/layoutzuordnung/index.html": {
        "title": "Layoutzuordnung",
        "intro": "Die Layoutzuordnung bestimmt, welche Vorlage oder welches Format für eine Ausgabe verwendet wird. Sie beeinflusst Darstellung, Sprache, Logo, Pflichtangaben und Corporate Design.",
        "flow": ("Layout zuordnen", [
            "Belegart prüfen: Layout nur dem passenden Dokumenttyp zuweisen.",
            "Sprache und Empfänger beachten: Varianten für Länder, Kunden oder Rollen berücksichtigen.",
            "Corporate Design kontrollieren: Logo, Farben, Fußtexte und Pflichtangaben prüfen.",
            "Druckbild testen: PDF oder Ausdruck mit realistischen Daten erzeugen.",
        ]),
        "bullets": [
            "Ein falsches Layout kann fachlich korrekte Daten falsch oder unvollständig darstellen.",
            "Änderungen sollten mit Reportdesigner und Ausgaberegeln abgestimmt werden.",
            "Pflichtangaben auf externen Dokumenten müssen sichtbar bleiben.",
        ],
        "box": ("Qualität", "Layoutzuordnung ist erst geprüft, wenn das erzeugte Dokument fachlich und optisch stimmt."),
    },
    "Customizing/desktopkonfigurator/index.html": {
        "title": "Desktopkonfigurator",
        "intro": "Der Desktopkonfigurator steuert, welche Menüs, Schaltflächen und Arbeitsbereiche Benutzer oder Rollen sehen. Damit lässt sich X-ERP auf Aufgaben, Teams und Verantwortlichkeiten zuschneiden.",
        "flow": ("Arbeitsoberfläche gestalten", [
            "Rolle klären: Festlegen, welche Benutzergruppe unterstützt werden soll.",
            "Navigation strukturieren: Häufige Funktionen sichtbar und seltene Funktionen erreichbar halten.",
            "Schaltflächen prüfen: Nur sinnvolle Aktionen anbieten.",
            "Rollenlayout testen: Mit einem passenden Benutzer anmelden und Arbeitsabläufe durchgehen.",
        ]),
        "bullets": [
            "Eine gute Oberfläche reduziert Suchaufwand und Fehlbedienung.",
            "Zu viele sichtbare Funktionen verlangsamen Anwender.",
            "Desktopkonfiguration sollte zu Berechtigungen und Rollen passen.",
        ],
        "box": ("Ergonomie", "Customizing der Oberfläche ist erfolgreich, wenn Anwender ihre täglichen Aufgaben schneller und sicherer erreichen."),
    },
    "Customizing/desktopkonfigurator/menue-und-navigation/index.html": {
        "title": "Menü und Navigation",
        "intro": "Menü und Navigation legen fest, welche Bereiche Anwender schnell erreichen. Sie sollten nach Arbeitsablauf und Häufigkeit strukturiert sein, nicht nur nach technischer Modulzuordnung.",
        "flow": ("Navigation prüfen", [
            "Hauptaufgaben sammeln: Die wichtigsten täglichen Arbeitsschritte einer Rolle erfassen.",
            "Menüpunkte ordnen: Häufig genutzte Funktionen prominent platzieren.",
            "Begriffe prüfen: Beschriftungen müssen für Anwender verständlich sein.",
            "Testlauf durchführen: Typische Aufgaben ohne Suche durchspielen.",
        ]),
        "bullets": [
            "Navigation sollte rollenbezogen und nicht überladen sein.",
            "Ähnliche Funktionen brauchen klare Namen.",
            "Änderungen an Menüs sollten Anwendern angekündigt werden.",
        ],
        "box": ("Bedienqualität", "Eine gute Navigation macht die richtige Funktion auffindbar, bevor der Anwender suchen muss."),
    },
    "Customizing/desktopkonfigurator/schaltflaechen/index.html": {
        "title": "Schaltflächen",
        "intro": "Schaltflächen stellen Aktionen bereit, die Anwender direkt ausführen können. Sie sollten klar benannt, sinnvoll gruppiert und nur dort sichtbar sein, wo sie fachlich passen.",
        "flow": ("Schaltfläche bewerten", [
            "Aktion klären: Prüfen, was beim Klick fachlich passieren soll.",
            "Kontext prüfen: Button nur in passenden Masken oder Rollen anzeigen.",
            "Beschriftung wählen: Kurz, eindeutig und handlungsorientiert formulieren.",
            "Auswirkung testen: Aktion mit Beispielvorgang ausführen und Ergebnis kontrollieren.",
        ]),
        "bullets": [
            "Schaltflächen sollten keine versteckten Nebenwirkungen haben.",
            "Zu viele Buttons machen Masken schwer lesbar.",
            "Kritische Aktionen brauchen klare Berechtigung und Prüfung.",
        ],
        "box": ("Sicherheit", "Ein Button ist nur dann gut konfiguriert, wenn Anwender seine Wirkung vor dem Klick verstehen."),
    },
    "Customizing/desktopkonfigurator/rollenlayout/index.html": {
        "title": "Rollenlayout",
        "intro": "Das Rollenlayout bündelt Menü, Schaltflächen und Arbeitsbereiche für eine Benutzerrolle. Es hilft, unterschiedliche Aufgabenprofile in X-ERP gezielt abzubilden.",
        "flow": ("Rollenlayout abstimmen", [
            "Rolle definieren: Aufgaben, Verantwortlichkeiten und benötigte Module klären.",
            "Oberfläche konfigurieren: Navigation, Buttons und Startbereiche passend auswählen.",
            "Berechtigungen abgleichen: Sichtbarkeit und tatsächliche Rechte gemeinsam prüfen.",
            "Mit Anwender testen: Typischen Arbeitstag der Rolle durchspielen.",
        ]),
        "bullets": [
            "Rollenlayouts sollten nicht für einzelne Personen, sondern für Aufgabenprofile entstehen.",
            "Layout und Berechtigung sind getrennt zu prüfen.",
            "Änderungen an Rollenlayouts wirken auf alle zugeordneten Benutzer.",
        ],
        "box": ("Rollendenken", "Ein gutes Rollenlayout zeigt genau die Funktionen, die eine Rolle regelmäßig braucht."),
    },
    "Customizing/reportdesigner/index.html": {
        "title": "Reportdesigner",
        "intro": "Der Reportdesigner gestaltet Belege, Listen und Ausgaben. Er beeinflusst, welche Daten sichtbar sind, wie Dokumente aussehen und ob Pflichtinformationen korrekt ausgegeben werden.",
        "flow": ("Report sicher ändern", [
            "Zielausgabe wählen: Beleg, Liste oder internes Dokument klar bestimmen.",
            "Datenquelle prüfen: Sicherstellen, dass alle benötigten Felder verfügbar sind.",
            "Layout bearbeiten: Struktur, Logo, Spalten, Texte und Pflichtangaben anpassen.",
            "Mit Echtdaten testen: Lange Texte, mehrere Positionen, Rabatte, Steuern und Seitenumbrüche prüfen.",
        ]),
        "bullets": [
            "Reportänderungen wirken direkt auf Ausdrucke, PDFs und E-Mail-Anhänge.",
            "Testen Sie Sonderfälle wie lange Bezeichnungen und mehrseitige Dokumente.",
            "Corporate Design darf Pflichtangaben nicht verdecken.",
        ],
        "box": ("Druckbild", "Ein Report ist erst fertig, wenn die erzeugte Ausgabe mit realistischen Daten geprüft wurde."),
    },
    "Customizing/reportdesigner/belegvorlage/index.html": {
        "title": "Belegvorlage",
        "intro": "Belegvorlagen definieren das Layout von Angeboten, Aufträgen, Lieferscheinen, Rechnungen oder anderen Dokumenten. Sie verbinden Dateninhalt und Außenwirkung.",
        "flow": ("Belegvorlage prüfen", [
            "Belegart festlegen: Vorlage nur dem passenden Dokumenttyp zuordnen.",
            "Pflichtfelder prüfen: Empfänger, Positionen, Preise, Steuer und Fußtexte sichtbar halten.",
            "Varianten beachten: Sprache, Land, Kunde oder Ausgabeweg berücksichtigen.",
            "PDF testen: Dokument mit realistischen Positionen erzeugen.",
        ]),
        "bullets": [
            "Belegvorlagen haben direkte Wirkung auf Kunden- und Lieferantendokumente.",
            "Mehrseitige Belege und lange Texte müssen geprüft werden.",
            "Änderungen sollten versioniert oder dokumentiert werden.",
        ],
        "box": ("Außenwirkung", "Die Belegvorlage ist oft das Gesicht des Unternehmens nach außen."),
    },
    "Customizing/reportdesigner/corporate-design/index.html": {
        "title": "Corporate Design",
        "intro": "Corporate Design im Reportdesigner sorgt dafür, dass Belege und Listen zur Unternehmensdarstellung passen. Dazu gehören Logo, Farben, Abstände, Schriften und wiederkehrende Fuß- oder Kopfbereiche.",
        "flow": ("Design konsistent halten", [
            "Grundlayout prüfen: Logo, Kopf, Fuß und Seitenränder abstimmen.",
            "Lesbarkeit sichern: Schriftgrößen, Kontraste und Tabellenbreiten testen.",
            "Pflichtangaben beachten: Design darf rechtliche Informationen nicht verdrängen.",
            "Mehrere Dokumenttypen vergleichen: Angebot, Rechnung und Liste konsistent halten.",
        ]),
        "bullets": [
            "Design darf die fachliche Verständlichkeit nicht verschlechtern.",
            "Einheitliche Dokumente stärken Wiedererkennung und Professionalität.",
            "Druck und PDF sollten getrennt geprüft werden.",
        ],
        "box": ("Balance", "Gutes Corporate Design unterstützt den Inhalt. Es ersetzt keine fachlich vollständige Ausgabe."),
    },
    "Customizing/reportdesigner/listendruck/index.html": {
        "title": "Listendruck",
        "intro": "Listendruck steuert, wie tabellarische Daten ausgegeben werden. Wichtig sind Spaltenauswahl, Sortierung, Gruppierung, Summen und Lesbarkeit.",
        "flow": ("Liste druckbar machen", [
            "Zweck klären: Entscheiden, welche Frage die Liste beantworten soll.",
            "Spalten wählen: Nur relevante Informationen aufnehmen.",
            "Sortierung und Gruppen prüfen: Daten für Anwender logisch ordnen.",
            "Ausgabe testen: Breite, Seitenumbrüche und Summen kontrollieren.",
        ]),
        "bullets": [
            "Zu viele Spalten machen Listen schwer lesbar.",
            "Summen und Gruppierungen müssen fachlich stimmen.",
            "Listen sollten sowohl am Bildschirm als auch als PDF geprüft werden.",
        ],
        "box": ("Nutzwert", "Eine gute Liste ist nicht möglichst vollständig, sondern beantwortet eine konkrete Arbeitsfrage."),
    },
    "Customizing/dashboarddesigner/index.html": {
        "title": "Dashboard konfigurieren",
        "intro": "Der Dashboarddesigner stellt Kennzahlen, Diagramme und Layouts für Überblicksseiten zusammen. Er unterstützt Steuerung, Kontrolle und schnelle Orientierung im Tagesgeschäft.",
        "flow": ("Dashboard aufbauen", [
            "Zielgruppe klären: Rolle, Aufgabe und Entscheidungsbedarf bestimmen.",
            "Kennzahlen auswählen: Nur Werte zeigen, die wirklich genutzt werden.",
            "Visualisierung wählen: Diagramm, Tabelle oder Kennzahl passend zur Frage einsetzen.",
            "Layout strukturieren: Wichtiges zuerst, Details darunter anordnen.",
            "Mit echten Daten prüfen: Leere Werte, Extremwerte und Filter testen.",
        ]),
        "bullets": [
            "Dashboards sollten Entscheidungen unterstützen, nicht nur Daten sammeln.",
            "Zu viele Kennzahlen reduzieren die Aussagekraft.",
            "Datenquelle, Aktualität und Filter müssen klar sein.",
        ],
        "box": ("Fokus", "Ein gutes Dashboard beantwortet wenige wichtige Fragen schnell und zuverlässig."),
    },
    "Customizing/dashboarddesigner/kennzahl/index.html": {
        "title": "Kennzahl",
        "intro": "Eine Kennzahl verdichtet Daten zu einem Wert, der eine Entscheidung oder Kontrolle unterstützt. Beispiele sind Umsatz, offene Posten, Lieferquote, Bestand oder Projektaufwand.",
        "flow": ("Kennzahl definieren", [
            "Fragestellung klären: Entscheiden, was die Kennzahl aussagen soll.",
            "Datenbasis prüfen: Quelle, Filter, Zeitraum und Berechnung nachvollziehbar festlegen.",
            "Grenzwerte setzen: Ampeln oder Warnungen nur mit fachlicher Grundlage verwenden.",
            "Ergebnis validieren: Kennzahl gegen bekannte Beispiele prüfen.",
        ]),
        "bullets": [
            "Eine Kennzahl ohne klare Definition führt zu Missverständnissen.",
            "Filter und Zeitraum sind genauso wichtig wie die Berechnung.",
            "Kennzahlen sollten regelmäßig auf Aussagekraft geprüft werden.",
        ],
        "box": ("Definition", "Schreiben Sie bei wichtigen Kennzahlen auf, wie sie berechnet werden und wofür sie genutzt werden."),
    },
    "Customizing/dashboarddesigner/diagramm/index.html": {
        "title": "Diagramm",
        "intro": "Diagramme machen Entwicklungen, Verteilungen und Vergleiche sichtbar. Sie sollten so gewählt werden, dass Anwender die Aussage ohne lange Erklärung verstehen.",
        "flow": ("Diagramm prüfen", [
            "Aussage bestimmen: Trend, Vergleich, Anteil oder Verteilung unterscheiden.",
            "Diagrammtyp wählen: Balken, Linie, Kreis oder Tabelle passend einsetzen.",
            "Achsen und Beschriftung prüfen: Werte verständlich und nicht irreführend darstellen.",
            "Daten testen: Leere Werte und Ausreißer kontrollieren.",
        ]),
        "bullets": [
            "Nicht jede Kennzahl braucht ein Diagramm.",
            "Beschriftungen müssen auch bei vielen Datenpunkten lesbar bleiben.",
            "Farben sollten Bedeutung haben und nicht nur dekorativ sein.",
        ],
        "box": ("Lesbarkeit", "Ein gutes Diagramm macht eine Entwicklung schneller verständlich als eine Tabelle."),
    },
    "Customizing/dashboarddesigner/dashboardlayout/index.html": {
        "title": "Dashboardlayout",
        "intro": "Das Dashboardlayout ordnet Kennzahlen, Diagramme und Listen auf einer Oberfläche an. Es entscheidet, welche Informationen zuerst gesehen werden und wie schnell Anwender reagieren können.",
        "flow": ("Layout strukturieren", [
            "Priorität festlegen: Wichtigste Informationen oben und sichtbar platzieren.",
            "Gruppen bilden: Zusammengehörige Kennzahlen und Diagramme bündeln.",
            "Dichte prüfen: Übersicht und Detailtiefe ausbalancieren.",
            "Rollen testen: Layout mit den Anwendern prüfen, die damit arbeiten.",
        ]),
        "bullets": [
            "Ein Dashboard sollte nicht wie eine Sammlung zufälliger Kacheln wirken.",
            "Abstände, Größen und Reihenfolge beeinflussen die Nutzbarkeit.",
            "Mobile oder kleinere Bildschirme sollten mitgedacht werden.",
        ],
        "box": ("Aufmerksamkeit", "Das Layout lenkt den Blick. Platzieren Sie die wichtigsten Entscheidungen dort, wo Anwender zuerst hinschauen."),
    },
    "Customizing/extra-tabellen/index.html": {
        "title": "Extra-Tabellen",
        "intro": "Extra-Tabellen erweitern X-ERP um eigene strukturierte Daten. Sie eignen sich, wenn zusätzliche Informationen nicht nur als Textnotiz, sondern suchbar, verknüpfbar und auswertbar gepflegt werden sollen.",
        "flow": ("Extra-Tabelle planen", [
            "Zweck beschreiben: Klären, welche eigene Information dauerhaft gespeichert werden soll.",
            "Tabellenstruktur entwerfen: Felder, Datentypen, Pflichtangaben und Schlüssel festlegen.",
            "Verknüpfung prüfen: Bezug zu Artikeln, Partnern, Belegen oder anderen Objekten planen.",
            "Datenpflege testen: Erfassen, suchen und auswerten mit Beispieldaten prüfen.",
        ]),
        "bullets": [
            "Extra-Tabellen sollten nur für wiederkehrende strukturierte Daten genutzt werden.",
            "Zu viele Sondertabellen erhöhen Pflege- und Schulungsaufwand.",
            "Verknüpfungen müssen stabil und für Anwender verständlich sein.",
        ],
        "box": ("Struktur statt Notiz", "Nutzen Sie Extra-Tabellen, wenn Informationen wiederverwendet, gefiltert oder ausgewertet werden sollen."),
    },
    "Customizing/extra-tabellen/tabellendefinition/index.html": {
        "title": "Tabellendefinition",
        "intro": "Die Tabellendefinition beschreibt Aufbau und Zweck einer Extra-Tabelle. Sie legt fest, welche Felder vorhanden sind, welche Werte erlaubt sind und wie Datensätze eindeutig werden.",
        "flow": ("Tabelle definieren", [
            "Datenobjekt benennen: Fachlichen Namen und Zweck festlegen.",
            "Felder planen: Datentypen, Pflichtfelder und Eingabehilfen bestimmen.",
            "Schlüssel prüfen: Eindeutigkeit und Dubletten vermeiden.",
            "Beispiele erfassen: Testdaten anlegen und Bedienung prüfen.",
        ]),
        "bullets": [
            "Eine Tabelle sollte ein klares fachliches Objekt beschreiben.",
            "Feldnamen müssen für Anwender verständlich sein.",
            "Nachträgliche Strukturänderungen können bestehende Daten beeinflussen.",
        ],
        "box": ("Planung", "Eine gute Tabellendefinition spart später viel Pflege- und Korrekturaufwand."),
    },
    "Customizing/extra-tabellen/datenpflege/index.html": {
        "title": "Datenpflege",
        "intro": "Datenpflege beschreibt, wie Datensätze in Extra-Tabellen erfasst, geändert und kontrolliert werden. Ziel ist eine einheitliche und zuverlässige Nutzung eigener Datenstrukturen.",
        "flow": ("Daten pflegen", [
            "Pflichtangaben erfassen: Alle notwendigen Werte vollständig pflegen.",
            "Eingaben prüfen: Datentypen, Wertebereiche und Dubletten kontrollieren.",
            "Änderungen nachvollziehen: Kritische Anpassungen dokumentieren.",
            "Auswertung testen: Prüfen, ob die Daten wie geplant gefunden und verwendet werden.",
        ]),
        "bullets": [
            "Freie Eingaben sollten möglichst durch strukturierte Auswahl unterstützt werden.",
            "Veraltete Datensätze sollten gekennzeichnet oder bereinigt werden.",
            "Pflegeverantwortung sollte klar geregelt sein.",
        ],
        "box": ("Datenqualität", "Eine Extra-Tabelle ist nur so wertvoll wie die Daten, die darin konsistent gepflegt werden."),
    },
    "Customizing/extra-tabellen/verknuepfung/index.html": {
        "title": "Verknüpfung",
        "intro": "Verknüpfungen verbinden Extra-Tabellen mit bestehenden X-ERP-Objekten wie Artikeln, Partnern, Belegen oder Projekten. Dadurch werden eigene Daten im richtigen Prozesskontext nutzbar.",
        "flow": ("Verknüpfung prüfen", [
            "Bezugsobjekt wählen: Festlegen, woran die Extra-Daten hängen sollen.",
            "Eindeutigkeit sichern: Mehrdeutige oder doppelte Beziehungen vermeiden.",
            "Anzeige prüfen: Daten dort sichtbar machen, wo Anwender sie brauchen.",
            "Folgeprozesse testen: Suche, Filter, Auswertung und Bearbeitung kontrollieren.",
        ]),
        "bullets": [
            "Eine falsche Verknüpfung macht Daten schwer auffindbar oder irreführend.",
            "Beziehungen sollten fachlich stabil sein.",
            "Prüfen Sie, ob eine 1:1- oder 1:n-Beziehung benötigt wird.",
        ],
        "box": ("Kontext", "Eigene Daten helfen nur dann, wenn sie am passenden Objekt und im passenden Prozess sichtbar sind."),
    },
    "Customizing/extra-felder/index.html": {
        "title": "Extra-Felder",
        "intro": "Extra-Felder erweitern bestehende Masken um eigene Informationen. Sie eignen sich für Daten, die direkt am Artikel, Partner, Beleg oder anderen Objekten benötigt werden.",
        "flow": ("Extra-Feld anlegen", [
            "Informationsbedarf klären: Prüfen, ob ein eigenes Feld wirklich nötig ist.",
            "Feldtyp wählen: Text, Zahl, Datum, Auswahl oder Verknüpfung passend festlegen.",
            "Maske zuordnen: Feld nur dort anzeigen, wo es fachlich gebraucht wird.",
            "Auswertung prüfen: Suche, Filter und Berichtsnutzung testen.",
        ]),
        "bullets": [
            "Extra-Felder sollten nicht als Ersatz für vorhandene Standardfelder dienen.",
            "Zu viele Felder machen Masken unübersichtlich.",
            "Pflichtfelder sollten sparsam und begründet eingesetzt werden.",
        ],
        "box": ("Einordnung", "Nutzen Sie Extra-Felder für einzelne Zusatzinformationen am bestehenden Objekt. Für eigene Datensammlungen sind Extra-Tabellen besser."),
    },
    "Customizing/extra-felder/feldtyp/index.html": {
        "title": "Feldtyp",
        "intro": "Der Feldtyp bestimmt, welche Art von Wert in einem Extra-Feld gespeichert wird. Er beeinflusst Eingabe, Prüfung, Suche, Auswertung und spätere Datenqualität.",
        "flow": ("Feldtyp auswählen", [
            "Inhalt beschreiben: Klären, ob Text, Zahl, Datum, Ja/Nein, Auswahl oder Verknüpfung benötigt wird.",
            "Auswertung mitdenken: Den Typ wählen, der Filter und Berechnung ermöglicht.",
            "Eingabefehler vermeiden: Auswahlfelder nutzen, wenn Werte standardisiert sein müssen.",
            "Testdaten erfassen: Prüfen, ob Anwender den Wert korrekt pflegen können.",
        ]),
        "bullets": [
            "Der falsche Feldtyp lässt sich später oft nur mit Aufwand korrigieren.",
            "Freitext ist flexibel, aber schlecht auswertbar.",
            "Auswahlfelder erhöhen Konsistenz, brauchen aber Pflege.",
        ],
        "box": ("Langfristigkeit", "Wählen Sie den Feldtyp so, wie die Information später genutzt werden soll, nicht nur wie sie heute eingegeben wird."),
    },
    "Customizing/extra-felder/maskenzuordnung/index.html": {
        "title": "Maskenzuordnung",
        "intro": "Die Maskenzuordnung steuert, wo ein Extra-Feld angezeigt wird. Sie entscheidet, welche Anwender das Feld in welchem Arbeitskontext sehen und pflegen können.",
        "flow": ("Maske zuordnen", [
            "Arbeitskontext wählen: Feld an der Maske anzeigen, in der die Information entsteht.",
            "Rolle prüfen: Nur relevante Anwender mit dem Feld belasten.",
            "Position bewerten: Feld so platzieren, dass es fachlich verständlich ist.",
            "Eingabe testen: Datensatz öffnen, ändern und speichern.",
        ]),
        "bullets": [
            "Ein Extra-Feld am falschen Ort wird nicht zuverlässig gepflegt.",
            "Zu viele Felder stören die Bedienung.",
            "Pflichtfelder müssen dort sichtbar sein, wo sie erfüllt werden können.",
        ],
        "box": ("Ergonomie", "Die richtige Maskenzuordnung entscheidet, ob ein Extra-Feld im Alltag wirklich genutzt wird."),
    },
    "Customizing/extra-felder/auswertbarkeit/index.html": {
        "title": "Auswertbarkeit",
        "intro": "Auswertbarkeit beschreibt, ob Extra-Felder später gesucht, gefiltert, berichtet oder in Regeln verwendet werden können. Diese Entscheidung sollte vor der Anlage des Feldes getroffen werden.",
        "flow": ("Auswertung vorbereiten", [
            "Fragestellung klären: Entscheiden, welche Auswertung das Feld unterstützen soll.",
            "Feldtyp prüfen: Datentyp und Eingabeform passend wählen.",
            "Werte standardisieren: Auswahl oder Regeln verwenden, wenn konsistente Filter nötig sind.",
            "Bericht testen: Beispielauswertung mit echten Daten prüfen.",
        ]),
        "bullets": [
            "Freitextfelder sind für Auswertungen nur eingeschränkt geeignet.",
            "Einheitliche Werte sind wichtiger als möglichst freie Eingabe.",
            "Auswertbarkeit muss vor produktiver Datenpflege getestet werden.",
        ],
        "box": ("Planung", "Ein Extra-Feld ist nur dann auswertbar, wenn Eingabe, Datentyp und Berichtsziel zusammenpassen."),
    },
    "Customizing/vorschlagswerte/index.html": {
        "title": "Vorschlagswerte",
        "intro": "Vorschlagswerte belegen Felder automatisch vor. Sie sparen Zeit und erhöhen Konsistenz, wenn in bestimmten Situationen regelmäßig derselbe Wert verwendet wird.",
        "flow": ("Vorschlagswert einrichten", [
            "Standardfall erkennen: Prüfen, welcher Wert in den meisten Fällen richtig ist.",
            "Gültigkeit festlegen: Bereich, Rolle, Belegart oder Kontext sauber begrenzen.",
            "Überschreibbarkeit prüfen: Entscheiden, ob Anwender den Wert ändern dürfen.",
            "Testfälle ausführen: Trefferfall und Ausnahmefall kontrollieren.",
        ]),
        "bullets": [
            "Vorschlagswerte dürfen keine falschen Automatismen erzeugen.",
            "Ein zu allgemeiner Vorschlag führt zu stillen Fehlern.",
            "Ausnahmen müssen für Anwender einfach korrigierbar bleiben.",
        ],
        "box": ("Kein Screenshot gesetzt", "Für Vorschlagswerte wurde kein eigener, eindeutig passender X-ERP-Screenshot gefunden. Deshalb wird hier bewusst kein fremdes Bild zugeordnet."),
    },
    "Customizing/vorschlagswerte/feldvorbelegung/index.html": {
        "title": "Feldvorbelegung",
        "intro": "Feldvorbelegung setzt einen Wert automatisch in ein Feld, bevor Anwender ihn manuell erfassen. Das ist hilfreich für häufige Standards, muss aber fachlich eng begrenzt sein.",
        "flow": ("Vorbelegung prüfen", [
            "Feld auswählen: Nur Felder vorbelegen, deren Standard klar ist.",
            "Wert definieren: Vorschlag fachlich begründen.",
            "Ausnahme prüfen: Sicherstellen, dass Sonderfälle möglich bleiben.",
            "Beispiel testen: Datensatz neu anlegen und Vorbelegung kontrollieren.",
        ]),
        "bullets": [
            "Vorbelegung spart Zeit, kann aber falsche Werte unbemerkt verbreiten.",
            "Kritische Felder wie Steuer, Konto oder Status brauchen besondere Prüfung.",
            "Anwender müssen erkennen können, ob ein Wert vorgeschlagen oder bewusst gesetzt wurde.",
        ],
        "box": ("Kein Screenshot gesetzt", "Für diese Detailseite wurde kein eindeutig passender Screenshot gefunden. Eine ungenaue Zuordnung wurde bewusst vermieden."),
    },
    "Customizing/vorschlagswerte/gueltigkeitsbereich/index.html": {
        "title": "Gültigkeitsbereich",
        "intro": "Der Gültigkeitsbereich legt fest, wann ein Vorschlagswert angewendet wird. Er kann sich nach Rolle, Maske, Belegart, Mandant, Zeitraum oder fachlichem Kontext richten.",
        "flow": ("Gültigkeit eingrenzen", [
            "Kontext bestimmen: Definieren, wo der Vorschlag wirklich richtig ist.",
            "Grenzen setzen: Rollen, Belege oder Masken möglichst präzise wählen.",
            "Konflikte prüfen: Überschneidungen mit anderen Vorschlagswerten vermeiden.",
            "Ausnahmen testen: Fälle außerhalb des Bereichs bewusst prüfen.",
        ]),
        "bullets": [
            "Je genauer der Gültigkeitsbereich, desto geringer das Risiko falscher Vorschläge.",
            "Mehrere überlappende Regeln können schwer nachvollziehbar werden.",
            "Gültigkeitsbereiche sollten dokumentiert und regelmäßig überprüft werden.",
        ],
        "box": ("Kein Screenshot gesetzt", "Für den Gültigkeitsbereich wurde keine eindeutig passende X-ERP-Ansicht gefunden. Deshalb bleibt diese Seite ohne Screenshot."),
    },
    "Customizing/formatierte-suche/index.html": {
        "title": "Formatierte Suche",
        "intro": "Formatierte Suche unterstützt dynamische Werte, Abfragen und automatische Vorschläge. Sie kann Felder füllen, Daten suchen oder Anwender bei komplexen Eingaben unterstützen.",
        "flow": ("Formatierte Suche einrichten", [
            "Ziel bestimmen: Festlegen, welches Feld oder welcher Prozess unterstützt werden soll.",
            "Auslöser wählen: Zeitpunkt und Kontext der Suche definieren.",
            "Formel oder Abfrage prüfen: Logik nachvollziehbar und testbar halten.",
            "Zielfeld kontrollieren: Ergebnis nur dort schreiben, wo es fachlich passt.",
            "Testfälle ausführen: Treffer, leeres Ergebnis und Fehlerfall prüfen.",
        ]),
        "bullets": [
            "Formatierte Suche kann sehr hilfreich sein, muss aber transparent bleiben.",
            "Unklare Abfragen sind schwer zu warten.",
            "Zielfelder dürfen nicht unbeabsichtigt überschrieben werden.",
        ],
        "box": ("Testpflicht", "Jede formatierte Suche braucht konkrete Testfälle, sonst bleibt ihre Wirkung im Alltag unsicher."),
    },
    "Customizing/formatierte-suche/ausloeser/index.html": {
        "title": "Auslöser",
        "intro": "Der Auslöser bestimmt, wann eine formatierte Suche ausgeführt wird. Das kann beim Öffnen, Ändern, Speichern oder in einem bestimmten Feldkontext passieren.",
        "flow": ("Auslöser wählen", [
            "Ereignis bestimmen: Zeitpunkt der Ausführung fachlich begründen.",
            "Kontext prüfen: Maske, Feld und Benutzerrolle berücksichtigen.",
            "Häufigkeit bewerten: Zu häufige Ausführung kann stören oder verlangsamen.",
            "Ausführung testen: Prüfen, ob der Auslöser genau im erwarteten Moment greift.",
        ]),
        "bullets": [
            "Der richtige Auslöser ist entscheidend für Bedienbarkeit und Performance.",
            "Automatische Ausführung sollte für Anwender nachvollziehbar bleiben.",
            "Fehlerfälle müssen sichtbar und verständlich sein.",
        ],
        "box": ("Timing", "Eine gute Suche läuft nicht möglichst oft, sondern im richtigen Moment."),
    },
    "Customizing/formatierte-suche/formel-und-abfrage/index.html": {
        "title": "Formel und Abfrage",
        "intro": "Formeln und Abfragen liefern das Ergebnis einer formatierten Suche. Sie können Werte berechnen, Datensätze finden oder Eingaben aus anderen Informationen ableiten.",
        "flow": ("Logik prüfen", [
            "Fragestellung formulieren: Ergebnis und Datenbasis klar beschreiben.",
            "Abfrage schreiben: Einfach, nachvollziehbar und wartbar halten.",
            "Grenzfälle testen: Kein Treffer, mehrere Treffer und ungültige Eingaben prüfen.",
            "Ergebnis kontrollieren: Wert im Zielfeld mit bekannten Beispielen vergleichen.",
        ]),
        "bullets": [
            "Komplexe Abfragen sollten kommentiert oder dokumentiert werden.",
            "Mehrdeutige Ergebnisse sind für Anwender gefährlich.",
            "Performance ist wichtig, wenn eine Suche häufig ausgelöst wird.",
        ],
        "box": ("Wartbarkeit", "Eine Formel ist gut, wenn ein anderer Administrator später versteht, was sie tut und warum sie existiert."),
    },
    "Customizing/formatierte-suche/zielfeld/index.html": {
        "title": "Zielfeld",
        "intro": "Das Zielfeld ist das Feld, in das das Ergebnis einer formatierten Suche geschrieben oder vorgeschlagen wird. Es muss fachlich zum Ergebnis passen und darf keine wichtigen Benutzereingaben überschreiben.",
        "flow": ("Zielfeld absichern", [
            "Feld auswählen: Ergebnis nur in ein fachlich passendes Feld schreiben.",
            "Überschreibung prüfen: Bestehende Werte nicht ungewollt ersetzen.",
            "Benutzerführung testen: Anwender müssen erkennen, was automatisch passiert.",
            "Speicherverhalten prüfen: Ergebnis nach Speichern und erneutem Öffnen kontrollieren.",
        ]),
        "bullets": [
            "Falsche Zielfelder verursachen schwer erkennbare Datenfehler.",
            "Automatische Werte sollten bei Bedarf korrigierbar bleiben.",
            "Pflichtfelder und Validierungen müssen mit der Suche harmonieren.",
        ],
        "box": ("Datenwirkung", "Das Zielfeld ist der Punkt, an dem eine Suche Daten verändert. Prüfen Sie diesen Schritt besonders sorgfältig."),
    },
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def markdown_for(page: dict[str, object]) -> str:
    parts = [f"# {page['title']}", "", str(page["intro"]).strip()]
    flow = page.get("flow")
    if flow:
        title, items = flow
        parts.extend(["", f":::flow {title}"])
        parts.extend(f"- {item}" for item in items)
        parts.append(":::")
    bullets = page.get("bullets") or []
    if bullets:
        parts.extend(["", "## Worauf Sie achten sollten"])
        parts.extend(f"- {item}" for item in bullets)
    box = page.get("box")
    if box:
        title, body = box
        parts.extend(["", f":::box {title}", str(body).strip(), ":::"])
    return "\n".join(parts).strip() + "\n"


def meta(markdown: str, title: str) -> str:
    clean = re.sub(r"^#{1,6}\s*", "", markdown, flags=re.M)
    clean = re.sub(r"^:::(?:flow|box)\s+.*$", "", clean, flags=re.M)
    clean = clean.replace(":::", "")
    clean = re.sub(r"^[\s>*-]*[-*]\s*", "", clean, flags=re.M)
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        clean = f"{title} in der X-ERP Hilfe."
    return clean[:297].rstrip(" ,.;") + ("..." if len(clean) > 297 else "")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-customizing-quality-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False, keep_links=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    by_path = {
        cell_text(ws.cell(row, headers["URL_PATH"]).value).replace("\\", "/"): row
        for row in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row, headers["URL_PATH"]).value)
    }

    changed = 0
    missing: list[str] = []
    for path, page in PAGES.items():
        row = by_path.get(path)
        if not row:
            missing.append(path)
            continue
        title = str(page["title"])
        body = markdown_for(page)
        ws.cell(row, headers["Beschreibung"]).value = body
        for column, value in {
            "H1": title,
            "NAV_TITLE": title,
            "TITLE": f"{title} | X-ERP ERP Hilfe",
            "META_DESCRIPTION": meta(body, title),
            "PRIMARY_KEYWORD": f"{title} X-ERP",
        }.items():
            if column in headers:
                ws.cell(row, headers[column]).value = value
        changed += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"changed={changed}")
    print(f"missing={len(missing)}")
    for path in missing:
        print(f"MISSING {path}")


if __name__ == "__main__":
    main()
