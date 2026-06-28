from __future__ import annotations

import locale
import shutil
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


@dataclass
class RowData:
    values: list
    styles: list
    height: float | None
    outline_level: int
    hidden: bool
    collapsed: bool


@dataclass
class DomainBlock:
    name: str
    source_row: int
    rows: list[RowData]


def capture_row(ws, row: int) -> RowData:
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    dim = ws.row_dimensions[row]
    return RowData(
        values=[ws.cell(row, col).value for col in range(1, ws.max_column + 1)],
        styles=styles,
        height=dim.height,
        outline_level=dim.outlineLevel,
        hidden=False,
        collapsed=False,
    )


def write_row(ws, row: int, data: RowData) -> None:
    for col, value in enumerate(data.values, start=1):
        cell = ws.cell(row, col)
        cell.value = value
        style = data.styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
    dim = ws.row_dimensions[row]
    dim.height = data.height
    dim.outlineLevel = data.outline_level
    dim.hidden = False
    dim.collapsed = False


def sort_key(name: str) -> str:
    name = name.replace("ß", "ss")
    normalized = unicodedata.normalize("NFKD", name)
    return "".join(char for char in normalized if not unicodedata.combining(char)).casefold()


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-sort-views-domains-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    ansichten_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")
    faq_row = next(row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")

    blocks: list[DomainBlock] = []
    row = ansichten_row + 1
    while row < faq_row:
        content_type = ws.cell(row, content_col).value
        level = ws.row_dimensions[row].outlineLevel
        name = ws.cell(row, thema_col).value
        if content_type == "HelpPage" and level == 1 and name:
            start = row
            row += 1
            while row < faq_row and not (
                ws.cell(row, content_col).value == "HelpPage" and ws.row_dimensions[row].outlineLevel == 1
            ):
                row += 1
            blocks.append(
                DomainBlock(
                    name=str(name),
                    source_row=start,
                    rows=[capture_row(ws, source_row) for source_row in range(start, row)],
                )
            )
        else:
            raise RuntimeError(f"Unexpected row in Ansichten domain area: row={row}, name={name}, type={content_type}, level={level}")

    sorted_blocks = sorted(blocks, key=lambda block: sort_key(block.name))
    rebuilt = [row_data for block in sorted_blocks for row_data in block.rows]

    delete_count = faq_row - (ansichten_row + 1)
    ws.delete_rows(ansichten_row + 1, delete_count)
    ws.insert_rows(ansichten_row + 1, len(rebuilt))
    for offset, row_data in enumerate(rebuilt):
        write_row(ws, ansichten_row + 1 + offset, row_data)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].collapsed = False

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    if ws.sheet_view:
        ws.sheet_view.showOutlineSymbols = True

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print("domain_order=")
    for index, block in enumerate(sorted_blocks, start=1):
        print(f"{index:02d} {block.name}")


if __name__ == "__main__":
    main()
