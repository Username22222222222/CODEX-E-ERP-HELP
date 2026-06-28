from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"


CONTENT = {
    "Stammdaten/artikel/index.html": """# Artikelstammdaten

Artikelstammdaten sind der zentrale Ort für alles, was in X-ERP verkauft, eingekauft, gelagert, produziert, als Set kombiniert oder als Makro automatisiert wird. Ein gut gepflegter Artikel ist nicht nur ein Name mit Preis, sondern ein Steuerungsobjekt für Belege, Lager, Einkauf, Produktion, Webshop, Auswertungen und Automatisierung.

Diese Seite hilft Ihnen zu entscheiden, welche Artikelart Sie benötigen und in welcher Reihenfolge Sie die Daten sinnvoll pflegen.

:::flow Artikel sicher anlegen
- Zweck klären: Wird der Artikel verkauft, eingekauft, gelagert, produziert, als Set genutzt oder als Makro verwendet?
- Grunddaten pflegen: Artikelnummer, Matchcode, Bezeichnung, Mengeneinheit, Warengruppe und Status festlegen.
- Prozessbereiche aktivieren: Verkauf, Beschaffung, Lagerung, Produktion, Set, Makro oder Zubehör nur dort pflegen, wo sie fachlich benötigt werden.
- Preise und Kontierung prüfen: Verkaufs- und Einkaufspreise, Finanzgruppen, Rabattlogik und Preismengen sauber hinterlegen.
- Abhängigkeiten ergänzen: Set-Komponenten, Produktionskomponenten, Katalognummern, Texte, Bilder und Zubehör erfassen.
- Testen: Artikel in Suche, Beleg, Lagerbewegung oder Produktion kurz prüfen, bevor er produktiv verwendet wird.
:::

## Welche Artikelart passt?

:::box Normaler Verkaufs- oder Einkaufsartikel
Verwenden Sie einen normalen Artikel für Waren, Leistungen oder Materialien, die auf Belegen erscheinen. Aktivieren Sie Verkauf, Beschaffung und Lagerung nur, wenn der Artikel in diesen Prozessen tatsächlich verwendet wird.
:::

:::box Set-Artikel
Ein Set fasst mehrere Artikel zu einem Paket zusammen. Das ist sinnvoll für Bundles, Aktionen, Zubehörpakete oder vorkonfigurierte Zusammenstellungen. Entscheidend ist, ob auf dem Beleg das Set als eine Position sichtbar bleiben soll oder ob die Komponenten aufgelöst werden.
:::

:::box Produktionsartikel
Ein Produktionsartikel beschreibt ein herzustellendes Endprodukt. Er verbindet Stammdaten mit Komponenten, Produktionsschritten, Ressourcen und Zeiten. Er ist die Grundlage für Fertigung, Materialbedarf und Nachkalkulation.
:::

:::box Makro-Artikel
Ein Makro-Artikel löst eine vorbereitete Logik aus. Er eignet sich, wenn beim Einfügen eines Artikels wiederkehrende Positionen, Texte oder Prozessregeln automatisch entstehen sollen.
:::

:::box Zubehör
Zubehörartikel werden einem Hauptartikel zugeordnet und können im Verkauf als ergänzende Artikel vorgeschlagen werden. Das verbessert Beratung, Warenkorbqualität und Vollständigkeit von Angeboten.
:::

## Gute Stammdatenqualität

- Verwenden Sie eindeutige, dauerhafte Artikelnummern.
- Pflegen Sie Bezeichnungen so, dass Anwender den Artikel auch ohne Zusatzwissen finden.
- Trennen Sie Artikelarten sauber: Set ist nicht dasselbe wie Produktion, Zubehör ist nicht dasselbe wie Set-Komponente.
- Vermeiden Sie Preis-, Steuer- oder Lieferanteninformationen in der Artikelnummer.
- Prüfen Sie bei lagergeführten Artikeln Mindestbestand, Standardlager und negative Bestände besonders sorgfältig.

## Typische Reihenfolge für neue Artikel

1. Artikel Übersicht pflegen.
2. Verkauf und Beschaffung aktivieren, wenn der Artikel dort verwendet wird.
3. Lagerung nur aktivieren, wenn Bestände geführt werden sollen.
4. Texte, Bild, Kategorien und Katalognummern ergänzen.
5. Bei Spezialfällen Set, Makro, Produktion oder Zubehör pflegen.
6. Artikel in einem Beispielbeleg oder Prozess testen.
""",
    "Stammdaten/artikel/artikel-uebersicht/index.html": """# Artikel Übersicht

Die Artikel-Übersicht enthält die wichtigsten Identifikationsdaten. Diese Angaben entscheiden, ob Anwender den Artikel schnell finden, eindeutig unterscheiden und korrekt in Belegen verwenden können.

:::flow Normalen Artikel anlegen
- Artikelnummer vergeben: Eindeutig, dauerhaft und möglichst regelbasiert.
- Bezeichnung pflegen: Name1 und Name2 so formulieren, dass Suche und Belege verständlich bleiben.
- Mengeneinheit wählen: Die Basiseinheit festlegen, in der Verkauf, Einkauf oder Lagerung rechnen.
- Typ und Gruppe setzen: Warengruppe, Artikeltyp und Status fachlich richtig zuordnen.
- Suchbarkeit verbessern: Matchcode, Kategorien und Favoriten nutzen, wenn viele ähnliche Artikel existieren.
- Speichern und prüfen: Artikel öffnen, suchen und in einem Testbeleg auswählen.
:::

:::box Empfehlung zur Artikelnummer
Die Artikelnummer sollte nicht von kurzfristigen Eigenschaften abhängen. Lieferant, Lagerort oder Preisgruppe können sich ändern und gehören deshalb in eigene Felder.
:::

## Was hier besonders wichtig ist

- Artikelnummer und Matchcode sollten eindeutig sein.
- Bezeichnungen sollten für Anwender, Belege und Suche verständlich sein.
- Die Mengeneinheit muss zur späteren Preis- und Lagerlogik passen.
- Der Status entscheidet, ob ein Artikel aktiv genutzt werden kann.

Eine saubere Übersicht verhindert Dubletten und reduziert Rückfragen in Verkauf, Einkauf und Lager.
""",
    "Stammdaten/artikel/artikel-details/index.html": """# Artikel Details

Die Details ergänzen die Grunddaten um Informationen, die für Handel, Webshop, Druckausgaben und Mengenlogik wichtig sind. Diese Felder wirken oft unscheinbar, haben aber direkte Auswirkungen auf Belege, Etiketten, Schnittstellen und Auswertungen.

## Typische Inhalte

- Hersteller und Hersteller-Bestellnummer für Beschaffung und Produktidentifikation.
- GTIN/EAN für Barcode, Handel, Kataloge und externe Systeme.
- Shop-Kennzeichen für die Anzeige im Webshop.
- Drucken zur Steuerung, ob der Artikel auf Belegen erscheint.
- Anzahl Nachkommastellen für die Mengenerfassung.
- Inhalt und Inhalts-ME für Verpackungs- oder Verkaufseinheiten.

:::box Praxisregel
Pflegen Sie GTIN, Herstellerdaten und Inhaltsangaben möglichst früh. Diese Informationen werden häufig erst beim Etikett, Webshop oder Lieferantenabgleich vermisst.
:::
""",
    "Stammdaten/artikel/artikel-verkauf/index.html": """# Artikel Verkauf

Der Bereich Verkauf steuert, wie ein Artikel in Angeboten, Aufträgen, Lieferscheinen und Rechnungen verwendet wird. Hier werden verkaufsrelevante Mengen, Lieferlogik, Kontierung und Preisgrundlagen gepflegt.

:::flow Verkauf vorbereiten
- Verkaufsartikel aktivieren: Nur Artikel aktivieren, die tatsächlich verkauft werden sollen.
- Mengenlogik festlegen: Mindestmenge, Losgröße und Preiseinheit prüfen.
- Lieferzeit pflegen: Liefertage realistisch angeben, damit Termine belastbar bleiben.
- Finanzgruppe wählen: Erlöskonten und Steuerlogik über die richtige Verkaufs-Finanzgruppe steuern.
- Preislisten prüfen: Verkaufspreise mit Gültigkeit und Mengeneinheiten kontrollieren.
:::

:::box Wichtig für Anwender
Ein Artikel kann existieren, ohne verkaufsfähig zu sein. Wenn er nicht in Verkaufsbelegen angeboten werden soll, sollte die Verkaufseinstellung bewusst deaktiviert bleiben.
:::
""",
    "Stammdaten/artikel/artikel-beschaffung/index.html": """# Artikel Beschaffung

Der Bereich Beschaffung beschreibt, wie ein Artikel eingekauft wird. Er ist wichtig für Lieferantenpreise, Beschaffungsplanung, Kostenbewertung und Einkaufsauswertungen.

## Wofür dieser Bereich genutzt wird

- Einkaufspreise und Einstandskosten.
- Beschaffungs-Finanzgruppe für Aufwandskonten und Steuerlogik.
- Lieferantenpreise, Katalognummern, Mindestmengen und Lieferzeiten.
- Grundlage für Beschaffungsassistent und Einkaufsvorgänge.

:::box Einkaufspreis ist nicht gleich Einstandspreis
Der Einkaufspreis ist der Preis des Lieferanten. Der Einstandspreis kann zusätzliche Kosten wie Logistik, Zoll oder Risiko enthalten und ist deshalb für Kalkulation und Bewertung oft aussagekräftiger.
:::
""",
    "Stammdaten/artikel/artikel-text/index.html": """# Artikel Text

Im Bereich Text werden beschreibende Artikeltexte gepflegt. Diese Texte helfen Anwendern intern und können auf Belegen, Bestellungen oder anderen Ausgaben verwendet werden.

## Gute Artikeltexte

- Beschreiben Sie den Artikel fachlich, nicht nur mit Schlagworten.
- Verwenden Sie Bestelltexte, wenn Lieferanten besondere Angaben benötigen.
- Nutzen Sie Textbausteine für wiederkehrende Formulierungen.
- Vermeiden Sie technische interne Kürzel, wenn der Text auf externen Belegen erscheinen kann.

:::box Unterschied zwischen Info- und Bestelltext
Der Infotext erklärt den Artikel aus Verkaufssicht oder für interne Hinweise. Der Bestelltext richtet sich an Lieferanten und sollte alle einkaufsrelevanten Angaben enthalten.
:::
""",
    "Stammdaten/artikel/artikel-bild/index.html": """# Artikel Bild

Artikelbilder unterstützen Anwender bei Suche, Auswahl, Verkauf und Webshop-Pflege. Sie sind besonders hilfreich bei ähnlichen Artikeln, Ersatzteilen, Varianten oder erklärungsbedürftigen Produkten.

## Einsatzbereiche

- Anzeige in Artikelansichten und Listen.
- Nutzung in Angeboten, Katalogen oder Webshop.
- Schnellere visuelle Kontrolle bei ähnlichen Artikeln.

:::box Bildqualität
Verwenden Sie klare, sachliche Produktbilder mit ausreichend Auflösung. Vermeiden Sie Bilder, die wichtige Details abschneiden oder nicht den tatsächlichen Artikel zeigen.
:::
""",
    "Stammdaten/artikel/artikel-lagerung/index.html": """# Artikel Lagerung

Die Lagerung steuert, ob und wie Bestände für einen Artikel geführt werden. Dieser Bereich ist nur relevant, wenn der Artikel physisch oder mengenmäßig im Lager verfolgt werden soll.

:::flow Lagerartikel einrichten
- Lagerführung aktivieren: Nur für Artikel, deren Bestand überwacht werden soll.
- Standardlager setzen: Das bevorzugte Lager für Zu- und Abgänge festlegen.
- Mindestbestand prüfen: Nachbestell- oder Warnlogik realistisch einstellen.
- Negative Bestände bewerten: Nur zulassen, wenn der Prozess das bewusst erfordert.
- Bestand testen: Eine Lagerbewegung oder Belegbuchung prüfen.
:::

:::box Achtung bei nachträglicher Lageraktivierung
Wenn ein Artikel bereits in Belegen verwendet wurde, sollte die Aktivierung der Lagerführung sorgfältig geprüft werden. Bestand, Bewertung und Historie müssen danach fachlich stimmen.
:::
""",
    "Stammdaten/artikel/artikel-lagerungshistorie/index.html": """# Artikel Lagerungshistorie

Die Lagerungshistorie zeigt die Lagerbewegungen eines Artikels. Sie hilft zu verstehen, warum ein Bestand entstanden ist und welche Belege oder Vorgänge ihn verändert haben.

## Typische Fragen

- Wann wurde Bestand eingebucht oder ausgebucht?
- Aus welchem Beleg stammt eine Bewegung?
- Warum stimmt der aktuelle Bestand nicht mit der Erwartung überein?
- Welche Lagerplätze oder Lager waren beteiligt?

Die Historie ist besonders wichtig bei Inventur, Reklamationen, Nachverfolgung und Bestandskorrekturen.
""",
    "Stammdaten/artikel/artikel-set/index.html": """# Set-Artikel anlegen

Ein Set-Artikel bündelt mehrere Artikel zu einem Paket. Anwender wählen im Beleg den Set-Artikel aus, X-ERP kennt aber die enthaltenen Komponenten. Sets eignen sich für Pakete, Bundles, Aktionsartikel, Zubehörkombinationen oder vorkonfigurierte Verkaufsartikel.

:::flow Set sauber aufbauen
- Set-Artikel anlegen: Einen eigenen Artikel für das Paket erstellen.
- Set-Bereich öffnen: Im Register Set die Komponenten pflegen.
- Komponenten hinzufügen: Jeden enthaltenen Artikel mit Mengenfaktor erfassen.
- Preislogik entscheiden: Preis am Set führen oder Komponenten separat bewerten.
- Auflösung festlegen: Entscheiden, ob Komponenten auf dem Beleg sichtbar werden.
- Testbeleg erstellen: Prüfen, ob Menge, Text, Preis und Druckbild stimmen.
:::

:::box Set oder Zubehör?
Ein Set ist ein Paket, das aus Komponenten besteht. Zubehör sind zusätzliche Artikel, die empfohlen werden, aber nicht automatisch Bestandteil des Hauptartikels sein müssen.
:::

:::box Set oder Produktion?
Ein Set beschreibt eine kaufmännische Zusammenstellung. Ein Produktionsartikel beschreibt eine Fertigung mit Material, Ressourcen und Arbeitsschritten.
:::

## Wichtige Felder

- Komponente: Der Artikel, der im Set enthalten ist.
- Mengenfaktor: Wie oft die Komponente pro Set benötigt wird.
- Kostenlos: Die Komponente wird mitgeliefert, aber nicht berechnet.
- Auflösen: Steuert, ob die Komponenten auf dem Beleg einzeln erscheinen.
""",
    "Stammdaten/artikel/artikel-makro/index.html": """# Makro-Artikel

Ein Makro-Artikel steht nicht nur für ein Produkt, sondern für vorbereitete Logik. Beim Verwenden des Artikels können wiederkehrende Positionen, Texte oder Prozessregeln automatisch entstehen.

:::flow Makro-Artikel verwenden
- Zweck definieren: Klären, welche wiederkehrende Aktion automatisiert werden soll.
- Auslöseartikel anlegen: Einen Artikel schaffen, den Anwender bewusst auswählen können.
- Makro hinterlegen: Die auszuführende Logik dem Artikel zuordnen.
- Ergebnis prüfen: Testen, welche Positionen, Texte oder Werte entstehen.
- Anwender schulen: Klar benennen, wann der Makro-Artikel verwendet werden soll.
:::

:::box Wann ist ein Makro sinnvoll?
Makros sind sinnvoll, wenn Anwender immer wieder dieselbe Struktur erfassen müssten. Sie ersetzen keine sauberen Stammdaten, sondern automatisieren wiederkehrende Abläufe.
:::
""",
    "Stammdaten/artikel/artikel-produktion/index.html": """# Produktionsartikel

Ein Produktionsartikel beschreibt ein herzustellendes Endprodukt. Er verbindet Verkaufs- oder Lagerartikel mit Materialbedarf, Produktionskomponenten, Produktionsschritten und Ressourcen.

:::flow Produktionsartikel vorbereiten
- Endprodukt anlegen: Den fertigen Artikel als eigenen Artikel pflegen.
- Komponenten zuordnen: Benötigte Materialien und Mengen erfassen.
- Produktionsschritte definieren: Arbeitsschritte fachlich in sinnvoller Reihenfolge abbilden.
- Ressourcen planen: Maschinen, Arbeitsplätze oder Mitarbeitergruppen zuordnen.
- Zeiten prüfen: Rüst-, Bearbeitungs- oder Produktionszeiten realistisch pflegen.
- Testproduktion prüfen: Materialbedarf, Zeiten und Ergebnis kontrollieren.
:::

:::box Produktionsartikel oder Set?
Ein Produktionsartikel wird gefertigt und benötigt eine Produktionslogik. Ein Set ist eine Zusammenstellung vorhandener Artikel und bildet meist keinen Fertigungsprozess ab.
:::
""",
    "Stammdaten/artikel/artikel-zubehoer/index.html": """# Zubehör pflegen

Zubehör verbindet einen Hauptartikel mit ergänzenden Artikeln. Diese werden Anwendern im Verkauf vorgeschlagen, damit Angebote vollständiger und Beratungsprozesse sicherer werden.

:::flow Zubehör sinnvoll zuordnen
- Hauptartikel wählen: Den Artikel öffnen, zu dem Zubehör empfohlen werden soll.
- Zubehörartikel ergänzen: Passende Ergänzungsartikel hinterlegen.
- Verkaufssituation prüfen: Nur Zubehör aufnehmen, das in der Praxis wirklich vorgeschlagen werden soll.
- Reihenfolge bewerten: Wichtiges Zubehör möglichst sichtbar halten.
- Testangebot erstellen: Prüfen, ob Anwender das Zubehör im Verkaufsprozess sinnvoll nutzen können.
:::

:::box Zubehör ist kein Set
Zubehör wird vorgeschlagen. Ein Set ist bereits eine feste Zusammenstellung. Nutzen Sie Zubehör für Empfehlungen und Sets für verbindliche Pakete.
:::
""",
    "Stammdaten/artikel/artikel-kategorien/index.html": """# Artikel Kategorien

Kategorien strukturieren Artikel für Suche, Filter, Auswertungen und Webshop. Sie helfen Anwendern, große Artikelmengen fachlich zu ordnen.

## Empfehlungen

- Kategorien nach Anwendersicht aufbauen, nicht nur nach interner Technik.
- Zu tiefe Strukturen vermeiden.
- Kategorien konsistent verwenden, damit Filter und Auswertungen zuverlässig bleiben.
- Webshop- und Kataloganforderungen früh berücksichtigen.
""",
    "Stammdaten/artikel/artikel-katalognummern/index.html": """# Artikel Katalognummern

Katalognummern speichern abweichende Nummern, unter denen ein Artikel bei Kunden, Lieferanten oder in Katalogen geführt wird. Dadurch bleibt die eigene Artikelnummer stabil, während externe Nummern sauber zugeordnet werden.

## Typische Nutzung

- Lieferantenartikelnummern für Bestellungen.
- Kundenartikelnummern für Rahmenverträge oder EDI.
- Katalognummern für Import, Suche und Abstimmung.

:::box Vorteil
Externe Nummern gehören nicht in die eigene Artikelnummer. So bleibt die interne Struktur stabil und trotzdem sind Lieferanten- oder Kundennummern schnell auffindbar.
:::
""",
    "Stammdaten/artikel/artikel-produktionsschritt-ressourcen/index.html": """# Produktionsschritt Ressourcen

Dieser Bereich ordnet einem Produktionsschritt die benötigten Ressourcen zu. Ressourcen können Maschinen, Arbeitsplätze, Werkzeuge oder andere Kapazitäten sein.

## Warum das wichtig ist

- Produktionszeiten werden planbarer.
- Engpässe werden sichtbar.
- Arbeitsschritte lassen sich besser kalkulieren.
- Ressourcen können später in Auswertungen und Planung berücksichtigt werden.
""",
    "Stammdaten/artikel/artikel-produktionsschritt-artikel/index.html": """# Produktionsschritt Artikel

Dieser Bereich beschreibt, welche Artikel oder Materialien in einem Produktionsschritt eingesetzt werden. Damit wird sichtbar, was nicht nur für das Endprodukt, sondern konkret für einen Arbeitsschritt benötigt wird.

## Gute Pflege

- Materialien dem richtigen Schritt zuordnen.
- Mengen realistisch und nachvollziehbar pflegen.
- Ersatz- oder Hilfsstoffe nicht mit Hauptkomponenten verwechseln.
- Nach Änderungen einen Test der Produktionsstruktur durchführen.
""",
    "Stammdaten/artikel/artikel-produktionszeit/index.html": """# Produktionszeit

Produktionszeiten helfen, Fertigung, Kapazitäten und Kosten realistisch zu planen. Sie können für Artikel oder produktionsnahe Leistungen relevant sein.

:::box Praxisregel
Pflegen Sie Produktionszeiten nur so genau, wie sie später auch genutzt und kontrolliert werden. Zu grobe Zeiten machen Planung ungenau, zu feine Zeiten erhöhen den Pflegeaufwand.
:::
""",
    "Stammdaten/artikel/artikel-verwendung-bei-sets/index.html": """# Verwendung bei Sets

Diese Seite zeigt, in welchen Set-Artikeln der aktuelle Artikel als Komponente verwendet wird. Sie ist wichtig, bevor ein Artikel geändert, deaktiviert oder ersetzt wird.

## Typische Fragen

- Welche Sets enthalten diesen Artikel?
- Welche Mengen werden dort verwendet?
- Wird eine Änderung Auswirkungen auf bestehende Pakete haben?
- Muss ein Ersatzartikel in mehreren Sets gepflegt werden?
""",
    "Stammdaten/artikel/artikel-verwendung-bei-produktion/index.html": """# Verwendung bei Produktion

Diese Seite zeigt, in welchen Produktionsartikeln der aktuelle Artikel als Bestandteil verwendet wird. Sie unterstützt die Kontrolle von Materialabhängigkeiten.

## Wann prüfen?

- Vor dem Deaktivieren eines Artikels.
- Vor Preis- oder Mengeneinheitsänderungen.
- Bei Lieferproblemen oder Materialersatz.
- Bei Änderungen an Stücklisten oder Produktionsstrukturen.
""",
    "Stammdaten/artikel/artikel-positionsliste/index.html": """# Artikel Positionsliste

Die Positionsliste zeigt, in welchen Belegen der Artikel verwendet wurde. Sie verbindet den Artikelstamm mit dem tatsächlichen Geschäftsverlauf.

## Nutzen

- Nachvollziehen, in welchen Angeboten, Aufträgen oder Rechnungen ein Artikel vorkam.
- Verkaufs- und Einkaufsaktivität prüfen.
- Auswirkungen von Artikeländerungen besser einschätzen.
- Historische Preise, Mengen und Partner kontrollieren.
""",
    "Stammdaten/artikel/artikel-umsatz/index.html": """# Artikel Umsatz

Der Umsatzbereich zeigt die wirtschaftliche Entwicklung eines Artikels. Er hilft bei Sortimentsbewertung, Preisprüfung und Vertriebssteuerung.

## Typische Auswertungsfragen

- Wie entwickelt sich der Umsatz?
- Gibt es saisonale Muster?
- Ist der Rohertrag plausibel?
- Welche Artikel sollten aktiv vertrieben, überarbeitet oder ausgelistet werden?
""",
    "Stammdaten/artikel/artikel-vertraege/index.html": """# Artikel Verträge

Dieser Bereich zeigt Verträge oder Abos, in denen der Artikel verwendet wird. Das ist wichtig, wenn Artikel geändert, ersetzt oder aus dem Sortiment genommen werden.

:::box Vor Änderungen prüfen
Wenn ein Artikel in Verträgen verwendet wird, können Preis-, Text- oder Statusänderungen Auswirkungen auf wiederkehrende Vorgänge haben.
:::
""",
    "Stammdaten/artikel/artikel-lokal/index.html": """# Artikel Lokal

Der Bereich Lokal enthält länder- und außenhandelsrelevante Angaben. Diese Daten sind wichtig für internationale Prozesse, Zoll, Intrastat, Auswertungen oder lokale Besonderheiten.

## Beispiele

- Ursprungsland.
- Warencode oder Zolltarifnummer.
- Zollbeschreibung.
- Transportweg oder Lieferzeit.
- Lokale Artikelklassifikation.
""",
    "Stammdaten/artikel/artikel-aenderungsprotokoll/index.html": """# Artikel Änderungsprotokoll

Das Änderungsprotokoll zeigt, welche Artikeldaten wann und von wem geändert wurden. Es unterstützt Nachvollziehbarkeit, Prüfung und Fehlersuche.

## Typische Nutzung

- Klären, warum ein Preis, Status oder Text anders ist.
- Änderungen vor Reklamationen oder Buchungsproblemen nachvollziehen.
- Verantwortlichkeiten bei Stammdatenänderungen prüfen.
""",
    "Stammdaten/artikel/artikel-anhaenge/index.html": """# Artikel Anhänge

Anhänge speichern Dateien und Dokumente direkt am Artikel. Dazu gehören Datenblätter, Zertifikate, Zeichnungen, Sicherheitsinformationen, Lieferantenunterlagen oder Produktfotos.

## Empfehlungen

- Benennen Sie Dateien eindeutig.
- Entfernen Sie veraltete Dokumente oder kennzeichnen Sie sie klar.
- Nutzen Sie Anhänge für Unterlagen, die Anwender direkt beim Artikel benötigen.
""",
    "Stammdaten/artikel/artikel-extra-felder/index.html": """# Artikel Extra-Felder

Extra-Felder erweitern den Artikelstamm um kundenspezifische Informationen. Welche Felder erscheinen, hängt von der Konfiguration ab.

## Gute Verwendung

- Nur Informationen als Extra-Feld anlegen, die regelmäßig gesucht, gefiltert oder ausgewertet werden.
- Feldnamen für Anwender verständlich formulieren.
- Pflichtfelder sparsam einsetzen.
- Bei Auswertungen auf einheitliche Werte achten.
""",
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def first_heading(markdown: str) -> str:
    for line in markdown.splitlines():
        line = line.strip()
        if line.startswith("#"):
            return line.lstrip("#").strip()
    return ""


def meta(markdown: str, title: str) -> str:
    clean = markdown.replace(":::flow", "").replace(":::box", "")
    clean = clean.replace(":::", "")
    lines = [line.strip("#- *0123456789. ").strip() for line in clean.splitlines()]
    text = " ".join(line for line in lines if line)
    text = " ".join(text.split())
    if not text:
        text = f"{title} in den X-ERP Artikelstammdaten."
    return text[:297].rstrip(" ,.;") + ("..." if len(text) > 297 else "")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-artikel-quality-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    by_path = {
        cell_text(ws.cell(row, headers["URL_PATH"]).value).replace("\\", "/"): row
        for row in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row, headers["URL_PATH"]).value)
    }

    changed = 0
    missing: list[str] = []
    for path, markdown in CONTENT.items():
        row = by_path.get(path)
        if not row:
            missing.append(path)
            continue
        title = first_heading(markdown) or cell_text(ws.cell(row, headers["Thema"]).value)
        ws.cell(row, headers["Beschreibung"]).value = markdown
        for column, value in {
            "H1": title,
            "NAV_TITLE": title,
            "TITLE": f"{title} | X-ERP ERP Hilfe",
            "META_DESCRIPTION": meta(markdown, title),
            "PRIMARY_KEYWORD": f"{title} X-ERP",
        }.items():
            if column in headers:
                ws.cell(row, headers[column]).value = value
        changed += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"changed={changed}")
    print(f"missing={len(missing)}")
    for path in missing:
        print(f"MISSING {path}")


if __name__ == "__main__":
    main()
