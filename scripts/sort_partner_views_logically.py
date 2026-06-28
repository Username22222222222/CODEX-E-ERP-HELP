from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"

PARTNER_ORDER = [
    "PartnerEdit",
    "PartnerContactPersonEdit",
    "PartnerDeliveryAddressEdit",
    "PartnerBillingAddressEdit",
    "PartnerBankAccountEdit",
    "PartnerCatalogNumberEdit",
    "PartnerPurchasePriceListEdit",
    "PartnerGroupEdit",
    "PartnerCategoryEdit",
    "PartnerCategoryNameEdit",
    "PartnerDiscountGroupEdit",
    "PartnerDiscountGroupNameEdit",
    "PartnerTermsOfPaymentEdit",
    "PartnerTermsOfDeliveryEdit",
    "PartnerDirectDebitTypeEdit",
    "PartnerDetailOutputEdit",
    "PartnerContactPersonEmailCategoryNameEdit",
    "PartnerContactPersonEmailCategoryContactPersonEdit",
    "PartnerContactPersonEmailCategoryDocTypeEdit",
    "PartnerSetFirstPartnerNumber",
    "PartnerSetFirstCustomerNumber",
    "PartnerSetFirstSupplierNumber",
]


@dataclass
class RowData:
    values: list
    styles: list
    height: float | None
    outline_level: int
    hidden: bool
    collapsed: bool


@dataclass
class Block:
    name: str
    source_row: int
    rows: list[RowData]


def capture_row(ws, row: int) -> RowData:
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    dim = ws.row_dimensions[row]
    return RowData(
        values=[ws.cell(row, col).value for col in range(1, ws.max_column + 1)],
        styles=styles,
        height=dim.height,
        outline_level=dim.outlineLevel,
        hidden=False,
        collapsed=False,
    )


def write_row(ws, row: int, data: RowData) -> None:
    for col, value in enumerate(data.values, start=1):
        cell = ws.cell(row, col)
        cell.value = value
        style = data.styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
    dim = ws.row_dimensions[row]
    dim.height = data.height
    dim.outlineLevel = data.outline_level
    dim.hidden = False
    dim.collapsed = False


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-sort-partner-views-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    ansichten_row = next(
        row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten"
    )
    partner_row = next(
        row
        for row in range(ansichten_row + 1, ws.max_row + 1)
        if ws.cell(row, thema_col).value == "Partner"
        and ws.cell(row, content_col).value == "HelpPage"
        and ws.row_dimensions[row].outlineLevel == 1
    )
    next_module_row = next(
        row
        for row in range(partner_row + 1, ws.max_row + 1)
        if ws.cell(row, content_col).value == "HelpPage" and ws.row_dimensions[row].outlineLevel == 1
    )

    module_row = capture_row(ws, partner_row)
    blocks: list[Block] = []
    row = partner_row + 1
    while row < next_module_row:
        content_type = ws.cell(row, content_col).value
        name = ws.cell(row, thema_col).value
        if content_type in {"View", "Wizard"} and name:
            start = row
            row += 1
            while row < next_module_row and ws.cell(row, content_col).value not in {"View", "Wizard"}:
                row += 1
            block_rows = [capture_row(ws, source_row) for source_row in range(start, row)]
            block_rows[0].outline_level = 2
            for child in block_rows[1:]:
                if child.outline_level < 3:
                    child.outline_level = 3
            blocks.append(Block(str(name), start, block_rows))
        else:
            row += 1

    order = {name: index for index, name in enumerate(PARTNER_ORDER)}
    blocks.sort(key=lambda block: (order.get(block.name, 10_000), block.source_row))

    rows_to_write = [module_row]
    rows_to_write[0].outline_level = 1
    rows_to_write.extend(row_data for block in blocks for row_data in block.rows)

    expected_row_count = next_module_row - partner_row
    if len(rows_to_write) != expected_row_count:
        raise RuntimeError(f"Row count changed unexpectedly: {len(rows_to_write)} != {expected_row_count}")

    for offset, row_data in enumerate(rows_to_write):
        write_row(ws, partner_row + offset, row_data)

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print("partner_order=")
    for index, block in enumerate(blocks, start=1):
        print(f"{index:02d} {block.name}")


if __name__ == "__main__":
    main()
