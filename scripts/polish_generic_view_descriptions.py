from __future__ import annotations

import datetime as dt
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"


def u(text: str) -> str:
    try:
        return text.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return text


def is_generic(text: str | None) -> bool:
    if not isinstance(text, str):
        return False
    patterns = [
        "ist eine fachliche Angabe",
        "unterst\u00fctzt Anwender dabei, den Datensatz korrekt zu pflegen",
        "unterstuetzt Anwender dabei, den Datensatz korrekt zu pflegen",
        "Beschreibt die Angabe `",
    ]
    return any(p in text for p in patterns)


def view_label(view: str | None) -> str:
    if not view:
        return "dieser Ansicht"
    label = re.sub(r"Edit$|Wizard$", "", str(view))
    label = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", label)
    return label


def section_text(name: str) -> str:
    return u("{name} buendelt zusammengehoerige Funktionen und Ansichten innerhalb der X-ERP-Hilfe. Der Abschnitt hilft Anwendern, die passende Seite schneller zu finden und den fachlichen Zusammenhang zu verstehen.").format(name=name)


def view_text(name: str, parent: str | None) -> str:
    label = view_label(name)
    parent_text = f" im Bereich {parent}" if parent else ""
    if name.endswith("Wizard"):
        return u("{name} fuehrt Anwender schrittweise durch einen gefuehrten ERP-Vorgang{parent_text}. Die Hilfe erklaert Voraussetzungen, Eingaben, Schaltflaechen und das Ergebnis des Assistenten.").format(name=name, parent_text=parent_text)
    return u("{name} ist die Bearbeitungsansicht fuer {label}{parent_text}. Die Hilfe erklaert Zweck, Register, Felder und Aktionen, damit Anwender die Maske sicher ausfuellen und den Vorgang richtig einordnen koennen.").format(name=name, label=label, parent_text=parent_text)


def register_text(name: str, view: str | None) -> str:
    clean = name
    if view and name.startswith(f"{view}-"):
        clean = name[len(view) + 1 :]
    return u("Das Register {clean} fasst die zugehoerigen Angaben der Ansicht {view} zusammen. Hier stehen die Felder und Aktionen, die fuer diesen Teil des Vorgangs gemeinsam bearbeitet oder geprueft werden.").format(clean=clean, view=view or "der Ansicht")


def field_text(field: str, view: str | None, tab: str | None) -> str:
    f = field.lower()
    context = f" in {view}" if view else ""
    tab_text = f" im Register {tab}" if tab else ""
    if field.startswith("Button: "):
        button = field.replace("Button: ", "", 1)
        b = button.lower()
        if "speichern" in b:
            return u("Speichert die aktuellen Eingaben{context}, sofern Pflichtfelder und Plausibilitaetspruefungen erfuellt sind.").format(context=context)
        if "abbrechen" in b:
            return u("Bricht die Bearbeitung ab oder verlaesst die Ansicht, ohne den aktuellen Schritt fortzusetzen.")
        if "vorschau" in b:
            return u("Oeffnet eine Vorschau des aktuellen Datensatzes oder der zugehoerigen Ausgabe.")
        if "neu" in b:
            return u("Startet die Neuanlage eines weiteren Datensatzes.")
        if "weiter" in b:
            return u("Fuehrt zum naechsten Schritt des Assistenten, wenn die bisherigen Eingaben gueltig sind.")
        if "erstellen" in b or "create" in b:
            return u("Erzeugt die vorgesehenen Folgedaten aus den eingegebenen Werten.")
        return u("Fuehrt die Aktion {button}{context} aus.").format(button=button, context=context)

    if any(x in f for x in ["status", "zustand"]):
        return u("Zeigt oder setzt den Bearbeitungsstand{context}. Der Status steuert, wie der Vorgang gefiltert, bewertet und weiterverarbeitet wird.").format(context=context)
    if any(x in f for x in ["partner", "kunde", "lieferant", "contact", "ansprechpartner"]):
        return u("Verknuepft den Vorgang{context} mit dem passenden Geschaeftspartner oder Ansprechpartner. Dadurch stimmen Kommunikation, Belege und Auswertungen mit dem richtigen Kontakt ueberein.").format(context=context)
    if any(x in f for x in ["datum", "date", "zeit", "time", "frist", "stichtag"]):
        return u("Speichert den zeitlichen Bezug{context}. Diese Angabe ist wichtig fuer Planung, Nachverfolgung, Wiedervorlage und Auswertung.").format(context=context)
    if any(x in f for x in ["menge", "quantity", "anzahl", "bestand"]):
        return u("Enthaelt eine Mengen- oder Bestandsangabe{context}. Der Wert beeinflusst Planung, Lager, Produktion oder Auswertung.").format(context=context)
    if any(x in f for x in ["preis", "betrag", "rabatt", "kosten", "konto", "finanz", "waehrung", "w\u00e4hrung", "steuer"]):
        return u("Enthaelt eine kaufmaennische Angabe{context}. Sie wirkt auf Bewertung, Buchhaltung, Preisfindung oder finanzielle Auswertungen.").format(context=context)
    if any(x in f for x in ["lager", "warehouse", "bin location", "lagerplatz"]):
        return u("Ordnet den Vorgang{context} einem Lager oder Lagerplatz zu. Diese Zuordnung ist fuer Bestand, Verfuegbarkeit und Lagerbewegungen entscheidend.").format(context=context)
    if any(x in f for x in ["name", "nummer", "matchcode", "id", "code"]):
        return u("Kennzeichnet den Datensatz{context} eindeutig oder sprechend. Anwender nutzen diese Angabe zum Suchen, Wiedererkennen und Zuordnen.").format(context=context)
    if any(x in f for x in ["beschreibung", "info", "text", "comment", "kommentar", "inhalt", "notiz"]):
        return u("Ergaenzt den Vorgang{context} um beschreibende Informationen. Der Text macht Entscheidungen, Hinweise oder fachliche Details nachvollziehbar.").format(context=context)
    if any(x in f for x in ["aktiv", "favorit", "gesperrt", "block", "anzeigen", "show"]):
        return u("Steuert eine Eigenschaft oder Sichtbarkeit{context}. Die Einstellung beeinflusst, ob der Datensatz aktiv genutzt, angezeigt oder blockiert wird.").format(context=context)
    if any(x in f for x in ["rolle", "benutzer", "recht", "authorization", "webapi"]):
        return u("Steuert Benutzer-, Rollen- oder Zugriffsbezug{context}. Diese Angabe ist fuer Berechtigungen und sichere ERP-Nutzung relevant.").format(context=context)
    return u("Das Feld {field}{tab_text}{context} speichert eine fachliche Information, die fuer Pflege, Suche, Auswertung oder Weiterverarbeitung dieses ERP-Vorgangs relevant ist.").format(field=field, tab_text=tab_text, context=context)


def main() -> None:
    ARCHIVE.mkdir(exist_ok=True)
    shutil.copy2(WORKBOOK, ARCHIVE / f"X-ERP-HELP-before-polish-generic-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx")
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}
    current_group = None
    current_view = None
    current_tab = None
    changed = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, col["Thema"]).value
        desc_cell = ws.cell(row, col["Beschreibung"])
        desc = desc_cell.value
        if not name:
            continue
        outline = ws.row_dimensions[row].outlineLevel
        content_type = ws.cell(row, col["CONTENT_TYPE"]).value if "CONTENT_TYPE" in col else None
        if outline == 1:
            current_group = str(name)
            current_view = None
            current_tab = None
        elif outline == 2:
            current_view = str(name)
            current_tab = None
        elif outline == 3:
            current_tab = str(name)

        if desc and not is_generic(desc):
            continue

        if outline == 1:
            new_desc = section_text(str(name))
        elif content_type in ("View", "Wizard") or outline == 2:
            new_desc = view_text(str(name), current_group)
        elif outline == 3 and current_view and (str(name).startswith(f"{current_view}-") or "-" in str(name)):
            new_desc = register_text(str(name), current_view)
        else:
            new_desc = field_text(str(name), current_view, current_tab)
        if new_desc != desc:
            desc_cell.value = new_desc
            changed += 1
    wb.save(WORKBOOK)
    print(f"changed={changed}")


if __name__ == "__main__":
    main()
