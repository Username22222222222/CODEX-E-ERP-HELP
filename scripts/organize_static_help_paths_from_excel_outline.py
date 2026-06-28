from __future__ import annotations

import re
import shutil
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"
ARCHIVE_DIR = Path(r"C:\Users\micha\Documents\X-ERP-HELP\ARCHIV")

TOP_LEVEL_SEGMENTS = {
    "Vorwort": "Vorwort",
    "Inhaltsverzeichnis": "Inhaltsverzeichnis",
    "Grundlagen": "Grundlagen",
    "Stammdaten": "Stammdaten",
    "Module": "Module",
    "Customizing": "Customizing",
    "Länderpakete": "Laenderpakete",
    "Portal-Apps": "Portal-Apps",
    "API-Schnittstellen": "API-Schnittstellen",
    "Branchenlösungen": "Branchenloesungen",
    "Firma": "Firma",
    "Setup": "Setup",
    "FAQ": "FAQ",
    "Glossar": "Glossar",
}


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def slugify(value: str) -> str:
    value = (
        value.replace("Ä", "Ae")
        .replace("Ö", "Oe")
        .replace("Ü", "Ue")
        .replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.lower()
    value = value.replace("&", " und ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-{2,}", "-", value).strip("-")
    return value or "seite"


def ensure_unique(path: str, seen: dict[str, int]) -> str:
    key = path.casefold()
    if key not in seen:
        seen[key] = 1
        return path
    seen[key] += 1
    return f"{path}-{seen[key]}"


def main() -> None:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE_DIR / f"X-ERP-HELP-before-path-outline-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    ws = wb[SHEET]
    headers = {
        text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if text(ws.cell(1, col).value)
    }

    required = {"Thema", "URL_PATH", "DIRECTORY_PATH", "FILE_NAME", "STORAGE_PATH", "NAV_TITLE", "BREADCRUMB", "TOC_PARENT", "TOC_LEVEL", "UNIQUE_PAGE_KEY"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"Missing required headers: {', '.join(missing)}")

    content_col = headers.get("CONTENT_TYPE")
    canonical_col = headers.get("CANONICAL_URL")
    seo_status_col = headers.get("SEO_STATUS")
    reviewed_col = headers.get("LAST_REVIEWED")

    stack_titles: dict[int, str] = {}
    stack_segments: dict[int, str] = {}
    seen_paths: dict[str, int] = {}
    updated = 0
    skipped_ansichten = 0

    in_ansichten = False
    for row in range(2, ws.max_row + 1):
        topic = text(ws.cell(row, headers["Thema"]).value)
        if not topic:
            continue

        level = int(ws.row_dimensions[row].outlineLevel or 0)
        if level == 0:
            in_ansichten = topic == "Ansichten"
        if in_ansichten:
            skipped_ansichten += 1
            continue

        # Remove deeper stale stack entries before assigning the current row.
        for old_level in list(stack_titles):
            if old_level >= level:
                stack_titles.pop(old_level, None)
                stack_segments.pop(old_level, None)

        nav_title = text(ws.cell(row, headers["NAV_TITLE"]).value) or topic
        segment = TOP_LEVEL_SEGMENTS.get(nav_title) if level == 0 else slugify(nav_title)
        if not segment:
            segment = slugify(topic)

        parent_segments = [stack_segments[i] for i in sorted(stack_segments) if i < level]
        directory = "/".join(parent_segments + [segment])
        directory = ensure_unique(directory, seen_paths)
        url_path = f"{directory}/index.html"

        stack_titles[level] = nav_title
        stack_segments[level] = directory.split("/")[-1]

        breadcrumb_parts = [stack_titles[i] for i in sorted(stack_titles) if i <= level]
        effective_level = max(0, len(breadcrumb_parts) - 1)
        breadcrumb = " > ".join(breadcrumb_parts)
        parent_title = breadcrumb_parts[-2] if len(breadcrumb_parts) >= 2 else None

        ws.cell(row, headers["DIRECTORY_PATH"]).value = directory
        ws.cell(row, headers["FILE_NAME"]).value = "index.html"
        ws.cell(row, headers["URL_PATH"]).value = url_path
        ws.cell(row, headers["STORAGE_PATH"]).value = url_path
        ws.cell(row, headers["BREADCRUMB"]).value = breadcrumb
        ws.cell(row, headers["TOC_PARENT"]).value = parent_title
        ws.cell(row, headers["TOC_LEVEL"]).value = effective_level
        ws.cell(row, headers["UNIQUE_PAGE_KEY"]).value = directory
        ws.row_dimensions[row].outlineLevel = effective_level

        if canonical_col:
            ws.cell(row, canonical_col).value = "https://x-erp.de/de/help/" + directory + "/"
        if seo_status_col and text(ws.cell(row, seo_status_col).value) == "draft":
            ws.cell(row, seo_status_col).value = "reviewed"
        if reviewed_col and not ws.cell(row, reviewed_col).value:
            ws.cell(row, reviewed_col).value = date(2026, 6, 28)
        if content_col and not text(ws.cell(row, content_col).value):
            ws.cell(row, content_col).value = "HelpPage"

        updated += 1

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"updated_static_rows={updated}")
    print(f"skipped_ansichten_rows={skipped_ansichten}")


if __name__ == "__main__":
    main()
