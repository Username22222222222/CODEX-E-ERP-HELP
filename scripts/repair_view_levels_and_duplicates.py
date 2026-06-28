from __future__ import annotations

import datetime as dt
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"


def copy_cell_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)


def apply_row_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed


def main() -> None:
    ARCHIVE.mkdir(exist_ok=True)
    shutil.copy2(WORKBOOK, ARCHIVE / f"X-ERP-HELP-before-repair-view-levels-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx")
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}
    view_style_row = 1142
    group_style_row = 1141

    removed = 0
    for row in range(ws.max_row, 2, -1):
        name = ws.cell(row, col["Thema"]).value
        prev_name = ws.cell(row - 1, col["Thema"]).value
        desc = ws.cell(row, col["Beschreibung"]).value
        prev_desc = ws.cell(row - 1, col["Beschreibung"]).value
        if name and name == prev_name and desc == prev_desc and ws.row_dimensions[row].outlineLevel == ws.row_dimensions[row - 1].outlineLevel:
            ws.delete_rows(row, 1)
            removed += 1

    fixed_levels = 0
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row, col["Thema"]).value
        content_type = ws.cell(row, col["CONTENT_TYPE"]).value if "CONTENT_TYPE" in col else None
        if content_type in ("View", "Wizard") and ws.row_dimensions[row].outlineLevel != 2:
            apply_row_style(ws, view_style_row, row)
            fixed_levels += 1
        if isinstance(name, str) and name.startswith("Nachtrag - "):
            apply_row_style(ws, group_style_row, row)
            ws.cell(row, col["Beschreibung"]).value = (
                f"{name} fasst nachgetragene Ansichten zusammen, die aus den Razor-Dateien "
                "und der Routeninventur in die Hilfe-Struktur uebernommen wurden."
            )

    wb.save(WORKBOOK)
    print(f"removed_duplicates={removed}")
    print(f"fixed_view_levels={fixed_levels}")


if __name__ == "__main__":
    main()
