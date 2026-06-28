from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"


MODULES = {
    "verkauf": {
        "label": "Verkauf",
        "intro": "Der Verkauf führt vom ersten Angebot bis zur Rechnung durch den kompletten Kundenprozess. Das Modul verbindet Partner, Artikel, Preise, Lieferinformationen, Belege und Folgeprozesse zu einem nachvollziehbaren Ablauf.",
        "flow": "Verkaufsprozess steuern",
        "steps": [
            "Kundenbedarf erfassen: Partner, Ansprechpartner, Lieferadresse und gewünschte Artikel sauber wählen.",
            "Konditionen prüfen: Preise, Rabatte, Zahlungsbedingungen und Liefertermine kontrollieren.",
            "Beleg fortführen: Angebot, Auftrag, Lieferschein, Rechnung und Gutschrift bewusst als Kette nutzen.",
            "Abweichungen klären: Mengen, Preise, Lieferfähigkeit und offene Rückfragen vor dem nächsten Schritt bereinigen.",
            "Nachverfolgen: Status, Belegkette und offene Vorgänge regelmäßig prüfen.",
        ],
        "bullets": [
            "Ein Verkaufsbeleg ist mehr als ein Ausdruck; er startet Folgeprozesse in Lager, Versand, Finanzwesen und Auswertung.",
            "Je sauberer Artikel, Partner und Preise gepflegt sind, desto weniger manuelle Korrekturen entstehen im Beleg.",
            "Die Belegkette hilft, Ursprung und Folgebelege später sicher nachzuvollziehen.",
        ],
        "box": ("Praxisregel", "Führen Sie Belege möglichst über die vorgesehenen Folgefunktionen fort. So bleiben Mengen, Preise, Adressen und Status nachvollziehbar verbunden."),
    },
    "einkauf": {
        "label": "Einkauf",
        "intro": "Der Einkauf organisiert Bedarf, Bestellung, Wareneingang und Eingangsrechnung. Er sorgt dafür, dass benötigte Artikel rechtzeitig, zum richtigen Preis und mit nachvollziehbarer Dokumentation beschafft werden.",
        "flow": "Einkauf kontrolliert abwickeln",
        "steps": [
            "Bedarf klären: Artikel, Menge, Termin und Lieferant prüfen.",
            "Bestellung auslösen: Konditionen, Preise, Lieferadresse und interne Freigaben kontrollieren.",
            "Wareneingang buchen: Gelieferte Mengen, Qualität und offene Restmengen prüfen.",
            "Rechnung abstimmen: Eingangsrechnung mit Bestellung und Wareneingang vergleichen.",
            "Abweichungen dokumentieren: Preis-, Mengen- oder Lieferdifferenzen sauber klären.",
        ],
        "bullets": [
            "Einkauf wirkt direkt auf Lagerbestand, Kosten, Lieferfähigkeit und Liquidität.",
            "Lieferantenpreise, Katalognummern und Lieferzeiten sollten vor der Bestellung plausibel sein.",
            "Bestellung, Wareneingang und Rechnung sollten als zusammenhängender Prozess geprüft werden.",
        ],
        "box": ("Dreifache Prüfung", "Bei kritischen Beschaffungen sollten Bestellung, Wareneingang und Eingangsrechnung zusammenpassen. Genau dort entstehen die meisten Klärfälle."),
    },
    "kommissionieren": {
        "label": "Kommissionieren",
        "intro": "Kommissionieren beschreibt das Zusammenstellen von Waren für Aufträge oder Lieferungen. Das Modul unterstützt Anwender dabei, die richtigen Artikel in der richtigen Menge aus dem richtigen Lagerbereich zu entnehmen.",
        "flow": "Kommissionierung sicher durchführen",
        "steps": [
            "Auftrag prüfen: Lieferpositionen, Lagerverfügbarkeit und Priorität kontrollieren.",
            "Ware entnehmen: Lagerplatz, Menge und Artikel eindeutig bestätigen.",
            "Abweichungen melden: Fehlmengen, Ersatzartikel oder beschädigte Ware sofort erfassen.",
            "Bereitstellung abschließen: Kommissionierte Ware für Packen oder Versand übergeben.",
        ],
        "bullets": [
            "Eine gute Lagerstruktur reduziert Suchaufwand und Fehler beim Entnehmen.",
            "Teilkommissionierungen sollten sichtbar bleiben, damit Restmengen nicht verloren gehen.",
            "Kommissionierung ist die Brücke zwischen Verkaufsauftrag und physischer Ware.",
        ],
        "box": ("Qualität", "Kommissionieren ist dann gut, wenn der Pack- oder Versandbereich ohne Rückfragen weiterarbeiten kann."),
    },
    "packen": {
        "label": "Packen",
        "intro": "Packen fasst kommissionierte Ware in Packstücke, Packlisten und versandfähige Einheiten zusammen. Der Bereich ist wichtig für Vollständigkeit, Gewichte, Verpackung und Übergabe an den Versand.",
        "flow": "Ware versandbereit packen",
        "steps": [
            "Kommissionierte Ware prüfen: Artikel, Menge und Auftrag kontrollieren.",
            "Packstück bilden: Verpackung, Gewicht und Inhalt nachvollziehbar erfassen.",
            "Packliste erzeugen: Inhalt und Empfängerangaben prüfen.",
            "An Versand übergeben: Packstücke vollständig und statusklar bereitstellen.",
        ],
        "bullets": [
            "Packstücke sollten eindeutig einem Auftrag oder Versandvorgang zugeordnet sein.",
            "Gewicht, Maße und Inhalt sind wichtig für Frachtführer und Reklamationen.",
            "Eine saubere Packliste reduziert Rückfragen beim Kunden und im Lager.",
        ],
        "box": ("Abgrenzung", "Kommissionieren stellt Ware zusammen. Packen bildet daraus transportfähige Einheiten."),
    },
    "versand": {
        "label": "Versand",
        "intro": "Der Versand organisiert Frachtführer, Versandaufträge, Sendungsverfolgung und Versandpapiere. Er macht aus fertig gepackter Ware einen nachvollziehbaren Transport zum Empfänger.",
        "flow": "Sendung vorbereiten",
        "steps": [
            "Empfänger prüfen: Lieferadresse, Ansprechpartner und Lieferbedingung kontrollieren.",
            "Frachtführer wählen: Versandart, Servicelevel und Kosten bewerten.",
            "Papiere erzeugen: Label, Versandpapiere und Begleitdokumente prüfen.",
            "Sendung verfolgen: Trackingdaten und Status aktuell halten.",
        ],
        "bullets": [
            "Versanddaten wirken direkt auf Zustellung, Kosten und Kundenerlebnis.",
            "Falsche Adressen oder fehlende Papiere verursachen Verzögerungen.",
            "Sendungsverfolgung hilft Vertrieb, Lager und Kunde bei Rückfragen.",
        ],
        "box": ("Nachvollziehbarkeit", "Ein guter Versandvorgang zeigt, was versendet wurde, mit welchem Dienstleister und in welchem Status sich die Sendung befindet."),
    },
    "produktion": {
        "label": "Produktion",
        "intro": "Die Produktion steuert Stücklisten, Fertigungsaufträge, Materialentnahmen und Rückmeldungen. Sie verbindet Artikelstammdaten, Lager, Ressourcen und Zeiten zu einem kontrollierten Herstellungsprozess.",
        "flow": "Fertigung abbilden",
        "steps": [
            "Produktionsartikel prüfen: Stückliste, Komponenten und Mengeneinheiten kontrollieren.",
            "Fertigungsauftrag planen: Menge, Termin, Ressourcen und Materialbedarf festlegen.",
            "Material entnehmen: Komponenten lagerwirksam und nachvollziehbar ausbuchen.",
            "Fortschritt rückmelden: Mengen, Zeiten, Ausschuss und Abschluss erfassen.",
            "Ergebnis kontrollieren: Bestand, Kosten und Abweichungen prüfen.",
        ],
        "bullets": [
            "Produktion benötigt saubere Artikel-, Lager- und Ressourcenstammdaten.",
            "Stücklistenänderungen können Materialbedarf, Kosten und Termine verändern.",
            "Rückmeldungen sind die Grundlage für Bestand, Nachkalkulation und Transparenz.",
        ],
        "box": ("Set oder Produktion?", "Ein Set ist eine kaufmännische Zusammenstellung. Produktion bildet die Herstellung mit Material, Ressourcen und Rückmeldung ab."),
    },
    "beschaffungsassistent": {
        "label": "Beschaffungsassistent",
        "intro": "Der Beschaffungsassistent hilft, Bedarf zu erkennen und daraus sinnvolle Bestellvorschläge abzuleiten. Er verbindet Lagerbestand, Mindestmengen, offene Vorgänge und Beschaffungsregeln.",
        "flow": "Bedarf in Bestellung überführen",
        "steps": [
            "Bedarf ermitteln: Bestand, Mindestmenge, Reservierungen und offene Aufträge prüfen.",
            "Vorschlag bewerten: Menge, Lieferant, Preis und Lieferzeit kontrollieren.",
            "Beschaffungsart wählen: Einkauf, Produktion oder interne Umlagerung unterscheiden.",
            "Bestellung erzeugen: Vorschläge bewusst übernehmen und offene Punkte klären.",
        ],
        "bullets": [
            "Der Assistent ist nur so gut wie Artikel-, Lager- und Lieferantendaten.",
            "Bestellvorschläge sollten geprüft, nicht blind übernommen werden.",
            "Beschaffungsarten müssen zum tatsächlichen Prozess passen.",
        ],
        "box": ("Entscheidungshilfe", "Der Beschaffungsassistent ersetzt keine fachliche Prüfung. Er macht Bedarf sichtbar und bereitet die Entscheidung vor."),
    },
    "aussendienst": {
        "label": "Außendienst",
        "intro": "Der Außendienst unterstützt Besuchsplanung, Besuchsberichte und mobilen Zugriff auf relevante Kundeninformationen. Ziel ist eine saubere Vorbereitung, Durchführung und Nachverfolgung von Kundenterminen.",
        "flow": "Kundenbesuch vorbereiten",
        "steps": [
            "Besuch planen: Kunde, Ansprechpartner, Termin und Ziel festlegen.",
            "Informationen prüfen: offene Angebote, Aufträge, Reklamationen und CRM-Historie ansehen.",
            "Gespräch dokumentieren: Ergebnis, Zusagen und nächste Schritte erfassen.",
            "Nachverfolgen: Aufgaben, Angebote oder Wiedervorlagen direkt ableiten.",
        ],
        "bullets": [
            "Gute Vorbereitung spart Gesprächszeit und erhöht Verbindlichkeit.",
            "Besuchsberichte sollten zeitnah nach dem Termin gepflegt werden.",
            "Mobiler Zugriff ist besonders wertvoll, wenn Stammdaten und CRM aktuell sind.",
        ],
        "box": ("Vertriebspraxis", "Der Wert des Außendienstmoduls entsteht nicht durch den Termin allein, sondern durch die konsequente Nachverfolgung danach."),
    },
    "geraeteservice": {
        "label": "Geräteservice",
        "intro": "Der Geräteservice verwaltet Geräteakten, Serviceaufträge, Wartungspläne und Garantieinformationen. Er ist für Unternehmen wichtig, die installierte Geräte betreuen, warten oder reparieren.",
        "flow": "Servicefall bearbeiten",
        "steps": [
            "Gerät identifizieren: Seriennummer, Standort, Partner und Historie prüfen.",
            "Serviceauftrag anlegen: Problem, Priorität, Termin und Verantwortliche erfassen.",
            "Wartung oder Reparatur durchführen: Material, Zeiten und Ergebnisse dokumentieren.",
            "Garantie prüfen: Ansprüche, Laufzeit und Abrechnung klären.",
            "Historie fortschreiben: Geräteakte für spätere Servicefälle aktualisieren.",
        ],
        "bullets": [
            "Die Geräteakte ist die zentrale Informationsquelle für Servicehistorie und Zustand.",
            "Wartungspläne reduzieren ungeplante Ausfälle.",
            "Garantie- und Vertragsdaten sollten vor der Abrechnung geprüft werden.",
        ],
        "box": ("Servicequalität", "Ein guter Servicefall ist nachvollziehbar: welches Gerät, welches Problem, welche Maßnahme, welches Ergebnis."),
    },
    "vertragsmanagement": {
        "label": "Vertragsmanagement",
        "intro": "Das Vertragsmanagement verwaltet Verträge, Laufzeiten und wiederkehrende Abrechnungen. Es unterstützt Anwender dabei, Verpflichtungen, Kündigungen, Leistungen und regelmäßige Rechnungen im Blick zu behalten.",
        "flow": "Vertrag kontrolliert führen",
        "steps": [
            "Vertragsdaten erfassen: Partner, Laufzeit, Leistung, Preis und Status pflegen.",
            "Fristen prüfen: Start, Ende, Verlängerung und Kündigungsfristen überwachen.",
            "Abrechnung vorbereiten: wiederkehrende Positionen, Intervalle und Preisänderungen kontrollieren.",
            "Änderungen dokumentieren: Anpassungen, Nachträge und Sondervereinbarungen nachvollziehbar halten.",
        ],
        "bullets": [
            "Verträge wirken oft über lange Zeiträume und müssen deshalb besonders klar gepflegt sein.",
            "Wiederkehrende Abrechnung braucht korrekte Artikel, Preise und Intervalle.",
            "Laufzeiten und Kündigungsfristen sollten aktiv überwacht werden.",
        ],
        "box": ("Risiko", "Unklare Vertragsdaten führen schnell zu verpassten Fristen oder falschen wiederkehrenden Rechnungen."),
    },
    "verknuepfungsplan": {
        "label": "Verknüpfungsplan",
        "intro": "Der Verknüpfungsplan macht Beziehungen zwischen Daten, Belegen oder Prozessen sichtbar. Er hilft Anwendern zu verstehen, wie ein Vorgang mit anderen Informationen zusammenhängt.",
        "flow": "Zusammenhänge lesen",
        "steps": [
            "Ausgangspunkt wählen: Datensatz oder Vorgang öffnen, dessen Beziehungen geprüft werden sollen.",
            "Verbindungen betrachten: Vorgänger, Folgeobjekte und fachliche Bezüge nachvollziehen.",
            "Auffälligkeiten klären: fehlende, doppelte oder unerwartete Verknüpfungen prüfen.",
            "Gezielt springen: Relevante verbundene Objekte direkt öffnen.",
        ],
        "bullets": [
            "Der Plan hilft besonders bei komplexen Vorgängen und Rückfragen.",
            "Verknüpfungen ersetzen keine saubere Belegkette, ergänzen sie aber um Kontext.",
            "Unklare Beziehungen sollten nicht ignoriert, sondern fachlich geprüft werden.",
        ],
        "box": ("Orientierung", "Nutzen Sie den Verknüpfungsplan, wenn Sie nicht nur einen Datensatz sehen wollen, sondern dessen Zusammenhang im System."),
    },
    "genehmigungen": {
        "label": "Genehmigungen",
        "intro": "Genehmigungen steuern Freigaberegeln, Genehmigungsschritte und Status. Sie sorgen dafür, dass kritische Vorgänge geprüft werden, bevor sie wirksam weiterlaufen.",
        "flow": "Freigabeprozess einrichten",
        "steps": [
            "Regel definieren: Auslöser, Betrag, Vorgangsart oder Bedingung klar festlegen.",
            "Schritte planen: Prüfer, Reihenfolge und Eskalation bestimmen.",
            "Status verfolgen: offene, genehmigte und abgelehnte Vorgänge kontrollieren.",
            "Ausnahme klären: Vertretung, Ablehnung und Nacharbeit nachvollziehbar dokumentieren.",
        ],
        "bullets": [
            "Freigaben sollten nur dort eingesetzt werden, wo Kontrolle wirklich nötig ist.",
            "Zu viele Genehmigungsschritte verlangsamen den Prozess.",
            "Status und Verantwortlichkeit müssen für Anwender eindeutig sein.",
        ],
        "box": ("Balance", "Ein guter Genehmigungsprozess schützt vor Risiken, ohne den Alltag unnötig zu blockieren."),
    },
    "crm": {
        "label": "CRM",
        "intro": "CRM bündelt Kontakte, Aktivitäten, Verkaufschancen und Kampagnen. Es macht Kundenbeziehungen nachvollziehbar und hilft Vertrieb, Service und Management bei konsequenter Nachverfolgung.",
        "flow": "Kundenbeziehung steuern",
        "steps": [
            "Kontaktbasis pflegen: Partner, Ansprechpartner und Zuständigkeit aktuell halten.",
            "Aktivitäten dokumentieren: Gespräche, Aufgaben und Termine zeitnah erfassen.",
            "Chancen bewerten: Potenzial, Phase, Wahrscheinlichkeit und nächsten Schritt prüfen.",
            "Kampagnen nutzen: Zielgruppen, Aktionen und Rückmeldungen strukturiert verfolgen.",
        ],
        "bullets": [
            "CRM lebt von aktuellen, kurzen und verwertbaren Einträgen.",
            "Jede Aktivität sollte einen nächsten Schritt oder ein klares Ergebnis haben.",
            "Verkaufschancen werden erst wertvoll, wenn Phase und Wahrscheinlichkeit realistisch gepflegt sind.",
        ],
        "box": ("Zusammenarbeit", "CRM verhindert Informationsinseln, wenn mehrere Personen mit denselben Kunden arbeiten."),
    },
    "intercom": {
        "label": "Intercom",
        "intro": "Intercom unterstützt interne Nachrichten, Erwähnungen und Benachrichtigungen. Es hilft, Informationen direkt im Arbeitskontext zu teilen, ohne sie aus dem ERP-Prozess herauszulösen.",
        "flow": "Kommunikation im Kontext halten",
        "steps": [
            "Nachricht erstellen: Thema, Bezug und Empfänger klar formulieren.",
            "Person erwähnen: Zuständige gezielt einbinden.",
            "Benachrichtigung prüfen: Relevante Hinweise zeitnah bearbeiten.",
            "Ergebnis dokumentieren: Entscheidung oder Aktion im passenden Vorgang festhalten.",
        ],
        "bullets": [
            "Nachrichten sollten kurz und handlungsorientiert sein.",
            "Erwähnungen sind sinnvoll, wenn eine konkrete Reaktion erwartet wird.",
            "Wichtige Entscheidungen gehören zusätzlich in den fachlichen Vorgang.",
        ],
        "box": ("Kontext", "Interne Kommunikation ist am stärksten, wenn sie direkt am richtigen Datensatz oder Vorgang stattfindet."),
    },
    "wiki": {
        "label": "Wiki",
        "intro": "Das Wiki sammelt internes Wissen in Artikeln, Kategorien und Suche. Es unterstützt Anwender dabei, Prozesse, Regeln und Hintergrundwissen zentral zu finden.",
        "flow": "Wissen pflegen",
        "steps": [
            "Artikel erstellen: Thema klar eingrenzen und verständlich beschreiben.",
            "Kategorie zuordnen: Inhalte auffindbar strukturieren.",
            "Suche berücksichtigen: Begriffe verwenden, nach denen Anwender tatsächlich suchen.",
            "Aktualität prüfen: Veraltete Inhalte regelmäßig überarbeiten.",
        ],
        "bullets": [
            "Wiki-Artikel sollten konkrete Fragen beantworten.",
            "Kategorien helfen nur, wenn sie konsequent und nicht zu tief aufgebaut sind.",
            "Eine gute Suche braucht klare Begriffe und verständliche Titel.",
        ],
        "box": ("Wissensqualität", "Ein Wiki ist kein Archiv für alles, sondern ein Werkzeug für verlässliche Antworten im Alltag."),
    },
    "bulletin": {
        "label": "Bulletin",
        "intro": "Bulletins veröffentlichen interne Aushänge, Hinweise und Informationen mit Gültigkeit und Kenntnisnahme. Sie eignen sich für verbindliche Mitteilungen an definierte Empfängergruppen.",
        "flow": "Aushang veröffentlichen",
        "steps": [
            "Inhalt formulieren: Kurz, verbindlich und zielgruppengerecht schreiben.",
            "Gültigkeit setzen: Zeitraum und Sichtbarkeit festlegen.",
            "Empfänger wählen: Zielgruppe oder Bereich sauber zuordnen.",
            "Kenntnisnahme prüfen: Rückmeldungen und offene Bestätigungen verfolgen.",
        ],
        "bullets": [
            "Bulletins sollten nicht für dauerhafte Prozessdokumentation missbraucht werden.",
            "Gültigkeit verhindert, dass veraltete Hinweise sichtbar bleiben.",
            "Kenntnisnahme ist wichtig, wenn Informationen verbindlich bestätigt werden müssen.",
        ],
        "box": ("Kommunikation", "Ein Bulletin ist ideal für wichtige interne Hinweise, die sichtbar und zeitlich gesteuert verteilt werden sollen."),
    },
    "zeiterfassung": {
        "label": "Zeiterfassung",
        "intro": "Die Zeiterfassung dokumentiert Arbeitszeiten, Projektzeiten, Zeitkonten und Abwesenheiten. Sie unterstützt Abrechnung, Planung, Projektcontrolling und Transparenz.",
        "flow": "Zeiten sauber erfassen",
        "steps": [
            "Zeit buchen: Beginn, Ende, Dauer und Tätigkeit korrekt erfassen.",
            "Bezug zuordnen: Projekt, Aufgabe, Kunde oder interne Kategorie wählen.",
            "Zeitkonto prüfen: Salden, Korrekturen und Abweichungen beobachten.",
            "Abwesenheit pflegen: Urlaub, Krankheit oder andere Abwesenheiten rechtzeitig eintragen.",
        ],
        "bullets": [
            "Zeitdaten sind nur belastbar, wenn sie zeitnah gepflegt werden.",
            "Projektzeiten sollten nachvollziehbar beschrieben sein.",
            "Abwesenheiten beeinflussen Planung, Ressourcen und Termine.",
        ],
        "box": ("Nachvollziehbarkeit", "Gute Zeiterfassung erklärt nicht nur wie lange gearbeitet wurde, sondern woran und in welchem Kontext."),
    },
    "archivierung": {
        "label": "Archivierung",
        "intro": "Archivierung sichert Dokumente, Aufbewahrungsfristen und revisionsrelevante Nachweise. Sie hilft, geschäftliche Unterlagen langfristig auffindbar und nachvollziehbar zu halten.",
        "flow": "Dokumente sicher archivieren",
        "steps": [
            "Dokument einordnen: Belegart, Partner, Datum und Vorgang prüfen.",
            "Archivieren: Dokument mit den richtigen Metadaten ablegen.",
            "Aufbewahrung beachten: Fristen und rechtliche Anforderungen berücksichtigen.",
            "Wiederfinden testen: Suche, Filter und Belegbezug prüfen.",
        ],
        "bullets": [
            "Archivierung dient nicht nur Ablage, sondern späterer Nachweisbarkeit.",
            "Revisionssicherheit hängt von Prozess, Berechtigung und Unveränderbarkeit ab.",
            "Gute Metadaten sind entscheidend für die spätere Suche.",
        ],
        "box": ("Sorgfalt", "Ein Dokument ist erst dann gut archiviert, wenn es später sicher gefunden und fachlich eingeordnet werden kann."),
    },
    "elektronische-haftnotizen": {
        "label": "Elektronische Haftnotizen",
        "intro": "Elektronische Haftnotizen speichern kurze Hinweise direkt im Arbeitskontext. Sie eignen sich für temporäre Informationen, Erinnerungen oder Klärpunkte an Datensätzen und Vorgängen.",
        "flow": "Notiz sinnvoll einsetzen",
        "steps": [
            "Hinweis formulieren: Kurz und eindeutig beschreiben, worauf zu achten ist.",
            "Richtig verknüpfen: Notiz am passenden Datensatz oder Vorgang speichern.",
            "Erledigung prüfen: Temporäre Hinweise nach Abschluss schließen oder entfernen.",
        ],
        "bullets": [
            "Haftnotizen sind für kurze Hinweise gedacht, nicht für dauerhafte Dokumentation.",
            "Verknüpfungen sorgen dafür, dass Hinweise im richtigen Kontext erscheinen.",
            "Erledigte Notizen sollten nicht dauerhaft den Arbeitsbereich überladen.",
        ],
        "box": ("Kurzfristiger Hinweis", "Nutzen Sie Haftnotizen für Dinge, die Anwender sofort sehen sollen, aber nicht als Stammdaten oder Dokument gelten."),
    },
    "e-mail": {
        "label": "E-Mail",
        "intro": "Das E-Mail-Modul verbindet Posteingang, Belegversand, Vorlagen und Zuordnung mit ERP-Vorgängen. So bleibt Kommunikation nachvollziehbar beim passenden Partner, Beleg oder Prozess.",
        "flow": "E-Mail im ERP-Kontext nutzen",
        "steps": [
            "Nachricht prüfen: Absender, Empfänger, Betreff und Bezug erkennen.",
            "Zuordnen: E-Mail dem richtigen Partner, Beleg oder Vorgang zuweisen.",
            "Vorlage nutzen: Standardtexte und Anhänge kontrolliert verwenden.",
            "Beleg versenden: Empfänger, Dokument und Versandstatus prüfen.",
        ],
        "bullets": [
            "E-Mails sollten nicht nur im persönlichen Postfach bleiben, wenn sie fachlich relevant sind.",
            "Vorlagen sparen Zeit und sorgen für einheitliche Kommunikation.",
            "Belegversand braucht korrekte Empfänger- und Dokumentdaten.",
        ],
        "box": ("Nachvollziehbarkeit", "Eine zugeordnete E-Mail kann später im Partner- oder Belegkontext wiedergefunden werden."),
    },
    "kalender": {
        "label": "Kalender",
        "intro": "Der Kalender organisiert Termine, Aufgaben, Wiedervorlagen und Freigaben. Er unterstützt persönliche Planung und gemeinsame Abstimmung im ERP-Kontext.",
        "flow": "Arbeit planen",
        "steps": [
            "Termin oder Aufgabe erfassen: Inhalt, Zeit, Verantwortliche und Bezug festlegen.",
            "Wiedervorlage setzen: Offene Themen rechtzeitig erneut sichtbar machen.",
            "Teilen oder freigeben: Beteiligte Personen passend einbinden.",
            "Status pflegen: Erledigung, Verschiebung oder Ergebnis nachführen.",
        ],
        "bullets": [
            "Kalendereinträge sind besonders nützlich, wenn sie mit Partnern, Projekten oder Vorgängen verbunden sind.",
            "Wiedervorlagen verhindern, dass offene Punkte im Alltag untergehen.",
            "Freigaben und Teilen sollten bewusst eingesetzt werden, damit Zuständigkeiten klar bleiben.",
        ],
        "box": ("Planung", "Der Kalender hilft nicht nur bei Terminen, sondern auch beim kontrollierten Wiederaufgreifen offener Aufgaben."),
    },
    "ki-assistent": {
        "label": "KI-Assistent",
        "intro": "Der KI-Assistent unterstützt Anwender bei Texten, Datenabfragen und Vorschlägen. Er soll Arbeit beschleunigen, ohne fachliche Prüfung und Verantwortung zu ersetzen.",
        "flow": "KI sinnvoll nutzen",
        "steps": [
            "Ziel formulieren: Möglichst konkret beschreiben, welches Ergebnis benötigt wird.",
            "Datenbezug prüfen: Nur passende und freigegebene Informationen verwenden.",
            "Vorschlag bewerten: Ergebnis fachlich lesen, korrigieren und freigeben.",
            "Übernehmen: Nur geprüfte Texte, Antworten oder Vorschläge weiterverwenden.",
        ],
        "bullets": [
            "KI-Ergebnisse müssen fachlich geprüft werden.",
            "Sensible oder unklare Daten sollten nicht unbedacht verarbeitet werden.",
            "Gute Prompts und klare Kontextinformationen verbessern die Qualität.",
        ],
        "box": ("Verantwortung", "Der KI-Assistent liefert Unterstützung. Die fachliche Entscheidung bleibt beim Anwender."),
    },
    "finanzwesen": {
        "label": "Finanzwesen",
        "intro": "Das Finanzwesen bündelt offene Posten, Zahlungsverkehr, Mahnwesen, Kostenrechnung, Anlagenbuchhaltung, Kontenfindung und Bestandswertbuchungen. Es verbindet operative Vorgänge mit buchhalterischer Kontrolle.",
        "flow": "Finanzprozesse kontrollieren",
        "steps": [
            "Vorgänge prüfen: Rechnungen, Zahlungen, offene Posten und Buchungsstatus kontrollieren.",
            "Kontierung sichern: Kontenfindung, Kostenstellen und Finanzgruppen plausibel halten.",
            "Zahlungen steuern: Fälligkeiten, Zahlungsverkehr und Mahnläufe überwachen.",
            "Werte abstimmen: Bestandswerte, Anlagen und Kostenrechnung regelmäßig prüfen.",
        ],
        "bullets": [
            "Finanzdaten entstehen aus Stammdaten, Belegen und Buchungsregeln.",
            "Fehler in Kontenfindung oder Finanzgruppen wirken später in Auswertungen und Buchhaltung.",
            "Offene Posten, Mahnungen und Zahlungen sollten im Zusammenhang betrachtet werden.",
        ],
        "box": ("Kontrolle", "Das Finanzwesen macht sichtbar, ob operative Vorgänge kaufmännisch korrekt abgeschlossen sind."),
    },
}


TOPICS = {
    "angebot": "Ein Angebot beschreibt eine verbindliche oder vorbereitende Verkaufssituation. Es sollte den Bedarf des Kunden klar abbilden und als belastbare Grundlage für den Auftrag dienen.",
    "auftrag": "Ein Auftrag ist der zentrale Verkaufsbeleg für bestätigte Kundenbestellungen. Er steuert Lieferfähigkeit, Termine, Folgebelege und interne Bearbeitung.",
    "lieferschein": "Der Lieferschein dokumentiert, welche Ware an den Kunden geliefert wird. Er verbindet Auftrag, Lagerbewegung und später häufig die Rechnung.",
    "rechnung": "Die Rechnung schließt den Verkauf kaufmännisch ab. Sie muss Adresse, Positionen, Preise, Steuer und Zahlungsbedingungen korrekt enthalten.",
    "gutschrift": "Eine Gutschrift korrigiert oder erstattet einen vorherigen Vorgang. Ursache, Bezug und Betrag sollten nachvollziehbar dokumentiert sein.",
    "bestellanforderung": "Eine Bestellanforderung sammelt internen Bedarf, bevor daraus eine Bestellung entsteht. Sie hilft, Beschaffung kontrolliert auszulösen.",
    "bestellung": "Die Bestellung ist der verbindliche Einkaufsbeleg an den Lieferanten. Sie sollte Artikel, Menge, Preis, Termin und Lieferbedingungen eindeutig enthalten.",
    "wareneingang": "Der Wareneingang bestätigt gelieferte Ware. Er beeinflusst Bestand, Restmengen, Qualitätssicherung und die spätere Rechnungsprüfung.",
    "eingangsrechnung": "Die Eingangsrechnung wird mit Bestellung und Wareneingang abgestimmt. Abweichungen bei Preis, Menge oder Steuer müssen vor der Freigabe geklärt werden.",
    "kommissionierung": "Die Kommissionierung stellt Ware für einen Auftrag zusammen. Entscheidend sind Artikel, Menge, Lagerplatz und vollständige Übergabe an Packen oder Versand.",
    "packstueck": "Ein Packstück beschreibt eine konkrete Versand- oder Verpackungseinheit. Inhalt, Gewicht, Maße und Bezug müssen nachvollziehbar sein.",
    "packliste": "Die Packliste zeigt, welche Artikel in welchen Packstücken enthalten sind. Sie unterstützt Kontrolle, Versand und spätere Rückfragen.",
    "versandauftrag": "Der Versandauftrag bündelt die Informationen für den Transport. Er verbindet Empfänger, Packstücke, Frachtführer, Papiere und Sendungsstatus.",
    "frachtfuehrer": "Der Frachtführer ist der Dienstleister für den Transport. Auswahl, Servicelevel, Kosten und Tracking müssen zum Versandfall passen.",
    "sendungsverfolgung": "Die Sendungsverfolgung zeigt den Transportstatus. Sie hilft bei Lieferauskunft, Reklamation und interner Nachverfolgung.",
    "versandpapiere": "Versandpapiere begleiten die Sendung. Sie müssen Empfänger, Inhalt, Frachtführer und gesetzliche oder organisatorische Angaben korrekt enthalten.",
    "stueckliste": "Die Stückliste beschreibt, aus welchen Komponenten ein Produkt hergestellt wird. Mengen, Einheiten und Gültigkeit beeinflussen Materialbedarf und Kosten.",
    "fertigungsauftrag": "Der Fertigungsauftrag steuert die Herstellung einer konkreten Menge. Er verbindet Material, Ressourcen, Termine und Rückmeldungen.",
    "materialentnahme": "Die Materialentnahme bucht Komponenten aus dem Lager in die Fertigung. Sie muss zur Stückliste und zum tatsächlichen Verbrauch passen.",
    "rueckmeldung": "Die Rückmeldung dokumentiert Fortschritt, gefertigte Menge, Zeiten und Abweichungen. Sie ist wichtig für Bestand und Nachkalkulation.",
    "bedarfsermittlung": "Die Bedarfsermittlung zeigt, welche Artikel beschafft oder produziert werden müssen. Bestand, Reservierungen und offene Vorgänge sind dafür entscheidend.",
    "bestellvorschlag": "Der Bestellvorschlag bereitet konkrete Beschaffungsmaßnahmen vor. Menge, Lieferant, Preis und Termin sollten vor Übernahme geprüft werden.",
    "beschaffungsart": "Die Beschaffungsart entscheidet, ob Bedarf eingekauft, produziert oder anders gedeckt wird. Sie muss zum Artikel und Prozess passen.",
    "besuchsplanung": "Die Besuchsplanung bereitet Kundentermine strukturiert vor. Ziel, Ansprechpartner, offene Vorgänge und Nachbereitung sollten bereits vor dem Termin klar sein.",
    "besuchsbericht": "Der Besuchsbericht dokumentiert Gesprächsergebnis, Zusagen und nächste Schritte. Er macht Außendienstwissen für das Team nutzbar.",
    "mobiler-zugriff": "Mobiler Zugriff stellt relevante Informationen unterwegs bereit. Entscheidend sind aktuelle Daten, passende Rechte und einfache Bedienung.",
    "geraeteakte": "Die Geräteakte sammelt Stammdaten, Standort, Historie und Serviceinformationen zu einem Gerät. Sie ist der Einstieg in jeden Servicefall.",
    "serviceauftrag": "Der Serviceauftrag beschreibt eine Wartung, Reparatur oder Störung. Problem, Termin, Techniker, Material und Ergebnis sollten klar dokumentiert sein.",
    "wartungsplan": "Der Wartungsplan legt regelmäßige Serviceeinsätze fest. Intervalle, Geräte, Verantwortliche und Fälligkeiten müssen realistisch gepflegt werden.",
    "garantie": "Garantieinformationen klären Ansprüche, Laufzeiten und Abrechnung. Vor Service oder Berechnung sollte geprüft werden, ob Garantie greift.",
    "vertrag": "Der Vertrag beschreibt Leistung, Partner, Preis, Laufzeit und Status. Er ist die Grundlage für Verpflichtungen und wiederkehrende Abrechnung.",
    "laufzeit": "Die Laufzeit steuert Beginn, Ende, Verlängerung und Kündigungsfristen. Sie muss aktiv überwacht werden.",
    "wiederkehrende-abrechnung": "Wiederkehrende Abrechnung erzeugt regelmäßig abrechenbare Vorgänge. Artikel, Preis, Intervall und Gültigkeit müssen exakt stimmen.",
    "freigaberegel": "Die Freigaberegel definiert, wann ein Vorgang genehmigt werden muss. Auslöser und Schwellenwerte sollten klar und prüfbar sein.",
    "genehmigungsschritt": "Ein Genehmigungsschritt beschreibt, wer in welcher Reihenfolge prüft. Vertretung, Eskalation und Ergebnis müssen nachvollziehbar bleiben.",
    "genehmigungsstatus": "Der Genehmigungsstatus zeigt, ob ein Vorgang offen, genehmigt, abgelehnt oder in Nacharbeit ist. Er ist wichtig für weitere Bearbeitung.",
    "kontakt": "Ein Kontakt verbindet Person, Partner und Kommunikation. Rollen, Zuständigkeiten und Kontaktdaten sollten aktuell sein.",
    "aktivitaet": "Eine Aktivität dokumentiert Gespräch, Aufgabe, Termin oder Nachverfolgung. Sie sollte Ergebnis und nächsten Schritt enthalten.",
    "verkaufschance": "Eine Verkaufschance bewertet mögliches Geschäft. Phase, Potenzial, Wahrscheinlichkeit und nächster Schritt machen sie steuerbar.",
    "kampagne": "Eine Kampagne bündelt Vertriebs- oder Marketingaktionen. Zielgruppe, Inhalt, Zeitraum und Ergebnis sollten nachvollziehbar sein.",
    "nachricht": "Eine Nachricht teilt Informationen im Arbeitskontext. Sie sollte kurz, eindeutig und mit klarem Bezug formuliert sein.",
    "erwaehnung": "Eine Erwähnung bindet eine Person gezielt ein. Sie sollte verwendet werden, wenn eine konkrete Reaktion erwartet wird.",
    "benachrichtigung": "Benachrichtigungen machen relevante Ereignisse sichtbar. Anwender sollten offene Hinweise zeitnah prüfen und abschließen.",
    "artikel": "Ein Wiki-Artikel erklärt ein Thema, eine Regel oder einen Ablauf. Er sollte eine konkrete Frage verständlich beantworten.",
    "kategorie": "Eine Kategorie ordnet Inhalte für Suche und Navigation. Sie hilft nur, wenn sie nachvollziehbar und konsequent verwendet wird.",
    "suche": "Die Suche hilft Anwendern, Wissen und Inhalte schnell zu finden. Gute Titel und Begriffe verbessern die Trefferqualität.",
    "aushang": "Ein Aushang veröffentlicht eine interne Information. Inhalt, Zielgruppe und Sichtbarkeit sollten klar festgelegt sein.",
    "gueltigkeit": "Gültigkeit steuert, wann ein Aushang oder Hinweis sichtbar ist. Sie verhindert, dass veraltete Informationen weiter angezeigt werden.",
    "kenntnisnahme": "Kenntnisnahme zeigt, wer eine Information bestätigt hat. Sie ist wichtig bei verbindlichen Mitteilungen.",
    "zeitbuchung": "Eine Zeitbuchung erfasst Arbeitszeit oder Tätigkeit. Dauer, Bezug und Beschreibung sollten nachvollziehbar sein.",
    "projektzeit": "Projektzeit ordnet Aufwand einem Projekt oder einer Aufgabe zu. Sie ist Grundlage für Nachkalkulation und Abrechnung.",
    "zeitkonto": "Das Zeitkonto zeigt Salden, Buchungen und Abweichungen. Es unterstützt Kontrolle und Korrektur von Arbeitszeiten.",
    "abwesenheit": "Abwesenheit dokumentiert Urlaub, Krankheit oder andere Ausfälle. Sie beeinflusst Planung, Ressourcen und Termine.",
    "dokumentenarchiv": "Das Dokumentenarchiv legt Unterlagen mit Bezug und Metadaten ab. Es soll Dokumente langfristig auffindbar machen.",
    "revisionssicherheit": "Revisionssicherheit beschreibt Anforderungen an unveränderbare, nachvollziehbare und prüfbare Archivierung.",
    "aufbewahrungsfrist": "Die Aufbewahrungsfrist legt fest, wie lange Dokumente gespeichert werden müssen. Fristen sollten fachlich und rechtlich abgestimmt sein.",
    "notiz": "Eine Notiz speichert einen kurzen Hinweis im Kontext. Sie sollte knapp sein und bei Erledigung geschlossen werden.",
    "verknuepfung": "Eine Verknüpfung verbindet die Notiz mit dem passenden Datensatz oder Vorgang. Dadurch erscheint sie an der richtigen Stelle.",
    "erledigung": "Erledigung markiert, dass ein Hinweis bearbeitet wurde. So bleiben nur relevante Notizen sichtbar.",
    "posteingang": "Der Posteingang zeigt eingehende E-Mails. Wichtige Nachrichten sollten zu Partnern, Belegen oder Vorgängen zugeordnet werden.",
    "belegversand": "Belegversand verschickt Angebote, Aufträge, Rechnungen oder andere Dokumente per E-Mail. Empfänger und Anhang müssen geprüft werden.",
    "vorlage": "Eine Vorlage stellt wiederverwendbare E-Mail-Texte bereit. Sie sorgt für einheitliche Sprache und weniger manuelle Arbeit.",
    "zuordnung": "Zuordnung verbindet E-Mails mit Partnern, Belegen oder Vorgängen. Sie macht Kommunikation später wieder auffindbar.",
    "termin": "Ein Termin plant einen Zeitpunkt mit Beteiligten und Bezug. Gute Termine enthalten Zweck, Ort oder Link und Verantwortlichkeit.",
    "aufgabe": "Eine Aufgabe beschreibt eine konkrete zu erledigende Arbeit. Zuständigkeit, Fälligkeit und Status müssen klar sein.",
    "wiedervorlage": "Eine Wiedervorlage bringt ein Thema zu einem späteren Zeitpunkt wieder in den Fokus. Sie verhindert vergessene Nachverfolgung.",
    "freigabe-und-teilen": "Freigabe und Teilen steuern, wer Kalenderinformationen sehen oder bearbeiten darf. Rechte sollten bewusst vergeben werden.",
    "textunterstuetzung": "Textunterstützung hilft beim Formulieren, Kürzen oder Strukturieren von Texten. Ergebnisse müssen fachlich geprüft werden.",
    "datenabfrage": "Datenabfragen helfen, Informationen aus dem System verständlich auszuwerten. Fragestellung, Datenbasis und Ergebnis müssen kontrolliert werden.",
    "vorschlaege": "Vorschläge unterstützen Anwender mit möglichen nächsten Schritten oder Inhalten. Sie sollten als Hilfe, nicht als automatische Entscheidung verstanden werden.",
    "op-verwaltung": "Die OP Verwaltung zeigt offene Posten. Sie unterstützt Zahlungsstatus, Klärfälle und Liquiditätskontrolle.",
    "zahlungsverkehr": "Zahlungsverkehr steuert Überweisungen, Lastschriften und Zahlungsdateien. Bankdaten, Freigaben und Fälligkeiten müssen stimmen.",
    "mahnwesen": "Mahnwesen verfolgt überfällige Forderungen. Mahnstufen, Sperren und Kundenkommunikation sollten sorgfältig geprüft werden.",
    "kostenrechnung": "Kostenrechnung ordnet Kosten und Erlöse nach Kostenstellen, Kostenträgern oder Projekten. Sie macht Wirtschaftlichkeit sichtbar.",
    "anlagenbuchhaltung": "Anlagenbuchhaltung verwaltet Anlagegüter, Abschreibung und Bestandswerte. Anschaffung, Nutzungsdauer und Buchung müssen zusammenpassen.",
    "kontenfindung": "Kontenfindung entscheidet, welche Buchungskonten automatisch verwendet werden. Stammdaten und Regeln müssen fachlich korrekt sein.",
    "kontinuierliche-bestandswertbuchungen": "Kontinuierliche Bestandswertbuchungen bilden Lagerwertänderungen laufend im Finanzwesen ab. Artikel, Lager, Bewertung und Konten müssen konsistent sein.",
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def title_from_path(path: str) -> str:
    slug = path.strip("/").split("/")[-2 if path.endswith("/index.html") else -1]
    special = {
        "crm": "CRM",
        "ki": "KI",
        "e-mail": "E-Mail",
        "op": "OP",
    }
    return " ".join(special.get(part, part[:1].upper() + part[1:]) for part in slug.split("-"))


def markdown(title: str, intro: str, flow_title: str | None, steps: list[str], bullets: list[str], box: tuple[str, str]) -> str:
    parts = [f"# {title}", "", intro.strip()]
    if flow_title and steps:
        parts.extend(["", f":::flow {flow_title}"])
        parts.extend(f"- {step}" for step in steps)
        parts.append(":::")
    if bullets:
        parts.extend(["", "## Worauf Sie achten sollten"])
        parts.extend(f"- {bullet}" for bullet in bullets)
    parts.extend(["", f":::box {box[0]}", box[1].strip(), ":::"])
    return "\n".join(parts).strip() + "\n"


def module_page(module_key: str, title: str) -> str:
    data = MODULES[module_key]
    return markdown(
        title=title,
        intro=data["intro"],
        flow_title=data["flow"],
        steps=data["steps"],
        bullets=data["bullets"],
        box=data["box"],
    )


def child_page(module_key: str, slug: str, title: str, content_type: str) -> str:
    module = MODULES[module_key]
    topic = TOPICS.get(slug, f"{title} ist ein Arbeitsbereich im Modul {module['label']}. Er unterstützt Anwender dabei, diesen Prozessschritt nachvollziehbar zu bearbeiten.")
    intro = f"{topic} Im Modul {module['label']} ist dieser Bereich besonders dann wichtig, wenn Daten nicht nur erfasst, sondern im Folgeprozess sicher weiterverwendet werden sollen."
    steps = [
        f"Kontext prüfen: Öffnen Sie {title} aus dem passenden Vorgang oder Modulbereich.",
        "Daten kontrollieren: Partner, Artikel, Menge, Status, Termin oder Zuständigkeit prüfen, soweit sie für diesen Bereich relevant sind.",
        "Abweichungen klären: Fehlende, falsche oder unplausible Angaben vor dem nächsten Prozessschritt bereinigen.",
        "Folgeprozess nutzen: Wenn möglich über die vorgesehenen Funktionen weiterarbeiten, damit Bezüge erhalten bleiben.",
    ]
    bullets = [
        f"{title} sollte immer im Zusammenhang mit dem übergeordneten Modul {module['label']} betrachtet werden.",
        "Achten Sie auf klare Zuständigkeit, nachvollziehbaren Status und vollständige Pflichtangaben.",
        "Prüfen Sie bei wichtigen Vorgängen die Belegkette, Historie oder Verknüpfungen, bevor Sie Änderungen übernehmen.",
    ]
    return markdown(
        title=title,
        intro=intro,
        flow_title=f"{title} sicher nutzen",
        steps=steps,
        bullets=bullets,
        box=("Praxisregel", f"Nutzen Sie {title} nicht isoliert. Entscheidend ist, ob die Informationen danach im Prozess {module['label']} korrekt weiterlaufen."),
    )


def root_page() -> str:
    return markdown(
        title="Module",
        intro="Die Module in X-ERP bilden die operativen Arbeitsbereiche ab: Verkauf, Einkauf, Lager- und Versandprozesse, Produktion, CRM, Service, Zeit, Archivierung, KI-Unterstützung und Finanzwesen. Diese Hilfe erklärt nicht nur einzelne Masken, sondern vor allem die fachlichen Abläufe, damit Anwender sicher vom ersten Vorgang bis zum Abschluss arbeiten können.",
        flow_title="Modulhilfe nutzen",
        steps=[
            "Prozess wählen: Starten Sie beim Modul, in dem die fachliche Arbeit beginnt.",
            "Ablauf verstehen: Lesen Sie zuerst die Modulübersicht, bevor Sie in Detailseiten wechseln.",
            "Folgebelege beachten: Viele Module arbeiten über Belegketten, Status und Übergaben zusammen.",
            "Stammdaten prüfen: Artikel, Partner, Preise, Lager, Mitarbeiter und Konten müssen zu den Modulprozessen passen.",
            "Ergebnis kontrollieren: Prüfen Sie am Ende Status, Belegkette, Auswertung oder Finanzwirkung.",
        ],
        bullets=[
            "Module sind Arbeitsprozesse, nicht nur Menüpunkte.",
            "Viele Fehler entstehen, wenn ein Modul isoliert betrachtet wird und Folgeprozesse ignoriert werden.",
            "Die Detailseiten erklären jeweils den Zweck, typische Prüfpunkte und den sicheren Umgang im Alltag.",
            "Für Google und Anwender sind klare Prozessbegriffe wichtiger als technische Feldlisten.",
        ],
        box=("Einordnung", "Starten Sie bei der Modulübersicht, wenn Sie den Ablauf verstehen wollen. Nutzen Sie die Unterseiten, wenn Sie einen konkreten Prozessschritt bearbeiten."),
    )


def meta(markdown_text: str, title: str) -> str:
    clean = re.sub(r"^#{1,6}\s*", "", markdown_text, flags=re.M)
    clean = re.sub(r"^:::(?:flow|box)\s+.*$", "", clean, flags=re.M)
    clean = clean.replace(":::", "")
    clean = re.sub(r"^[\s>*-]*[-*]\s*", "", clean, flags=re.M)
    clean = re.sub(r"\s+", " ", clean).strip()
    if not clean:
        clean = f"{title} in der X-ERP Hilfe."
    return clean[:297].rstrip(" ,.;") + ("..." if len(clean) > 297 else "")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-module-quality-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False, keep_links=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }

    changed = 0
    unknown_modules: set[str] = set()
    for row in range(2, ws.max_row + 1):
        path = cell_text(ws.cell(row, headers["URL_PATH"]).value).replace("\\", "/")
        if not path.startswith("Module/"):
            continue
        content_type = cell_text(ws.cell(row, headers["CONTENT_TYPE"]).value)
        if content_type not in {"HelpPage", "FAQ"}:
            continue

        parts = path.split("/")
        original_title = cell_text(ws.cell(row, headers["H1"]).value) or cell_text(ws.cell(row, headers["Thema"]).value) or title_from_path(path)
        if path == "Module/index.html":
            title = "Module"
            body = root_page()
        elif len(parts) >= 3 and parts[2] == "index.html":
            module_key = parts[1]
            if module_key not in MODULES:
                unknown_modules.add(module_key)
                continue
            title = MODULES[module_key]["label"]
            body = module_page(module_key, title)
        else:
            module_key = parts[1]
            if module_key not in MODULES:
                unknown_modules.add(module_key)
                continue
            slug = parts[2]
            title = original_title
            body = child_page(module_key, slug, title, content_type)

        ws.cell(row, headers["Beschreibung"]).value = body
        for column, value in {
            "H1": title,
            "NAV_TITLE": title,
            "TITLE": f"{title} | X-ERP ERP Hilfe",
            "META_DESCRIPTION": meta(body, title),
            "PRIMARY_KEYWORD": f"{title} X-ERP",
        }.items():
            if column in headers:
                ws.cell(row, headers[column]).value = value
        changed += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"changed={changed}")
    print(f"unknown_modules={len(unknown_modules)}")
    for module_key in sorted(unknown_modules):
        print(f"UNKNOWN {module_key}")


if __name__ == "__main__":
    main()
