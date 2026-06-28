from __future__ import annotations

import copy
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")


SPECS = [
    {
        "parent": "ArticleEdit-Verkauf",
        "list": "ArticleSalesPriceListList",
        "purpose": "zeigt die für diesen Artikel gepflegten Verkaufspreise je Preisliste und Gültigkeitsdatum.",
        "edit": "ArticleSalesPriceListEdit",
        "columns": [
            ("Artikel", "Article", "ArticleId"),
            ("Preisliste", "Price List", "PriceListId"),
            ("Gültig ab", "Valid From", "ValidFrom"),
            ("Verkaufspreis", "Sales Price", "SalesPrice"),
        ],
    },
    {
        "parent": "ArticleEdit-Beschaffung",
        "list": "PartnerPurchasePriceListList",
        "purpose": "zeigt lieferantenbezogene Einkaufspreise, Katalognummern, Lieferzeiten und Beschaffungskosten.",
        "edit": "PartnerPurchasePriceListEdit",
        "columns": [
            ("Lieferant", "Supplier", "PartnerMatchcode"),
            ("Artikel", "Article", "ArticleMatchcode"),
            ("Katalognummer", "Catalog Number", "CatalogNumber"),
            ("Lieferzeit", "Delivery Days", "DeliveryTime"),
            ("Mindestmenge", "Minimum", "MinimumQuantity"),
            ("Losgröße", "Lot Size", "LotSize"),
            ("Einkaufspreis", "Purchase Price", "PurchasePrice"),
            ("Rabatt %", "Discount %", "DiscountPercent"),
            ("Preiseinheit", "Price Unit", "PriceUnitId"),
            ("Logistik", "Logistic", "LogisticCost"),
            ("Zoll", "Customs", "CustomsCost"),
            ("Risiko", "Risk", "RiskCost"),
            ("Gemeinkosten", "Overhead", "OverheadCost"),
            ("Einstandskosten", "Landed Cost", "LandedCost"),
        ],
    },
    {
        "parent": "ArticleEdit-Lagerung",
        "list": "WarehouseQuantityList",
        "purpose": "zeigt Lager, Lagerplätze, Bestände und verfügbare Mengen dieses Artikels.",
        "edit": "WarehouseQuantityEdit",
        "columns": [
            ("Lager", "Warehouse", "WarehouseName"),
            ("Lagerplatz", "Bin Location", "WarehouseBinLocationName"),
            ("Artikel", "Article", "ArticleId"),
            ("Bestand", "Stock", "Stock"),
            ("Gepackt", "Packed", "Packed"),
            ("Jetzt verfügbar", "Now Available", "NowAvailable"),
            ("Mindestbestand", "Minimum Stock", "MinimumStock"),
            ("Meldeschwelle", "Message Threshold", "MessageThreshold"),
            ("Maximalbestand", "Maximum Stock", "MaximumStock"),
        ],
    },
    {
        "parent": "ArticleEdit-Makro",
        "list": "ArticleMacroList",
        "purpose": "zeigt Makro-Komponenten, die beim Artikel automatisch oder strukturiert verwendet werden.",
        "edit": "ArticleMacroEdit",
        "columns": [
            ("Position", "Position", "Position"),
            ("Makro", "Macro", "MacroMatchcode"),
            ("Menge", "Quantity", "QuantityFactor"),
            ("Komponente drucken", "Print Component", "PrintComponent"),
        ],
    },
    {
        "parent": "ArticleEdit-Produktion",
        "list": "ArticleProductList",
        "purpose": "zeigt die Komponenten und Kalkulationswerte eines Produktionsartikels.",
        "edit": "ArticleProductEdit",
        "columns": [
            ("Position", "Position", "Position"),
            ("Komponente", "Component", "ComponentMatchcode"),
            ("Mengenfaktor", "Quantity Factor", "QuantityFactor"),
            ("Mengeneinheit", "UoM", "UoMId"),
            ("Preiseinheit", "Price Unit", "PriceUnitId"),
            ("Einstandspreis", "Cost Price", "CostPrice"),
            ("Kalkulierter Verkaufspreis", "Calculated Sales Price", "CalculatedSalesPrice"),
            ("Einstandspreis gesamt", "Total Cost Price", "TotalCostPrice"),
            ("Kalkulierter Verkaufspreis gesamt", "Total Calculated Sales Price", "TotalCalculatedSalesPrice"),
            ("Rohertrag", "Profit", "Profit"),
        ],
    },
    {
        "parent": "ArticleEdit-Zubehör",
        "list": "ArticleAccessoryList",
        "purpose": "zeigt Zubehörartikel, die zusammen mit diesem Artikel angeboten oder verwendet werden können.",
        "edit": "ArticleAccessoryEdit",
        "columns": [
            ("Position", "Position", "Position"),
            ("Zubehör", "Accessory", "ComponentMatchcode"),
            ("Mengenfaktor", "Quantity Factor", "QuantityFactor"),
            ("Kostenlos", "For Free", "ForFree"),
            ("Kundenseite", "Customer Side", "CustomerSide"),
            ("Lieferantenseite", "Supplier Side", "SupplierSide"),
            ("Einfügen", "Inserting", "Inserting"),
        ],
    },
    {
        "parent": "ArticleEdit-Kategorien",
        "list": "ArticleCategoryList",
        "purpose": "zeigt die Kategorien, denen dieser Artikel zugeordnet ist.",
        "edit": "ArticleCategoryEdit",
        "columns": [
            ("Kategorie", "Category", "Name"),
            ("Artikel", "Article", "ArticleMatchcode"),
        ],
    },
    {
        "parent": "ArticleEdit-Katalognummern",
        "list": "PartnerCatalogNumberList",
        "purpose": "zeigt partnerbezogene Katalog- und Bestellnummern zu diesem Artikel.",
        "edit": "PartnerCatalogNumberEdit",
        "columns": [
            ("Artikel", "Article", "ArticleId"),
            ("Partner", "Partner", "PartnerId"),
            ("Katalognummer", "Catalog Number", "CatalogNumber"),
        ],
    },
    {
        "parent": "ArticleEdit-Produktionsschritt-Ressourcen",
        "list": "ArticleProductionStepResourceList",
        "purpose": "zeigt Ressourcen, die diesem Produktionsschritt zugeordnet sind.",
        "edit": "ArticleProductionStepResourceEdit",
        "columns": [
            ("Position", "Position", "Position"),
            ("Ressource", "Resource", "ResourceId"),
            ("Produktionsschritt", "Production Step", "ArticleProductionStepId"),
            ("Ressourcengruppe", "Resource Group", "ResourceGroupId"),
            ("Sekunden pro Zyklus", "Seconds Per Cycle", "SecondsPerCycle"),
            ("Menge pro Zyklus", "Quantity Per Cycle", "QuantityPerCycle"),
            ("Preiseinheit", "Price Unit", "PriceUnitId"),
            ("Kosten pro Zyklus", "Costs Per Cycle", "CostsPerCycle"),
            ("Rüstkosten", "Setup Costs", "SetupCost"),
            ("Standard", "Default", "IsDefault"),
        ],
    },
    {
        "parent": "ArticleEdit-Produktionsschritt-Artikel",
        "list": "ArticleProductionStepComponentList",
        "purpose": "zeigt Artikelkomponenten, die in diesem Produktionsschritt verwendet werden.",
        "edit": "ArticleProductionStepComponentEdit",
        "columns": [
            ("Komponente", "Component", "ArticleComponentId"),
            ("Produktionsschritt", "Production Step", "ArticleProductionStepId"),
            ("Typ", "Type", "ArticleProductionStepComponentTypeId"),
            ("Position", "Position", "Position"),
            ("Mengenfaktor", "Quantity Factor", "QuantityFactor"),
            ("Mengenformel", "Quantity Formular", "QuantityFormular"),
        ],
    },
    {
        "parent": "ArticleEdit-Positionsliste",
        "list": "ArticleDocPositionList",
        "purpose": "zeigt Belegpositionen, in denen dieser Artikel verwendet wurde.",
        "edit": "",
        "columns": [
            ("Belegdatum", "Doc Date", "DocDate"),
            ("Belegnummer", "Doc Number", "DocNumber"),
            ("Partner", "Partner", "Matchcode"),
            ("Nummer", "Number", "BusinessPartnerNumber"),
            ("Menge", "Quantity", "Quantity"),
            ("Offene Menge", "Quantity Open", "QuantityOpen"),
            ("Preis", "Price", "Price"),
            ("Gesamtpreis", "Total Price", "TotalPrice"),
            ("Lagerstatus", "Inventory Status", "InventoryStatusId"),
            ("Lagermenge", "Warehousing Quantity", "ViewWarehousingQuantity"),
            ("Umsatz", "Revenue", "ViewRevenue"),
            ("Rohertrag Betrag", "Profit Amount", "ProfitAmount"),
            ("Rohertrag %", "Profit Percent", "ProfitPercent"),
        ],
    },
    {
        "parent": "ArticleEdit-Umsatz",
        "list": "ArticleRevenueList",
        "purpose": "zeigt Einkaufs- und Verkaufswerte des Artikels nach Jahr und Monat.",
        "edit": "",
        "columns": [
            ("Jahr", "Year", "YEAR"),
            ("Monat", "Month", "MONTH"),
            ("Einkaufsmenge", "Purchase Quantity", "PurchaseQuantity"),
            ("Einkaufsbetrag", "Purchase Amount", "PurchaseAmount"),
            ("Verkaufsmenge", "Sales Quantity", "SalesQuantity"),
            ("Verkaufsumsatz", "Sales Revenue", "SalesRevenue"),
            ("Rohertrag Betrag", "Profit Amount", "ProfitAmount"),
            ("Rohertrag %", "Profit Percent", "ProfitPercent"),
        ],
    },
    {
        "parent": "ArticleEdit-Verträge",
        "list": "ArticleDocPositionList",
        "purpose": "zeigt Vertragspositionen, in denen dieser Artikel verwendet wird.",
        "edit": "",
        "columns": [
            ("Belegdatum", "Doc Date", "DocDate"),
            ("Belegnummer", "Doc Number", "DocNumber"),
            ("Partner", "Partner", "Matchcode"),
            ("Nummer", "Number", "BusinessPartnerNumber"),
            ("Menge", "Quantity", "Quantity"),
            ("Offene Menge", "Quantity Open", "QuantityOpen"),
            ("Preis", "Price", "Price"),
            ("Gesamtpreis", "Total Price", "TotalPrice"),
        ],
    },
    {
        "parent": "ArticleEdit-Änderungsprotokoll",
        "list": "ArticleHistoryList",
        "purpose": "zeigt die Änderungshistorie dieses Artikels.",
        "edit": "",
        "columns": [
            ("Matchcode", "Matchcode", "Matchcode"),
            ("Name1", "Name1", "Name1"),
            ("Name2", "Name2", "Name2"),
            ("Artikelgruppe", "Article Group", "ArticleGroupName"),
            ("Mengeneinheit", "UoM", "ArticleUoMName"),
            ("Hersteller", "Manufacturer", "ManufacturerName"),
            ("Lagerung", "Warehousing", "Warehousing"),
            ("Shop", "Shop", "Shop"),
            ("Drucken", "Printing", "Printing"),
        ],
    },
    {
        "parent": "ArticleEdit-Anhänge",
        "list": "AttachmentList",
        "purpose": "zeigt Dateien, Bilder und sonstige Anlagen zu diesem Artikel.",
        "edit": "AttachmentEdit",
        "columns": [
            ("Info", "Info", "Info"),
            ("Wiedervorlage", "Follow-Up", "FollowUp"),
            ("Status", "Status", "AttachmentStatusId"),
            ("Erstellt am", "Date Created", "DateCreated"),
            ("Erstellt von", "Created By", "CreatedBy"),
        ],
    },
    {
        "parent": "ArticleEdit-Verwendung in Sets",
        "list": "ArticleUsageSetList",
        "purpose": "zeigt Sets, in denen dieser Artikel als Bestandteil verwendet wird.",
        "edit": "",
        "columns": [
            ("Set-Matchcode", "Set Matchcode", "Id"),
            ("Name 1", "Name 1", "Name1"),
            ("Name 2", "Name 2", "Name2"),
            ("Artikelgruppe", "Article Group", "ArticleGroupId"),
            ("Lagerung", "Warehousing", "Warehousing"),
            ("Aktiv", "Active", "Active"),
        ],
    },
    {
        "parent": "ArticleEdit-Verwendung in Produktion",
        "list": "ArticleUsageProductionList",
        "purpose": "zeigt Produktionsartikel, in denen dieser Artikel als Komponente verwendet wird.",
        "edit": "",
        "columns": [
            ("Produktions-Matchcode", "Production Matchcode", "Id"),
            ("Name 1", "Name 1", "Name1"),
            ("Name 2", "Name 2", "Name2"),
            ("Artikelgruppe", "Article Group", "ArticleGroupId"),
            ("Lagerung", "Warehousing", "Warehousing"),
            ("Aktiv", "Active", "Active"),
        ],
    },
    {
        "parent": "ArticleEdit-Lagerhistorie",
        "list": "ArticleQuantityHistoryTableList",
        "purpose": "zeigt die historische Entwicklung von Bestand und Verfügbarkeit.",
        "edit": "",
        "columns": [
            ("Jetzt verfügbar", "Now Available", "NowAvailable"),
            ("Bestand", "Stock", "Stock"),
            ("Bestätigt", "Confirmed", "Confirmed"),
            ("Gepackt", "Packed", "Packed"),
            ("Bestellt", "Ordered", "Ordered"),
            ("Angekündigt", "Announced", "Announced"),
            ("Später verfügbar", "Later Available", "LaterAvailable"),
            ("Mindestbestand", "Minimum Stock", "MinimumStock"),
            ("Meldeschwelle", "Message Threshold", "MessageThreshold"),
            ("Maximalbestand", "Maximum Stock", "MaximumStock"),
            ("Inventur", "Inventory Counting", "WarehouseInventoryCountingId"),
            ("Beleg", "Doc", "DocNumber"),
            ("Produktion", "Production", "ProductionNumber"),
            ("Datum", "Date", "SysStartTime"),
        ],
    },
]


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.font:
        dst.font = copy.copy(src.font)
    if src.fill:
        dst.fill = copy.copy(src.fill)
    if src.border:
        dst.border = copy.copy(src.border)


def copy_row_style(ws, style_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[style_row].height
    for col in range(1, min(ws.max_column, 20) + 1):
        copy_cell_style(ws.cell(style_row, col), ws.cell(target_row, col))


def find_article_block(ws) -> tuple[int, int, int]:
    start = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "ArticleEdit")
    level = ws.row_dimensions[start].outlineLevel
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if ws.row_dimensions[row].outlineLevel <= level:
            end = row - 1
            break
    return start, end, level


def find_parent_row(ws, parent_name: str) -> int | None:
    start, end, article_level = find_article_block(ws)
    for row in range(start + 1, end + 1):
        if ws.row_dimensions[row].outlineLevel == article_level + 1 and ws.cell(row, 1).value == parent_name:
            return row
    return None


def child_range_end(ws, parent_row: int) -> int:
    start, end, _ = find_article_block(ws)
    parent_level = ws.row_dimensions[parent_row].outlineLevel
    last = parent_row
    for row in range(parent_row + 1, end + 1):
        if ws.row_dimensions[row].outlineLevel <= parent_level:
            break
        last = row
    return last


def describe_column(label: str, field: str, list_name: str) -> str:
    key = field.lower()
    if any(part in key for part in ("price", "cost", "amount", "revenue", "profit", "landed")):
        return f"{label} zeigt einen Wert für Kalkulation, Preisbewertung oder Auswertung in {list_name}."
    if any(part in key for part in ("quantity", "stock", "available", "minimum", "maximum", "threshold", "packed", "ordered", "announced")):
        return f"{label} zeigt eine Menge oder Bestandsinformation in {list_name}."
    if any(part in key for part in ("date", "time", "year", "month")):
        return f"{label} zeigt den zeitlichen Bezug des Listeneintrags."
    if any(part in key for part in ("id", "number", "matchcode", "code")):
        return f"{label} dient zur eindeutigen Zuordnung, Suche oder Identifikation des Listeneintrags."
    return f"{label} zeigt eine fachliche Information des Listeneintrags in {list_name}."


def build_rows(spec: dict) -> list[tuple[str, str, str, str]]:
    rows = [
        (
            spec["list"],
            spec["list"],
            "",
            f"Die eingebettete Liste {spec['list']} {spec['purpose']} Sie macht die zugehörigen Datensätze direkt im Register sichtbar.",
        )
    ]
    rows.extend(
        (label, original, field, describe_column(label, field, spec["list"]))
        for label, original, field in spec["columns"]
    )
    if spec["edit"]:
        rows.append(
            (
                f"{spec['edit']} anlegen/bearbeiten",
                spec["edit"],
                spec["edit"],
                f"Über Neu oder Bearbeiten öffnet X-ERP die Maske {spec['edit']}. Dort werden die einzelnen Datensätze dieser Liste gepflegt.",
            )
        )
    return rows


def set_row(ws, row: int, level: int, values: tuple[str, str, str, str]) -> None:
    topic, original, field, description = values
    ws.row_dimensions[row].outlineLevel = level
    ws.cell(row, 1).value = topic
    ws.cell(row, 2).value = original
    ws.cell(row, 3).value = None
    ws.cell(row, 4).value = None
    ws.cell(row, 5).value = None
    ws.cell(row, 6).value = field
    ws.cell(row, 7).value = None
    ws.cell(row, 8).value = description
    ws.cell(row, 9).value = None


def validate(ws) -> list[str]:
    problems: list[str] = []
    start, end, article_level = find_article_block(ws)
    for row in range(start + 1, end + 1):
        value = ws.cell(row, 1).value
        level = ws.row_dimensions[row].outlineLevel
        if isinstance(value, str) and value.startswith("ArticleEdit-") and level != article_level + 1:
            problems.append(f"{row}: {value} hat Level {level}, erwartet {article_level + 1}")
    for spec in SPECS:
        parent = find_parent_row(ws, spec["parent"])
        if parent is None:
            problems.append(f"{spec['parent']} fehlt")
            continue
        end_row = child_range_end(ws, parent)
        found = False
        for row in range(parent + 1, end_row + 1):
            if ws.cell(row, 1).value == spec["list"]:
                found = True
                if ws.row_dimensions[row].outlineLevel != ws.row_dimensions[parent].outlineLevel + 1:
                    problems.append(f"{row}: {spec['list']} hat falsches Level")
                break
        if not found and spec["list"] != "ArticleSetList":
            problems.append(f"{spec['list']} unter {spec['parent']} fehlt")
    return problems


def main() -> None:
    archive_dir = WORKBOOK.parent / "ARCHIV"
    archive_dir.mkdir(exist_ok=True)
    backup = archive_dir / f"X-ERP-HELP-before-article-edit-lists-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]

    # Use the already verified ArticleSetList rows as visual templates.
    list_template = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "ArticleSetList")
    child_template = list_template + 1

    planned: list[tuple[int, dict]] = []
    for spec in SPECS:
        parent = find_parent_row(ws, spec["parent"])
        if parent is None:
            continue
        end = child_range_end(ws, parent)
        if any(ws.cell(row, 1).value == spec["list"] for row in range(parent + 1, end + 1)):
            continue
        planned.append((parent, spec))

    added = 0
    for parent, spec in sorted(planned, key=lambda item: item[0], reverse=True):
        parent_level = ws.row_dimensions[parent].outlineLevel
        insert_at = child_range_end(ws, parent) + 1
        rows = build_rows(spec)
        ws.insert_rows(insert_at, len(rows))
        for index, values in enumerate(rows):
            target = insert_at + index
            copy_row_style(ws, list_template if index == 0 else child_template, target)
            set_row(ws, target, parent_level + (1 if index == 0 else 2), values)
        added += len(rows)

    problems = validate(ws)
    if problems:
        raise RuntimeError("Validierung fehlgeschlagen:\n" + "\n".join(problems[:30]))

    wb.save(WORKBOOK)
    print(f"added={added}")
    print(f"backup={backup}")


if __name__ == "__main__":
    from add_article_edit_embedded_lists_xml_checked import main as xml_main

    xml_main()
