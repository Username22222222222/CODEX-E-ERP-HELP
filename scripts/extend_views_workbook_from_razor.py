from __future__ import annotations

import csv
import datetime as dt
import json
import re
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = ROOT / "X-ERP-HELP.xlsx"
INVENTORY = ROOT / "outputs" / "program-audit" / "route-inventory.csv"
SCREENSHOT_DIR = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help\Screenshots")
ARCHIVE = ROOT / "ARCHIV"
REPORT = ROOT / "outputs" / "program-audit" / "view-structure-extension-report.json"


def u(text: str) -> str:
    try:
        return text.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return text


TRANSLATIONS = {
    "Overview": u("\u00dcbersicht"),
    "Details": "Details",
    "Attachments": u("Anh\u00e4nge"),
    "ExtraFields": "Extra-Felder",
    "Extra Fields": "Extra-Felder",
    "Dialog": "Dialog",
    "History": "Historie",
    "Articles": "Artikel",
    "BinLocations": u("Lagerpl\u00e4tze"),
    "Positions": "Positionen",
    "Picture": "Bild",
    "Info": "Info",
    "Warehouse": "Lager",
    "Warehouse Name": "Lagername",
    "Bin Locations": u("Lagerpl\u00e4tze"),
    "Location": "Standort",
    "Type": "Typ",
    "Incoming Block": "Eingangssperre",
    "Outgoing Block": "Ausgangssperre",
    "Counting List Number": u("Z\u00e4hllistennummer"),
    "Active": "Aktiv",
    "Partner": "Partner",
    "Contact Person": "Ansprechpartner",
    "Helpdesk Date": "Helpdesk-Datum",
    "Helpdesk Number": "Helpdesk-Nummer",
    "Supporter": "Supporter",
    "Status": "Status",
    "Category": "Kategorie",
    "Topic": "Thema",
    "Prio": "Prioritaet",
    "Estimated Processing Time": u("Gesch\u00e4tzte Bearbeitungszeit"),
    "Processing Time": "Bearbeitungszeit",
    "Date Created": "Erstellt am",
    "Stars": "Bewertung",
    "Quality Comment": "Qualitaetskommentar",
    "Name": "Name",
    "Description": "Beschreibung",
    "Subject": "Betreff",
    "Report Name": "Reportname",
    "Report Template XML": "Reportvorlagen-XML",
    "Report Type": "Reporttyp",
    "Doc Type": "Dok-Typ",
    "Number of Dimensions": "Anzahl Dimensionen",
    "Dimension": "Dimension",
    "From": "Von",
    "To": "Bis",
    "Continue": "Weiter",
    "Create": "Erstellen",
    "Send": "Senden",
    "Start": "Start",
    "Running": "Laeuft",
    "Please enter a value greater than 0": u("Bitte einen Wert gr\u00f6\u00dfer als 0 eingeben"),
    "Entered values are invalid": u("Eingegebene Werte sind ung\u00fcltig"),
    "Bin Location Wizard": "Lagerplatz-Assistent",
}


MODULE_LABELS = {
    "Dashboard": "Dashboard",
    "Doc": "Belege und Service",
    "Helpdesk": "Helpdesk",
    "Marketplace": "Marketplace",
    "Partner": "Partner",
    "Phone": "Telefon",
    "Production": "Produktion",
    "SystemViews": "Systemansichten",
    "UserLoginSecurity": "Benutzer und Sicherheit",
    "Warehouse": "Lager",
    "Webshop": "Webshop",
}


MODULE_PURPOSES = {
    "Dashboard": u("Auswertungen, Kennzahlen und Dashboard-Konfigurationen im ERP."),
    "Doc": u("Beleg-, Service- und Dokumentprozesse im ERP."),
    "Helpdesk": u("Servicefaelle, Tickets, Kategorien, Themen, Prioritaeten und Dialoge."),
    "Marketplace": u("Marketplace- und Lieferantenportalprozesse im Webshop-Umfeld."),
    "Partner": u("partnerbezogene Preislisten und Einkaufsinformationen."),
    "Phone": u("Telefonie, Anrufe, Leitungen, Ordner und Telefonkommunikation."),
    "Production": u("Produktion, BDE, Arbeitsplaetze, Rueckmeldungen und Ressourcen."),
    "SystemViews": u("technische Zuordnungen und Formate fuer Datenimport, Banking und Schnittstellen."),
    "UserLoginSecurity": u("Autorisierung, Seitenrechte und WebApi-Rechte."),
    "Warehouse": u("Lager, Lagerplaetze, Bestandsfuehrung, Inventur, Bewertung und Umlagerung."),
    "Webshop": u("Webshop-Protokolle, Gestaltung, Symbole und Onlinevertriebsfunktionen."),
}


def de_label(text: str) -> str:
    if not text:
        return text
    text = text.strip().strip('"')
    if text in TRANSLATIONS:
        return TRANSLATIONS[text]
    if "-" in text:
        parts = text.split("-")
        return "-".join(TRANSLATIONS.get(p, p) for p in parts)
    return text


def slugify(name: str) -> str:
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", name).lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def humanize_name(name: str) -> str:
    label = re.sub(r"Edit$", "", name)
    label = re.sub(r"Wizard$", " Wizard", label)
    label = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", label)
    label = label.replace("Doc Xml", "Doc XML").replace("Web Api", "WebApi")
    return label


def read_inventory() -> list[dict[str, str]]:
    with INVENTORY.open(encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=";,")
        return list(csv.DictReader(handle, dialect=dialect))


def screenshot_for(name: str) -> str | None:
    exact = SCREENSHOT_DIR / f"{name}.webp"
    if exact.exists():
        return f"Screenshots/{exact.name}"
    matches = sorted(SCREENSHOT_DIR.glob(f"{name}*.webp"))
    if matches:
        return f"Screenshots/{matches[0].name}"
    return None


def extract_registers(text: str, name: str) -> list[str]:
    regs: list[str] = []
    for match in re.findall(r'"([A-Za-z0-9]+-[A-Za-z0-9_-]+)"', text):
        if match not in regs and not match.endswith(".razor"):
            regs.append(match)
    for match in re.findall(r'const\s+string\s+REGISTER_[A-Za-z0-9_]+\s*=\s*"([^"]+)"', text):
        if match not in regs:
            regs.append(match)
    shots = sorted(SCREENSHOT_DIR.glob(f"{name}_*.webp"))
    for shot in shots:
        suffix = shot.stem[len(name) + 1 :]
        if suffix and suffix not in regs:
            regs.append(f"{name}-{suffix}")
    return regs[:20]


def extract_fields(text: str) -> list[str]:
    fields: list[str] = []
    for raw in re.findall(r'\bCaption\s*=\s*"([^"]+)"', text):
        label = de_label(raw)
        if label and label not in fields:
            fields.append(label)
    for raw in re.findall(r'\bBindField\s*=\s*"([^"]+)"', text):
        label = de_label(raw)
        if label and label not in fields:
            fields.append(label)
    return fields[:40]


def extract_buttons(text: str) -> list[str]:
    buttons: list[str] = []
    toolbar_map = {
        "NewClicked": "Neu",
        "SaveClicked": "Speichern",
        "SaveExitClicked": u("Speichern & Schlie\u00dfen"),
        "ExitCancelClicked": "Abbrechen",
        "Preview_Clicked": "Vorschau",
    }
    for token, label in toolbar_map.items():
        if token in text and label not in buttons:
            buttons.append(label)
    if "XTBWizard" in text and "Abbrechen" not in buttons:
        buttons.append("Abbrechen")
    for raw in re.findall(r'lzr\["([^"]+)"\]', text):
        label = de_label(raw)
        if label and label not in buttons and re.search(r"weiter|erstellen|start|senden|abbrechen|wizard|ausw", label, re.I):
            buttons.append(label)
    for raw in re.findall(r'\bText\s*=\s*"([^"]+)"', text):
        label = de_label(raw)
        if label and label not in buttons and re.search(r"send|save|create|continue|start|ok|cancel|senden|erstellen|weiter", label, re.I):
            buttons.append(label)
    return buttons[:20]


def purpose_for(name: str, module: str, fields: list[str], kind: str) -> str:
    visible = ", ".join(fields[:8]) if fields else "die relevanten Eingaben"
    label = humanize_name(name)
    module_purpose = MODULE_PURPOSES.get(module, u("den jeweiligen ERP-Prozess."))
    if kind == "Wizard":
        return u("{label} fuehrt Anwender schrittweise durch eine gefuehrte ERP-Aktion. Die Maske reduziert freie Eingaben auf die notwendigen Parameter und erzeugt daraus die benoetigten Folgedaten. Wichtige Angaben sind {visible}.").format(label=label, visible=visible)
    return u("{name} ist die Bearbeitungsansicht fuer {label}. Sie dient im Bereich {module_label} dazu, {module_purpose} Anwender pruefen und pflegen hier insbesondere {visible}.").format(
        name=name,
        label=label,
        module_label=MODULE_LABELS.get(module, module),
        module_purpose=module_purpose,
        visible=visible,
    )


def field_desc(field: str, view_name: str, module: str) -> str:
    f = field.lower()
    if "status" in f:
        return u("Ordnet den Datensatz einem Bearbeitungsstand zu und macht sichtbar, wie der Vorgang im ERP weiterverarbeitet wird.")
    if "partner" in f or "kunde" in f or "lieferant" in f:
        return u("Verknuepft den Vorgang mit dem passenden Geschaeftspartner, damit Folgeprozesse, Auswertungen und Kommunikation korrekt zugeordnet werden.")
    if "datum" in f or "date" in f or "zeit" in f:
        return u("Speichert den zeitlichen Bezug des Vorgangs und ist wichtig fuer Planung, Nachverfolgung und Auswertung.")
    if "name" in f:
        return u("Bezeichnet den Datensatz so, dass Anwender ihn in Listen, Auswahlfeldern und Auswertungen eindeutig erkennen.")
    if "beschreibung" in f or "description" in f or "info" in f or "comment" in f:
        return u("Enthaelt ergaenzende Informationen, die den Vorgang fachlich beschreiben und fuer andere Anwender nachvollziehbar machen.")
    if "xml" in f:
        return u("Enthaelt eine technische XML-Definition. Aenderungen sollten nur durch berechtigte Anwender mit Kenntnis der jeweiligen Ausgabe- oder Schnittstellenstruktur erfolgen.")
    if "lager" in f or "warehouse" in f:
        return u("Ordnet den Vorgang einem Lager oder Lagerbereich zu und beeinflusst Bestand, Verfuegbarkeit und lagerbezogene Auswertungen.")
    if "prio" in f or "prior" in f:
        return u("Legt die Prioritaet fest, damit dringende Vorgaenge im Arbeitsablauf schneller erkannt und bearbeitet werden.")
    if "rolle" in f or "recht" in f or "authorization" in f:
        return u("Steuert Berechtigung oder Zugriff und ist fuer sichere ERP-Nutzung durch Rollen und Benutzer relevant.")
    if "von" == f or "bis" == f or "from" == f or "to" == f:
        return u("Definiert eine Grenze oder einen Bereich, aus dem der Assistent die folgenden Datensaetze erzeugt.")
    return u("Beschreibt die Angabe `{field}` in {view_name}. Das Feld unterstuetzt Anwender dabei, den Vorgang im Bereich {module_label} korrekt zu pflegen und spaeter wiederzufinden.").format(
        field=field,
        view_name=view_name,
        module_label=MODULE_LABELS.get(module, module),
    )


def button_desc(button: str, view_name: str) -> str:
    b = button.lower()
    if "speichern" in b:
        return u("Speichert die Eingaben dieser Ansicht, sofern Pflichtfelder und Validierungen erfuellt sind.")
    if "abbrechen" in b:
        return u("Verlaesst den Assistenten oder die Bearbeitung ohne den aktuellen Schritt fortzusetzen.")
    if "neu" == b:
        return u("Startet das Anlegen eines neuen Datensatzes.")
    if "vorschau" in b:
        return u("Oeffnet die Druck- oder Reportvorschau fuer den aktuellen Datensatz.")
    if "weiter" in b:
        return u("Fuehrt zum naechsten Schritt des Assistenten, wenn die bisherigen Eingaben plausibel sind.")
    if "erstellen" in b or "create" in b:
        return u("Fuehrt die Aktion aus und erzeugt die vorgesehenen Folgedaten.")
    if "start" in b:
        return u("Startet den zugehoerigen Vorgang, beispielsweise eine Zeiterfassung oder einen Prozessschritt.")
    return u("Fuehrt die Aktion `{button}` in {view_name} aus.").format(button=button, view_name=view_name)


def copy_cell_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for c in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(src_row, c), ws.cell(dst_row, c))
        ws.cell(dst_row, c).value = None
    src_dim = ws.row_dimensions[src_row]
    dst_dim = ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed


def main() -> None:
    ARCHIVE.mkdir(exist_ok=True)
    shutil.copy2(WORKBOOK, ARCHIVE / f"X-ERP-HELP-before-extend-views-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx")
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}
    existing = {str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[0] is not None}

    inventory = read_inventory()
    wanted: dict[str, dict[str, str]] = {}
    for item in inventory:
        name = item.get("Name") or ""
        if item.get("Kind") not in ("Edit", "Wizard") or not name or name in existing:
            continue
        wanted.setdefault(name, item)

    insert_at = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, col["Thema"]).value == "FAQ (H\u00e4ufig gestellte Fragen)")
    generated: list[dict[str, str]] = []
    row_defs: list[dict[str, str]] = []
    current_module = None
    toc_order = 12140

    for name, item in sorted(wanted.items(), key=lambda x: ((x[1].get("Module") or ""), x[0])):
        module = item.get("Module") or "Views"
        kind = item.get("Kind") or "View"
        file_path = Path(item.get("File") or "")
        text = file_path.read_text(encoding="utf-8-sig", errors="replace") if file_path.exists() else ""
        fields = extract_fields(text)
        buttons = extract_buttons(text)
        registers = extract_registers(text, name)
        screenshot = screenshot_for(name)
        slug = slugify(name)

        if module != current_module:
            current_module = module
            row_defs.append({"kind": "group", "Thema": f"Nachtrag - {MODULE_LABELS.get(module, module)}"})

        content_type = "Wizard" if kind == "Wizard" else "View"
        title_label = humanize_name(name)
        row_defs.append(
            {
                "kind": "view",
                "Thema": name,
                "Original Text": name,
                "Ordner": "screenshots" if screenshot else None,
                "Screenshot": screenshot,
                "Beschreibung": purpose_for(name, module, fields, kind),
                "PAGE_ID": f"views/{slug}",
                "SLUG": slug,
                "TITLE": f"{title_label} erklaert | X-ERP ERP Hilfe",
                "META_DESCRIPTION": f"{name} in X-ERP: Register, Felder, Buttons und Zweck der ERP-Ansicht verstaendlich erklaert.",
                "H1": title_label,
                "PRIMARY_KEYWORD": "ERP",
                "CONTENT_TYPE": content_type,
                "STRUCTURED_DATA_TYPE": "TechArticle",
                "CANONICAL_URL": f"https://x-erp.de/de/help/views/{slug}.html",
                "HREFLANG_GROUP": f"views/{slug}",
                "SEO_STATUS": "draft",
                "DIRECTORY_PATH": "views",
                "FILE_NAME": f"{slug}.html",
                "URL_PATH": f"views/{slug}.html",
                "STORAGE_PATH": f"views/{slug}.html",
                "NAV_TITLE": title_label,
                "BREADCRUMB": f"Ansichten > {MODULE_LABELS.get(module, module)} > {title_label}",
                "TOC_PARENT": "views",
                "TOC_LEVEL": "2",
                "TOC_ORDER": str(toc_order),
                "UNIQUE_PAGE_KEY": f"views/{slug}",
                "SCREENSHOT_REL_PATH": screenshot,
                "SCREENSHOT_WEB_PATH": f"/de/help/{screenshot}" if screenshot else None,
                "IMAGE_ALT": f"{name} Ansicht im ERP-System X-ERP",
                "IMAGE_CAPTION": f"{name} in X-ERP.",
                "IMAGE_STATUS": "vorhanden" if screenshot else "kein Screenshot",
            }
        )
        toc_order += 10

        if registers:
            for reg in registers[:8]:
                reg_label = de_label(reg.split("-", 1)[1] if "-" in reg else reg)
                row_defs.append(
                    {
                        "kind": "tab",
                        "Thema": f"{name}-{reg_label}",
                        "Original Text": reg,
                        "Beschreibung": f"Das Register {reg_label} buendelt die fachlich zusammengehoerigen Informationen der Ansicht {name}.",
                        "IMAGE_STATUS": "kein Screenshot",
                    }
                )
            parent_has_tabs = True
        else:
            parent_has_tabs = False

        field_kind = "field4" if parent_has_tabs else "field3"
        for field in fields[:18]:
            row_defs.append(
                {
                    "kind": field_kind,
                    "Thema": field,
                    "Original Text": f"Eingabefeld fuer: {field}.",
                    "Feld": field,
                    "Beschreibung": field_desc(field, name, module),
                    "IMAGE_STATUS": "kein Screenshot",
                }
            )
        for button in buttons[:10]:
            row_defs.append(
                {
                    "kind": field_kind,
                    "Thema": f"Button: {button}",
                    "Original Text": f"Schaltflaeche: {button}.",
                    "Feld": button,
                    "Beschreibung": button_desc(button, name),
                    "IMAGE_STATUS": "kein Screenshot",
                }
            )
        generated.append({"name": name, "module": module, "kind": kind, "fields": len(fields), "buttons": len(buttons), "registers": len(registers), "screenshot": screenshot})

    ws.insert_rows(insert_at, len(row_defs))
    style_source = {"group": 1141, "view": 1142, "tab": 1192, "field4": 1193, "field3": 1206}

    for offset, data in enumerate(row_defs):
        row = insert_at + offset
        src = style_source[data["kind"]]
        copy_row_style(ws, src, row)
        for key, value in data.items():
            if key != "kind" and key in col:
                ws.cell(row, col[key]).value = value

    wb.save(WORKBOOK)
    REPORT.write_text(json.dumps({"inserted_rows": len(row_defs), "inserted_views": generated}, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"inserted_rows": len(row_defs), "inserted_views": len(generated), "report": str(REPORT)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
