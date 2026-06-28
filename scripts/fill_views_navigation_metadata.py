from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")

REQUIRED_HEADERS = [
    "LAST_REVIEWED",
    "DIRECTORY_PATH",
    "FILE_NAME",
    "URL_PATH",
    "STORAGE_PATH",
    "NAV_TITLE",
    "BREADCRUMB",
    "TOC_PARENT",
    "TOC_LEVEL",
    "TOC_ORDER",
    "UNIQUE_PAGE_KEY",
    "SCREENSHOT_REL_PATH",
    "SCREENSHOT_WEB_PATH",
    "IMAGE_ALT",
    "IMAGE_CAPTION",
    "IMAGE_STATUS",
]


def slugify(value: str) -> str:
    replacements = {
        "Ä": "Ae",
        "Ö": "Oe",
        "Ü": "Ue",
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }
    for src, dst in replacements.items():
        value = value.replace(src, dst)
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def nav_title(topic: str) -> str:
    if " anlegen/bearbeiten" in topic:
        return topic
    if "-" in topic:
        left, right = topic.split("-", 1)
        if left.endswith(("Edit", "Wizard", "Clicked")) or "_" in left:
            return right.strip() or topic
    return topic


def ensure_headers(ws) -> dict[str, int]:
    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1) if ws.cell(1, col).value}
    col = ws.max_column
    for header in REQUIRED_HEADERS:
        if header not in headers:
            col += 1
            ws.cell(1, col).value = header
            headers[header] = col
    return headers


def find_section(ws, title: str) -> tuple[int, int, int]:
    start = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == title)
    level = ws.row_dimensions[start].outlineLevel
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if ws.row_dimensions[row].outlineLevel <= level:
            end = row - 1
            break
    return start, end, level


def should_have_page(level: int, ansichten_level: int, topic: str) -> bool:
    if level <= ansichten_level + 3:
        return True
    if topic.endswith("List") or " anlegen/bearbeiten" in topic:
        return True
    return False


def main() -> None:
    archive_dir = WORKBOOK.parent / "ARCHIV"
    archive_dir.mkdir(exist_ok=True)
    backup = archive_dir / f"X-ERP-HELP-before-views-navigation-metadata-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = ensure_headers(ws)
    start, end, ansichten_level = find_section(ws, "Ansichten")

    stack: dict[int, str] = {}
    slug_stack: dict[int, str] = {}
    sibling_counter: dict[tuple[int, str], int] = {}
    order_stack: dict[int, str] = {}
    updated = 0

    for row in range(start, end + 1):
        topic = ws.cell(row, 1).value
        if not isinstance(topic, str) or not topic.strip():
            continue
        level = ws.row_dimensions[row].outlineLevel

        for old_level in list(stack):
            if old_level >= level:
                stack.pop(old_level, None)
                slug_stack.pop(old_level, None)
                order_stack.pop(old_level, None)

        parent_topic = stack.get(level - 1, "")
        parent_order = order_stack.get(level - 1, "")
        key = (level, parent_order)
        sibling_counter[key] = sibling_counter.get(key, 0) + 1
        order = f"{parent_order}.{sibling_counter[key]:04d}" if parent_order else f"{sibling_counter[key]:04d}"

        title = nav_title(topic)
        current_slugs = [slug_stack[i] for i in sorted(slug_stack)]
        own_slug = slugify(topic)
        breadcrumb_items = [stack[i] for i in sorted(stack)] + [topic]
        breadcrumb = " > ".join(breadcrumb_items)
        unique_page_key = "/".join(current_slugs + [own_slug])

        ws.cell(row, headers["NAV_TITLE"]).value = title
        ws.cell(row, headers["BREADCRUMB"]).value = breadcrumb
        ws.cell(row, headers["TOC_PARENT"]).value = parent_topic
        ws.cell(row, headers["TOC_LEVEL"]).value = level - ansichten_level
        ws.cell(row, headers["TOC_ORDER"]).value = order
        ws.cell(row, headers["UNIQUE_PAGE_KEY"]).value = unique_page_key

        screenshot = ws.cell(row, 5).value
        if screenshot:
            ws.cell(row, headers["SCREENSHOT_REL_PATH"]).value = screenshot
            ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value = "/" + str(screenshot).replace("\\", "/")
            ws.cell(row, headers["IMAGE_ALT"]).value = f"Screenshot {topic} in X-ERP"
            ws.cell(row, headers["IMAGE_CAPTION"]).value = f"{topic} in der X-ERP Hilfe"
            ws.cell(row, headers["IMAGE_STATUS"]).value = "assigned"
        elif should_have_page(level, ansichten_level, topic):
            ws.cell(row, headers["IMAGE_STATUS"]).value = "missing"

        if should_have_page(level, ansichten_level, topic):
            path_parts = current_slugs + [own_slug]
            directory = "/".join(path_parts[:-1]) if len(path_parts) > 1 else path_parts[0]
            filename = f"{path_parts[-1]}.md"
            url_path = "/de/help/" + "/".join(path_parts)
            ws.cell(row, headers["DIRECTORY_PATH"]).value = directory
            ws.cell(row, headers["FILE_NAME"]).value = filename
            ws.cell(row, headers["URL_PATH"]).value = url_path
            ws.cell(row, headers["STORAGE_PATH"]).value = f"D:/DATEN/HOMEPAGES/x-erp.de/de/help/{directory}/{filename}".replace("//", "/")

        stack[level] = topic
        slug_stack[level] = own_slug
        order_stack[level] = order
        updated += 1

    wb.save(WORKBOOK)
    print(f"updated={updated}")
    print(f"backup={backup}")


if __name__ == "__main__":
    main()
