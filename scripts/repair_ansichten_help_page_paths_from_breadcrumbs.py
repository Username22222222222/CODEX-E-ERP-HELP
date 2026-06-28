from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def slug(value: str) -> str:
    value = value.replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue")
    value = value.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
    value = value.replace("%", "prozent")
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", value)
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "seite"


def normalize_parent_url(path: str) -> str:
    path = path.replace("\\", "/").strip()
    if path.startswith("/de/help/"):
        path = path[len("/de/help/") :]
    if path.startswith("de/help/"):
        path = path[len("de/help/") :]
    if path.startswith("ansichten/"):
        path = "Ansichten/" + path[len("ansichten/") :]
    if path.endswith("index.html"):
        path = path[: -len("index.html")]
    return path.rstrip("/")


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-ansichten-path-repair-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    topic_col = headers["Thema"]
    breadcrumb_col = headers["BREADCRUMB"]
    url_col = headers["URL_PATH"]
    content_type_col = headers["CONTENT_TYPE"]

    by_breadcrumb = {
        cell_text(ws.cell(row, breadcrumb_col).value): row
        for row in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row, breadcrumb_col).value)
    }

    changed = 0
    for row in range(2, ws.max_row + 1):
        breadcrumb = cell_text(ws.cell(row, breadcrumb_col).value)
        if not breadcrumb.startswith("Ansichten > "):
            continue
        if cell_text(ws.cell(row, content_type_col).value) != "HelpPage":
            continue
        parts = [part.strip() for part in breadcrumb.split(">") if part.strip()]
        if len(parts) <= 4:
            continue
        parent_breadcrumb = " > ".join(parts[:-1])
        parent_row = by_breadcrumb.get(parent_breadcrumb)
        if not parent_row:
            continue
        parent_url = normalize_parent_url(cell_text(ws.cell(parent_row, url_col).value))
        if not parent_url:
            continue
        topic = cell_text(ws.cell(row, topic_col).value)
        new_path = f"{parent_url}/{slug(topic)}/index.html"
        if cell_text(ws.cell(row, url_col).value) != new_path:
            ws.cell(row, url_col).value = new_path
            changed += 1

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"paths_repaired={changed}")


if __name__ == "__main__":
    main()
