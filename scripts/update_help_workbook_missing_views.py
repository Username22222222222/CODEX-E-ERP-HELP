from __future__ import annotations

import datetime as dt
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
INSERT_AT = 1141


def u(text: str) -> str:
    try:
        return text.encode("ascii").decode("unicode_escape")
    except UnicodeEncodeError:
        return text


def copy_cell_style(src_cell, dst_cell) -> None:
    if src_cell.has_style:
        dst_cell._style = copy(src_cell._style)
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)


def copy_row_dimension(src_ws, dst_ws, src_row: int, dst_row: int) -> None:
    src_dim = src_ws.row_dimensions[src_row]
    dst_dim = dst_ws.row_dimensions[dst_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = src_dim.hidden
    dst_dim.outlineLevel = src_dim.outlineLevel
    dst_dim.collapsed = src_dim.collapsed


def build_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []

    def group(name: str) -> None:
        rows.append({"kind": "group", "Thema": name})

    def view(
        thema: str,
        screenshot: str,
        desc: str,
        slug: str,
        title: str,
        meta: str,
        h1: str,
        nav: str,
        breadcrumb: str,
        order: int,
        alt: str,
        caption: str,
    ) -> None:
        rows.append(
            {
                "kind": "view",
                "Thema": thema,
                "Original Text": thema,
                "Ordner": "screenshots",
                "Screenshot": screenshot,
                "Beschreibung": desc,
                "PAGE_ID": f"views/{slug}",
                "SLUG": slug,
                "TITLE": title,
                "META_DESCRIPTION": meta,
                "H1": h1,
                "PRIMARY_KEYWORD": "ERP",
                "CONTENT_TYPE": "View",
                "STRUCTURED_DATA_TYPE": "TechArticle",
                "CANONICAL_URL": f"https://x-erp.de/de/help/views/{slug}.html",
                "HREFLANG_GROUP": f"views/{slug}",
                "SEO_STATUS": "draft",
                "DIRECTORY_PATH": "views",
                "FILE_NAME": f"{slug}.html",
                "URL_PATH": f"views/{slug}.html",
                "STORAGE_PATH": f"views/{slug}.html",
                "NAV_TITLE": nav,
                "BREADCRUMB": breadcrumb,
                "TOC_PARENT": "views",
                "TOC_LEVEL": "2",
                "TOC_ORDER": str(order),
                "UNIQUE_PAGE_KEY": f"views/{slug}",
                "SCREENSHOT_REL_PATH": screenshot,
                "SCREENSHOT_WEB_PATH": f"/de/help/{screenshot}",
                "IMAGE_ALT": alt,
                "IMAGE_CAPTION": caption,
                "IMAGE_STATUS": "vorhanden",
            }
        )

    def tab(name: str, desc: str) -> None:
        rows.append(
            {
                "kind": "tab",
                "Thema": name,
                "Original Text": name,
                "Beschreibung": desc,
                "IMAGE_STATUS": "kein Screenshot",
            }
        )

    def field(name: str, desc: str) -> None:
        rows.append(
            {
                "kind": "field",
                "Thema": name,
                "Original Text": u(f"Eingabefeld f\\u00fcr: {name}."),
                "Feld": name,
                "Beschreibung": desc,
                "IMAGE_STATUS": "kein Screenshot",
            }
        )

    def button(name: str, desc: str) -> None:
        rows.append(
            {
                "kind": "field",
                "Thema": name,
                "Original Text": u(f"Schaltfl\\u00e4che: {name}."),
                "Feld": name,
                "Beschreibung": desc,
                "IMAGE_STATUS": "kein Screenshot",
            }
        )

    group("Portal und Webshop")
    view(
        "WebshopSettingEdit",
        "Screenshots/WebshopSettingLive.webp",
        u("WebshopSettingEdit steuert die zentralen Einstellungen des X-ERP.webshop. Die Ansicht verbindet Shop-Name, Kontakt, Rechtstexte, Sprache, Standort, Zahlungs- und Lieferbedingungen, Kundenanlage, Benachrichtigungen, Bestelltexte und Bilddarstellung. Sie ist wichtig, damit Onlinebestellungen ohne doppelte Pflege in den ERP-Verkaufsprozess einflie\\u00dfen."),
        "webshop-setting-edit",
        u("Webshop-Einstellungen bearbeiten | X-ERP ERP Hilfe"),
        u("WebshopSettingEdit in X-ERP: Webshop, Kundenanlage, Rechtstexte, Sprache, Zahlung, Lieferung, Lageranzeige und Bestell-E-Mails im ERP konfigurieren."),
        "Webshop-Einstellungen bearbeiten",
        "Webshop-Einstellungen",
        "Ansichten > Portal und Webshop > Webshop-Einstellungen",
        9991,
        u("WebshopSettingEdit Ansicht im ERP-System X-ERP f\\u00fcr X-ERP.webshop Einstellungen"),
        "WebshopSettingEdit konfiguriert den X-ERP.webshop.",
    )
    for name, desc in [
        ("Name", u("Bezeichnet den Webshop oder Portalbereich und hilft Anwendern, die richtige Shop-Konfiguration zu erkennen.")),
        ("Beschreibung", u("Beschreibt Zweck und Einsatzbereich der Webshop-Konfiguration. Dieser Text sollte nicht als Testplatzhalter stehen bleiben.")),
        ("Ansprechpartner E-Mail", u("Legt die Kontaktadresse fest, die Kunden f\\u00fcr R\\u00fcckfragen oder Shop-Kommunikation verwenden.")),
        ("Telefonnummer", u("Erg\\u00e4nzt die sichtbaren Kontaktdaten f\\u00fcr Kundenkommunikation im Webshop.")),
        ("Impressum", u("Enth\\u00e4lt die rechtlichen Anbieterinformationen f\\u00fcr den Webshop. Der Inhalt muss vor Livebetrieb fachlich und rechtlich gepr\\u00fcft werden.")),
        ("Datenschutzrichtlinie", u("Enth\\u00e4lt die Datenschutzhinweise f\\u00fcr Portal- und Webshopnutzer. Der Inhalt darf kein Platzhalter bleiben.")),
        ("Url", u("Speichert die Webadresse, unter der der Shop oder Portalbereich erreichbar ist.")),
        ("Kalender", u("Ordnet Shop-Vorg\\u00e4nge optional einem Kalender f\\u00fcr Wiedervorlagen oder Termine zu.")),
        ("Bestell-E-Mail-Text", u("Definiert den Text f\\u00fcr Bestellbest\\u00e4tigungen oder Bestellkommunikation.")),
        ("Kontoerstellung aktivieren", u("Steuert, ob neue Portal- oder Webshopkonten angelegt werden k\\u00f6nnen.")),
        ("Standardstadt", u("Gibt einen Vorschlagswert f\\u00fcr neu angelegte Webshop-Kunden oder Adressen vor.")),
        ("Land", u("Ordnet neue Webshop-Adressen einem Land zu. Feldwerte k\\u00f6nnen technisch vorgegeben sein.")),
        ("Partnergruppe", u("Legt fest, welcher Partnergruppe neu angelegte Webshop-Kunden zugeordnet werden.")),
        ("Sprache", u("Bestimmt die Standardsprache f\\u00fcr Webshop- und Portalvorg\\u00e4nge.")),
        ("Standort", u("Ordnet Bestellungen und Kundenanlage einem Standort zu.")),
        ("Standard-Rechnungsadressen-Matchcode", u("Legt den Matchcode f\\u00fcr automatisch erzeugte Rechnungsadressen fest.")),
        ("Zahlungskonditionen", u("Bestimmt die Standard-Zahlungsbedingungen f\\u00fcr Webshop-Auftr\\u00e4ge.")),
        ("Lieferkondition", u("Bestimmt die Standard-Lieferbedingungen f\\u00fcr Webshop-Auftr\\u00e4ge.")),
        ("Finanzgruppe", u("Ordnet Webshop-Kunden oder Belege einer Finanzgruppe f\\u00fcr Buchhaltung und Auswertung zu.")),
        ("Nutzerbewertungen aktivieren", u("Schaltet Produktbewertungen im Webshop frei, sofern diese Funktion genutzt wird.")),
        ("Gastmodus aktivieren", u("Erlaubt Bestellungen oder Nutzung ohne vollst\\u00e4ndiges Kundenkonto, wenn der Prozess dies vorsieht.")),
        (u("K\\u00fcrzlich-Bestellt Benachrichtigungen"), u("Steuert Benachrichtigungen zu zuletzt bestellten Artikeln oder Bestellereignissen.")),
        ("Show Stock", u("Legt fest, ob Lagerbest\\u00e4nde im Webshop angezeigt werden.")),
        ("E-Mail-Text zur Kontoerstellung", u("Definiert den Nachrichtentext f\\u00fcr neu erstellte Webshop- oder Portalkonten.")),
        (u("E-Mail zur Bestellungsbenachrichtigung"), u("W\\u00e4hlt die E-Mail-Vorlage f\\u00fcr Bestellbenachrichtigungen aus.")),
        ("Reportvorlage", u("Legt die Vorlage fest, die f\\u00fcr Shop-bezogene Dokumente oder Ausgaben verwendet wird.")),
        ("Bild", u("Hinterlegt ein Bild oder Logo f\\u00fcr die Darstellung im Shop oder Portalbereich.")),
    ]:
        field(name, desc)

    group("Reportvorlagen")
    view(
        "ReportTemplateEdit",
        "Screenshots/ReportTemplateLive.webp",
        u("ReportTemplateEdit verwaltet Reportvorlagen, mit denen X-ERP Ausgaben, Druckdokumente und Belege formatiert. Anwender pflegen Reportname, XML-Layout, Reporttyp und Dok-Typ, damit Druck, Vorschau und Ausgabesteuerung die richtige Vorlage verwenden."),
        "report-template-edit",
        u("Reportvorlage bearbeiten | X-ERP ERP Hilfe"),
        u("ReportTemplateEdit in X-ERP: Reportvorlagen, XML-Layout, Reporttyp und Dok-Typ f\\u00fcr ERP-Druck, Vorschau und Ausgabesteuerung pflegen."),
        "Reportvorlage bearbeiten",
        "Reportvorlage",
        "Ansichten > Reportvorlagen > Reportvorlage",
        9992,
        u("ReportTemplateEdit Ansicht im ERP-System X-ERP f\\u00fcr Reportvorlagen"),
        "ReportTemplateEdit verwaltet Reportvorlagen in X-ERP.",
    )
    for name, desc in [
        ("Report Name", u("Bezeichnet die Reportvorlage eindeutig f\\u00fcr Auswahl, Druckvorschau und Ausgabesteuerung.")),
        ("Report Template XML", u("Enth\\u00e4lt die technische Layoutdefinition der Vorlage. \\u00c4nderungen sollten nur durch berechtigte Anwender erfolgen.")),
        ("Report Type", u("Ordnet die Vorlage einer Reportart zu. Diese Zuordnung steuert, wo die Vorlage angeboten wird.")),
        ("Dok-Typ", u("Verkn\\u00fcpft die Vorlage mit einem Dokumenttyp, damit Belege die passende Ausgabe verwenden.")),
    ]:
        field(name, desc)

    group("Benutzer und Rollen")
    view(
        "UserEdit",
        "Screenshots/UserManagerLive.webp",
        u("UserEdit verwaltet ERP-Benutzer, Anmeldung, Sprache, Rolle, Passwortregeln und die Verkn\\u00fcpfung zum Mitarbeiter. Die Ansicht ist sicherheitsrelevant, weil sie festlegt, wer sich anmelden darf, welche Rolle verwendet wird und ob ein Konto aktiv, gesperrt oder zur Passwort\\u00e4nderung verpflichtet ist."),
        "user-edit",
        u("Benutzer bearbeiten | X-ERP ERP Hilfe"),
        u("UserEdit in X-ERP: ERP-Benutzer, Rolle, Sprache, Passwort, Mitarbeiterzuordnung, Aktiv-Status und Sperre sicher verwalten."),
        "Benutzer bearbeiten",
        "Benutzer",
        "Ansichten > Benutzer und Rollen > Benutzer",
        9993,
        u("UserEdit Ansicht im ERP-System X-ERP f\\u00fcr Benutzerverwaltung"),
        "UserEdit verwaltet Anmeldung, Rolle und Sprache eines Benutzers.",
    )
    for name, desc in [
        ("Aktiv", u("Steuert, ob der Benutzer sich anmelden und im ERP arbeiten kann.")),
        (u("Verf\\u00fcgbare aktive Benutzer"), u("Zeigt die genutzten und verf\\u00fcgbaren aktiven Benutzerlizenzen an.")),
        ("Benutzername (E-Mail)", u("Dient als Anmeldename des Benutzers und sollte eindeutig sein.")),
        (u("Passwort \\u00e4ndern"), u("Aktiviert die Eingabe eines neuen Passworts f\\u00fcr den Benutzer.")),
        ("Passwort (Pflichtfeld)", u("Nimmt das neue Passwort auf, wenn eine Passwort\\u00e4nderung durchgef\\u00fchrt wird.")),
        (u("Passwort best\\u00e4tigen (Pflichtfeld)"), u("Wiederholt das Passwort zur Kontrolle, damit Eingabefehler vermieden werden.")),
        ("Vorname", u("Speichert den Vornamen des Benutzers f\\u00fcr Anzeige und Zuordnung.")),
        ("Nachname", u("Speichert den Nachnamen des Benutzers f\\u00fcr Anzeige und Zuordnung.")),
        ("Verbundener Mitarbeiter", u("Verkn\\u00fcpft den Benutzer mit einem Mitarbeiterstammsatz und verbindet Anmeldung mit internen Prozessen.")),
        (u("Benutzer muss das Passwort bei der n\\u00e4chsten Anmeldung \\u00e4ndern"), u("Erzwingt beim n\\u00e4chsten Login eine Passwort\\u00e4nderung.")),
        ("Gesperrt", u("Sperrt den Zugriff des Benutzers, ohne den Datensatz zu l\\u00f6schen.")),
        ("Rolle", u("Ordnet dem Benutzer eine Berechtigungsrolle zu.")),
        ("Sprache und Formate", u("Legt Sprache, Zahlen-, Datums- und Formatdarstellung f\\u00fcr den Benutzer fest.")),
    ]:
        field(name, desc)

    view(
        "RoleEdit",
        "Screenshots/RoleSalesLive.webp",
        u("RoleEdit verwaltet Berechtigungen und Systemmen\\u00fcs einer Benutzerrolle. Administratoren legen je Seite und WebApi fest, ob Lesen, Erstellen, \\u00c4ndern oder L\\u00f6schen erlaubt ist. Dadurch wird X-ERP auf Aufgabenbereiche zugeschnitten und sensible ERP-Funktionen werden kontrolliert freigegeben."),
        "role-edit",
        u("Rolle bearbeiten | X-ERP ERP Hilfe"),
        u("RoleEdit in X-ERP: Rollenname, Seitenrechte, WebApi-Rechte, Systemmen\\u00fcs und Berechtigungen f\\u00fcr ERP-Benutzer verwalten."),
        "Rolle bearbeiten",
        "Rolle",
        "Ansichten > Benutzer und Rollen > Rolle",
        9994,
        u("RoleEdit Ansicht im ERP-System X-ERP f\\u00fcr Berechtigungen und Systemmen\\u00fcs"),
        "RoleEdit verwaltet Rollenrechte in X-ERP.",
    )
    tab("RoleEdit-Berechtigungen", u("Das Register Berechtigungen zeigt Seiten- und WebApi-Rechte der Rolle. Hier wird festgelegt, welche Funktionen gelesen, erstellt, ge\\u00e4ndert oder gel\\u00f6scht werden d\\u00fcrfen."))
    for name, desc in [
        ("Rollenname", u("Bezeichnet die Rolle, die Benutzern zugewiesen wird.")),
        ("Seitenname", u("Nennt die X-ERP-Seite, f\\u00fcr die Berechtigungen gesetzt werden.")),
        ("WebApi Name", u("Nennt die technische API-Funktion, f\\u00fcr die Berechtigungen gesetzt werden.")),
        ("Lesen", u("Erlaubt das Anzeigen von Daten oder Funktionen.")),
        ("Erstellen", u("Erlaubt das Anlegen neuer Datens\\u00e4tze.")),
        (u("\\u00c4ndern"), u("Erlaubt das Bearbeiten vorhandener Datens\\u00e4tze.")),
        (u("L\\u00f6schen"), u("Erlaubt das Entfernen von Datens\\u00e4tzen oder Rechten.")),
    ]:
        field(name, desc)
    button(u("Alle ausw\\u00e4hlen"), u("Setzt die sichtbaren Berechtigungen gesammelt. Diese Aktion sollte bewusst verwendet werden."))
    button("Alle entfernen", u("Entfernt die sichtbaren Berechtigungen gesammelt. Diese Aktion ist f\\u00fcr Rollenbereinigung und Tests hilfreich."))
    tab(u("RoleEdit-Systemmen\\u00fcs"), u("Das Register Systemmen\\u00fcs ordnet einer Rolle die Men\\u00fc- und Navigationsbereiche zu, die im Arbeitsplatz sichtbar sein sollen."))
    field(u("Systemmen\\u00fcs"), u("Legt fest, welche Men\\u00fceintr\\u00e4ge der Rolle im X-ERP-Arbeitsplatz zur Verf\\u00fcgung stehen."))

    return rows


def main() -> None:
    ARCHIVE.mkdir(exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-missing-view-script-{dt.datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    src_wb = load_workbook(WORKBOOK)
    src_ws = src_wb["de-DE"]
    wb = load_workbook(WORKBOOK)
    ws = wb["de-DE"]
    headers = [cell.value for cell in ws[1]]
    col = {name: i + 1 for i, name in enumerate(headers) if name}

    repairs = {
        861: u("Dashboards fassen Kennzahlen, Auswertungen und Diagramme f\\u00fcr Steuerung und Kontrolle zusammen. In X-ERP werden Dashboarddefinitionen als Konfiguration gepflegt und im Arbeitsplatz ausgewertet. Die Hilfe sollte erkl\\u00e4ren, welche Datenquelle hinter einem Dashboard steht, wie es benannt wird und wann ein Designer statt einer normalen Eingabemaske verwendet wird."),
        1098: u("Dashboards fassen Kennzahlen, Auswertungen und Diagramme f\\u00fcr Steuerung und Kontrolle zusammen. In X-ERP werden Dashboarddefinitionen als Konfiguration gepflegt und im Arbeitsplatz ausgewertet. Die Hilfe sollte erkl\\u00e4ren, welche Datenquelle hinter einem Dashboard steht, wie es benannt wird und wann ein Designer statt einer normalen Eingabemaske verwendet wird."),
        1099: u("Dashboards fassen Kennzahlen, Auswertungen und Diagramme f\\u00fcr Steuerung und Kontrolle zusammen. In X-ERP werden Dashboarddefinitionen als Konfiguration gepflegt und im Arbeitsplatz ausgewertet. Die Hilfe sollte erkl\\u00e4ren, welche Datenquelle hinter einem Dashboard steht, wie es benannt wird und wann ein Designer statt einer normalen Eingabemaske verwendet wird."),
        1111: u("WorkspaceRoleDesktopEdit konfiguriert Desktop- und Startbereichselemente f\\u00fcr Benutzerrollen. Anwender erhalten dadurch einen Arbeitsplatz, der zu ihren Aufgaben passt und unn\\u00f6tige Funktionen ausblendet. Die Ansicht verbindet Berechtigung, Navigation und Bedienkomfort im ERP-Alltag."),
        1132: u("WorkspaceRoleMenuEdit steuert, welche Men\\u00fcpunkte eine Rolle im X-ERP-Arbeitsplatz sieht. Damit lassen sich Arbeitsbereiche nach Aufgabe, Abteilung oder Berechtigung reduzieren und klar strukturieren. Die Ansicht ist besonders wichtig f\\u00fcr produktive ERP-Einf\\u00fchrungen, weil sie Navigation und Zugriff der Anwender pr\\u00e4gt."),
    }
    for row, text in repairs.items():
        ws.cell(row, col["Beschreibung"]).value = text

    rows = build_rows()
    old_max_row = ws.max_row
    ws.insert_rows(INSERT_AT, amount=len(rows))

    style_source = {"group": 1098, "view": 1111, "tab": 1161, "field": 1162}

    for old_row in range(INSERT_AT, old_max_row + 1):
        new_row = old_row + len(rows)
        copy_row_dimension(src_ws, ws, old_row, new_row)
        for c in range(1, ws.max_column + 1):
            copy_cell_style(src_ws.cell(old_row, c), ws.cell(new_row, c))

    for offset, data in enumerate(rows):
        row = INSERT_AT + offset
        src_row = style_source[data["kind"]]
        copy_row_dimension(src_ws, ws, src_row, row)
        for c in range(1, ws.max_column + 1):
            copy_cell_style(src_ws.cell(src_row, c), ws.cell(row, c))
            ws.cell(row, c).value = None
        for key, value in data.items():
            if key != "kind" and key in col:
                ws.cell(row, col[key]).value = value

    shifted_screenshots = {
        1111: "Screenshots/WorkspaceRoleDesktopLive.webp",
        1132: "Screenshots/WorkspaceRoleMenuLive.webp",
        1689: "Screenshots/CrmClicked_Uebersicht.webp",
        2209: "Screenshots/FinancePostingDraftEdit_1.webp",
        2486: "Screenshots/IntercomClicked_Uebersicht.webp",
        2733: "Screenshots/EmployeeClicked_Uebersicht.webp",
        3201: "Screenshots/ProjectClicked_Uebersicht.webp",
        3232: "Screenshots/ResourceClicked_Uebersicht.webp",
        3367: "Screenshots/TimeTrackingClicked_Uebersicht.webp",
    }
    for old_row, rel in shifted_screenshots.items():
        row = old_row + (len(rows) if old_row >= INSERT_AT else 0)
        ws.cell(row, col["SCREENSHOT_REL_PATH"]).value = rel
        ws.cell(row, col["SCREENSHOT_WEB_PATH"]).value = f"/de/help/{rel}"
        ws.cell(row, col["IMAGE_STATUS"]).value = "vorhanden"

    wb.save(WORKBOOK)
    print(f"saved={WORKBOOK}")
    print(f"backup={backup}")
    print(f"inserted_rows={len(rows)}")


if __name__ == "__main__":
    main()
