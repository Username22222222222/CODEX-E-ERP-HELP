from __future__ import annotations

import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"
HELP_ROOT = Path(r"D:\DATEN\HOMEPAGES\x-erp.de\de\help")

TARGET_BY_STAMMDATEN_PATH = {
    "Stammdaten/artikel/index.html": ("Artikel", "ArticleEdit"),
    "Stammdaten/artikel/artikel-uebersicht/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Übersicht"),
    "Stammdaten/artikel/artikel-details/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Details"),
    "Stammdaten/artikel/artikel-verkauf/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Verkauf"),
    "Stammdaten/artikel/artikel-beschaffung/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Beschaffung"),
    "Stammdaten/artikel/artikel-text/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Text"),
    "Stammdaten/artikel/artikel-bild/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Bild"),
    "Stammdaten/artikel/artikel-lagerung/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Lagerung"),
    "Stammdaten/artikel/artikel-lagerungshistorie/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Lagerhistorie"),
    "Stammdaten/artikel/artikel-set/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Set"),
    "Stammdaten/artikel/artikel-makro/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Makro"),
    "Stammdaten/artikel/artikel-produktion/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Produktion"),
    "Stammdaten/artikel/artikel-zubehoer/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Zubehör"),
    "Stammdaten/artikel/artikel-kategorien/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Kategorien"),
    "Stammdaten/artikel/artikel-katalognummern/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Katalognummern"),
    "Stammdaten/artikel/artikel-produktionsschritt-ressourcen/index.html": (
        "Artikel",
        "ArticleEdit",
        "ArticleEdit-Produktionsschritt-Ressourcen",
    ),
    "Stammdaten/artikel/artikel-produktionsschritt-artikel/index.html": (
        "Artikel",
        "ArticleEdit",
        "ArticleEdit-Produktionsschritt-Artikel",
    ),
    "Stammdaten/artikel/artikel-produktionszeit/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Produktionszeit"),
    "Stammdaten/artikel/artikel-verwendung-bei-sets/index.html": (
        "Artikel",
        "ArticleEdit",
        "ArticleEdit-Verwendung in Sets",
    ),
    "Stammdaten/artikel/artikel-verwendung-bei-produktion/index.html": (
        "Artikel",
        "ArticleEdit",
        "ArticleEdit-Verwendung in Produktion",
    ),
    "Stammdaten/artikel/artikel-positionsliste/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Positionsliste"),
    "Stammdaten/artikel/artikel-umsatz/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Umsatz"),
    "Stammdaten/artikel/artikel-vertraege/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Verträge"),
    "Stammdaten/artikel/artikel-lokal/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Lokal"),
    "Stammdaten/artikel/artikel-aenderungsprotokoll/index.html": (
        "Artikel",
        "ArticleEdit",
        "ArticleEdit-Änderungsprotokoll",
    ),
    "Stammdaten/artikel/artikel-anhaenge/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Anhänge"),
    "Stammdaten/artikel/artikel-extra-felder/index.html": ("Artikel", "ArticleEdit", "ArticleEdit-Extra-Felder"),
    "Stammdaten/partner/index.html": ("Partner", "PartnerEdit"),
    "Stammdaten/partner/partner-adresse/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Adresse"),
    "Stammdaten/partner/partner-optionen/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Optionen"),
    "Stammdaten/partner/partner-kundendetails/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Kundendetails"),
    "Stammdaten/partner/partner-lieferantendetails/index.html": (
        "Partner",
        "PartnerEdit",
        "PartnerEdit-Lieferantendetails",
    ),
    "Stammdaten/partner/partner-ansprechpartner/index.html": (
        "Partner",
        "PartnerContactPersonEdit",
        "PartnerContactPersonEdit-Übersicht",
    ),
    "Stammdaten/partner/partner-lieferadressen/index.html": (
        "Partner",
        "PartnerDeliveryAddressEdit",
        "PartnerDeliveryAddressEdit-Neuer Matchcode",
    ),
    "Stammdaten/partner/partner-rechnungsadressen/index.html": (
        "Partner",
        "PartnerBillingAddressEdit",
        "PartnerBillingAddressEdit-Neuer Matchcode",
    ),
    "Stammdaten/partner/partner-katalognummern/index.html": (
        "Partner",
        "PartnerCatalogNumberEdit",
        "PartnerCatalogNumberEdit-Übersicht",
    ),
    "Stammdaten/partner/partner-einkaufspreislisten/index.html": (
        "Partner",
        "PartnerPurchasePriceListEdit",
        "PartnerPurchasePriceListEdit-Übersicht",
    ),
    "Stammdaten/partner/partner-umsatz/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Umsatz"),
    "Stammdaten/partner/partner-info/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Info"),
    "Stammdaten/partner/partner-bild/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Bild"),
    "Stammdaten/partner/partner-kategorie/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Kategorie"),
    "Stammdaten/partner/partner-crm/index.html": ("Partner", "PartnerEdit", "PartnerEdit-CRM"),
    "Stammdaten/partner/partner-e-mail/index.html": ("Partner", "PartnerEdit", "PartnerEdit-E-Mail"),
    "Stammdaten/partner/partner-dok-liste/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Belegliste"),
    "Stammdaten/partner/partner-positionsliste/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Positionsliste"),
    "Stammdaten/partner/partner-mahnungsliste/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Mahnungsliste"),
    "Stammdaten/partner/partner-offene-posten/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Offene Posten"),
    "Stammdaten/partner/partner-banking/index.html": ("Partner", "PartnerBankAccountEdit", "PartnerBankAccountEdit-Übersicht"),
    "Stammdaten/partner/partner-portal/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Portal"),
    "Stammdaten/partner/partner-kontenblatt/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Kontenblatt"),
    "Stammdaten/partner/partner-lokal/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Lokal"),
    "Stammdaten/partner/partner-aenderungsprotokoll/index.html": (
        "Partner",
        "PartnerEdit",
        "PartnerEdit-Änderungsprotokoll",
    ),
    "Stammdaten/partner/partner-anhaenge/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Anhänge"),
    "Stammdaten/partner/partner-extra-felder/index.html": ("Partner", "PartnerEdit", "PartnerEdit-Extra-Felder"),
    "Stammdaten/preismanagement/preise-einkauf/index.html": (
        "Partner",
        "PartnerPurchasePriceListEdit",
        "PartnerPurchasePriceListEdit-Übersicht",
    ),
    "Stammdaten/preismanagement/preise-verkauf/index.html": (
        "Artikel",
        "ArticleSalesPriceListEdit",
        "ArticleSalesPriceListEdit-Fremdwährung",
    ),
    "Stammdaten/preismanagement/dok/index.html": ("Belege", "DocEdit_Overview"),
    "Stammdaten/preismanagement/dok-positionen/index.html": ("Belege", "DocEdit", "DocEdit-Positionen"),
    "Stammdaten/preismanagement/dok-kette/index.html": ("Belege", "DocEdit", "DocEdit-Dok-Kette"),
    "Stammdaten/preismanagement/belegkette/index.html": ("Belege", "DocEdit", "DocEdit-Belegkette"),
    "Stammdaten/lagermanagement/index.html": ("Lager", "WarehouseEdit"),
    "Stammdaten/lagermanagement/lager/index.html": ("Lager", "WarehouseEdit", "WarehouseEdit-Übersicht"),
    "Stammdaten/lagermanagement/lagerplatz/index.html": ("Lager", "WarehouseBinLocationEdit", "WarehouseBinLocationEdit-Übersicht"),
    "Stammdaten/lagermanagement/lagertyp/index.html": ("Lager", "WarehouseTypeEdit", "WarehouseTypeEdit-Übersicht"),
    "Stammdaten/lagermanagement/lagermedium/index.html": ("Lager", "WarehouseMediumEdit", "WarehouseMediumEdit-Übersicht"),
    "Stammdaten/mitarbeiter/index.html": ("Mitarbeiter", "EmployeeEdit"),
    "Stammdaten/mitarbeiter/mitarbeiter-stammdaten/index.html": ("Mitarbeiter", "EmployeeEdit", "EmployeeEdit-Übersicht"),
    "Stammdaten/mitarbeiter/mitarbeiter-gruppe/index.html": ("Mitarbeiter", "EmployeeGroupEdit", "EmployeeGroupEdit-Übersicht"),
    "Stammdaten/mitarbeiter/mitarbeiter-provisionsgruppe/index.html": (
        "Mitarbeiter",
        "EmployeeCommissionGroupEdit",
        "EmployeeCommissionGroupEdit-Übersicht",
    ),
    "Stammdaten/mitarbeiter/mitarbeiter-urlaub/index.html": ("Mitarbeiter", "EmployeeVacationEntitlementEdit"),
    "Stammdaten/textbloecke/index.html": ("Textblöcke", "TextBlockEdit"),
    "Stammdaten/textbloecke/textbloecke-stammdaten/index.html": ("Textblöcke", "TextBlockEdit", "TextBlockEdit-Übersicht"),
    "Stammdaten/textbloecke/textbloecke-gruppe/index.html": ("Textblöcke", "TextBlockGroupEdit", "TextBlockGroupEdit-Übersicht"),
    "Stammdaten/textbloecke/textbloecke-variablen/index.html": ("Textblöcke", "TextVariableEdit", "TextVariableEdit-Versand"),
    "Stammdaten/projektmanagement/index.html": ("Projektmanagement", "ProjectEdit"),
    "Stammdaten/projektmanagement/projekt/index.html": ("Projektmanagement", "ProjectEdit", "ProjectEdit-Übersicht"),
    "Stammdaten/projektmanagement/aufgabe/index.html": ("Projektmanagement", "ProjectEdit"),
    "Stammdaten/projektmanagement/projektzeiten/index.html": ("Projektmanagement", "ProjectEdit"),
    "Stammdaten/projektmanagement/projektbudget/index.html": ("Projektmanagement", "ProjectEdit"),
    "Stammdaten/ressourcen/index.html": ("Ressourcen", "ResourceEdit"),
    "Stammdaten/anhaenge/index.html": ("Anhänge", "AttachmentEdit"),
    "Stammdaten/anhaenge/anhaenge-stammdaten/index.html": ("Anhänge", "AttachmentEdit"),
    "Stammdaten/anhaenge/anhaenge-dateien/index.html": ("Anhänge", "AttachmentEdit", "AttachmentEdit-Hochgeladene Dateien"),
}


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_path(path: str) -> str:
    return path.replace("\\", "/").lstrip("/")


def file_safe_stem(value: str) -> str:
    stem = Path(value.replace("\\", "/")).stem
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", stem).strip("_")
    return stem or "screenshot"


def help_file_from_web_path(path: str) -> Path:
    normalized = normalize_path(path)
    if normalized.startswith("de/help/"):
        normalized = normalized[len("de/help/") :]
    return HELP_ROOT / normalized


def find_source_row(
    target_bc: str,
    by_breadcrumb: dict[str, int],
    rows: list[int],
    ws,
    headers: dict[str, int],
) -> tuple[int | None, str]:
    exact = by_breadcrumb.get(target_bc)
    if exact and cell_text(ws.cell(exact, headers["SCREENSHOT_WEB_PATH"]).value):
        return exact, "exact"

    descendant_candidates: list[tuple[int, int]] = []
    prefix = target_bc + " > "
    for row in rows:
        bc = cell_text(ws.cell(row, headers["BREADCRUMB"]).value)
        shot = cell_text(ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value)
        if bc.startswith(prefix) and shot:
            descendant_candidates.append((bc.count(">"), row))
    if descendant_candidates:
        descendant_candidates.sort()
        return descendant_candidates[0][1], "descendant"

    parts = [part.strip() for part in target_bc.split(">") if part.strip()]
    for end in range(len(parts) - 1, 1, -1):
        ancestor_bc = " > ".join(parts[:end])
        row = by_breadcrumb.get(ancestor_bc)
        if row and cell_text(ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value):
            return row, "ancestor"

    return None, "missing"


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-all-stammdaten-screenshots-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }
    by_breadcrumb = {
        cell_text(ws.cell(row, headers["BREADCRUMB"]).value): row
        for row in range(2, ws.max_row + 1)
        if cell_text(ws.cell(row, headers["BREADCRUMB"]).value)
    }
    all_rows = list(range(2, ws.max_row + 1))

    updated = 0
    copied = 0
    skipped: list[str] = []
    missing: list[str] = []
    for row in range(2, ws.max_row + 1):
        url_path = normalize_path(cell_text(ws.cell(row, headers["URL_PATH"]).value))
        content_type = cell_text(ws.cell(row, headers["CONTENT_TYPE"]).value)
        if content_type != "HelpPage" or not url_path.startswith("Stammdaten/"):
            continue
        target_parts = TARGET_BY_STAMMDATEN_PATH.get(url_path)
        if not target_parts:
            skipped.append(url_path)
            continue

        target_bc = "Ansichten > " + " > ".join(target_parts)
        source_row, source_kind = find_source_row(target_bc, by_breadcrumb, all_rows, ws, headers)
        if source_row is None:
            missing.append(f"{url_path} -> {target_bc}")
            continue

        source_web_path = cell_text(ws.cell(source_row, headers["SCREENSHOT_WEB_PATH"]).value)
        source_file = help_file_from_web_path(source_web_path)
        if not source_file.is_file():
            missing.append(f"{url_path} -> missing file {source_web_path}")
            continue

        target_dir = HELP_ROOT / url_path.removesuffix("index.html").rstrip("/")
        target_dir.mkdir(parents=True, exist_ok=True)
        target_name = file_safe_stem(source_file.name) + "_Stammdaten.webp"
        target_file = target_dir / target_name
        if not target_file.exists() or target_file.read_bytes() != source_file.read_bytes():
            shutil.copy2(source_file, target_file)
            copied += 1

        target_rel = target_file.relative_to(HELP_ROOT).as_posix()
        web_path = "/de/help/" + target_rel
        topic = cell_text(ws.cell(row, headers["Thema"]).value)
        ws.cell(row, headers["Screenshot"]).value = target_rel
        ws.cell(row, headers["SCREENSHOT_WEB_PATH"]).value = web_path
        ws.cell(row, headers["IMAGE_ALT"]).value = f"Screenshot {topic} in X-ERP"
        ws.cell(row, headers["IMAGE_CAPTION"]).value = f"{topic} in den X-ERP Stammdaten"
        ws.cell(row, headers["IMAGE_STATUS"]).value = f"local-copy:{source_kind}"
        updated += 1

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"updated_rows={updated}")
    print(f"copied_files={copied}")
    print(f"skipped_unmapped={len(skipped)}")
    for item in skipped:
        print(f"SKIPPED {item}")
    print(f"missing_source={len(missing)}")
    for item in missing:
        print(f"MISSING {item}")


if __name__ == "__main__":
    main()
