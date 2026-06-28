from copy import copy
from pathlib import Path

from openpyxl import load_workbook


BOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
SHEET = "de-DE"

TARGETS = {
    "ArticleCategoryNameEdit",
    "ArticleSalesPriceListNameEdit",
    "CalendarNameEdit",
    "EmployeeCommissionGroupNameEdit",
    "PartnerCategoryNameEdit",
    "PartnerContactPersonEmailCategoryNameEdit",
    "PartnerDiscountGroupNameEdit",
}

MAIN_TYPES = {"View", "Wizard", "EditView"}


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)
    src_dim = ws.row_dimensions[source_row]
    dst_dim = ws.row_dimensions[target_row]
    dst_dim.height = src_dim.height
    dst_dim.hidden = False
    dst_dim.collapsed = False


def nearest_main_style_row(ws, row: int, content_col: int) -> int | None:
    for distance in range(1, 80):
        for candidate in (row - distance, row + distance):
            if 2 <= candidate <= ws.max_row:
                content_type = ws.cell(candidate, content_col).value
                if content_type in {"View", "Wizard"} and ws.row_dimensions[candidate].outlineLevel == 2:
                    return candidate
    return None


def main() -> None:
    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: idx + 1 for idx, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    changed = []
    for row in range(2, ws.max_row + 1):
        thema = ws.cell(row, thema_col).value
        if thema not in TARGETS:
            continue

        style_row = nearest_main_style_row(ws, row, content_col)
        if style_row:
            copy_row_style(ws, style_row, row)

        ws.cell(row, content_col).value = "View"
        ws.row_dimensions[row].outlineLevel = 2
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].collapsed = False

        child_rows = []
        for child in range(row + 1, ws.max_row + 1):
            child_thema = ws.cell(child, thema_col).value
            child_type = ws.cell(child, content_col).value
            if child_thema and child_type in MAIN_TYPES:
                break
            if child_thema:
                ws.row_dimensions[child].outlineLevel = 3
                ws.row_dimensions[child].hidden = False
                ws.row_dimensions[child].collapsed = False
                child_rows.append(child)

        changed.append(
            {
                "row": row,
                "thema": thema,
                "copied_style_from": style_row,
                "child_rows": child_rows,
            }
        )

    wb.save(BOOK)
    wb.close()

    print(f"changed={len(changed)}")
    for item in changed:
        print(
            f"{item['row']}: {item['thema']} -> View level 2; "
            f"children={item['child_rows']}; style_from={item['copied_style_from']}"
        )


if __name__ == "__main__":
    main()
