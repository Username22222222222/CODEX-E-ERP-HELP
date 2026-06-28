from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
WORKBOOK = PROJECT_ROOT / "X-ERP-HELP.xlsx"
SHEET = "de-DE"

SOURCE_PATH = "Stammdaten/artikel/artikel-uebersicht/name1/index.html"
TARGET_BREADCRUMB = "Ansichten > Artikel > ArticleEdit > ArticleEdit-Übersicht > Name1"
TARGET_PATH = "ansichten/artikel/article-edit/article-edit-uebersicht/name1/index.html"


def cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def headers_for(ws) -> dict[str, int]:
    return {
        cell_text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if cell_text(ws.cell(1, col).value)
    }


def find_row_by_value(ws, headers: dict[str, int], column: str, value: str) -> int:
    col = headers[column]
    for row in range(2, ws.max_row + 1):
        if cell_text(ws.cell(row, col).value) == value:
            return row
    raise RuntimeError(f"Could not find {column}={value!r}")


def set_if_present(ws, headers: dict[str, int], row: int, column: str, value: object) -> None:
    col = headers.get(column)
    if col:
        ws.cell(row, col).value = value


def main() -> None:
    backup_dir = WORKBOOK.parent / "ARCHIV"
    backup_dir.mkdir(exist_ok=True)
    backup = backup_dir / f"X-ERP-HELP-before-move-article-name1-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, data_only=False, read_only=False)
    ws = wb[SHEET]
    headers = headers_for(ws)

    source_row = find_row_by_value(ws, headers, "URL_PATH", SOURCE_PATH)
    target_row = find_row_by_value(ws, headers, "BREADCRUMB", TARGET_BREADCRUMB)

    set_if_present(ws, headers, source_row, "URL_PATH", None)
    set_if_present(ws, headers, source_row, "CONTENT_TYPE", "InlineSection")
    set_if_present(
        ws,
        headers,
        source_row,
        "Beschreibung",
        "Verschoben nach Ansichten > Artikel > ArticleEdit > Übersicht > Name1.",
    )
    set_if_present(ws, headers, source_row, "H1", None)
    set_if_present(ws, headers, source_row, "TITLE", None)
    set_if_present(ws, headers, source_row, "META_DESCRIPTION", None)

    set_if_present(ws, headers, target_row, "URL_PATH", TARGET_PATH)
    set_if_present(ws, headers, target_row, "CONTENT_TYPE", "HelpPage")
    set_if_present(ws, headers, target_row, "H1", "Name1")
    set_if_present(ws, headers, target_row, "NAV_TITLE", "Name1")
    set_if_present(ws, headers, target_row, "TITLE", "Name1 in ArticleEdit | X-ERP ERP Hilfe")
    set_if_present(
        ws,
        headers,
        target_row,
        "META_DESCRIPTION",
        "Name1 in der Artikelansicht ArticleEdit: erste Artikelbezeichnung fuer Suche, Listen, Belege und Wiedererkennung.",
    )
    set_if_present(
        ws,
        headers,
        target_row,
        "Beschreibung",
        """# Name1

Name1 ist die erste und führende Bezeichnung des Artikels. Dieses Feld erscheint in der Artikelansicht `ArticleEdit` im Bereich `Übersicht` und wird verwendet, damit Anwender den Artikel in Listen, Suchfeldern, Belegen und Auswertungen schnell wiedererkennen.

## Wann Name1 wichtig ist

- Name1 sollte den Artikel kurz, eindeutig und verständlich beschreiben.
- Die Bezeichnung sollte für Verkauf, Einkauf, Lager und Service gleichermaßen lesbar sein.
- Ergänzende Angaben gehören je nach Aufbau in Name2, Texte, Kategorien oder weitere Detailfelder.

:::box Empfehlung
Verwenden Sie Name1 für die wichtigste Artikelbezeichnung. Interne Sortiermerkmale, Lieferantenangaben oder wechselnde Preise sollten nicht in Name1 gepflegt werden, sondern in den passenden Stammdatenfeldern.
:::
""",
    )

    wb.save(WORKBOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"source_row={source_row}")
    print(f"target_row={target_row}")
    print(f"target_path={TARGET_PATH}")


if __name__ == "__main__":
    main()
