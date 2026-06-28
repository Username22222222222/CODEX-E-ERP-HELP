from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def set_level(ws, row: int, level: int, templates: dict[int, int]) -> None:
    ws.row_dimensions[row].outlineLevel = level
    ws.row_dimensions[row].hidden = False
    ws.row_dimensions[row].collapsed = False
    if level in templates:
        copy_row_style(ws, templates[level], row)


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-repair-outline-indent-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}

    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]
    field_col = col["Feld"]

    ansichten_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")
    faq_row = next(row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")
    glossar_row = next(row for row in range(faq_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Glossar")

    # Current workbook already contains the desired visual styles, but some rows
    # have the wrong outline level. Use stable rows as visual templates.
    templates = {
        0: faq_row,          # dark main section
        1: faq_row + 1,      # domain / subsection
        2: 4313,             # view / wizard style
        3: 4314,             # register tab style
        4: 4310,             # field / button style
    }

    # Root of views.
    set_level(ws, ansichten_row, 0, templates)

    # Ansichten hierarchy: Ansichten > Domain > View/Wizard > RegisterTab > Field/Button.
    row = ansichten_row + 1
    while row < faq_row:
        content_type = ws.cell(row, content_col).value
        thema = ws.cell(row, thema_col).value
        field = ws.cell(row, field_col).value

        if content_type == "HelpPage":
            set_level(ws, row, 1, templates)
        elif content_type in {"View", "Wizard"}:
            set_level(ws, row, 2, templates)
        elif thema or field is not None:
            if field is None:
                set_level(ws, row, 3, templates)
            else:
                set_level(ws, row, 4, templates)
        row += 1

    # FAQ and Glossar are own top-level sections with children.
    set_level(ws, faq_row, 0, templates)
    for row in range(faq_row + 1, glossar_row):
        if ws.cell(row, thema_col).value:
            set_level(ws, row, 1, templates)

    set_level(ws, glossar_row, 0, templates)
    for row in range(glossar_row + 1, ws.max_row + 1):
        if ws.cell(row, thema_col).value:
            set_level(ws, row, 1, templates)

    # Keep all rows visible; the outline symbols remain available for manual
    # expanding/collapsing in Excel.
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].collapsed = False

    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.freeze_panes = "A2"
    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False
    if ws.sheet_view:
        ws.sheet_view.showOutlineSymbols = True

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print("repaired=True")


if __name__ == "__main__":
    main()
