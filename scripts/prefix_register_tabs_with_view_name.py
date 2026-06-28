from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-prefix-register-tabs-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]
    field_col = col["Feld"]
    nav_col = col.get("NAV_TITLE")
    breadcrumb_col = col.get("BREADCRUMB")

    ansichten_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")
    faq_row = next(row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")

    current_view = None
    changed = []
    for row in range(ansichten_row + 1, faq_row):
        content_type = ws.cell(row, content_col).value
        if content_type in {"View", "Wizard"}:
            current_view = str(ws.cell(row, thema_col).value)
            continue
        if content_type == "HelpPage":
            current_view = None
            continue

        if (
            current_view
            and ws.row_dimensions[row].outlineLevel == 3
            and ws.cell(row, field_col).value is None
            and ws.cell(row, thema_col).value
        ):
            old = str(ws.cell(row, thema_col).value)
            prefix = f"{current_view}-"
            if old.startswith(prefix):
                continue
            new = f"{prefix}{old}"
            ws.cell(row, thema_col).value = new
            if nav_col:
                ws.cell(row, nav_col).value = new
            if breadcrumb_col:
                ws.cell(row, breadcrumb_col).value = new
            changed.append((row, old, new))

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"changed={len(changed)}")
    for row, old, new in changed[:80]:
        print(f"{row}: {old} -> {new}")
    if len(changed) > 80:
        print(f"... {len(changed) - 80} more")


if __name__ == "__main__":
    main()
