from __future__ import annotations

import shutil
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\micha\Documents\X-ERP-HELP")
BOOK = ROOT / "X-ERP-HELP.xlsx"
ARCHIVE = ROOT / "ARCHIV"
SHEET = "de-DE"


MAIN_VIEW_BY_DOMAIN = {
    "Artikel": "ArticleEdit",
    "Belege": "DocEdit",
    "Benutzer und Sicherheit": "UserEdit",
    "Beschaffungsassistent": "ProcurementWizard",
    "Bulletin": "BulletinEdit",
    "CRM": "CrmEdit",
    "Dashboard": "DashboardEdit",
    "Desktopkonfigurator": "WorkspaceRoleDesktopEdit",
    "E-Mail": "EmailEdit",
    "Extra-Tabellen und Extra-Felder": "ExtraTableEdit",
    "Finanzwesen": "FinancePostingEdit",
    "Helpdesk": "HelpdeskEdit",
    "Kalender": "CalendarEdit",
    "Lager": "WarehouseEdit",
    "Marketplace": "MarketplaceSettingEdit",
    "Mitarbeiter": "EmployeeEdit",
    "Partner": "PartnerEdit",
    "Produktion": "ProductionEdit",
    "Reportdesigner": "Reportdesigner",
    "Systemansichten": "SystemViewEdit",
    "Webshop": "WebshopSettingEdit",
    "Wiki": "WikiEdit",
    "Zeiterfassung": "TimeTrackingEdit",
}


@dataclass
class RowData:
    values: list
    styles: list
    height: float | None
    outline_level: int
    hidden: bool
    collapsed: bool


@dataclass
class ViewBlock:
    name: str
    source_row: int
    rows: list[RowData]


@dataclass
class DomainBlock:
    name: str
    source_row: int
    domain_row: RowData
    views: list[ViewBlock]


def capture_row(ws, row: int) -> RowData:
    styles = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row, col)
        styles.append(
            {
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    dim = ws.row_dimensions[row]
    return RowData(
        values=[ws.cell(row, col).value for col in range(1, ws.max_column + 1)],
        styles=styles,
        height=dim.height,
        outline_level=dim.outlineLevel,
        hidden=False,
        collapsed=False,
    )


def write_row(ws, row: int, data: RowData) -> None:
    for col, value in enumerate(data.values, start=1):
        cell = ws.cell(row, col)
        cell.value = value
        style = data.styles[col - 1]
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])
    dim = ws.row_dimensions[row]
    dim.height = data.height
    dim.outlineLevel = data.outline_level
    dim.hidden = False
    dim.collapsed = False


def parse_domain(ws, start: int, end: int, thema_col: int, content_col: int) -> DomainBlock:
    name = str(ws.cell(start, thema_col).value)
    domain_row = capture_row(ws, start)
    views: list[ViewBlock] = []
    row = start + 1
    while row < end:
        content_type = ws.cell(row, content_col).value
        if content_type in {"View", "Wizard"}:
            view_start = row
            view_name = str(ws.cell(row, thema_col).value)
            row += 1
            while row < end and ws.cell(row, content_col).value not in {"View", "Wizard"}:
                row += 1
            views.append(
                ViewBlock(
                    name=view_name,
                    source_row=view_start,
                    rows=[capture_row(ws, source_row) for source_row in range(view_start, row)],
                )
            )
        else:
            raise RuntimeError(f"Unexpected non-view row inside domain {name}: {row}")
    return DomainBlock(name=name, source_row=start, domain_row=domain_row, views=views)


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-main-views-first-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(BOOK, backup)

    wb = load_workbook(BOOK)
    ws = wb[SHEET]
    headers = [cell.value for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers) if header}
    thema_col = col["Thema"]
    content_col = col["CONTENT_TYPE"]

    ansichten_row = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, thema_col).value == "Ansichten")
    faq_row = next(row for row in range(ansichten_row + 1, ws.max_row + 1) if ws.cell(row, thema_col).value == "FAQ")

    domain_starts = [
        row
        for row in range(ansichten_row + 1, faq_row)
        if ws.cell(row, content_col).value == "HelpPage" and ws.row_dimensions[row].outlineLevel == 1
    ]
    domain_starts.append(faq_row)

    domains: list[DomainBlock] = []
    for idx in range(len(domain_starts) - 1):
        domains.append(parse_domain(ws, domain_starts[idx], domain_starts[idx + 1], thema_col, content_col))

    changed = []
    rebuilt: list[RowData] = []
    for domain in domains:
        main_view = MAIN_VIEW_BY_DOMAIN.get(domain.name)
        view_names = [view.name for view in domain.views]
        ordered_views = list(domain.views)
        if main_view and main_view in view_names:
            ordered_views.sort(key=lambda view: (0 if view.name == main_view else 1, view.source_row))
            if ordered_views[0].name != domain.views[0].name:
                changed.append((domain.name, main_view, domain.views[0].name))

        rebuilt.append(domain.domain_row)
        for view in ordered_views:
            rebuilt.extend(view.rows)

    delete_count = faq_row - (ansichten_row + 1)
    ws.delete_rows(ansichten_row + 1, delete_count)
    ws.insert_rows(ansichten_row + 1, len(rebuilt))
    for offset, row_data in enumerate(rebuilt):
        write_row(ws, ansichten_row + 1 + offset, row_data)

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].hidden = False
        ws.row_dimensions[row].collapsed = False

    wb.save(BOOK)
    wb.close()

    print(f"backup={backup}")
    print(f"changed_domains={len(changed)}")
    for domain, main_view, previous_first in changed:
        print(f"{domain}: {main_view} moved before {previous_first}")


if __name__ == "__main__":
    main()
