from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
ARCHIVE = Path(r"C:\Users\micha\Documents\X-ERP-HELP\ARCHIV")
SHEET = "de-DE"


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def main() -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    backup = ARCHIVE / f"X-ERP-HELP-before-outline-repair-{datetime.now():%Y%m%d-%H%M%S}.xlsx"
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK, read_only=False, data_only=False)
    ws = wb[SHEET]
    headers = {
        text(ws.cell(1, col).value): col
        for col in range(1, ws.max_column + 1)
        if text(ws.cell(1, col).value)
    }
    toc_col = headers["TOC_LEVEL"]
    topic_col = headers["Thema"]

    levels: dict[int, int] = {}
    for row in range(2, ws.max_row + 1):
        if not text(ws.cell(row, topic_col).value):
            continue
        raw = text(ws.cell(row, toc_col).value)
        try:
            level = int(float(raw))
        except ValueError:
            level = int(ws.row_dimensions[row].outlineLevel or 0)
        levels[row] = max(0, min(7, level))

    sorted_rows = sorted(levels)
    for index, row in enumerate(sorted_rows):
        level = levels[row]
        next_level = levels.get(sorted_rows[index + 1], -1) if index + 1 < len(sorted_rows) else -1
        dim = ws.row_dimensions[row]
        dim.outlineLevel = level
        dim.hidden = level > 0
        dim.collapsed = next_level > level

    wb.save(WORKBOOK)
    wb.close()
    print(f"backup={backup}")
    print(f"outline_rows={len(levels)}")


if __name__ == "__main__":
    main()
