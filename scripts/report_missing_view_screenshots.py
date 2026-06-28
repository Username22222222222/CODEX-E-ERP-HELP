from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\micha\Documents\X-ERP-HELP\X-ERP-HELP.xlsx")
REPORT = Path(r"C:\Users\micha\Documents\X-ERP-HELP\reports\missing-view-screenshots.csv")


def main() -> None:
    REPORT.parent.mkdir(exist_ok=True)
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["de-DE"]
    start = next(row for row in range(1, ws.max_row + 1) if ws.cell(row, 1).value == "Ansichten")
    ansichten_level = ws.row_dimensions[start].outlineLevel
    end = ws.max_row
    for row in range(start + 1, ws.max_row + 1):
        if ws.row_dimensions[row].outlineLevel <= ansichten_level:
            end = row - 1
            break

    rows = []
    for row in range(start + 1, end + 1):
        level = ws.row_dimensions[row].outlineLevel
        topic = ws.cell(row, 1).value
        if not isinstance(topic, str) or not topic.strip():
            continue
        if level < ansichten_level + 2 or level > ansichten_level + 5:
            continue
        if level >= ansichten_level + 4 and not (topic.endswith("List") or " anlegen/bearbeiten" in topic):
            continue
        if ws.cell(row, 5).value:
            continue
        rows.append(
            {
                "row": row,
                "outline_level": level,
                "thema": topic,
                "original_text": ws.cell(row, 2).value or "",
                "page_id": ws.cell(row, 10).value or "",
                "slug": ws.cell(row, 11).value or "",
            }
        )

    with REPORT.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=["row", "outline_level", "thema", "original_text", "page_id", "slug"], delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    print(f"missing={len(rows)}")
    print(REPORT)


if __name__ == "__main__":
    main()
